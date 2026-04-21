from pathlib import Path
from openpyxl import load_workbook
import duckdb
from duckdb import DuckDBPyConnection
import pandas as pd
pd.set_option('display.float_format', '{:,.2f}'.format)

root = Path("/Users/pavelustenko/Downloads/import_xlsx")
duck_file = "/Users/pavelustenko/ts/data/analytics.duckdb"

def connect_duck_db(file) -> DuckDBPyConnection:
    "conn establish"
    return duckdb.connect(file)

def get_data_from_xlsx(conn):
    rows = []

    files = [f for f in root.rglob("*") if f.is_file() and f.suffix.lower() == ".xlsx"]

    for f in files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)

            ws = None
            for name in wb.sheetnames:
                if "счет" in name.lower():
                    ws = wb[name]
                    break

            if ws is None:
                print(f"FAIL: {f} -> нет листа со словом 'Счет'")
                continue

            year = f.parts[-2]
            file = f.name

            e3 = ws["E3"].value
            f3 = ws["F3"].value
            g3 = ws["G3"].value
            h3 = ws["H3"].value

            number = e3 if e3 not in (None, "") else f3
            date = h3 if isinstance(g3, str) and g3.strip().lower() == "от" else g3
            article = ws["D17"].value
            amount = str(ws["L17"].value).strip().split(' ')[0]
            first_col = str(ws["B19"].value).strip().split(' ')[0]
            count = 0

            for row in ws.iter_rows(min_row=19, min_col=2, max_col=2):
                val = row[0].value

                if val in (None, ""):
                    break

                count += 1
            
            
            

            rows.append({
                "path": f,
                "year": year,
                "file": file,
                "sheet": ws.title,
                "number": number,
                "date": date,
                "article": article,
                "amount":amount,
                "rows":count,
                "range":f"B19:Z{count}"
               
            })

        except Exception as e:
            print(f"FAIL: {f} -> {e}")

    df = pd.DataFrame(rows)
    
    return df

def write_files_table(conn:DuckDBPyConnection):
    df = get_data_from_xlsx(conn)

    df.index.name = "id"
    df = df.reset_index()
    conn.execute("DROP TABLE IF EXISTS sheets")
    conn.commit()

    conn.register("df_temp", df)
    conn.execute("""CREATE OR REPLACE TABLE sheets AS SELECT 
                 id,
                 path,
                 year,
                 file,
                 sheet,
                 number,
                 date as date_from,
                 case when article = 'Артикул' then true else false end is_article,
                 rows,
                 range,
                 'D' as article_col,
                 'F' as qty_col,
                 'E' as amount_col                
                   
                 
                 FROM df_temp""")
    conn.commit()

def write_raw_data(conn):
    conn.execute("INSTALL excel;")
    conn.execute("LOAD excel;")

    sheets = conn.execute("""
        SELECT id, path, sheet, range
        FROM sheets
        ORDER BY id
    """).fetchall()

    first = True

    for sheet_id, path, sheet, rng in sheets:
        try:
            if first:
                conn.execute("""
                    CREATE OR REPLACE TABLE raw AS
                    SELECT
                        ? AS id,
                        *
                    FROM read_xlsx(
                        ?,
                        sheet = ?,
                        range = ?,
                        header = false,
                        all_varchar = true,
                        stop_at_empty = true
                    )
                """, [sheet_id, path, sheet, rng])
                first = False
            else:
                conn.execute("""
                    INSERT INTO raw
                    SELECT
                        ? AS id,
                        *
                    FROM read_xlsx(
                        ?,
                        sheet = ?,
                        range = ?,
                        header = false,
                        all_varchar = true,
                        stop_at_empty = true
                    )
                """, [sheet_id, path, sheet, rng])

            print(f"OK: {sheet_id}")

        except Exception as e:
            print(f"FAIL: {sheet_id} -> {e}")
    
    
def make_clean(conn: DuckDBPyConnection):
    sheets = conn.sql("""
        SELECT id, article_col, qty_col, amount_col
        FROM sheets
        ORDER BY id
    """).fetchall()

    first = True

    for sheet_id, article_col, qty_col, amount_col in sheets:
        query = f"""
            SELECT
                id AS sheet_id,
                B AS upd_position,
                C AS name_in_upd,
                LEFT(TRIM("{article_col}"), 10)::text sa_name,
                "{qty_col}"::BIGINT AS qty,                
                ROUND("{amount_col}"::double / "{qty_col}"::double * 100,0)::bigint AS price
            FROM raw
            WHERE id = ?
        """

        try:
            if first:
                conn.execute(f"""
                    CREATE OR REPLACE TABLE clean AS
                    {query}
                """, [sheet_id])
                first = False
            else:
                conn.execute(f"""
                    INSERT INTO clean
                    {query}
                """, [sheet_id])

            print(f"OK: {sheet_id}")

        except Exception as e:
            print(f"FAIL: {sheet_id} -> {e}")
    
    
def merge_cards(conn:DuckDBPyConnection):
    
    source = conn.sql(
    """
    SELECT DISTINCT
        s.is_article,
        TRIM(sa_name) AS sa_name,
        name_in_upd
    FROM clean
    join sheets s on s.id = clean.sheet_id
    WHERE sa_name IS NOT NULL and s.is_article = true
    """
)

    compare = conn.sql(
        """
        SELECT
            t.sa_name::text as sa_name,
            t.name_in_upd as name_in_upd,
            s.sa_name::text AS product_sa_name,
            s.nm_id::bigint as nm_id,
            s.title::text as produc_name,
            CASE
                WHEN t.sa_name IS NOT NULL AND s.sa_name IS NOT NULL THEN 'matched'
                WHEN t.sa_name IS NOT NULL AND s.sa_name IS NULL THEN 'only_in_upd'
                WHEN t.sa_name IS NULL AND s.sa_name IS NOT NULL THEN 'only_in_product'
            END AS match_status
        FROM source t
        FULL OUTER JOIN product s
            ON TRIM(s.sa_name) LIKE TRIM(t.sa_name) || '%'
        """
    )
    
    conn.execute("""
                    CREATE OR REPLACE TABLE merge AS 
                    SELECT * from compare
    """)
    

def analyse_data(conn:DuckDBPyConnection):
    mixed_up = conn.sql(
        """ 
        SELECT 
        sa_name,
        list(DISTINCT TRIM(name_in_upd)) as upd_names,
        array_length(list(DISTINCT TRIM(name_in_upd))) as cnt
        from clean
        join sheets s on s.id = clean.sheet_id
        where s.is_article = true
        group by sa_name
        having array_length(list(DISTINCT TRIM(name_in_upd))) > 1
        """
    ).df().to_csv('/Users/pavelustenko/Downloads/import_xlsx/mixedup.csv',index=False)

    
def analyse_matches(conn:DuckDBPyConnection):
    rel = conn.sql("select * from merge") #.df().to_csv('/Users/pavelustenko/Downloads/import_xlsx/compare.csv',index=False)
    
    match_stats = conn.sql(
        """ 
        SELECT
        match_status,
        count(*) as match_count
        from rel
        group by match_status        
        """
    )
    mtm = conn.sql(
        """ 
        SELECT
        sa_name, name_in_upd,
        list(DISTINCT product_sa_name) as product_sa_name,
        list(DISTINCT nm_id) as nm_ids,
        list(DISTINCT produc_name) as product_name,
        array_length(list(DISTINCT nm_id)) as cnt
        from rel
        where match_status = 'matched'
        group by sa_name, name_in_upd
        """
    )
    
    upd_product = conn.sql("select * from mtm").df().to_csv("/Users/pavelustenko/Downloads/import_xlsx/upd_noms.csv",index=False)
    
    utp = conn.sql(
        """ 
        SELECT
        nm_id, product_sa_name,  produc_name, 
        list(DISTINCT sa_name) as sa_names,
        list(DISTINCT name_in_upd) as upd_names,
        array_length(list(DISTINCT sa_name)) as cnt,
        array_length(list(DISTINCT name_in_upd)) as name_cnt
        from rel
        where match_status = 'matched'
        group by nm_id, product_sa_name, produc_name
        """
    )
    
    utp_product = conn.sql("select * from utp where cnt > 1").df().to_csv("/Users/pavelustenko/Downloads/import_xlsx/noms_upd.csv",index=False)
    oto_product = conn.sql("select * from utp where cnt = 1")
    oto_product.df().to_csv("/Users/pavelustenko/Downloads/import_xlsx/ont_to_one.csv",index=False)
    
    unset_oto = conn.sql(
        """ 
        SELECT 
        nm_id,
        product_sa_name,
        produc_name,
        UNNEST(sa_names)::text as sa_name        
        from 
        oto_product
        where name_cnt = 1
        """
    )
    
    sales_sales = conn.sql(
        """ 
        SELECT
            nm_id,
            sa_name,
            COALESCE(count(value) FILTER (WHERE dtn_id = 2 AND field = 'retail_price'), 0) -
            COALESCE(count(value) FILTER (WHERE dtn_id = 1 AND field = 'retail_price'), 0) AS qty,

            COALESCE(sum(value) FILTER (WHERE dtn_id = 2 AND field = 'retail_price'), 0) -
            COALESCE(sum(value) FILTER (WHERE dtn_id = 1 AND field = 'retail_price'), 0) AS amount
        FROM sales
        WHERE date_from >= '2024-01-01'
        AND date_from < '2026-01-01'
        GROUP BY nm_id, sa_name
        """
    )
    
    
    calcs_oto = conn.sql(
        """ 
        select
        t.nm_id,
        t.product_sa_name,
        t.produc_name,
        t.sa_name,
        sum(c.qty)::bigint as qty_dt,
        sum(c.qty*c.price/100.0)::double as amount_dt,
        s.qty as qty_cr,
        s.amount/100.0 as amount_cr        
        from unset_oto t
        join clean as c on c.sa_name = t.sa_name
        join sales_sales as s on s.nm_id = t.nm_id
        group by t.nm_id,
        t.product_sa_name,
        t.produc_name,
        t.sa_name,
        s.qty,
        amount_cr
        """
    ) #.df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/ont_to_one_stat.xlsx",index=False)
    
    final_write_off = conn.sql(
        """ 
        SELECT
        nm_id,
        product_sa_name,
        sa_name as upd_article,
        produc_name,
        qty_dt,
        amount_dt,
        qty_cr,
        amount_cr,
        amount_dt / qty_dt as wa_costs,
        amount_cr / qty_cr as wa_price,
        qty_dt - qty_cr as qty_balance,
        (qty_dt - qty_cr) * (amount_dt / qty_dt) as write_off  
        from calcs_oto
        """
    ).df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/ont_to_one_stat.xlsx",index=False)
    
def try_to_merge(conn:DuckDBPyConnection):
    rel = conn.sql("SELECT distinct sa_name from clean")
    rel2 = conn.sql("select distinct nm_id, sa_name, title from product")
    gr = conn.sql(
        """ 
        SELECT 
            CASE 
                WHEN length(sa_name) = 10 THEN left(sa_name, 7) 
                WHEN length(sa_name) = 6 THEN '0' || sa_name
                ELSE left(sa_name, 8)  
            END AS pid,
            sa_name::TEXT AS sa_name        
        FROM rel t
        
        """
    ) # .df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/sa_names.xlsx",index=False)
    
    sales_gr = conn.sql(
    """
    WITH rel2 AS (
        SELECT DISTINCT
            nm_id::BIGINT AS nm_id,
            sa_name::TEXT AS sa_name,
            title
        FROM product
        WHERE sa_name IS NOT NULL
    )
    SELECT
        CASE 
        WHEN regexp_matches(sa_name, '^[0-9]') THEN LEFT(sa_name,7) ELSE  LEFT(sa_name,8)
        END AS pid,       
        sa_name,
        nm_id,
        title
    FROM rel2
    ORDER BY pid, sa_name
    """
    )
    sales_gr.df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/sa_names.xlsx",index=False)
    
    inventories = conn.sql(
    """ 
    SELECT         
        t.pid,
        t.sa_name,
        SUM(c.qty) AS qty,
        SUM(c.qty * c.price / 100) AS value
    FROM gr t
    JOIN clean AS c 
      ON c.sa_name = t.sa_name
    JOIN sheets AS s 
      ON s.id = c.sheet_id 
    WHERE s.is_article = true
    GROUP BY       
        t.pid,
        t.sa_name
    """
    )
    
    unit_costs = conn.sql(
        """ 
        SELECT 
        pid,
        sa_name,
        qty,
        value,
        value / qty as unit_costs
        from inventories
        """
    ).df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/inventorie.xlsx",index=False)
    
    pid_unit_costs = conn.sql(
        """ 
        SELECT 
        pid::text as pid,
        sum(qty) as qty,
        sum(value) as value,
        sum(value) / sum(qty) as unit_costs
        from inventories
        group by pid
        """
    )
    
    
    pid_unit_costs.df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/pid_unit_cost.xlsx",index=False)
    
    merge_it = conn.sql(
        """ 
        SELECT
        t.pid,
        t.sa_name,
        t.nm_id,
        t.title,
        c.pid as upd_pid,
        c.unit_costs
        from sales_gr t
        left join pid_unit_costs as c on c.pid = t.pid
        
        """
    )
    merge_it.df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/comparison.xlsx",index=False)
      
       
    sales_sales = conn.sql(
        """ 
        SELECT
        m.pid,
        x.nm_id,
        x.sa_name,
        m.title,
        COALESCE(m.unit_costs,0) as unit_costs,
        x.qty,
        x.amount / 100.0 as amount      
        
        from ( 
        
        SELECT
            nm_id,
            sa_name,
            COALESCE(count(value) FILTER (WHERE dtn_id = 2 AND field = 'retail_price'), 0) -
            COALESCE(count(value) FILTER (WHERE dtn_id = 1 AND field = 'retail_price'), 0) AS qty,

            COALESCE(sum(value) FILTER (WHERE dtn_id = 2 AND field = 'retail_price'), 0) -
            COALESCE(sum(value) FILTER (WHERE dtn_id = 1 AND field = 'retail_price'), 0) AS amount
        FROM sales
        WHERE date_from >= '2025-01-01'
        AND date_from < '2026-01-01'
        GROUP BY nm_id, sa_name
        ) x 
        left join merge_it as m on m.nm_id::bigint = x.nm_id::bigint
        """
    )
    
    write_offs = conn.sql(
        """ 
        SELECT t.*,
        t.unit_costs * t.qty as write_off,
        t.amount -  t.unit_costs * t.qty as margin,
        (t.amount -  t.unit_costs * t.qty) / t.amount * 100 as margin_prc
        
        from sales_sales as t
        where t.qty > 0
        """
    )
    write_offs.df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/write_offs_2025.xlsx",index=False)
    
    invetorie_movments = conn.sql(
    """ 
    SELECT
            t.pid,
            LIST(t.title) as group_goods,
            p.qty as dt_qty,
            p.value as dt_amount,
            p.unit_costs,
            sum(t.qty) as cr_qty,
            sum(t.write_off) as cr_amount
        FROM write_offs t
        LEFT JOIN pid_unit_costs AS p 
            ON p.pid = t.pid
        GROUP BY t.pid, p.qty,   p.value, p.unit_costs      
        """
    )
    invetorie_mvn = conn.sql(
        """ 
        SELECT *,
        dt_qty - cr_qty as mv_qty,
        dt_amount - cr_amount as mvn_amount
        from invetorie_movments
        """
    )   
    invetorie_mvn.df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/inventor_mvm.xlsx",index=False)
    
    
    
    
    # .df().to_excel("/Users/pavelustenko/Downloads/import_xlsx/inventorie.xlsx",index=False)
    
def for_report(conn:DuckDBPyConnection):
    
    
    upd_db = conn.sql(
        """ 
        SELECT        
        t.sheet_id,
        s.file,
        s.year,
        s.number,
        s.date_from,        
        t.upd_position as pos,
        t.name_in_upd,
        t.sa_name,
        t.qty,        
        t.price / 100.0 as price_rub,
        t.qty * t.price / 100.0 as value
        from clean t
        join sheets s on s.id = t.sheet_id
        """
    ).df()
    
    df = upd_db.pivot_table(
        index="year",
        values=["qty","value"],
        aggfunc='sum'
    )
    # ) .to_excel("/Users/pavelustenko/Downloads/import_xlsx/pivot_years.xlsx")
    print(upd_db['name_in_upd'].nunique(), upd_db['sa_name'].nunique() )
    
    

def main():
    conn = connect_duck_db(duck_file)
    # conn.sql("CALL start_ui();")
    # write_files_table(conn)
    # conn.sql("select * from sheets").show()
    # write_raw_data(conn)
    # make_clean(conn)
    # merge_cards(conn)
    # analyse_matches(conn)
    # analyse_data(conn)
    try_to_merge(conn)
    # for_report(conn)
    conn.close()

main()