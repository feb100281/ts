# reporting/excel/engine.py
from pathlib import Path
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from django.db import connection

from .styles.style_helpers import draw_nav_button
from .styles.drill_down import style_drilldown_sheet

from .styles.treasury import style_sheet_2_1
from .styles.toc import build_toc_sheet

from .treasure import get_treasury_report, get_wb_balance

from .pl_data import get_pl_report
from .styles.pl import style_pl_sheet

from .cogs_realization_data import get_cogs_realization_report
from .styles.cogs_realization import style_cogs_realization_sheet

from .overhead_expenses_data import get_overhead_expenses_report
from .styles.overhead_expenses import style_overhead_expenses_sheet

from .corporate_expenses_data import get_corporate_expenses_report
from .styles.corporate_expenses import style_corporate_expenses_sheet

from .other_income_expenses_data import get_other_income_expenses_report
from .styles.other_income_expenses import style_other_income_expenses_sheet

from .financial_expenses_data import get_financial_expenses_report
from .styles.financial_expenses import style_financial_expenses_sheet


TEMPLATE_PATH = Path("reporting/excel/template.xlsx")

DATA_UPDATE = {
    "raw_cf": "raw_cf",
    "raw_pl": "raw_pl",
}

SQL = {
    "raw_cf": (
        "SELECT * "
        "FROM public.cf_to_csv "
        "WHERE date_from <= %s "
    ),
    "raw_pl": (
        "SELECT * "
        "FROM public.pl_for_csv "
        "WHERE date_from <= %s "
    ),
}


def fetch_data(sql, date_to):
    with connection.cursor() as cur:
        cur.execute(sql, [date_to])
        rows = cur.fetchall()
        columns = [col[0] for col in cur.description]

    return columns, rows


def update_table(ws, tbl_name, sql, date_to):
    columns, rows = fetch_data(sql, date_to)

    if tbl_name not in ws.tables:
        raise ValueError(f"Таблица {tbl_name!r} не найдена на листе {ws.title!r}")

    table = ws.tables[tbl_name]

    start_cell, end_cell = table.ref.split(":")
    start_col_idx = ws[start_cell].column
    start_row_idx = ws[start_cell].row
    end_col_idx = ws[end_cell].column
    old_end_row_idx = ws[end_cell].row

    for i, col_name in enumerate(columns, start=start_col_idx):
        ws.cell(row=start_row_idx, column=i, value=col_name)

    data_start_row = start_row_idx + 1

    for r_idx, row in enumerate(rows, start=data_start_row):
        for c_idx, value in enumerate(row, start=start_col_idx):
            ws.cell(row=r_idx, column=c_idx, value=value)

    new_last_row = data_start_row + len(rows) - 1 if rows else data_start_row
    clear_to_row = max(old_end_row_idx, ws.max_row)

    for r in range(new_last_row + 1, clear_to_row + 1):
        for c in range(start_col_idx, end_col_idx + 1):
            ws.cell(row=r, column=c, value=None)

    last_col = start_col_idx + len(columns) - 1
    table.ref = (
        f"{get_column_letter(start_col_idx)}{start_row_idx}:"
        f"{get_column_letter(last_col)}{new_last_row}"
    )


def write_df(ws, start_cell, df):
    col_letter, row_start = coordinate_from_string(start_cell)
    col_start = column_index_from_string(col_letter)

    for j, col_name in enumerate(df.columns):
        ws.cell(row=row_start, column=col_start + j + 1, value=col_name)

    for i, idx in enumerate(df.index):
        ws.cell(row=row_start + i + 1, column=col_start, value=idx)

    for i, row in enumerate(df.itertuples(index=False)):
        for j, value in enumerate(row):
            ws.cell(
                row=row_start + i + 1,
                column=col_start + j + 1,
                value=value
            )


def build_manpack(date_to=None, output_path=None):
    wb_date = date_to
    date_to = pd.to_datetime(date_to).date() if date_to else date.today()

    wb = load_workbook(TEMPLATE_PATH)

    for sheet_name, table_name in DATA_UPDATE.items():
        ws = wb[sheet_name]
        sql = SQL[table_name]
        update_table(ws, table_name, sql, date_to)

    ba_df = get_treasury_report(date_to)
    wb_df = get_wb_balance(wb_date)


    ws = wb["2.1"]
    write_df(ws, "A15", ba_df)
    write_df(ws, "A35", wb_df)
    ba_ballance = float(ba_df["Остаток"].fillna(0).sum() or 0)
    ws["E7"] = ba_ballance

    money_in_trasfer = float(wb_df["Деньги в пути (ДВП)"].fillna(0).max() or 0)
    wb_ballance = float(wb_df["Конечный баланс"].fillna(0).max() or 0)

    ws["E8"] = money_in_trasfer
    ws["E9"] = wb_ballance
    ws["E10"] = 0
    ws["E11"] = ba_ballance + money_in_trasfer + wb_ballance

    ws["A7"] = "БАНКОВСКИЕ СЧЕТА"
    ws["A8"] = "ДЕНЬГИ В ПУТИ"
    ws["A9"] = "БАЛАНС WB"
    ws["A10"] = "БЕССРОЧНЫЕ ДЕПОЗИТЫ"
    ws["A11"] = "ИТОГО:"

    style_sheet_2_1(ws, ba_df, wb_df)

    # TOC
    ws_toc = wb["TOC"]
    build_toc_sheet(
        ws_toc,
        report_date=date_to.strftime("%d.%m.%Y"),
        version="Stand alone",
    )
    
    # PL
    pl_df = get_pl_report(date_to)
    ws_pl = wb["PL"]
    style_pl_sheet(ws_pl, pl_df, date_to=date_to)
    
    # 1.3 COGS realization
    cogs_real_payload = get_cogs_realization_report(date_to)
    ws_cogs_real = wb["1.3"]
    style_cogs_realization_sheet(
        ws_cogs_real,
        cogs_real_payload,
        date_to=date_to,
    )
    
    
    # 1.4 Накладные расходы
    overhead_payload = get_overhead_expenses_report(date_to)
    ws_overhead = wb["1.4"]
    style_overhead_expenses_sheet(
        ws_overhead,
        overhead_payload,
        date_to=date_to,
    )
    
    # 1.5 Корпоративные расходы (G&A)
    corporate_payload = get_corporate_expenses_report(date_to)
    ws_corporate = wb["1.5"]
    style_corporate_expenses_sheet(
        ws_corporate,
        corporate_payload,
        date_to=date_to,
    )
    
    # 1.6 Прочие доходы и расходы
    other_ie_payload = get_other_income_expenses_report(date_to)
    ws_other_ie = wb["1.6"]
    style_other_income_expenses_sheet(
        ws_other_ie,
        other_ie_payload,
        date_to=date_to,
    )
    
    # 1.7 Финансовые расходы
    financial_expenses_payload = get_financial_expenses_report(date_to)
    ws_financial_expenses = wb["1.7"]
    style_financial_expenses_sheet(
        ws_financial_expenses,
        financial_expenses_payload,
        date_to=date_to,
    )
    
    
    
    # Навигация на drill-down листах
    ws_13_drill = wb["1.3_drill_down"]
    draw_nav_button(
        ws_13_drill,
        cell="C1",
        text="← Вернуться в 1.3",
        target_sheet="1.3",
        target_cell="A1",
    )
    draw_nav_button(
        ws_13_drill,
        cell="D1",
        text="← Вернуться в P&L",
        target_sheet="PL",
        target_cell="A1",
    )

    ws_14_drill = wb["1.4_drill_down"]
    draw_nav_button(
        ws_14_drill,
        cell="C1",
        text="← Вернуться в 1.4",
        target_sheet="1.4",
        target_cell="A1",
    )
    draw_nav_button(
        ws_14_drill,
        cell="D1",
        text="← Вернуться в P&L",
        target_sheet="PL",
        target_cell="A1",
    )
    
    # стиль для drill-down 1.3
    ws_13_drill = wb["1.3_drill_down"]
    style_drilldown_sheet(
        ws_13_drill,
        title="1.3 ДЕТАЛИЗАЦИЯ",
        subtitle="Себестоимость реализации",
    )

    # стиль для drill-down 1.4
    ws_14_drill = wb["1.4_drill_down"]
    style_drilldown_sheet(
        ws_14_drill,
        title="1.4 ДЕТАЛИЗАЦИЯ",
        subtitle="Накладные расходы",
    )
    
    ws_15_drill = wb["1.5_drill_down"]
    draw_nav_button(
        ws_15_drill,
        cell="C1",
        text="← Вернуться в 1.5",
        target_sheet="1.5",
        target_cell="A1",
    )
    draw_nav_button(
        ws_15_drill,
        cell="D1",
        text="← Вернуться в P&L",
        target_sheet="PL",
        target_cell="A1",
    )

    style_drilldown_sheet(
        ws_15_drill,
        title="1.5 ДЕТАЛИЗАЦИЯ",
        subtitle="Корпоративные расходы (G&A)",
    )
    
    ws_16_drill = wb["1.6_drill_down"]
    draw_nav_button(
        ws_16_drill,
        cell="C1",
        text="← Вернуться в 1.6",
        target_sheet="1.6",
        target_cell="A1",
    )
    draw_nav_button(
        ws_16_drill,
        cell="D1",
        text="← Вернуться в P&L",
        target_sheet="PL",
        target_cell="A1",
    )

    style_drilldown_sheet(
        ws_16_drill,
        title="1.6 ДЕТАЛИЗАЦИЯ",
        subtitle="Прочие доходы и расходы",
    )
    
    ws_17_drill = wb["1.7_drill_down"]
    draw_nav_button(
        ws_17_drill,
        cell="C1",
        text="← Вернуться в 1.7",
        target_sheet="1.7",
        target_cell="A1",
    )
    draw_nav_button(
        ws_17_drill,
        cell="D1",
        text="← Вернуться в P&L",
        target_sheet="PL",
        target_cell="A1",
    )

    style_drilldown_sheet(
        ws_17_drill,
        title="1.7 ДЕТАЛИЗАЦИЯ",
        subtitle="Финансовые расходы",
    )
    
    
    # Открывать книгу с листа TOC
    if "TOC" in wb.sheetnames:
        ws_toc = wb["TOC"]

        # снять выделение с других листов
        for ws in wb.worksheets:
            ws.sheet_view.tabSelected = False

        # сделать TOC активным
        wb.active = wb.index(ws_toc)
        ws_toc.sheet_view.tabSelected = True
        ws_toc.sheet_view.selection[0].activeCell = "A1"
        ws_toc.sheet_view.selection[0].sqref = "A1"


    wb.security = None

    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    return wb


def main(date_to="2026-03-16", output_path="reporting/excel/manpack.xlsx"):
    output_path = build_manpack(date_to=date_to, output_path=output_path)
    print(f"Файл сохранен: {output_path}")


if __name__ == "__main__":
    main()