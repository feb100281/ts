# # reporting/excel/styles/pl.py
# from datetime import date, datetime
# from numbers import Number

# from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
# from openpyxl.utils import get_column_letter
# from openpyxl.utils.cell import quote_sheetname

# from reporting.excel.styles.theme import (
#     FILLS,
#     FONTS,
#     BORDERS,
#     ALIGNMENTS,
#     FORMATS,
#     COLORS,
# )
# from reporting.excel.styles.style_helpers import (
#     clear_range,
#     set_row_heights,
#     draw_toc_button,
#     draw_sheet_header,
#     draw_table_header,
# )


# PL_ROW_ORDER = [
#     "Выручка от основной деятельности",
#     "Себестоимость проданных товаров",
#     "Себестоимость реализации",
#     "Валовая прибыль",
#     "Рентабельность продаж",
#     "Накладные расходы",
#     "Корпоративные расходы (G&A)",
#     "EBITDA",
#     "EBITDA MARGIN",
#     "Прочие доходы и расходы",
#     "EBITDA Adjusted",
#     "Финансовые расходы",
#     "Налог на прибыль",
#     "Чистая прибыль / убыток",
# ]

# NOTE_MAP = {
#     "Выручка от основной деятельности": {"note": "1.1", "sheet": "1.1"},
#     "Себестоимость проданных товаров": {"note": "1.2", "sheet": "1.2"},
#     "Себестоимость реализации": {"note": "1.3", "sheet": "1.3"},
#     "Накладные расходы": {"note": "1.4", "sheet": "1.4"},
#     "Корпоративные расходы (G&A)": {"note": "1.5", "sheet": "1.5"},
#     "Прочие доходы и расходы": {"note": "1.6", "sheet": "1.6"},
#     "Финансовые расходы": {"note": "1.7", "sheet": "1.7"},
# }

# PERCENT_ROWS = {
#     "Рентабельность продаж",
#     "EBITDA MARGIN",
# }

# SUBTOTAL_ROWS = {
#     "Валовая прибыль",
#     "EBITDA",
#     "EBITDA Adjusted",
#     "Чистая прибыль / убыток",
# }

# BREAK_AFTER_ROWS = {
#     "Рентабельность продаж",
#     "Корпоративные расходы (G&A)",
#     "EBITDA MARGIN",
#     "EBITDA Adjusted",
# }


# # -----------------------------
# # Дополнительные стили
# # -----------------------------

# # badge note — мягкий, деловой
# NOTE_BADGE_FILL = PatternFill("solid", fgColor="E7F1ED")
# NOTE_BADGE_FONT = Font(
#     name="Roboto",
#     size=9,
#     bold=True,
#     color="3E6E5E",
# )
# NOTE_BADGE_ALIGN = Alignment(horizontal="center", vertical="center")

# # subtotal / net — мягкие финансовые акценты
# SUBTOTAL_FILL = PatternFill("solid", fgColor="EEF5F2")
# NET_FILL = PatternFill("solid", fgColor="E4EFEA")
# SEP_FILL = PatternFill("solid", fgColor=COLORS["white"])

# # отрицательные — приглушённый бордовый, без "ядовитого" red
# NEGATIVE_COLOR = "6B4C4C"
# NEGATIVE_FONT = Font(name="Roboto", size=10, color=NEGATIVE_COLOR)
# NEGATIVE_BOLD_FONT = Font(name="Roboto", size=10, bold=True, color=NEGATIVE_COLOR)

# POSITIVE_FONT = FONTS["normal"]
# POSITIVE_BOLD_FONT = FONTS["total"]

# THIN_GRAY_SIDE = Side(style="thin", color=COLORS["border_gray"])
# MEDIUM_DARK_SIDE = Side(style="medium", color=COLORS["black"])

# SUBTOTAL_BORDER = Border(
#     top=THIN_GRAY_SIDE,
#     bottom=MEDIUM_DARK_SIDE,
# )

# NET_BORDER = Border(
#     top=MEDIUM_DARK_SIDE,
#     bottom=MEDIUM_DARK_SIDE,
# )

# NO_BORDER = Border()


# def _safe_str_date(dt):
#     if dt is None:
#         return ""
#     if isinstance(dt, datetime):
#         return dt.strftime("%d.%m.%Y")
#     if isinstance(dt, date):
#         return dt.strftime("%d.%m.%Y")
#     return str(dt)


# def _unmerge_all(ws):
#     merged_ranges = list(ws.merged_cells.ranges)
#     for merged in merged_ranges:
#         ws.unmerge_cells(str(merged))


# def _clear_pl_sheet(ws):
#     _unmerge_all(ws)
#     clear_range(ws, row_start=1, row_end=400, col_start=1, col_end=25)


# def _prepare_layout(ws):
#     row_heights = {
#         1: 20,
#         2: 24,
#         3: 18,
#         4: 18,
#         6: 10,
#         7: 24,
#     }
#     set_row_heights(ws, row_heights)

#     ws.sheet_view.showGridLines = False
#     ws.freeze_panes = "C8"


# def _apply_dynamic_widths(ws, headers):
#     for col_idx, header in enumerate(headers, start=1):
#         col_letter = get_column_letter(col_idx)

#         if header == "Name":
#             width = 42
#         elif header == "Note":
#             width = 9
#         elif header == "":
#             width = 2.5
#         elif header in {"Diff rel", "MTD Diff rel"}:
#             width = 12
#         elif header in {"Diff abs", "MTD Diff abs"}:
#             width = 14
#         elif str(header).startswith("MTD "):
#             width = 15
#         else:
#             width = 15

#         ws.column_dimensions[col_letter].width = width


# def _draw_pl_header(ws, date_to=None):
#     draw_toc_button(ws)

#     subtitle = "Управленческая отчетность (management pack)"
#     currency = "Российский рубль (RUB)"
#     if date_to:
#         currency = f"{currency} • дата отчета: {_safe_str_date(date_to)}"

#     draw_sheet_header(
#         ws,
#         title="ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ (P&L)",
#         subtitle=subtitle,
#         currency=currency,
#     )


# def _get_ordered_columns(pl_df):
#     fye_cols = sorted(
#         [
#             c for c in pl_df.columns
#             if str(c).startswith("FYE ") and str(c).split()[1].isdigit()
#         ],
#         key=lambda x: int(str(x).split()[1])
#     )
#     ytd_cols = sorted(
#         [
#             c for c in pl_df.columns
#             if str(c).startswith("YTD ") and str(c).split()[1].isdigit()
#         ],
#         key=lambda x: int(str(x).split()[1]),
#         reverse=True
#     )
#     mtd_cols = sorted(
#         [
#             c for c in pl_df.columns
#             if str(c).startswith("MTD ") and str(c).split()[1].isdigit()
#         ],
#         key=lambda x: int(str(x).split()[1]),
#         reverse=True
#     )

#     ordered = []
#     if "Note" in pl_df.columns:
#         ordered.append("Note")

#     ordered.extend(fye_cols)
#     ordered.append("__SEP__")

#     ordered.extend(ytd_cols)
#     if "Diff abs" in pl_df.columns:
#         ordered.append("Diff abs")
#     if "Diff rel" in pl_df.columns:
#         ordered.append("Diff rel")

#     ordered.append("__SEP2__")

#     ordered.extend(mtd_cols)
#     if "MTD Diff abs" in pl_df.columns:
#         ordered.append("MTD Diff abs")
#     if "MTD Diff rel" in pl_df.columns:
#         ordered.append("MTD Diff rel")

#     return ordered

# def _excel_headers_from_ordered_cols(ordered_cols):
#     headers = ["Name"]
#     for c in ordered_cols:
#         if c in ("__SEP__", "__SEP2__"):
#             headers.append("")
#         else:
#             headers.append(c)
#     return headers


# def _is_negative_number(value):
#     if value is None:
#         return False

#     if isinstance(value, Number):
#         return value < 0

#     if isinstance(value, str):
#         v = value.strip().replace(" ", "")
#         if not v:
#             return False
#         return v.startswith("(") and v.endswith(")")

#     return False


# def _apply_negative_font(cell, bold=False):
#     cell.font = NEGATIVE_BOLD_FONT if bold else NEGATIVE_FONT


# def _apply_positive_font(cell, bold=False):
#     cell.font = POSITIVE_BOLD_FONT if bold else POSITIVE_FONT


# def _style_separator_cols(ws, sep_cols, row_start, row_end):
#     for col_idx in sep_cols:
#         for row in range(row_start, row_end + 1):
#             cell = ws.cell(row=row, column=col_idx)
#             cell.value = None
#             cell.fill = SEP_FILL
#             cell.border = NO_BORDER
#             cell.alignment = ALIGNMENTS["center"]


# def _draw_note_badge(ws, row, row_name):
#     cfg = NOTE_MAP.get(row_name)
#     cell = ws.cell(row=row, column=2)

#     if not cfg:
#         cell.value = None
#         cell.fill = FILLS["none"]
#         cell.border = BORDERS["thin"]
#         cell.alignment = ALIGNMENTS["center"]
#         cell.font = FONTS["normal"]
#         return

#     cell.value = cfg["note"]
#     cell.hyperlink = f"#{quote_sheetname(cfg['sheet'])}!A1"
#     cell.alignment = NOTE_BADGE_ALIGN
#     cell.fill = NOTE_BADGE_FILL
#     cell.border = BORDERS["thin"]
#     cell.font = NOTE_BADGE_FONT


# def _style_header_with_separator(ws, headers, row=7):
#     for i, header in enumerate(headers, start=1):
#         cell = ws.cell(row=row, column=i)

#         if header == "":
#             cell.value = None
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]
#             cell.alignment = ALIGNMENTS["center"]
#             continue

#         cell.value = header
#         cell.fill = FILLS["header"]
#         cell.font = FONTS["header_white"]
#         cell.alignment = ALIGNMENTS["center"]
#         cell.border = BORDERS["thin"]


# def _row_fill(row_num):
#     return FILLS["alt"] if row_num % 2 == 0 else FILLS["none"]


# def _style_regular_row_cell(cell, row_num):
#     cell.fill = _row_fill(row_num)
#     cell.border = BORDERS["thin"]
#     cell.font = FONTS["normal"]


# def _style_subtotal_row_cell(cell, is_net=False):
#     cell.fill = NET_FILL if is_net else SUBTOTAL_FILL
#     cell.border = NET_BORDER if is_net else SUBTOTAL_BORDER
#     cell.font = FONTS["total"]


# def _set_alignment_by_col(cell, col_idx, sep_cols):
#     if col_idx == 1:
#         cell.alignment = ALIGNMENTS["left"]
#     elif col_idx == 2:
#         cell.alignment = ALIGNMENTS["center"]
#     elif col_idx in sep_cols:
#         cell.alignment = ALIGNMENTS["center"]
#     else:
#         cell.alignment = ALIGNMENTS["right"]


# def _set_number_format(cell, row_name, header_name):
#     if cell.value is None:
#         return

#     if row_name in PERCENT_ROWS or header_name in {"Diff rel", "MTD Diff rel"}:
#         cell.number_format = "0.0%;(0.0%)"
#     else:
#         cell.number_format = '#,##0;(#,##0)'


# def _style_name_cell(cell, row_name, is_subtotal, row_num):
#     if is_subtotal:
#         cell.fill = NET_FILL if row_name == "Чистая прибыль / убыток" else SUBTOTAL_FILL
#         cell.border = NET_BORDER if row_name == "Чистая прибыль / убыток" else SUBTOTAL_BORDER
#         cell.font = FONTS["total"]
#     else:
#         cell.fill = _row_fill(row_num)
#         cell.border = BORDERS["thin"]
#         cell.font = FONTS["normal"]

#     cell.alignment = ALIGNMENTS["left"]


# def _style_note_cell(ws, row_num, row_name, is_subtotal):
#     note_cell = ws.cell(row=row_num, column=2)

#     if note_cell.value:
#         # Если это строка с note badge — сохраняем badge и для subtotal тоже
#         note_cell.alignment = NOTE_BADGE_ALIGN
#         note_cell.border = BORDERS["thin"]
#         note_cell.fill = NOTE_BADGE_FILL
#         note_cell.font = NOTE_BADGE_FONT
#     else:
#         if is_subtotal:
#             note_cell.fill = NET_FILL if row_name == "Чистая прибыль / убыток" else SUBTOTAL_FILL
#             note_cell.border = NET_BORDER if row_name == "Чистая прибыль / убыток" else SUBTOTAL_BORDER
#             note_cell.font = FONTS["total"]
#         else:
#             note_cell.fill = _row_fill(row_num)
#             note_cell.border = BORDERS["thin"]
#             note_cell.font = FONTS["normal"]

#         note_cell.alignment = ALIGNMENTS["center"]


# def _style_data_row(ws, row_num, row_name, header_map, last_col, sep_cols):
#     is_subtotal = row_name in SUBTOTAL_ROWS
#     is_net = row_name == "Чистая прибыль / убыток"

#     # отдельно стилизуем Name и Note, чтобы они выглядели аккуратнее
#     _style_name_cell(ws.cell(row=row_num, column=1), row_name, is_subtotal, row_num)
#     _style_note_cell(ws, row_num, row_name, is_subtotal)

#     for col in range(3, last_col + 1):
#         cell = ws.cell(row=row_num, column=col)

#         if col in sep_cols:
#             cell.value = None
#             cell.fill = SEP_FILL
#             cell.border = NO_BORDER
#             cell.alignment = ALIGNMENTS["center"]
#             continue

#         if is_subtotal:
#             _style_subtotal_row_cell(cell, is_net=is_net)
#         else:
#             _style_regular_row_cell(cell, row_num)

#         _set_alignment_by_col(cell, col, sep_cols)

#         header_name = header_map.get(col)
#         _set_number_format(cell, row_name, header_name)

#         if _is_negative_number(cell.value):
#             _apply_negative_font(cell, bold=is_subtotal)
#         else:
#             _apply_positive_font(cell, bold=is_subtotal)


# def _draw_blank_row(ws, row_num, last_col, sep_cols):
#     ws.row_dimensions[row_num].height = 8
#     for col in range(1, last_col + 1):
#         cell = ws.cell(row=row_num, column=col)
#         cell.value = None
#         if col in sep_cols:
#             cell.fill = SEP_FILL
#             cell.border = NO_BORDER
#         else:
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]
#         cell.alignment = ALIGNMENTS["left"]


# def _draw_footer_note(ws, row, last_col, sep_cols):
#     for col in range(1, last_col + 1):
#         cell = ws.cell(row=row, column=col)
#         cell.value = None
#         if col in sep_cols:
#             cell.fill = SEP_FILL
#             cell.border = NO_BORDER
#         else:
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]

#     ws.cell(row=row, column=1, value="Note — ссылки на листы с расшифровками.")
#     ws.cell(row=row, column=1).font = FONTS["subtitle"]
#     ws.cell(row=row, column=1).alignment = ALIGNMENTS["left"]


# def style_pl_sheet(ws, pl_df, date_to=None):
#     _clear_pl_sheet(ws)
#     ws.auto_filter.ref = None
#     _prepare_layout(ws)
#     _draw_pl_header(ws, date_to=date_to)

#     start_header_row = 7
#     start_data_row = 8

#     if pl_df is None or pl_df.empty:
#         headers = ["Name", "Note"]
#         draw_table_header(ws, row=start_header_row, headers=headers, start_col=1, wrap=False)
#         _apply_dynamic_widths(ws, headers)
#         ws["A8"] = "Нет данных"
#         ws["A8"].font = FONTS["normal"]
#         ws["A8"].alignment = ALIGNMENTS["left"]
#         return

#     pl_df = pl_df.copy()
#     pl_df = pl_df.reindex(PL_ROW_ORDER)

#     ordered_cols = _get_ordered_columns(pl_df)
#     headers = _excel_headers_from_ordered_cols(ordered_cols)

#     _apply_dynamic_widths(ws, headers)
#     _style_header_with_separator(ws, headers, row=start_header_row)

#     last_col = len(headers)
#     sep_cols = [idx for idx, hdr in enumerate(headers, start=1) if hdr == ""]
#     header_map = {idx: hdr for idx, hdr in enumerate(headers, start=1)}

#     current_row = start_data_row

#     for row_name in pl_df.index:
#         ws.cell(row=current_row, column=1, value=row_name)
#         ws.cell(row=current_row, column=2, value=None)
#         _draw_note_badge(ws, current_row, row_name)

#         excel_col = 3
#         for col_name in ordered_cols:
#             if col_name == "Note":
#                 continue

#             if col_name in ("__SEP__", "__SEP2__"):
#                 ws.cell(row=current_row, column=excel_col, value=None)
#                 excel_col += 1
#                 continue

#             value = pl_df.loc[row_name, col_name]
#             ws.cell(row=current_row, column=excel_col, value=value)
#             excel_col += 1

#         _style_data_row(
#             ws=ws,
#             row_num=current_row,
#             row_name=row_name,
#             header_map=header_map,
#             last_col=last_col,
#             sep_cols=sep_cols,
#         )
#         ws.row_dimensions[current_row].height = 20

#         if row_name in BREAK_AFTER_ROWS:
#             current_row += 1
#             _draw_blank_row(ws, current_row, last_col, sep_cols)

#         current_row += 1

#     end_row = current_row - 1

#     if sep_cols:
#         _style_separator_cols(ws, sep_cols, start_header_row, end_row + 2)

#     _draw_footer_note(ws, end_row + 2, last_col, sep_cols)







# reporting/excel/styles/pl.py
from datetime import date, datetime
from numbers import Number

from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname

from reporting.excel.styles.theme import (
    FILLS,
    FONTS,
    BORDERS,
    ALIGNMENTS,
    COLORS,
)
from reporting.excel.styles.style_helpers import (
    clear_range,
    set_row_heights,
    draw_toc_button,
    draw_sheet_header,
    draw_table_header,
    draw_nav_link_row, 
)


PL_ROW_ORDER = [
    "Выручка от основной деятельности",
    "Себестоимость проданных товаров",
    "Себестоимость реализации",
    "Валовая прибыль",
    "Рентабельность продаж",
    "Накладные расходы",
    "Корпоративные расходы (G&A)",
    "EBITDA",
    "EBITDA MARGIN",
    "Прочие доходы и расходы",
    "EBITDA Adjusted",
    "Финансовые расходы",
    "Налог на прибыль",
    "Чистая прибыль / убыток",
]

NOTE_MAP = {
    "Выручка от основной деятельности": {"note": "1.1", "sheet": "1.1"},
    "Себестоимость проданных товаров": {"note": "1.2", "sheet": "1.2"},
    "Себестоимость реализации": {"note": "1.3", "sheet": "1.3"},
    "Накладные расходы": {"note": "1.4", "sheet": "1.4"},
    "Корпоративные расходы (G&A)": {"note": "1.5", "sheet": "1.5"},
    "Прочие доходы и расходы": {"note": "1.6", "sheet": "1.6"},
    "Финансовые расходы": {"note": "1.7", "sheet": "1.7"},
}

PERCENT_ROWS = {
    "Рентабельность продаж",
    "EBITDA MARGIN",
}

SUBTOTAL_ROWS = {
    "Валовая прибыль",
    "EBITDA",
    "EBITDA Adjusted",
    "Чистая прибыль / убыток",
}

BREAK_AFTER_ROWS = {
    "Рентабельность продаж",
    "Корпоративные расходы (G&A)",
    "EBITDA MARGIN",
    "EBITDA Adjusted",
}


NOTE_BADGE_FILL = PatternFill("solid", fgColor="E7F1ED")
NOTE_BADGE_FONT = Font(
    name="Roboto",
    size=9,
    bold=True,
    color="3E6E5E",
)
NOTE_BADGE_ALIGN = Alignment(horizontal="center", vertical="center")

SUBTOTAL_FILL = PatternFill("solid", fgColor="EEF5F2")
NET_FILL = PatternFill("solid", fgColor="E4EFEA")
SEP_FILL = PatternFill("solid", fgColor=COLORS["white"])

NEGATIVE_COLOR = "6B4C4C"
NEGATIVE_FONT = Font(name="Roboto", size=10, color=NEGATIVE_COLOR)
NEGATIVE_BOLD_FONT = Font(name="Roboto", size=10, bold=True, color=NEGATIVE_COLOR)

POSITIVE_FONT = FONTS["normal"]
POSITIVE_BOLD_FONT = FONTS["total"]

THIN_GRAY_SIDE = Side(style="thin", color=COLORS["border_gray"])
MEDIUM_DARK_SIDE = Side(style="medium", color=COLORS["black"])

SUBTOTAL_BORDER = Border(
    top=THIN_GRAY_SIDE,
    bottom=MEDIUM_DARK_SIDE,
)

NET_BORDER = Border(
    top=MEDIUM_DARK_SIDE,
    bottom=MEDIUM_DARK_SIDE,
)

NO_BORDER = Border()


def _safe_str_date(dt):
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%d.%m.%Y")
    if isinstance(dt, date):
        return dt.strftime("%d.%m.%Y")
    return str(dt)


def _unmerge_all(ws):
    merged_ranges = list(ws.merged_cells.ranges)
    for merged in merged_ranges:
        ws.unmerge_cells(str(merged))


def _clear_pl_sheet(ws):
    _unmerge_all(ws)
    clear_range(ws, row_start=1, row_end=400, col_start=1, col_end=120)


def _prepare_layout(ws):
    row_heights = {
        1: 20,
        2: 24,
        3: 18,
        4: 18,
        6: 10,
        7: 24,
    }
    set_row_heights(ws, row_heights)

    ws.sheet_view.showGridLines = False
    ws.sheet_view.showOutlineSymbols = True
    ws.freeze_panes = "C8"

    # Плюсик слева от сгруппированных месяцев, рядом с FYE
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.sheet_properties.outlinePr.summaryBelow = False


def _apply_dynamic_widths(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        if header == "Name":
            width = 42
        elif header == "Note":
            width = 9
        elif header == "":
            width = 2.5
        elif header in {"Diff rel", "MTD Diff rel"}:
            width = 12
        elif header in {"Diff abs", "MTD Diff abs"}:
            width = 14
        elif str(header).startswith("MTD "):
            width = 15
        else:
            width = 15

        ws.column_dimensions[col_letter].width = width


def _draw_pl_header(ws, date_to=None):
    draw_toc_button(ws)

    subtitle = "Управленческая отчетность (management pack)"
    currency = "Российский рубль (RUB)"
    if date_to:
        currency = f"{currency} • дата отчета: {_safe_str_date(date_to)}"

    draw_sheet_header(
        ws,
        title="ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ (P&L)",
        subtitle=subtitle,
        currency=currency,
    )


def _get_ordered_columns(pl_df):
    month_names = {
        "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
        "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"
    }

    month_order = {
        "Янв": 1,
        "Фев": 2,
        "Мар": 3,
        "Апр": 4,
        "Май": 5,
        "Июн": 6,
        "Июл": 7,
        "Авг": 8,
        "Сен": 9,
        "Окт": 10,
        "Ноя": 11,
        "Дек": 12,
    }

    fye_cols = sorted(
        [
            c for c in pl_df.columns
            if str(c).startswith("FYE ") and str(c).split()[1].isdigit()
        ],
        key=lambda x: int(str(x).split()[1])
    )

    ytd_cols = sorted(
        [
            c for c in pl_df.columns
            if str(c).startswith("YTD ") and str(c).split()[1].isdigit()
        ],
        key=lambda x: int(str(x).split()[1]),
        reverse=True
    )

    mtd_cols = sorted(
        [
            c for c in pl_df.columns
            if str(c).startswith("MTD ") and str(c).split()[1].isdigit()
        ],
        key=lambda x: int(str(x).split()[1]),
        reverse=True
    )

    month_cols = [
        c for c in pl_df.columns
        if len(str(c).split()) == 2
        and str(c).split()[0] in month_names
        and str(c).split()[1].isdigit()
    ]

    month_cols = sorted(
        month_cols,
        key=lambda x: (
            int(str(x).split()[1]),
            month_order[str(x).split()[0]]
        )
    )

    ordered = []

    if "Note" in pl_df.columns:
        ordered.append("Note")

    # FYE + месячная расшифровка именно этого года
    for fye_col in fye_cols:
        year = int(str(fye_col).split()[1])
        ordered.append(fye_col)

        year_month_cols = [
            c for c in month_cols
            if int(str(c).split()[1]) == year
        ]
        ordered.extend(year_month_cols)

    if fye_cols or month_cols:
        ordered.append("__SEP__")

    # YTD без месячной расшифровки
    ordered.extend(ytd_cols)

    if "Diff abs" in pl_df.columns:
        ordered.append("Diff abs")
    if "Diff rel" in pl_df.columns:
        ordered.append("Diff rel")

    if mtd_cols or "MTD Diff abs" in pl_df.columns or "MTD Diff rel" in pl_df.columns:
        ordered.append("__SEP2__")

    # MTD не прячем
    ordered.extend(mtd_cols)

    if "MTD Diff abs" in pl_df.columns:
        ordered.append("MTD Diff abs")
    if "MTD Diff rel" in pl_df.columns:
        ordered.append("MTD Diff rel")

    return ordered


def _excel_headers_from_ordered_cols(ordered_cols):
    headers = ["Name"]
    for c in ordered_cols:
        if c in ("__SEP__", "__SEP2__"):
            headers.append("")
        else:
            headers.append(c)
    return headers


def _is_negative_number(value):
    if value is None:
        return False

    if isinstance(value, Number):
        return value < 0

    if isinstance(value, str):
        v = value.strip().replace(" ", "")
        if not v:
            return False
        return v.startswith("(") and v.endswith(")")

    return False


def _apply_negative_font(cell, bold=False):
    cell.font = NEGATIVE_BOLD_FONT if bold else NEGATIVE_FONT


def _apply_positive_font(cell, bold=False):
    cell.font = POSITIVE_BOLD_FONT if bold else POSITIVE_FONT


def _style_separator_cols(ws, sep_cols, row_start, row_end):
    for col_idx in sep_cols:
        for row in range(row_start, row_end + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = None
            cell.fill = SEP_FILL
            cell.border = NO_BORDER
            cell.alignment = ALIGNMENTS["center"]


def _draw_note_badge(ws, row, row_name):
    cfg = NOTE_MAP.get(row_name)
    cell = ws.cell(row=row, column=2)

    if not cfg:
        cell.value = None
        cell.fill = FILLS["none"]
        cell.border = BORDERS["thin"]
        cell.alignment = ALIGNMENTS["center"]
        cell.font = FONTS["normal"]
        return

    cell.value = cfg["note"]
    cell.hyperlink = f"#{quote_sheetname(cfg['sheet'])}!A1"
    cell.alignment = NOTE_BADGE_ALIGN
    cell.fill = NOTE_BADGE_FILL
    cell.border = BORDERS["thin"]
    cell.font = NOTE_BADGE_FONT


def _style_header_with_separator(ws, headers, row=7):
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i)

        if header == "":
            cell.value = None
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]
            cell.alignment = ALIGNMENTS["center"]
            continue

        cell.value = header
        cell.fill = FILLS["header"]
        cell.font = FONTS["header_white"]
        cell.alignment = ALIGNMENTS["center"]
        cell.border = BORDERS["thin"]


def _row_fill(row_num):
    return FILLS["alt"] if row_num % 2 == 0 else FILLS["none"]


def _style_regular_row_cell(cell, row_num):
    cell.fill = _row_fill(row_num)
    cell.border = BORDERS["thin"]
    cell.font = FONTS["normal"]


def _style_subtotal_row_cell(cell, is_net=False):
    cell.fill = NET_FILL if is_net else SUBTOTAL_FILL
    cell.border = NET_BORDER if is_net else SUBTOTAL_BORDER
    cell.font = FONTS["total"]


def _set_alignment_by_col(cell, col_idx, sep_cols):
    if col_idx == 1:
        cell.alignment = ALIGNMENTS["left"]
    elif col_idx == 2:
        cell.alignment = ALIGNMENTS["center"]
    elif col_idx in sep_cols:
        cell.alignment = ALIGNMENTS["center"]
    else:
        cell.alignment = ALIGNMENTS["right"]


def _set_number_format(cell, row_name, header_name):
    if cell.value is None:
        return

    if row_name in PERCENT_ROWS or header_name in {"Diff rel", "MTD Diff rel"}:
        cell.number_format = "0.0%;(0.0%)"
    else:
        cell.number_format = '#,##0;(#,##0)'


def _style_name_cell(cell, row_name, is_subtotal, row_num):
    if is_subtotal:
        cell.fill = NET_FILL if row_name == "Чистая прибыль / убыток" else SUBTOTAL_FILL
        cell.border = NET_BORDER if row_name == "Чистая прибыль / убыток" else SUBTOTAL_BORDER
        cell.font = FONTS["total"]
    else:
        cell.fill = _row_fill(row_num)
        cell.border = BORDERS["thin"]
        cell.font = FONTS["normal"]

    cell.alignment = ALIGNMENTS["left"]


def _style_note_cell(ws, row_num, row_name, is_subtotal):
    note_cell = ws.cell(row=row_num, column=2)

    if note_cell.value:
        note_cell.alignment = NOTE_BADGE_ALIGN
        note_cell.border = BORDERS["thin"]
        note_cell.fill = NOTE_BADGE_FILL
        note_cell.font = NOTE_BADGE_FONT
    else:
        if is_subtotal:
            note_cell.fill = NET_FILL if row_name == "Чистая прибыль / убыток" else SUBTOTAL_FILL
            note_cell.border = NET_BORDER if row_name == "Чистая прибыль / убыток" else SUBTOTAL_BORDER
            note_cell.font = FONTS["total"]
        else:
            note_cell.fill = _row_fill(row_num)
            note_cell.border = BORDERS["thin"]
            note_cell.font = FONTS["normal"]

        note_cell.alignment = ALIGNMENTS["center"]


def _style_data_row(ws, row_num, row_name, header_map, last_col, sep_cols):
    is_subtotal = row_name in SUBTOTAL_ROWS
    is_net = row_name == "Чистая прибыль / убыток"

    _style_name_cell(ws.cell(row=row_num, column=1), row_name, is_subtotal, row_num)
    _style_note_cell(ws, row_num, row_name, is_subtotal)

    for col in range(3, last_col + 1):
        cell = ws.cell(row=row_num, column=col)

        if col in sep_cols:
            cell.value = None
            cell.fill = SEP_FILL
            cell.border = NO_BORDER
            cell.alignment = ALIGNMENTS["center"]
            continue

        if is_subtotal:
            _style_subtotal_row_cell(cell, is_net=is_net)
        else:
            _style_regular_row_cell(cell, row_num)

        _set_alignment_by_col(cell, col, sep_cols)

        header_name = header_map.get(col)
        _set_number_format(cell, row_name, header_name)

        if _is_negative_number(cell.value):
            _apply_negative_font(cell, bold=is_subtotal)
        else:
            _apply_positive_font(cell, bold=is_subtotal)


# def _draw_blank_row(ws, row_num, last_col, sep_cols):
#     ws.row_dimensions[row_num].height = 8

#     for col in range(1, last_col + 1):
#         cell = ws.cell(row=row_num, column=col)
#         cell.value = None

#         if col in sep_cols:
#             cell.fill = SEP_FILL
#             cell.border = NO_BORDER
#         else:
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]

#         cell.alignment = ALIGNMENTS["left"]


def _draw_blank_row(ws, row_num, last_col, sep_cols, has_button=False):
    ws.row_dimensions[row_num].height = 8

    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = None

        if col in sep_cols:
            cell.fill = SEP_FILL
            cell.border = NO_BORDER
        else:
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]

        cell.alignment = ALIGNMENTS["left"]
    
    # Если это строка для кнопки, проставляем высоту побольше
    if has_button:
        ws.row_dimensions[row_num].height = 12


# def _draw_footer_note(ws, row, last_col, sep_cols):
#     for col in range(1, last_col + 1):
#         cell = ws.cell(row=row, column=col)
#         cell.value = None

#         if col in sep_cols:
#             cell.fill = SEP_FILL
#             cell.border = NO_BORDER
#         else:
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]

#     ws.cell(row=row, column=1, value="Note — ссылки на листы с расшифровками.")
#     ws.cell(row=row, column=1).font = FONTS["subtitle"]
#     ws.cell(row=row, column=1).alignment = ALIGNMENTS["left"]



def _draw_footer_note(ws, row, last_col, sep_cols, has_button=False):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = None

        if col in sep_cols:
            cell.fill = SEP_FILL
            cell.border = NO_BORDER
        else:
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]

    ws.cell(row=row, column=1, value="Note — ссылки на листы с расшифровками.")
    ws.cell(row=row, column=1).font = FONTS["subtitle"]
    ws.cell(row=row, column=1).alignment = ALIGNMENTS["left"]
    
    # Добавляем кнопку перехода к детализации
    if has_button:
        button_row = row + 2
        draw_nav_link_row(
            ws,
            row=button_row,
            text="Перейти к детализации →",
            target_sheet="pl_drill_down",
            target_cell="A1",
        )


def _apply_pl_column_outline(ws, headers):
    """
    Логика группировки:
    - FYE-колонки всегда видимы;
    - месяцы каждого FYE-года скрываются отдельной группой;
    - плюсик отображается рядом с FYE соответствующего года;
    - YTD, Diff, MTD не скрываются.
    """

    month_order = {
        "Янв": 1,
        "Фев": 2,
        "Мар": 3,
        "Апр": 4,
        "Май": 5,
        "Июн": 6,
        "Июл": 7,
        "Авг": 8,
        "Сен": 9,
        "Окт": 10,
        "Ноя": 11,
        "Дек": 12,
    }

    def is_month_header(header):
        parts = str(header).split()
        return (
            len(parts) == 2
            and parts[0] in month_order
            and parts[1].isdigit()
        )

    def group_cols(cols):
        if not cols:
            return

        start_col = min(cols)
        end_col = max(cols)

        ws.column_dimensions.group(
            start=get_column_letter(start_col),
            end=get_column_letter(end_col),
            outline_level=1,
            hidden=True,
        )

        # Так как summaryRight=False, итоговая колонка находится слева от месяцев.
        # Именно на ней Excel показывает плюсик.
        summary_col = start_col - 1
        if summary_col >= 1:
            ws.column_dimensions[get_column_letter(summary_col)].collapsed = True

    month_years = sorted({
        int(str(header).split()[1])
        for header in headers
        if is_month_header(header)
    })

    for year in month_years:
        year_month_cols = [
            idx for idx, header in enumerate(headers, start=1)
            if is_month_header(header)
            and int(str(header).split()[1]) == year
        ]

        group_cols(year_month_cols)


def style_pl_sheet(ws, pl_df, date_to=None):
    _clear_pl_sheet(ws)
    ws.auto_filter.ref = None
    _prepare_layout(ws)
    _draw_pl_header(ws, date_to=date_to)

    start_header_row = 7
    start_data_row = 8

    if pl_df is None or pl_df.empty:
        headers = ["Name", "Note"]
        draw_table_header(ws, row=start_header_row, headers=headers, start_col=1, wrap=False)
        _apply_dynamic_widths(ws, headers)

        ws["A8"] = "Нет данных"
        ws["A8"].font = FONTS["normal"]
        ws["A8"].alignment = ALIGNMENTS["left"]
        return

    pl_df = pl_df.copy()
    pl_df = pl_df.reindex(PL_ROW_ORDER)

    ordered_cols = _get_ordered_columns(pl_df)
    headers = _excel_headers_from_ordered_cols(ordered_cols)

    _apply_dynamic_widths(ws, headers)
    _style_header_with_separator(ws, headers, row=start_header_row)

    last_col = len(headers)
    sep_cols = [idx for idx, hdr in enumerate(headers, start=1) if hdr == ""]
    header_map = {idx: hdr for idx, hdr in enumerate(headers, start=1)}

    current_row = start_data_row

    for row_name in pl_df.index:
        ws.cell(row=current_row, column=1, value=row_name)
        ws.cell(row=current_row, column=2, value=None)
        _draw_note_badge(ws, current_row, row_name)

        excel_col = 3

        for col_name in ordered_cols:
            if col_name == "Note":
                continue

            if col_name in ("__SEP__", "__SEP2__"):
                ws.cell(row=current_row, column=excel_col, value=None)
                excel_col += 1
                continue

            value = pl_df.loc[row_name, col_name]
            ws.cell(row=current_row, column=excel_col, value=value)
            excel_col += 1

        _style_data_row(
            ws=ws,
            row_num=current_row,
            row_name=row_name,
            header_map=header_map,
            last_col=last_col,
            sep_cols=sep_cols,
        )

        ws.row_dimensions[current_row].height = 20

        if row_name in BREAK_AFTER_ROWS:
            current_row += 1
            _draw_blank_row(ws, current_row, last_col, sep_cols)

        current_row += 1

    end_row = current_row - 1

    if sep_cols:
        _style_separator_cols(ws, sep_cols, start_header_row, end_row + 2)

    _apply_pl_column_outline(ws, headers)


    _draw_footer_note(ws, end_row + 2, last_col, sep_cols, has_button=True)