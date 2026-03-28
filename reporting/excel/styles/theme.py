# reporting/excel/styles/theme.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED",
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "back_fill": "EAF2FB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "blue": "2F75B5",
    "back_text_green": "1F5E4E",
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "back": PatternFill("solid", fgColor=COLORS["light_green"]),
    "none": PatternFill(fill_type=None),
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=10, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "back": Font(name="Roboto", size=10, bold=True, color=COLORS["back_text_green"]),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium_dark = Side(style="medium", color=COLORS["black"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium_dark),
    "none": Border(),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
}

FORMATS = {
    "money": '#,##0',
    "date": 'yyyy-mm-dd',
}