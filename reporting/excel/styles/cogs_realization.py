# reporting/excel/styles/cogs_realization.py

from datetime import date, datetime
from numbers import Number

import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reporting.excel.styles.theme import (
    FILLS,
    FONTS,
    BORDERS,
    ALIGNMENTS,
    COLORS,
)
from reporting.excel.styles.style_helpers import (
    clear_range,
    set_row_heights,
    draw_back_to_pl_button,
    draw_sheet_header,
    draw_nav_link_row

)


TOTAL_ROW_KEY = "__TOTAL__"

YEAR_FILL = PatternFill("solid", fgColor="DCEBE4")
TOTAL_COL_FILL = PatternFill("solid", fgColor="EEF5F2")
TOTAL_ROW_FILL = PatternFill("solid", fgColor="E4EFEA")
PARENT_FILL = PatternFill("solid", fgColor="F4F7F6")
CHILD_FILL_EVEN = FILLS["alt"]
CHILD_FILL_ODD = FILLS["none"]
SEP_FILL = PatternFill("solid", fgColor=COLORS["white"])

THIN_GRAY_SIDE = Side(style="thin", color=COLORS["border_gray"])
MEDIUM_DARK_SIDE = Side(style="medium", color=COLORS["black"])

TOTAL_BORDER = Border(
    top=THIN_GRAY_SIDE,
    bottom=MEDIUM_DARK_SIDE,
)

PARENT_BORDER = Border(
    top=THIN_GRAY_SIDE,
    bottom=THIN_GRAY_SIDE,
)

NO_BORDER = Border()

NEGATIVE_COLOR = "6B4C4C"
NEGATIVE_FONT = Font(name="Roboto", size=10, color=NEGATIVE_COLOR)
NEGATIVE_BOLD_FONT = Font(name="Roboto", size=10, bold=True, color=NEGATIVE_COLOR)


def _safe_str_date(dt):
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%d.%m.%Y")
    if isinstance(dt, date):
        return dt.strftime("%d.%m.%Y")
    return str(dt)


def _unmerge_all(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def _clear_sheet(ws):
    _unmerge_all(ws)
    clear_range(ws, row_start=1, row_end=700, col_start=1, col_end=160)


def _prepare_layout(ws):
    row_heights = {
        1: 20,
        2: 24,
        3: 18,
        4: 18,
        6: 10,
        7: 24,
        8: 22,
    }
    set_row_heights(ws, row_heights)

    ws.sheet_view.showGridLines = False
    ws.sheet_view.showOutlineSymbols = True
    ws.freeze_panes = "B9"
    ws.sheet_properties.outlinePr.summaryRight = True
    ws.sheet_properties.outlinePr.summaryBelow = False


def _draw_header(ws, date_to=None):
    draw_back_to_pl_button(ws)

    subtitle = "Управленческая отчетность (management pack)"
    currency = "Российский рубль (RUB)"
    if date_to:
        currency = f"{currency} • дата отчета: {_safe_str_date(date_to)}"

    draw_sheet_header(
        ws,
        title="1.3 СЕБЕСТОИМОСТЬ РЕАЛИЗАЦИИ",
        subtitle=subtitle,
        currency=currency,
    )


def _build_visual_columns(year_groups):
    """
    Возвращает:
    visual_columns = [
        {"type": "data", "name": "...", "year": 2025, "is_total_col": False},
        {"type": "data", "name": "Итого 2025", "year": 2025, "is_total_col": True},
        {"type": "sep"},
        ...
    ]
    """
    visual_columns = []

    for idx, group in enumerate(year_groups):
        for month_col in group["month_cols"]:
            visual_columns.append({
                "type": "data",
                "name": month_col,
                "year": group["year"],
                "is_total_col": False,
            })

        visual_columns.append({
            "type": "data",
            "name": group["total_col"],
            "year": group["year"],
            "is_total_col": True,
        })

        if idx < len(year_groups) - 1:
            visual_columns.append({"type": "sep"})

    return visual_columns


def _draw_two_level_header(ws, year_groups, visual_columns, start_row=7):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)

    cell = ws.cell(row=start_row, column=1, value="Статья")
    cell.fill = FILLS["header"]
    cell.font = FONTS["header_white"]
    cell.alignment = ALIGNMENTS["center"]
    cell.border = BORDERS["thin"]

    ws.cell(row=start_row + 1, column=1).fill = FILLS["header"]
    ws.cell(row=start_row + 1, column=1).border = BORDERS["thin"]

    current_col = 2
    visual_idx = 0

    for group_idx, group in enumerate(year_groups):
        group_data_cols = len(group["month_cols"]) + 1

        start_col = current_col
        end_col = current_col + group_data_cols - 1

        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=end_col,
        )

        year_cell = ws.cell(row=start_row, column=start_col, value=str(group["year"]))
        year_cell.fill = YEAR_FILL
        year_cell.font = FONTS["total"]
        year_cell.alignment = ALIGNMENTS["center"]
        year_cell.border = BORDERS["thin"]

        for col in range(start_col, end_col + 1):
            ws.cell(row=start_row, column=col).fill = YEAR_FILL
            ws.cell(row=start_row, column=col).border = BORDERS["thin"]

        for month_col in group["month_cols"]:
            c = ws.cell(row=start_row + 1, column=current_col, value=month_col.split()[0])
            c.fill = FILLS["header"]
            c.font = FONTS["header_white"]
            c.alignment = ALIGNMENTS["center"]
            c.border = BORDERS["thin"]
            current_col += 1
            visual_idx += 1

        c = ws.cell(row=start_row + 1, column=current_col, value=group["total_col"])
        c.fill = TOTAL_COL_FILL
        c.font = FONTS["total"]
        c.alignment = ALIGNMENTS["center"]
        c.border = BORDERS["thin"]
        current_col += 1
        visual_idx += 1

        if group_idx < len(year_groups) - 1:
            # вертикальный разделитель
            top_sep = ws.cell(row=start_row, column=current_col)
            bot_sep = ws.cell(row=start_row + 1, column=current_col)
            top_sep.value = None
            bot_sep.value = None
            top_sep.fill = SEP_FILL
            bot_sep.fill = SEP_FILL
            top_sep.border = NO_BORDER
            bot_sep.border = NO_BORDER
            current_col += 1
            visual_idx += 1


def _build_col_map(visual_columns):
    data_col_map = {}
    sep_cols = []

    excel_col = 2
    for col_cfg in visual_columns:
        if col_cfg["type"] == "sep":
            sep_cols.append(excel_col)
            excel_col += 1
            continue

        data_col_map[col_cfg["name"]] = excel_col
        excel_col += 1

    return data_col_map, sep_cols


def _apply_month_outline(ws, year_groups, data_col_map, current_year=None):
    for group in year_groups:
        if not group["month_cols"]:
            continue

        start_col = data_col_map[group["month_cols"][0]]
        end_col = data_col_map[group["month_cols"][-1]]

        is_current_year = (group["year"] == current_year)

        ws.column_dimensions.group(
            start=get_column_letter(start_col),
            end=get_column_letter(end_col),
            outline_level=1,
            hidden=not is_current_year,
        )

        # чтобы у итоговой колонки года появлялось "плечико" / значок сворачивания
        total_col_letter = get_column_letter(data_col_map[group["total_col"]])
        ws.column_dimensions[total_col_letter].collapsed = not is_current_year


def _set_widths(ws, visual_columns):
    ws.column_dimensions["A"].width = 46

    excel_col = 2
    for col_cfg in visual_columns:
        letter = get_column_letter(excel_col)

        if col_cfg["type"] == "sep":
            ws.column_dimensions[letter].width = 2.5
        else:
            if col_cfg["is_total_col"]:
                ws.column_dimensions[letter].width = 14
            else:
                ws.column_dimensions[letter].width = 12

        excel_col += 1


def _is_negative(value):
    return isinstance(value, Number) and value < 0


def _child_fill(row_num):
    return CHILD_FILL_EVEN if row_num % 2 == 0 else CHILD_FILL_ODD


def _style_name_cell(cell, level, is_total_row, row_num):
    if is_total_row:
        cell.fill = TOTAL_ROW_FILL
        cell.border = TOTAL_BORDER
        cell.font = FONTS["total"]
        cell.alignment = ALIGNMENTS["left"]
        return

    if level == 1:
        cell.fill = PARENT_FILL
        cell.border = PARENT_BORDER
        cell.font = FONTS["total"]
        cell.alignment = ALIGNMENTS["left"]
        return

    cell.fill = _child_fill(row_num)
    cell.border = BORDERS["thin"]
    cell.font = FONTS["normal"]
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=2,
    )


def _style_number_cell(cell, is_total_col, is_total_row, level, row_num):
    if is_total_row:
        cell.fill = TOTAL_ROW_FILL
        cell.border = TOTAL_BORDER
        cell.font = FONTS["total"]
    elif level == 1:
        if is_total_col:
            cell.fill = TOTAL_COL_FILL
            cell.border = PARENT_BORDER
        else:
            cell.fill = PARENT_FILL
            cell.border = PARENT_BORDER
        cell.font = FONTS["total"]
    elif is_total_col:
        cell.fill = TOTAL_COL_FILL
        cell.border = BORDERS["thin"]
        cell.font = FONTS["total"]
    else:
        cell.fill = _child_fill(row_num)
        cell.border = BORDERS["thin"]
        cell.font = FONTS["normal"]

    cell.alignment = ALIGNMENTS["right"]
    cell.number_format = '#,##0;(#,##0)'

    if _is_negative(cell.value):
        if is_total_row or is_total_col or level == 1:
            cell.font = NEGATIVE_BOLD_FONT
        else:
            cell.font = NEGATIVE_FONT


def _display_name(display_name, code, level):
    label = "" if display_name is None else str(display_name)

    if code and label.startswith(f"{code} "):
        label = label[len(code) + 1:]

    if code:
        return f"{code} {label}"
    return label


def _apply_row_outline(ws, row_positions, row_meta):
    """
    Группируем детей под родителем.
    row_meta:
    - index = row_key
    - display_name
    - code
    - level
    - parent_row
    """
    if row_meta is None or row_meta.empty:
        return

    meta = row_meta.copy()
    if "parent_row" not in meta.columns:
        return

    meta = meta.reset_index().rename(columns={"index": "row_key"})
    parents = meta[meta["level"] == 1]["row_key"].tolist()

    for parent_key in parents:
        child_rows = meta[
            (meta["level"] == 2) &
            (meta["parent_row"] == parent_key)
        ]["row_key"].tolist()

        excel_rows = [row_positions[r] for r in child_rows if r in row_positions]
        if not excel_rows:
            continue

        for r in excel_rows:
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = False


def _draw_blank_separator_row(ws, row_num, last_col, sep_cols):
    ws.row_dimensions[row_num].height = 6

    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = None

        if col in sep_cols:
            cell.fill = SEP_FILL
            cell.border = NO_BORDER
        else:
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]

        cell.alignment = ALIGNMENTS["left"]


def _need_blank_after_row(current_key, next_key, row_meta):
    if current_key == TOTAL_ROW_KEY:
        return False
    if next_key is None:
        return False

    if row_meta is None or row_meta.empty:
        return False
    if current_key not in row_meta.index:
        return False

    current_level = int(row_meta.loc[current_key, "level"])
    if current_level != 2:
        return False

    current_parent = row_meta.loc[current_key, "parent_row"]

    if next_key == TOTAL_ROW_KEY:
        return True

    if next_key not in row_meta.index:
        return False

    next_parent = row_meta.loc[next_key, "parent_row"]
    return current_parent != next_parent


def style_cogs_realization_sheet(ws, payload, date_to=None):
    _clear_sheet(ws)
    ws.auto_filter.ref = None

    _prepare_layout(ws)
    _draw_header(ws, date_to)

    df = payload.get("data") if payload else None
    year_groups = payload.get("year_groups", []) if payload else []
    row_meta = payload.get("row_meta") if payload else None

    start_header_row = 7
    start_data_row = 9

    visual_columns = _build_visual_columns(year_groups)
    data_col_map, sep_cols = _build_col_map(visual_columns)

    if df is None or df.empty:
        _draw_two_level_header(ws, year_groups, visual_columns, start_row=start_header_row)
        _set_widths(ws, visual_columns)
        ws["A9"] = "Нет данных"
        ws["A9"].font = FONTS["normal"]
        ws["A9"].alignment = ALIGNMENTS["left"]
        return

    if row_meta is None or row_meta.empty:
        row_meta = pd.DataFrame(
            index=df.index,
            data={
                "display_name": df.index,
                "code": "",
                "level": 2,
                "parent_row": None,
            }
        )
    else:
        row_meta = row_meta.copy()

    _draw_two_level_header(ws, year_groups, visual_columns, start_row=start_header_row)
    _set_widths(ws, visual_columns)

    row_positions = {}
    current_row = start_data_row
    row_keys = list(df.index)

    for idx, row_key in enumerate(row_keys):
        next_key = row_keys[idx + 1] if idx + 1 < len(row_keys) else None
        is_total_row = str(row_key).strip() == TOTAL_ROW_KEY

        if is_total_row:
            code = ""
            level = 0
            display_name = "Итого"
        else:
            meta = row_meta.loc[row_key] if row_key in row_meta.index else None
            code = "" if meta is None else str(meta.get("code", "") or "")
            level = 2 if meta is None else int(meta.get("level", 2) or 2)
            raw_name = row_key if meta is None else meta.get("display_name", row_key)
            display_name = _display_name(raw_name, code, level)

        name_cell = ws.cell(row=current_row, column=1, value=display_name)
        _style_name_cell(
            name_cell,
            level=level,
            is_total_row=is_total_row,
            row_num=current_row,
        )

        # separator columns in row
        for sep_col in sep_cols:
            sep_cell = ws.cell(row=current_row, column=sep_col, value=None)
            sep_cell.fill = SEP_FILL
            sep_cell.border = NO_BORDER
            sep_cell.alignment = ALIGNMENTS["center"]

        for col_name in df.columns:
            col_idx = data_col_map[col_name]
            val = df.loc[row_key, col_name]

            data_cell = ws.cell(row=current_row, column=col_idx, value=val)
            _style_number_cell(
                data_cell,
                is_total_col=str(col_name).startswith("Итого "),
                is_total_row=is_total_row,
                level=level,
                row_num=current_row,
            )

        ws.row_dimensions[current_row].height = 20

        if not is_total_row:
            row_positions[row_key] = current_row

        current_row += 1

        if _need_blank_after_row(row_key, next_key, row_meta):
            last_col = 1 + len(visual_columns)
            _draw_blank_separator_row(ws, current_row, last_col, sep_cols)
            current_row += 1

    end_row = current_row - 1
    
    button_row = end_row + 3

    draw_nav_link_row(
        ws,
        row=button_row,
        text="Перейти к детализации →",
        target_sheet="1.3_drill_down",
        target_cell="A1",
    )
    
    total_cols = 1 + len(visual_columns)

    current_year = None
    if year_groups:
        current_year = max(g["year"] for g in year_groups)

    _apply_month_outline(ws, year_groups, data_col_map, current_year=current_year)
    
    _apply_row_outline(ws, row_positions, row_meta)

    ws.auto_filter.ref = None