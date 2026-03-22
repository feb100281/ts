# reporting/excel/styles/style_helpers.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reporting.excel.styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, FORMATS, COLORS


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_heights(ws, heights: dict):
    for row_idx, height in heights.items():
        ws.row_dimensions[row_idx].height = height


def clear_range(ws, row_start, row_end, col_start, col_end):
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["left"]


def draw_section_title(ws, row, col_start, col_end, title):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILLS["section"]
        cell.border = BORDERS["bottom_thin"]
        if col == col_start:
            cell.value = title
            cell.font = FONTS["section"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.value = None


def draw_table_header(ws, row, headers, start_col=1, wrap=False):
    alignment = ALIGNMENTS["center_wrap"] if wrap else ALIGNMENTS["center"]

    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i)
        cell.value = header
        cell.fill = FILLS["header"]
        cell.font = FONTS["header_white"]
        cell.alignment = alignment
        cell.border = BORDERS["thin"]


def style_zebra_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONTS["normal"]
        cell.border = BORDERS["thin"]
        cell.fill = FILLS["alt"] if row % 2 == 0 else FILLS["none"]
        
        
def clear_range_below_table(ws, data_end_row, row_end, col_start, col_end):
    for row in range(data_end_row + 1, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["left"]
            


#### ---- КНОПКА ОГЛАВЛЕНИЕ -----######
def draw_toc_button(ws, cell="A1", text="← Оглавление"):
    ws[cell] = text

    # ссылка на лист TOC
    ws[cell].hyperlink = "#'TOC'!A1"

    ws[cell].font = FONTS["back"]
    ws[cell].alignment = ALIGNMENTS["left"]

    ws[cell].fill = FILLS["back"]
    ws[cell].border = Border(
        left=Side(style="thin", color=COLORS["border_gray"]),
        right=Side(style="thin", color=COLORS["border_gray"]),
        top=Side(style="thin", color=COLORS["border_gray"]),
        bottom=Side(style="thin", color=COLORS["border_gray"]),
    )
    
#### ---- НАЗВАНИЕ ЛИСТА -----######
def draw_sheet_header(ws, title, subtitle, currency):
    ws["A2"] = title
    ws["A3"] = subtitle
    ws["A4"] = currency
    ws["A5"] = None

    ws["A2"].font = FONTS["title"]
    ws["A3"].font = FONTS["subtitle"]
    ws["A4"].font = FONTS["subtitle"]

    ws["A2"].alignment = ALIGNMENTS["left"]
    ws["A3"].alignment = ALIGNMENTS["left"]
    ws["A4"].alignment = ALIGNMENTS["left"]

    # линия под заголовком (A:E)
    for col in range(1, 6):
        ws.cell(row=6, column=col).border = BORDERS["bottom_medium"]
        
#### ---- САММАРИ БЛОК -----######       
def draw_summary_block(
    ws,
    items,
    start_row=7,
    label_col=1,
    value_col=5,
    col_start=1,
    col_end=5,
):
    """
    items: список словарей вида
    [
        {"label": "БАНКОВСКИЕ СЧЕТА", "value": 57021874},
        {"label": "ДЕНЬГИ В ПУТИ", "value": 15700000},
        {"label": "ИТОГО:", "value": 74852316, "is_total": True},
    ]
    """
    for idx, item in enumerate(items, start=start_row):
        is_total = item.get("is_total", False)

        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=idx, column=col)

            if is_total:
                cell.fill = FILLS["total"]
                cell.border = BORDERS["bottom_medium"]
            else:
                cell.fill = FILLS["summary"]
                cell.border = BORDERS["bottom_thin"]

        # текст слева
        label_cell = ws.cell(row=idx, column=label_col)
        label_cell.value = item.get("label")
        label_cell.font = FONTS["bold"] if is_total else FONTS["normal"]
        label_cell.alignment = ALIGNMENTS["left"]

        # значение справа
        value_cell = ws.cell(row=idx, column=value_col)
        value_cell.value = item.get("value")
        value_cell.font = FONTS["total"] if is_total else FONTS["normal"]
        value_cell.alignment = ALIGNMENTS["right"]
        value_cell.number_format = FORMATS["money"]

        # промежуточные пустые ячейки
        for col in range(label_col + 1, value_col):
            mid_cell = ws.cell(row=idx, column=col)
            mid_cell.value = None
            mid_cell.alignment = ALIGNMENTS["left"]