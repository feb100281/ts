# reporting/excel/styles/toc.py
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side

from reporting.excel.styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, COLORS
from reporting.excel.styles.style_helpers import set_column_widths, set_row_heights


TOC_ITEMS = [
    {"section": "content", "num": "1", "title": "Отчет о прибылях и убытках (PL)", "sheet": "PL"},
    {"section": "content", "num": "2", "title": "Отчет о движении денежных средств (Cash Flow)", "sheet": "CF"},
    {"section": "content", "num": "3", "title": "Аналитический баланс (Trial BS)", "sheet": "BS"},

    {"section": "detail", "num": "1.1", "title": "Выручка от реализации", "sheet": "1.1"},
    {"section": "detail", "num": "1.2", "title": "Себестоимость проданных товаров", "sheet": "1.2"},
    {"section": "detail", "num": "1.3", "title": "Себестоимость реализации", "sheet": "1.3"},
    {"section": "detail", "num": "1.4", "title": "Накладные расходы (Overheads)", "sheet": "1.4"},
    {"section": "detail", "num": "1.5", "title": "Корпоративные расходы (G&A)", "sheet": "1.5"},
    {"section": "detail", "num": "1.6", "title": "Прочие доходы и расходы", "sheet": "1.6"},
    {"section": "detail", "num": "1.7", "title": "Финансовые расходы", "sheet": "1.7"},
    {"section": "detail", "num": "2.1", "title": "Казначейство (Treasure report)", "sheet": "2.1"},
    {"section": "detail", "num": "2.2", "title": "Обороты по балансу маркетплейсов", "sheet": "2.2"},
    {"section": "detail", "num": "3.1", "title": "Взаиморасчеты", "sheet": "3.1"},
]


def _draw_toc_item_row(
    ws,
    row,
    num,
    title,
    sheet,
    workbook,
    *,
    is_detail=False,
    zebra=False,
    font_num=None,
    font_link=None,
    fill_alt=None,
    fill_default=None,
):
    row_fill = fill_alt if zebra else fill_default
    dotted_border = Border(bottom=Side(style="dotted", color=COLORS["border_gray"]))

    for col in range(2, 5):  
        cell = ws.cell(row=row, column=col)
        cell.fill = row_fill
        cell.border = dotted_border
        if col not in (2, 4):
            cell.value = None

    num_cell = ws.cell(row=row, column=2)
    title_cell = ws.cell(row=row, column=4)

    num_cell.value = str(num)
    num_cell.number_format = "@"
    num_cell.font = font_num
    num_cell.alignment = ALIGNMENTS["center"]

    title_cell.value = title
    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=1 if is_detail else 0
    )
    _set_link(title_cell, sheet, font_link, workbook)


def _clear_toc_area(ws, max_row=250, max_col=8):
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row <= max_row and merged_range.max_col <= max_col:
            ws.unmerge_cells(str(merged_range))

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.hyperlink = None
            cell.alignment = Alignment(horizontal="general", vertical="bottom")
            cell.number_format = "General"


def _write_section_bar(ws, row, title, fill_section, font_section, align_left):
    for col in range(2, 5):  # B:F
        cell = ws.cell(row=row, column=col)
        cell.fill = fill_section
        cell.border = BORDERS["bottom_thin"]
        cell.alignment = align_left
        if col == 2:
            cell.value = title
            cell.font = font_section
        else:
            cell.value = None


def _set_link(cell, target_sheet, font_link, workbook):
    if target_sheet in workbook.sheetnames:
        cell.hyperlink = f"#'{target_sheet}'!A1"
    cell.font = font_link


def build_toc_sheet(ws, report_date, version="Stand alone", toc_items=None):
    if toc_items is None:
        toc_items = TOC_ITEMS

    align_left = ALIGNMENTS["left"]
    align_center = ALIGNMENTS["center"]

    fill_section = FILLS["section"]
    fill_summary = FILLS["summary"]

    font_title = Font(name="Roboto", size=18, bold=True, color=COLORS["black"])
    font_subtitle = FONTS["subtitle"]
    font_section = Font(name="Roboto", size=13, bold=True, color=COLORS["black"])

    font_meta_label = Font(name="Roboto", size=10, bold=True, color=COLORS["black"])
    font_meta_value = Font(name="Roboto", size=10, bold=False, color=COLORS["black"])

    font_num_main = Font(name="Roboto", size=11, bold=True, color=COLORS["black"])
    font_num_detail = Font(name="Roboto", size=10, bold=True, color=COLORS["black"])

    font_link_main = Font(
        name="Roboto",
        size=11,
        bold=True,
        color="1F6F43",
        underline="single",
    )
    font_link_detail = Font(
        name="Roboto",
        size=10,
        bold=False,
        color="1F6F43",
        underline="single",
    )

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    _clear_toc_area(ws)

    widths = {
        "A": 4,
        "B": 9,
        "C": 4,
        "D": 68,
        "E": 22,
        "F": 22,
        "G": 18,
        "H": 18,
    }
    set_column_widths(ws, widths)

    base_row_heights = {
        1: 30,
        2: 22,
        3: 10,
        4: 20,
        5: 20,
        6: 12,
        7: 25,
        8: 8,
    }
    set_row_heights(ws, base_row_heights)

    ws["B1"] = "ТРЕНДСЕТТЕР"
    ws["B1"].font = font_title
    ws["B1"].alignment = align_left

    ws["B2"] = "Управленческая отчетность (management pack)"
    ws["B2"].font = font_subtitle
    ws["B2"].alignment = align_left

    for col in range(2, 5): 
        ws.cell(row=3, column=col).border = BORDERS["bottom_medium"]

    for cell_ref in ["B4", "C4", "D4", "B5", "C5", "D5"]:
        ws[cell_ref].fill = fill_summary
        ws[cell_ref].alignment = align_left

    ws["B4"] = "Версия:"
    ws["B4"].font = font_meta_label

    ws["D4"] = version
    ws["D4"].font = font_meta_value

    ws["B5"] = "Дата:"
    ws["B5"].font = font_meta_label

    ws["D5"] = report_date
    ws["D5"].font = font_meta_value

    content_items = [x for x in toc_items if x["section"] == "content"]
    detail_items = [x for x in toc_items if x["section"] == "detail"]

    content_section_row = 7
    current_row = content_section_row
    _write_section_bar(ws, current_row, "СОДЕРЖАНИЕ", fill_section, font_section, align_left)
    ws.row_dimensions[current_row].height = 25
    current_row += 2

    # for item in content_items:
    #     num_cell = ws[f"B{current_row}"]
    #     title_cell = ws[f"D{current_row}"]

    #     num_cell.value = str(item["num"])
    #     num_cell.number_format = "@"
    #     num_cell.font = font_num_main
    #     num_cell.alignment = align_center

    #     title_cell.value = item["title"]
    #     title_cell.alignment = align_left
    #     _set_link(title_cell, item["sheet"], font_link_main, ws.parent)

    #     ws.row_dimensions[current_row].height = 24
    #     current_row += 1
    
    
    for idx, item in enumerate(content_items):
        _draw_toc_item_row(
            ws,
            current_row,
            item["num"],
            item["title"],
            item["sheet"],
            ws.parent,
            is_detail=False,
            zebra=False,   # для основных разделов лучше без зебры
            font_num=font_num_main,
            font_link=font_link_main,
            fill_alt=FILLS["alt"],
            fill_default=FILLS["none"],
        )
        ws.row_dimensions[current_row].height = 24
        current_row += 1

    current_row += 1

    detail_section_row = current_row
    _write_section_bar(ws, current_row, "РАСШИФРОВКИ", fill_section, font_section, align_left)
    ws.row_dimensions[current_row].height = 25
    current_row += 2

    # for item in detail_items:
    #     num_cell = ws[f"B{current_row}"]
    #     title_cell = ws[f"D{current_row}"]

    #     num_cell.value = str(item["num"])
    #     num_cell.number_format = "@"
    #     num_cell.font = font_num_detail
    #     num_cell.alignment = align_center

    #     title_cell.value = item["title"]
    #     title_cell.alignment = align_left_indent
    #     _set_link(title_cell, item["sheet"], font_link_detail, ws.parent)

    #     ws.row_dimensions[current_row].height = 22
    #     current_row += 1
    
    
    for idx, item in enumerate(detail_items):
        _draw_toc_item_row(
            ws,
            current_row,
            item["num"],
            item["title"],
            item["sheet"],
            ws.parent,
            is_detail=False,
            zebra=False,
            font_num=font_num_detail,
            font_link=font_link_detail,
            fill_alt=FILLS["alt"],
            fill_default=FILLS["none"],
        )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    protected_rows = {3, content_section_row, detail_section_row}

    for row in range(1, current_row + 5):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value is None and row not in protected_rows:
                cell.fill = PatternFill(fill_type=None)