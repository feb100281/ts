# reporting/excel/styles/treasury.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import date, datetime

from reporting.excel.styles.theme import FILLS, FONTS, BORDERS, ALIGNMENTS, FORMATS
from reporting.excel.styles.style_helpers import (
    set_column_widths,
    set_row_heights,
    clear_range,
    clear_range_below_table,
    draw_section_title,
    draw_table_header,
    style_zebra_row,
    draw_toc_button,
    draw_sheet_header,
    draw_summary_block,
)


def _safe_str_date(dt):
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%d.%m.%Y")
    if isinstance(dt, date):
        return dt.strftime("%d.%m.%Y")
    return str(dt)


def style_sheet_2_1(ws, ba_df, wb_df):
    align_left = ALIGNMENTS["left"]
    align_center = ALIGNMENTS["center"]
    align_right = ALIGNMENTS["right"]

    money_fmt = FORMATS["money"]
    date_fmt = FORMATS["date"]

    # =========================
    # ШИРИНЫ КОЛОНОК
    # =========================
    widths = {
        "A": 54,
        "B": 14,
        "C": 14,
        "D": 10,
        "E": 17,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 15,
    }
    set_column_widths(ws, widths)

    # =========================
    # ВЫСОТЫ СТРОК
    # =========================
    row_heights = {
        1: 20,
        2: 26,
        3: 18,
        4: 18,
        6: 8,
        7: 21,
        8: 21,
        9: 21,
        10: 21,
        11: 23,
        14: 22,
        15: 24,
        34: 22,
        35: 40,
    }
    set_row_heights(ws, row_heights)

    # =========================
    # HEADER
    # =========================
    draw_toc_button(ws)

    draw_sheet_header(
        ws,
        title="ОТЧЕТ ОБ ОСТАТКАХ",
        subtitle="Управленческая отчетность (management pack)",
        currency = "Российский рубль (RUB)"

)

    
    # =========================
    # SUMMARY-БЛОК
    # =========================
    summary_items = [
        {"label": "БАНКОВСКИЕ СЧЕТА", "value": ws["E7"].value},
        {"label": "ДЕНЬГИ В ПУТИ", "value": ws["E8"].value},
        {"label": "БАЛАНС WB", "value": ws["E9"].value},
        {"label": "БЕССРОЧНЫЕ ДЕПОЗИТЫ", "value": ws["E10"].value},
        {"label": "ИТОГО:", "value": ws["E11"].value, "is_total": True},
    ]

    draw_summary_block(
        ws,
        items=summary_items,
        start_row=7,
        label_col=1,
        value_col=5,
        col_start=1,
        col_end=5,
    )

    
    # =========================
    # ЗАГОЛОВОК СЕКЦИИ БАНКОВ
    # =========================
    draw_section_title(ws, row=14, col_start=1, col_end=8, title="БАНКОВСКИЕ СЧЕТА")

    # =========================
    # ШАПКА ТАБЛИЦЫ БАНКОВ
    # =========================
    bank_headers = [
        "Банковский счет",
        "Начальная дата",
        "Последняя дата",
        "Валюта",
        "Действующий счет",
        "Поступления",
        "Расход",
        "Остаток",
    ]
    draw_table_header(ws, row=15, headers=bank_headers, start_col=1, wrap=False)

    # =========================
    # ТАБЛИЦА БАНКОВСКИХ СЧЕТОВ
    # =========================
    ba_start_row = 16
    ba_end_row = ba_start_row + len(ba_df) - 1 if len(ba_df) > 0 else ba_start_row

    clear_range_below_table(
        ws,
        data_end_row=ba_end_row,
        row_end=max(ba_end_row + 5, 35),
        col_start=1,
        col_end=8,
    )

    for row in range(ba_start_row, ba_end_row + 1):
        style_zebra_row(ws, row=row, col_start=1, col_end=8)

        ws[f"A{row}"].alignment = align_left
        ws[f"B{row}"].alignment = align_center
        ws[f"C{row}"].alignment = align_center
        ws[f"D{row}"].alignment = align_center
        ws[f"E{row}"].alignment = align_center
        ws[f"F{row}"].alignment = align_right
        ws[f"G{row}"].alignment = align_right
        ws[f"H{row}"].alignment = align_right

        ws[f"B{row}"].number_format = date_fmt
        ws[f"C{row}"].number_format = date_fmt
        ws[f"F{row}"].number_format = money_fmt
        ws[f"G{row}"].number_format = money_fmt
        ws[f"H{row}"].number_format = money_fmt

    # =========================
    # БЛОК WB
    # =========================
    wb_header_row = 35
    wb_start_row = 36
    wb_end_row = wb_start_row + len(wb_df) - 1 if len(wb_df) > 0 else wb_start_row

    wb_headers = [
        "Дата",
        "К перечислению\nпродавцу\nза весь период",
        "Вывод\nсредств\nза весь период",
        "Конечный\nбаланс\nбез ДВП",
        "Деньги в пути\n(ДВП)",
        "Конечный\nбаланс",
    ]
    wb_last_col = len(wb_headers)  # A:F

    wb_widths = {
        "A": 54,
        "B": 18,
        "C": 16,
        "D": 13,
        "E": 16,
        "F": 14,
    }
    set_column_widths(ws, wb_widths)

    # WB очищаем полностью, потому что этот блок перезаписывается заново
    clear_range(ws, row_start=34, row_end=45, col_start=1, col_end=9)

    draw_section_title(ws, row=34, col_start=1, col_end=wb_last_col, title="ОБОРОТЫ WB")
    draw_table_header(ws, row=wb_header_row, headers=wb_headers, start_col=1, wrap=True)

    # Записываем данные WB заново
    for row_offset, idx in enumerate(wb_df.index, start=wb_start_row):
        ws.cell(row=row_offset, column=1, value=idx)

    for row_offset, row_data in enumerate(wb_df.itertuples(index=False), start=wb_start_row):
        for col_offset, value in enumerate(row_data, start=2):
            if col_offset <= wb_last_col:
                ws.cell(row=row_offset, column=col_offset, value=value)
                
    for row in range(wb_start_row, wb_end_row + 1):
        style_zebra_row(ws, row=row, col_start=1, col_end=wb_last_col)

        for col in range(1, wb_last_col + 1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.alignment = align_center
                cell.number_format = date_fmt
            else:
                cell.alignment = align_right
                cell.number_format = money_fmt

    # =========================
    # ОБЩЕЕ
    # =========================
    ws.freeze_panes = None
    ws.sheet_view.showGridLines = False