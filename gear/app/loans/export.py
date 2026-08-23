# # gear/app/loans/export.py
# from __future__ import annotations

# from io import BytesIO

# import pandas as pd
# from openpyxl.styles import Alignment, Font, PatternFill
# from openpyxl.utils import get_column_letter


# def _prepare_registry(
#     df: pd.DataFrame,
# ) -> pd.DataFrame:
#     if df.empty:
#         return df.copy()

#     columns = {
#         "status": "Статус",
#         "counterparty_name": "Контрагент",
#         "inn": "ИНН",
#         "contract_number": "Договор",
#         "contract_date": "Дата договора",
#         "contract_type": "Тип договора",
#         "currency": "Валюта",
#         "contract_amount": "Сумма договора",
#         "total_drawdown": "Выдано / привлечено",
#         "total_repaid": "Погашено",
#         "ending_balance": "Основной долг",
#         "interest_balance": "Долг по процентам",
#         "total_debt": "Общий долг",
#         "rate": "Ставка, %",
#         "repayment_date": "Дата погашения",
#         "days_to_maturity": "Дней до погашения",
#         "repayment_profile": "Профиль погашения",
#         "compounding": "Компаундинг",
#         "penalty_rate": "Штрафная ставка, %",
#         "total_interest_accrued": "Начислено процентов",
#         "total_interest_repaid": "Погашено процентов",
#         "contract_id": "ID договора",
#     }

#     result = df[
#         [
#             column
#             for column in columns
#             if column in df.columns
#         ]
#     ].copy()

#     result = result.rename(columns=columns)

#     for column in (
#         "Дата договора",
#         "Дата погашения",
#     ):
#         if column in result.columns:
#             result[column] = pd.to_datetime(
#                 result[column],
#                 errors="coerce",
#             ).dt.date

#     return result


# def build_excel_bytes(
#     registry_df: pd.DataFrame,
#     transactions_df: pd.DataFrame | None = None,
# ) -> bytes:
#     output = BytesIO()

#     with pd.ExcelWriter(
#         output,
#         engine="openpyxl",
#     ) as writer:
#         registry = _prepare_registry(
#             registry_df
#         )
#         registry.to_excel(
#             writer,
#             sheet_name="Реестр займов",
#             index=False,
#         )

#         if (
#             transactions_df is not None
#             and not transactions_df.empty
#         ):
#             transactions = (
#                 transactions_df.copy()
#             )
#             if "date_from" in transactions.columns:
#                 transactions["date_from"] = (
#                     pd.to_datetime(
#                         transactions["date_from"],
#                         errors="coerce",
#                     ).dt.date
#                 )

#             transactions.to_excel(
#                 writer,
#                 sheet_name="Операции",
#                 index=False,
#             )

#         for sheet in writer.book.worksheets:
#             sheet.freeze_panes = "B2"
#             sheet.sheet_view.showGridLines = False

#             header_fill = PatternFill(
#                 fill_type="solid",
#                 fgColor="E7F1ED",
#             )

#             for cell in sheet[1]:
#                 cell.font = Font(
#                     name="Roboto Light",
#                     size=10,
#                     bold=True,
#                     color="22312D",
#                 )
#                 cell.fill = header_fill
#                 cell.alignment = Alignment(
#                     horizontal="center",
#                     vertical="center",
#                     wrap_text=True,
#                 )

#             for row in sheet.iter_rows(
#                 min_row=2
#             ):
#                 for cell in row:
#                     cell.font = Font(
#                         name="Roboto Light",
#                         size=10,
#                     )
#                     cell.alignment = Alignment(
#                         vertical="center",
#                     )

#                     if isinstance(
#                         cell.value,
#                         (int, float),
#                     ):
#                         cell.number_format = (
#                             '#,##0.00'
#                         )

#             for index, column_cells in enumerate(
#                 sheet.columns,
#                 start=1,
#             ):
#                 max_length = 0

#                 for cell in column_cells:
#                     value = (
#                         ""
#                         if cell.value is None
#                         else str(cell.value)
#                     )
#                     max_length = max(
#                         max_length,
#                         len(value),
#                     )

#                 sheet.column_dimensions[
#                     get_column_letter(index)
#                 ].width = min(
#                     max(max_length + 2, 12),
#                     34,
#                 )

#     return output.getvalue()


from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DIRECTION_LABELS = {
    "borrowed": "Мы должны",
    "issued": "Нам должны",
}


def _prepare_registry(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    work = df.copy()

    if "loan_direction_label" not in work.columns:
        direction = work.get(
            "loan_direction",
            pd.Series(index=work.index, dtype="object"),
        )

        work["loan_direction_label"] = (
            direction.astype("string")
            .map(DIRECTION_LABELS)
        )

    columns = {
        "loan_direction_label": "Направление",
        "status": "Статус",
        "counterparty_name": "Контрагент",
        "inn": "ИНН",
        "contract_number": "Договор",
        "contract_date": "Дата договора",
        "contract_type": "Тип договора",
        "currency": "Валюта",
        "contract_amount": "Сумма договора",
        "total_drawdown": "Выдано / привлечено",
        "total_repaid": "Погашено",
        "ending_balance": "Основной долг",
        "interest_balance": "Долг по процентам",
        "total_debt": "Задолженность по договору",
        "rate": "Ставка, %",
        "repayment_date": "Дата погашения",
        "days_to_maturity": "Дней до погашения",
        "repayment_profile": "Профиль погашения",
        "compounding": "Компаундинг",
        "penalty_rate": "Штрафная ставка, %",
        "total_interest_accrued": "Начислено процентов",
        "total_interest_repaid": "Погашено процентов",
        "contract_id": "ID договора",
    }

    selected_columns = [
        column
        for column in columns
        if column in work.columns
    ]

    result = work[selected_columns].copy()
    result = result.rename(columns=columns)

    for column in ("Дата договора", "Дата погашения"):
        if column in result.columns:
            result[column] = (
                pd.to_datetime(
                    result[column],
                    errors="coerce",
                )
                .dt.date
            )

    # ID договора должен быть целым числом.
    if "ID договора" in result.columns:
        result["ID договора"] = (
            pd.to_numeric(
                result["ID договора"],
                errors="coerce",
            )
            .round()
            .astype("Int64")
        )

    return result


def _summary(registry_df: pd.DataFrame) -> pd.DataFrame:
    def total(
        frame: pd.DataFrame,
        column: str,
    ) -> float:
        if column not in frame.columns:
            return 0.0

        return float(
            pd.to_numeric(
                frame[column],
                errors="coerce",
            )
            .fillna(0)
            .sum()
        )

    direction = registry_df.get(
        "loan_direction",
        pd.Series("unknown", index=registry_df.index),
    ).fillna("unknown").astype(str)

    debt = pd.to_numeric(
        registry_df.get(
            "total_debt",
            pd.Series(0, index=registry_df.index),
        ),
        errors="coerce",
    ).fillna(0)

    # В саммари попадают только договоры,
    # где задолженность не равна нулю с точностью до копеек.
    has_debt = debt.round(2).ne(0)

    rows = []

    for direction_code, title in (
        ("borrowed", "Полученные займы — мы должны"),
        ("issued", "Выданные займы — нам должны"),
    ):
        part = registry_df[
            direction.eq(direction_code) & has_debt
        ].copy()

        if "contract_id" in part.columns:
            contracts_count = int(
                part["contract_id"].dropna().nunique()
            )
        else:
            contracts_count = int(len(part))

        rows.append({
            "Позиция": title,
            "Договоров": contracts_count,
            "Основной долг": total(part, "ending_balance"),
            "Проценты": total(part, "interest_balance"),
            "Всего": total(part, "total_debt"),
            "Начислено процентов": total(
                part,
                "total_interest_accrued",
            ),
            "Погашено процентов": total(
                part,
                "total_interest_repaid",
            ),
        })

    borrowed_total = float(rows[0]["Всего"])
    issued_total = float(rows[1]["Всего"])

    rows.append({
        "Позиция": (
            "Чистые обязательства "
            "(мы должны − нам должны)"
        ),
        "Договоров": None,
        "Основной долг": None,
        "Проценты": None,
        "Всего": borrowed_total - issued_total,
        "Начислено процентов": None,
        "Погашено процентов": None,
    })

    return pd.DataFrame(rows)


def _style_workbook(writer) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="E7F1ED",
    )

    loan_sheets = {
        "Полученные займы",
        "Выданные займы",
    }

    for sheet in writer.book.worksheets:
        if sheet.title in loan_sheets:
            # Закрепляем первые три колонки и строку заголовков.
            sheet.freeze_panes = "D2"
        elif sheet.title == "Саммари":
            sheet.freeze_panes = None
        else:
            sheet.freeze_panes = "B2"

        sheet.sheet_view.showGridLines = False

        headers = {
            cell.value: cell.column
            for cell in sheet[1]
            if cell.value is not None
        }

        contract_id_column = headers.get("ID договора")
        contracts_count_column = headers.get("Договоров")

        for cell in sheet[1]:
            cell.font = Font(
                name="Helvetica Light",
                size=10,
                bold=True,
                color="22312D",
            )
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(
                    name="Helvetica Light",
                    size=10,
                )
                cell.alignment = Alignment(
                    vertical="center",
                )

                if not isinstance(cell.value, (int, float)):
                    continue

                # ID договора и количество договоров —
                # только целые числа, без .00.
                if cell.column in {
                    contract_id_column,
                    contracts_count_column,
                }:
                    cell.number_format = "0"
                else:
                    cell.number_format = (
                        '#,##0.00;[Red]-#,##0.00'
                    )

        for index, column_cells in enumerate(
            sheet.columns,
            start=1,
        ):
            max_length = max(
                (
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                )
                for cell in column_cells
            )

            sheet.column_dimensions[
                get_column_letter(index)
            ].width = min(
                max(max_length + 2, 12),
                38,
            )


def build_excel_bytes(
    registry_df: pd.DataFrame,
    transactions_df: pd.DataFrame | None = None,
) -> bytes:
    output = BytesIO()
    work = registry_df.copy()

    direction = work.get(
        "loan_direction",
        pd.Series("unknown", index=work.index),
    ).fillna("unknown").astype(str)

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        _summary(work).to_excel(
            writer,
            sheet_name="Саммари",
            index=False,
        )

        # Лист «Не определено» больше не создаём.
        for direction_code, sheet_name in (
            ("borrowed", "Полученные займы"),
            ("issued", "Выданные займы"),
        ):
            part = work[
                direction.eq(direction_code)
            ].copy()

            _prepare_registry(part).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        if (
            transactions_df is not None
            and not transactions_df.empty
        ):
            transactions = transactions_df.copy()

            if {
                "contract_id",
                "loan_direction_label",
            }.issubset(work.columns):
                mapping = (
                    work[
                        [
                            "contract_id",
                            "loan_direction_label",
                        ]
                    ]
                    .drop_duplicates("contract_id")
                )

                if "contract_id" in transactions.columns:
                    transactions = transactions.merge(
                        mapping,
                        on="contract_id",
                        how="left",
                    )

            if "date_from" in transactions.columns:
                transactions["date_from"] = (
                    pd.to_datetime(
                        transactions["date_from"],
                        errors="coerce",
                    )
                    .dt.date
                )

            transactions.to_excel(
                writer,
                sheet_name="Операции",
                index=False,
            )

        _style_workbook(writer)

    output.seek(0)
    return output.getvalue()