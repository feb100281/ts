# # gear/app/daily_sales/stocks/excel.py
# from io import BytesIO
# from datetime import datetime
# import re

# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.utils import get_column_letter

# from .styles import (
#     COLORS,
#     FONT_NAME,
#     FONT_NAME_BOLD,
#     THIN_BORDER,
#     HEADER_FONT,
#     BODY_FONT,
#     TITLE_FONT,
#     SUBTITLE_FONT,
#     SMALL_MUTED_FONT,
#     BUTTON_FONT,
#     CENTER,
#     LEFT,
# )


# TOC_SHEET_NAME = "Оглавление"

# SHEET_TAB_COLORS = [
#     "2F6656",
#     "4F7F70",
#     "7A9E92",
#     "A33A3A",
#     "8C6A3F",
#     "5B6F8C",
#     "6F5B8C",
# ]


# # ---------------------------------------------------------------------
# # Helpers
# # ---------------------------------------------------------------------

# def _safe_sheet_name(name: str) -> str:
#     name = str(name or "Без названия").strip()
#     name = re.sub(r"[\\/*?:\[\]]", " ", name)
#     name = re.sub(r"\s+", " ", name).strip()
#     return name[:31] or "Без названия"


# def _fmt_date(value) -> str:
#     if value is None or pd.isna(value):
#         return ""
#     return pd.to_datetime(value).strftime("%d.%m.%Y")


# def _first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
#     for col in candidates:
#         if col in df.columns:
#             return col
#     return None


# def _sum_col(df: pd.DataFrame, col_name: str):
#     if col_name not in df.columns:
#         return 0
#     return pd.to_numeric(df[col_name], errors="coerce").fillna(0).sum()


# def _safe_div(num, den):
#     if den is None or pd.isna(den) or den == 0:
#         return None
#     return num / den


# # ---------------------------------------------------------------------
# # Подготовка данных
# # ---------------------------------------------------------------------

# def _prepare_stocks_df(
#     df: pd.DataFrame,
#     report_date=None,
# ) -> pd.DataFrame:
#     """
#     Подготавливает детальную таблицу остатков.

#     Ожидаемые поля из data.py могут называться как по-русски, так и raw-именами.

#     Основная логика:
#     - last_price = текущая розничная цена;
#     - розничная стоимость остатка = Итого количество * текущая розничная цена;
#     - оборачиваемость за 7 дней = запас в днях:
#           Итого количество / (Продажи за 7 дней / 7)
#     - возраст товара = дата отчёта - дата последнего прихода.

#     Если data.py уже передал готовые поля:
#     - "Оборачиваемость 7 дней"
#     - "Возраст товара, дней"
#     они не пересчитываются повторно.
#     """
#     df = df.copy()
#     report_date = pd.to_datetime(
#         report_date or datetime.today().date()
#     ).normalize()

#     # --------------------------------------------------------------
#     # Нормализуем названия новых полей
#     # --------------------------------------------------------------
#     rename_map = {}

#     aliases = {
#         "Последняя наша розничная цена": [
#             "last_price",
#             "Last Price",
#             "Последняя розничная цена",
#             "Розничная цена",
#         ],
#         "Продажи за 7 дней": [
#             "sales_7d",
#             "sold_7d",
#             "quantity_sold_7d",
#             "Продано за 7 дней",
#         ],
#         "Дата последнего прихода": [
#             "last_income_date",
#             "last_receipt_date",
#             "last_supply_date",
#             "Последний приход",
#         ],
#         "Оборачиваемость 7 дней": [
#             "turnover_7d",
#             "turnover_days_7d",
#             "Оборачиваемость за 7 дней",
#         ],
#         "Возраст товара, дней": [
#             "product_age_days",
#             "age_days",
#             "Возраст, дней",
#         ],
#     }

#     for target, candidates in aliases.items():
#         if target in df.columns:
#             continue

#         source = _first_existing_column(df, candidates)
#         if source:
#             rename_map[source] = target

#     if rename_map:
#         df = df.rename(columns=rename_map)

#     if "Дата" in df.columns:
#         df = df.drop(columns=["Дата"])

#     # --------------------------------------------------------------
#     # Денежные поля себестоимости приходят в копейках
#     # --------------------------------------------------------------
#     cost_money_cols = [
#         "Бух. с/с за ед.",
#         "Упр. с/с за ед.",
#         "Бух. с/с всего",
#         "Упр. с/с всего",
#     ]

#     for col in cost_money_cols:
#         if col in df.columns:
#             df[col] = (
#                 pd.to_numeric(df[col], errors="coerce")
#                 .fillna(0)
#                 / 100
#             )

#     # last_price — розничная цена.
#     # Предполагаем, что она уже приходит в рублях.
#     if "Последняя наша розничная цена" in df.columns:
#         df["Последняя наша розничная цена"] = (
#             pd.to_numeric(
#                 df["Последняя наша розничная цена"],
#                 errors="coerce",
#             )
#             .fillna(0)
#         )

#     # --------------------------------------------------------------
#     # Дельта бухгалтерской / управленческой себестоимости
#     # --------------------------------------------------------------
#     if {"Бух. с/с за ед.", "Упр. с/с за ед."}.issubset(df.columns):
#         df["Δ с/с за ед."] = (
#             df["Упр. с/с за ед."] - df["Бух. с/с за ед."]
#         )

#         df["Δ с/с за ед., %"] = df.apply(
#             lambda x: (
#                 x["Δ с/с за ед."] / x["Бух. с/с за ед."]
#                 if x["Бух. с/с за ед."]
#                 else 0
#             ),
#             axis=1,
#         )

#     # --------------------------------------------------------------
#     # Остатки в текущих розничных ценах
#     # --------------------------------------------------------------
#     if (
#         "Итого количество" in df.columns
#         and "Последняя наша розничная цена" in df.columns
#     ):
#         qty = pd.to_numeric(
#             df["Итого количество"],
#             errors="coerce",
#         ).fillna(0)

#         retail_price = pd.to_numeric(
#             df["Последняя наша розничная цена"],
#             errors="coerce",
#         ).fillna(0)

#         df["Остатки в розничных ценах"] = qty * retail_price

#     # --------------------------------------------------------------
#     # Контроль убыточной розничной цены
#     # Если управленческая себестоимость ВЫШЕ последней нашей
#     # розничной цены, товар продаётся ниже управленческой себестоимости.
#     # --------------------------------------------------------------
#     if {"Упр. с/с за ед.", "Последняя наша розничная цена"}.issubset(df.columns):
#         man_cost = pd.to_numeric(
#             df["Упр. с/с за ед."],
#             errors="coerce",
#         ).fillna(0)

#         retail_price = pd.to_numeric(
#             df["Последняя наша розничная цена"],
#             errors="coerce",
#         ).fillna(0)

#         df["Упр. с/с выше розничной цены"] = (
#             (retail_price > 0)
#             & (man_cost > retail_price)
#         )

#         # Положительная дельта = управленческая себестоимость
#         # выше розничной цены, то есть потенциально убыточная продажа.
#         df["Δ упр. с/с к розничной цене"] = man_cost - retail_price

#     # --------------------------------------------------------------
#     # Продажи за последние 7 дней
#     # --------------------------------------------------------------
#     if "Продажи за 7 дней" in df.columns:
#         df["Продажи за 7 дней"] = pd.to_numeric(
#             df["Продажи за 7 дней"],
#             errors="coerce",
#         ).fillna(0)

#     # --------------------------------------------------------------
#     # Оборачиваемость за 7 дней в днях запаса
#     #
#     # Пример:
#     # остаток 70 шт, за 7 дней продано 35 шт
#     # средние продажи в день = 5 шт
#     # оборачиваемость = 70 / 5 = 14 дней
#     # --------------------------------------------------------------
#     if (
#         "Оборачиваемость 7 дней" not in df.columns
#         and "Итого количество" in df.columns
#         and "Продажи за 7 дней" in df.columns
#     ):
#         qty = pd.to_numeric(
#             df["Итого количество"],
#             errors="coerce",
#         ).fillna(0)

#         sold_7d = pd.to_numeric(
#             df["Продажи за 7 дней"],
#             errors="coerce",
#         ).fillna(0)

#         avg_daily_sales = sold_7d / 7

#         df["Оборачиваемость 7 дней"] = [
#             _safe_div(stock, daily_sales)
#             for stock, daily_sales in zip(qty, avg_daily_sales)
#         ]

#     if "Оборачиваемость 7 дней" in df.columns:
#         df["Оборачиваемость 7 дней"] = pd.to_numeric(
#             df["Оборачиваемость 7 дней"],
#             errors="coerce",
#         )

#     # --------------------------------------------------------------
#     # Последний приход и возраст товара
#     # --------------------------------------------------------------
#     if "Дата последнего прихода" in df.columns:
#         df["Дата последнего прихода"] = pd.to_datetime(
#             df["Дата последнего прихода"],
#             errors="coerce",
#         )

#     if (
#         "Возраст товара, дней" not in df.columns
#         and "Дата последнего прихода" in df.columns
#     ):
#         df["Возраст товара, дней"] = (
#             report_date - df["Дата последнего прихода"].dt.normalize()
#         ).dt.days

#         # На случай некорректной даты прихода из будущего
#         df.loc[
#             df["Возраст товара, дней"] < 0,
#             "Возраст товара, дней",
#         ] = 0

#     if "Возраст товара, дней" in df.columns:
#         df["Возраст товара, дней"] = pd.to_numeric(
#             df["Возраст товара, дней"],
#             errors="coerce",
#         )

#     # --------------------------------------------------------------
#     # ID как текст, чтобы Excel не переводил длинные ID в scientific
#     # --------------------------------------------------------------
#     for col in ["USK", "NM ID", "Chrt ID"]:
#         if col in df.columns:
#             df[col] = df[col].astype("string").fillna("")

#     # --------------------------------------------------------------
#     # Красивый и логичный порядок колонок
#     # --------------------------------------------------------------
#     preferred_order = [
#         "USK",
#         "Бренд",
#         "Категория",
#         "Пол",
#         "Артикул",
#         "Наименование",
#         "Размер",

#         # Остатки
#         "Итого количество",
#         "Остаток на складе",
#         "В пути от клиента",
#         "В пути к клиенту",

#         # Розница
#         "Последняя наша розничная цена",
#         "Упр. с/с за ед.",
#         "Δ упр. с/с к розничной цене",
#         "Упр. с/с выше розничной цены",
#         "Остатки в розничных ценах",

#         # Продажи / оборачиваемость
#         "Продажи за 7 дней",
#         "Оборачиваемость 7 дней",

#         # Возраст товара
#         "Дата последнего прихода",
#         "Возраст товара, дней",

#         # Себестоимость
#         "Бух. с/с за ед.",
#         "Δ с/с за ед.",
#         "Δ с/с за ед., %",
#         "Бух. с/с всего",
#         "Упр. с/с всего",

#         # Технические ID в конце
#         "NM ID",
#         "Chrt ID",
#     ]

#     existing_order = [
#         col for col in preferred_order
#         if col in df.columns
#     ]
#     other_cols = [
#         col for col in df.columns
#         if col not in existing_order
#     ]

#     df = df[existing_order + other_cols]

#     sort_cols = [
#         col
#         for col in [
#             "Бренд",
#             "Категория",
#             "Наименование",
#             "Размер",
#         ]
#         if col in df.columns
#     ]

#     if sort_cols:
#         df = df.sort_values(
#             sort_cols,
#             na_position="last",
#         )

#     # --------------------------------------------------------------
#     # Продажи и оборачиваемость считаются на уровне USK
#     #
#     # Один USK может быть представлен несколькими строками:
#     # - разные размеры;
#     # - разные Chrt ID.
#     #
#     # Поэтому значения "Продажи за 7 дней" и
#     # "Оборачиваемость 7 дней" показываем только
#     # в первой строке каждого USK.
#     #
#     # Это важно не только визуально:
#     # так продажи не будут повторно суммироваться
#     # в сводках по категориям и брендам.
#     # --------------------------------------------------------------
#     if "USK" in df.columns:
#         usk_values = (
#             df["USK"]
#             .astype("string")
#             .fillna("")
#             .str.strip()
#         )

#         duplicate_usk_mask = (
#             usk_values.ne("")
#             & usk_values.duplicated(
#                 keep="first"
#             )
#         )

#         # Продажи одного USK показываем только один раз.
#         if "Продажи за 7 дней" in df.columns:
#             df.loc[
#                 duplicate_usk_mask,
#                 "Продажи за 7 дней",
#             ] = pd.NA

#         # Оборачиваемость одного USK показываем только один раз.
#         if "Оборачиваемость 7 дней" in df.columns:
#             df.loc[
#                 duplicate_usk_mask,
#                 "Оборачиваемость 7 дней",
#             ] = pd.NA

#     return df


# # ---------------------------------------------------------------------
# # Excel styles
# # ---------------------------------------------------------------------

# def _autosize_columns(ws, min_width=10, max_width=45):
#     for col_idx in range(1, ws.max_column + 1):
#         letter = get_column_letter(col_idx)
#         max_len = 0

#         for row_idx in range(1, ws.max_row + 1):
#             value = ws.cell(
#                 row=row_idx,
#                 column=col_idx,
#             ).value

#             if value is not None:
#                 max_len = max(
#                     max_len,
#                     len(str(value)),
#                 )

#         ws.column_dimensions[letter].width = max(
#             min_width,
#             min(max_len + 2, max_width),
#         )


# def _add_back_button(ws, last_col):
#     if ws.title == TOC_SHEET_NAME:
#         return

#     btn_last_col = min(3, last_col)

#     ws.merge_cells(
#         start_row=1,
#         start_column=1,
#         end_row=1,
#         end_column=btn_last_col,
#     )

#     cell = ws.cell(
#         row=1,
#         column=1,
#         value="← ОГЛАВЛЕНИЕ",
#     )
#     cell.hyperlink = f"#'{TOC_SHEET_NAME}'!A1"
#     cell.font = BUTTON_FONT
#     cell.fill = PatternFill(
#         "solid",
#         fgColor=COLORS["success"],
#     )
#     cell.alignment = CENTER
#     cell.border = THIN_BORDER

#     ws.row_dimensions[1].height = 24


# def _style_title(
#     ws,
#     title,
#     subtitle,
#     report_date,
#     last_col,
# ):
#     ws.sheet_view.showGridLines = False

#     _add_back_button(
#         ws,
#         last_col,
#     )

#     ws.merge_cells(
#         start_row=3,
#         start_column=1,
#         end_row=3,
#         end_column=last_col,
#     )
#     cell = ws.cell(
#         row=3,
#         column=1,
#         value=title,
#     )
#     cell.font = TITLE_FONT
#     cell.alignment = LEFT
#     ws.row_dimensions[3].height = 30

#     ws.merge_cells(
#         start_row=4,
#         start_column=1,
#         end_row=4,
#         end_column=last_col,
#     )
#     cell = ws.cell(
#         row=4,
#         column=1,
#         value=(
#             f"{subtitle} · "
#             f"дата остатков: {_fmt_date(report_date)}"
#         ),
#     )
#     cell.font = SUBTITLE_FONT
#     cell.alignment = LEFT
#     ws.row_dimensions[4].height = 22

#     ws.merge_cells(
#         start_row=5,
#         start_column=1,
#         end_row=5,
#         end_column=last_col,
#     )
#     cell = ws.cell(
#         row=5,
#         column=1,
#         value=(
#             "Сформировано: "
#             f"{datetime.now().strftime('%d.%m.%Y в %H:%M')}"
#         ),
#     )
#     cell.font = SMALL_MUTED_FONT
#     cell.alignment = LEFT
#     ws.row_dimensions[5].height = 20


# def _add_sheet_summary_cards(
#     ws,
#     df,
#     start_row,
#     last_col,
# ):
#     total_qty = _sum_col(
#         df,
#         "Итого количество",
#     )
#     total_retail = _sum_col(
#         df,
#         "Остатки в розничных ценах",
#     )
#     total_buh = _sum_col(
#         df,
#         "Бух. с/с всего",
#     )
#     total_man = _sum_col(
#         df,
#         "Упр. с/с всего",
#     )

#     cards = [
#         ("Строк", len(df), "SKU"),
#         ("Количество", total_qty, "шт"),
#         ("Розничная стоимость", total_retail, "₽"),
#         ("Бух. стоимость", total_buh, "₽"),
#         ("Упр. стоимость", total_man, "₽"),
#     ]

#     max_cards = min(
#         len(cards),
#         max(1, last_col // 2),
#     )

#     row = start_row
#     col = 1

#     for idx, (title, value, subtitle) in enumerate(
#         cards[:max_cards]
#     ):
#         c1 = col + idx * 2
#         c2 = c1 + 1

#         ws.merge_cells(
#             start_row=row,
#             start_column=c1,
#             end_row=row,
#             end_column=c2,
#         )

#         title_cell = ws.cell(
#             row=row,
#             column=c1,
#             value=title,
#         )
#         title_cell.font = Font(
#             name=FONT_NAME_BOLD,
#             size=9,
#             bold=True,
#             color=COLORS["muted"],
#         )
#         title_cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_green"],
#         )
#         title_cell.alignment = CENTER

#         ws.merge_cells(
#             start_row=row + 1,
#             start_column=c1,
#             end_row=row + 1,
#             end_column=c2,
#         )

#         value_cell = ws.cell(
#             row=row + 1,
#             column=c1,
#             value=value,
#         )
#         value_cell.font = Font(
#             name=FONT_NAME_BOLD,
#             size=13,
#             bold=True,
#             color=COLORS["dark_green"],
#         )
#         value_cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_gray"],
#         )
#         value_cell.alignment = CENTER

#         if "стоимость" in title.lower():
#             value_cell.number_format = '#,##0.00 ₽'
#         else:
#             value_cell.number_format = "#,##0"

#         ws.merge_cells(
#             start_row=row + 2,
#             start_column=c1,
#             end_row=row + 2,
#             end_column=c2,
#         )

#         sub_cell = ws.cell(
#             row=row + 2,
#             column=c1,
#             value=subtitle,
#         )
#         sub_cell.font = Font(
#             name=FONT_NAME,
#             size=8,
#             color=COLORS["muted"],
#         )
#         sub_cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_gray"],
#         )
#         sub_cell.alignment = CENTER

#         for rr in range(row, row + 3):
#             for cc in range(c1, c2 + 1):
#                 ws.cell(
#                     rr,
#                     cc,
#                 ).border = THIN_BORDER

#     return start_row + 5


# def _style_body_cell(
#     cell,
#     col_name,
#     row_idx,
# ):
#     cell.font = BODY_FONT
#     cell.alignment = LEFT
#     cell.border = THIN_BORDER

#     if col_name in [
#         "USK",
#         "NM ID",
#         "Chrt ID",
#     ]:
#         cell.number_format = "@"
#         cell.alignment = Alignment(
#             horizontal="left",
#             vertical="center",
#         )

#     elif col_name in [
#         "Итого количество",
#         "Остаток на складе",
#         "В пути от клиента",
#         "В пути к клиенту",
#         "Продажи за 7 дней",
#     ]:
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["qty"],
#         )
#         cell.number_format = "#,##0"
#         cell.alignment = Alignment(
#             horizontal="right",
#             vertical="center",
#         )

#     elif col_name in [
#         "Последняя наша розничная цена",
#         "Остатки в розничных ценах",
#         "Δ упр. с/с к розничной цене",
#         "Бух. с/с за ед.",
#         "Упр. с/с за ед.",
#         "Бух. с/с всего",
#         "Упр. с/с всего",
#         "Δ с/с за ед.",
#     ]:
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["money"],
#         )
#         cell.number_format = '#,##0.00 ₽'
#         cell.alignment = Alignment(
#             horizontal="right",
#             vertical="center",
#         )

#     elif col_name == "Упр. с/с выше розничной цены":
#         is_loss = bool(cell.value)
#         cell.value = "ДА" if is_loss else "Нет"
#         cell.alignment = CENTER
#         cell.font = Font(
#             name=FONT_NAME_BOLD,
#             size=10,
#             bold=True,
#             color=(
#                 COLORS["discount"]
#                 if is_loss
#                 else COLORS["dark_green"]
#             ),
#         )
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=(
#                 COLORS["warning"]
#                 if is_loss
#                 else COLORS["success"]
#             ),
#         )

#     elif col_name == "Δ с/с за ед., %":
#         value = cell.value or 0

#         cell.fill = PatternFill(
#             "solid",
#             fgColor=(
#                 COLORS["warning"]
#                 if value > 0
#                 else COLORS["success"]
#             ),
#         )
#         cell.font = Font(
#             name=FONT_NAME_BOLD,
#             size=10,
#             bold=True,
#             color=(
#                 COLORS["discount"]
#                 if value > 0
#                 else COLORS["dark_green"]
#             ),
#         )
#         cell.number_format = "0.00%"
#         cell.alignment = Alignment(
#             horizontal="right",
#             vertical="center",
#         )

#     elif col_name == "Оборачиваемость 7 дней":
#         value = cell.value

#         if value is None:
#             cell.value = "Нет продаж"
#             cell.fill = PatternFill(
#                 "solid",
#                 fgColor=COLORS["warning"],
#             )
#             cell.alignment = CENTER
#         else:
#             cell.number_format = '0.0 "дн."'
#             cell.alignment = Alignment(
#                 horizontal="right",
#                 vertical="center",
#             )

#             # Чем больше дней запаса — тем внимательнее надо смотреть товар
#             if value >= 90:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["warning"],
#                 )
#             elif value <= 30:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["success"],
#                 )
#             else:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["light_green"],
#                 )

#     elif col_name == "Дата последнего прихода":
#         if cell.value is not None and not pd.isna(cell.value):
#             cell.value = pd.to_datetime(
#                 cell.value,
#             ).to_pydatetime()
#             cell.number_format = "dd.mm.yyyy"

#         cell.alignment = CENTER
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_green"],
#         )

#     elif col_name == "Возраст товара, дней":
#         value = cell.value

#         cell.number_format = '0 "дн."'
#         cell.alignment = Alignment(
#             horizontal="right",
#             vertical="center",
#         )

#         if value is not None and not pd.isna(value):
#             if value >= 365:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["warning"],
#                 )
#             elif value <= 90:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["success"],
#                 )
#             else:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["light_green"],
#                 )

#     elif row_idx % 2 == 0:
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_gray"],
#         )


# # ---------------------------------------------------------------------
# # Детальные листы
# # ---------------------------------------------------------------------

# def _write_dataframe_sheet(
#     wb,
#     sheet_name,
#     df,
#     report_date,
#     title,
#     subtitle,
# ):
#     ws = wb.create_sheet(
#         _safe_sheet_name(sheet_name)
#     )

#     df = df.copy()
#     last_col = max(
#         len(df.columns),
#         1,
#     )

#     _style_title(
#         ws,
#         title,
#         subtitle,
#         report_date,
#         last_col,
#     )

#     header_row = _add_sheet_summary_cards(
#         ws=ws,
#         df=df,
#         start_row=7,
#         last_col=last_col,
#     )

#     for col_idx, col_name in enumerate(
#         df.columns,
#         start=1,
#     ):
#         cell = ws.cell(
#             row=header_row,
#             column=col_idx,
#             value=col_name,
#         )
#         cell.font = HEADER_FONT
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["dark_green"],
#         )
#         cell.alignment = CENTER
#         cell.border = THIN_BORDER

#     for row_idx, row in enumerate(
#         df.itertuples(index=False),
#         start=header_row + 1,
#     ):
#         ws.row_dimensions[row_idx].height = 20

#         for col_idx, value in enumerate(
#             row,
#             start=1,
#         ):
#             col_name = df.columns[
#                 col_idx - 1
#             ]

#             # pandas NaN / NaT в Excel лучше писать как None
#             if pd.isna(value):
#                 value = None

#             cell = ws.cell(
#                 row=row_idx,
#                 column=col_idx,
#                 value=value,
#             )

#             _style_body_cell(
#                 cell,
#                 col_name,
#                 row_idx,
#             )

#         # Подсветка контроля: если управленческая себестоимость
#         # выше последней нашей розничной цены, выделяем ключевые ячейки красным.
#         if {
#             "Упр. с/с за ед.",
#             "Последняя наша розничная цена",
#             "Упр. с/с выше розничной цены",
#         }.issubset(df.columns):
#             risk_col_idx = df.columns.get_loc("Упр. с/с выше розничной цены") + 1
#             risk_value = ws.cell(row=row_idx, column=risk_col_idx).value

#             if risk_value == "ДА":
#                 for risk_col_name in [
#                     "Последняя наша розничная цена",
#                     "Упр. с/с за ед.",
#                     "Δ упр. с/с к розничной цене",
#                     "Упр. с/с выше розничной цены",
#                 ]:
#                     if risk_col_name not in df.columns:
#                         continue

#                     risk_cell = ws.cell(
#                         row=row_idx,
#                         column=df.columns.get_loc(risk_col_name) + 1,
#                     )
#                     risk_cell.fill = PatternFill(
#                         "solid",
#                         fgColor=COLORS["warning"],
#                     )
#                     risk_cell.font = Font(
#                         name=FONT_NAME_BOLD,
#                         size=10,
#                         bold=True,
#                         color=COLORS["discount"],
#                     )

#     last_row = (
#         header_row + len(df)
#     )

#     # Замораживаем строку заголовка и первые 3 колонки
#     freeze_col = 4
#     ws.freeze_panes = (
#         f"{get_column_letter(freeze_col)}"
#         f"{header_row + 1}"
#     )

#     if len(df) > 0:
#         ws.auto_filter.ref = (
#             f"A{header_row}:"
#             f"{get_column_letter(last_col)}"
#             f"{last_row}"
#         )

#     # Детализацию по движению WB прячем в outline-группу
#     detail_cols = [
#         "Остаток на складе",
#         "В пути от клиента",
#         "В пути к клиенту",
#     ]

#     detail_indexes = [
#         df.columns.get_loc(col) + 1
#         for col in detail_cols
#         if col in df.columns
#     ]

#     if detail_indexes:
#         first = min(
#             detail_indexes
#         )
#         last = max(
#             detail_indexes
#         )

#         for col_idx in range(
#             first,
#             last + 1,
#         ):
#             letter = get_column_letter(
#                 col_idx
#             )
#             ws.column_dimensions[
#                 letter
#             ].outlineLevel = 1
#             ws.column_dimensions[
#                 letter
#             ].hidden = True

#         ws.sheet_properties.outlinePr.summaryRight = False

#     _autosize_columns(
#         ws
#     )

#     widths = {
#         "USK": 16,
#         "Бренд": 20,
#         "Категория": 22,
#         "Пол": 12,
#         "Артикул": 18,
#         "Наименование": 42,
#         "Размер": 14,

#         "Итого количество": 15,
#         "Остаток на складе": 15,
#         "В пути от клиента": 15,
#         "В пути к клиенту": 15,

#         "Последняя наша розничная цена": 24,
#         "Δ упр. с/с к розничной цене": 24,
#         "Упр. с/с выше розничной цены": 25,
#         "Остатки в розничных ценах": 24,

#         "Продажи за 7 дней": 17,
#         "Оборачиваемость 7 дней": 22,

#         "Дата последнего прихода": 21,
#         "Возраст товара, дней": 20,

#         "Бух. с/с за ед.": 16,
#         "Упр. с/с за ед.": 16,
#         "Δ с/с за ед.": 16,
#         "Δ с/с за ед., %": 16,
#         "Бух. с/с всего": 17,
#         "Упр. с/с всего": 17,

#         "NM ID": 16,
#         "Chrt ID": 16,
#     }

#     for col_name, width in widths.items():
#         if col_name in df.columns:
#             letter = get_column_letter(
#                 df.columns.get_loc(
#                     col_name
#                 ) + 1
#             )
#             ws.column_dimensions[
#                 letter
#             ].width = width

#     return ws.title


# # ---------------------------------------------------------------------
# # Сводка
# # ---------------------------------------------------------------------

# def _build_summary_sheet(
#     wb,
#     df,
#     report_date,
# ):
#     ws = wb.create_sheet(
#         "Сводка"
#     )
#     ws.sheet_view.showGridLines = False

#     last_col = 10

#     _add_back_button(
#         ws,
#         last_col,
#     )

#     total_qty = _sum_col(
#         df,
#         "Итого количество",
#     )
#     on_hand = _sum_col(
#         df,
#         "Остаток на складе",
#     )
#     in_way_client = _sum_col(
#         df,
#         "В пути к клиенту",
#     )
#     in_way_from = _sum_col(
#         df,
#         "В пути от клиента",
#     )
#     total_retail = _sum_col(
#         df,
#         "Остатки в розничных ценах",
#     )
#     total_buh = _sum_col(
#         df,
#         "Бух. с/с всего",
#     )
#     total_man = _sum_col(
#         df,
#         "Упр. с/с всего",
#     )

#     ws.merge_cells(
#         "A3:J3"
#     )
#     cell = ws["A3"]
#     cell.value = (
#         "ОТЧЕТ ПО ОСТАТКАМ ТОВАРОВ"
#     )
#     cell.font = Font(
#         name=FONT_NAME_BOLD,
#         size=18,
#         bold=True,
#         color=COLORS["dark_green"],
#     )
#     cell.alignment = LEFT

#     ws.merge_cells(
#         "A4:J4"
#     )
#     cell = ws["A4"]
#     cell.value = (
#         f"Дата остатков: "
#         f"{_fmt_date(report_date)}"
#     )
#     cell.font = Font(
#         name=FONT_NAME,
#         size=11,
#         bold=True,
#         color=COLORS["muted"],
#     )

#     ws.merge_cells(
#         "A5:J5"
#     )
#     cell = ws["A5"]
#     cell.value = (
#         "Сформировано: "
#         f"{datetime.now().strftime('%d.%m.%Y в %H:%M')}"
#     )
#     cell.font = SMALL_MUTED_FONT

#     cards = [
#         ("SKU / строк", len(df), "позиций в выгрузке"),
#         ("Итого количество", total_qty, "шт"),
#         ("На складе", on_hand, "шт"),
#         ("В пути к клиенту", in_way_client, "шт"),
#         ("В пути от клиента", in_way_from, "шт"),
#         ("Розничная стоимость", total_retail, "в текущих розничных ценах"),
#         ("Бух. стоимость", total_buh, "₽"),
#         ("Упр. стоимость", total_man, "₽"),
#     ]

#     row = 7

#     for idx, (
#         title,
#         value,
#         subtitle,
#     ) in enumerate(cards):
#         c1 = 1 + (
#             idx % 4
#         ) * 2
#         r1 = row + (
#             idx // 4
#         ) * 4

#         ws.merge_cells(
#             start_row=r1,
#             start_column=c1,
#             end_row=r1,
#             end_column=c1 + 1,
#         )

#         title_cell = ws.cell(
#             r1,
#             c1,
#             title,
#         )
#         title_cell.font = Font(
#             name=FONT_NAME_BOLD,
#             size=9,
#             bold=True,
#             color=COLORS["muted"],
#         )
#         title_cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_green"],
#         )
#         title_cell.alignment = CENTER

#         ws.merge_cells(
#             start_row=r1 + 1,
#             start_column=c1,
#             end_row=r1 + 1,
#             end_column=c1 + 1,
#         )

#         value_cell = ws.cell(
#             r1 + 1,
#             c1,
#             value,
#         )
#         value_cell.font = Font(
#             name=FONT_NAME_BOLD,
#             size=14,
#             bold=True,
#             color=COLORS["dark_green"],
#         )
#         value_cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_gray"],
#         )
#         value_cell.alignment = CENTER

#         if "стоимость" in title.lower():
#             value_cell.number_format = '#,##0.00 ₽'
#         else:
#             value_cell.number_format = "#,##0"

#         ws.merge_cells(
#             start_row=r1 + 2,
#             start_column=c1,
#             end_row=r1 + 2,
#             end_column=c1 + 1,
#         )

#         sub_cell = ws.cell(
#             r1 + 2,
#             c1,
#             subtitle,
#         )
#         sub_cell.font = Font(
#             name=FONT_NAME,
#             size=8,
#             color=COLORS["muted"],
#         )
#         sub_cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["light_gray"],
#         )
#         sub_cell.alignment = CENTER

#         for rr in range(
#             r1,
#             r1 + 3,
#         ):
#             for cc in range(
#                 c1,
#                 c1 + 2,
#             ):
#                 ws.cell(
#                     rr,
#                     cc,
#                 ).border = THIN_BORDER

#     ws.column_dimensions[
#         "A"
#     ].width = 18

#     for col_letter in [
#         "B",
#         "C",
#         "D",
#         "E",
#         "F",
#         "G",
#         "H",
#         "I",
#         "J",
#     ]:
#         ws.column_dimensions[
#             col_letter
#         ].width = 16

#     return ws.title


# # ---------------------------------------------------------------------
# # Оглавление
# # ---------------------------------------------------------------------

# def _build_toc_sheet(
#     wb,
#     sheets_info,
#     report_date,
# ):
#     ws = wb.create_sheet(
#         TOC_SHEET_NAME,
#         0,
#     )
#     ws.sheet_view.showGridLines = False

#     ws.merge_cells(
#         "B2:D2"
#     )
#     cell = ws["B2"]
#     cell.value = (
#         "ОГЛАВЛЕНИЕ ОТЧЕТА ПО ОСТАТКАМ"
#     )
#     cell.font = Font(
#         name=FONT_NAME_BOLD,
#         size=18,
#         bold=True,
#         color=COLORS["dark_green"],
#     )

#     ws.merge_cells(
#         "B3:D3"
#     )
#     cell = ws["B3"]
#     cell.value = (
#         f"Дата остатков: "
#         f"{_fmt_date(report_date)}"
#     )
#     cell.font = Font(
#         name=FONT_NAME,
#         size=11,
#         bold=True,
#         color=COLORS["muted"],
#     )

#     ws.merge_cells(
#         "B4:D4"
#     )
#     cell = ws["B4"]
#     cell.value = (
#         "Для перехода к листу нажмите на название раздела"
#     )
#     cell.font = SMALL_MUTED_FONT

#     headers = [
#         "№",
#         "Лист",
#         "Описание",
#     ]
#     start_row = 7

#     for col_idx, header in enumerate(
#         headers,
#         start=2,
#     ):
#         cell = ws.cell(
#             start_row,
#             col_idx,
#             header,
#         )
#         cell.font = HEADER_FONT
#         cell.fill = PatternFill(
#             "solid",
#             fgColor=COLORS["dark_green"],
#         )
#         cell.alignment = CENTER
#         cell.border = THIN_BORDER

#     row = (
#         start_row + 1
#     )

#     for idx, item in enumerate(
#         sheets_info,
#         start=1,
#     ):
#         sheet_name = item[
#             "name"
#         ]

#         values = [
#             f"{idx:02d}",
#             sheet_name,
#             item.get(
#                 "description",
#                 "",
#             ),
#         ]

#         for col_idx, value in enumerate(
#             values,
#             start=2,
#         ):
#             cell = ws.cell(
#                 row,
#                 col_idx,
#                 value,
#             )
#             cell.font = BODY_FONT
#             cell.border = THIN_BORDER
#             cell.alignment = LEFT

#             if col_idx == 3:
#                 cell.font = Font(
#                     name=FONT_NAME_BOLD,
#                     size=10,
#                     bold=True,
#                     color=COLORS["link"],
#                 )
#                 cell.hyperlink = (
#                     f"#'{sheet_name}'!A1"
#                 )

#             if idx % 2 == 0:
#                 cell.fill = PatternFill(
#                     "solid",
#                     fgColor=COLORS["light_gray"],
#                 )

#         ws.row_dimensions[
#             row
#         ].height = 26
#         row += 1

#     ws.column_dimensions[
#         "A"
#     ].width = 3
#     ws.column_dimensions[
#         "B"
#     ].width = 8
#     ws.column_dimensions[
#         "C"
#     ].width = 34
#     ws.column_dimensions[
#         "D"
#     ].width = 70


# # ---------------------------------------------------------------------
# # Сводка по категориям
# # ---------------------------------------------------------------------

# def _build_category_summary(
#     wb,
#     df,
#     report_date,
# ):
#     group_cols = [
#         col
#         for col in [
#             "Категория",
#             "Бренд",
#         ]
#         if col in df.columns
#     ]

#     if not group_cols:
#         return None

#     value_cols = [
#         col
#         for col in [
#             "Итого количество",
#             "Остаток на складе",
#             "В пути от клиента",
#             "В пути к клиенту",
#             "Продажи за 7 дней",
#             "Остатки в розничных ценах",
#             "Бух. с/с всего",
#             "Упр. с/с всего",
#         ]
#         if col in df.columns
#     ]

#     summary = (
#         df.groupby(
#             group_cols,
#             dropna=False,
#         )[value_cols]
#         .sum()
#         .reset_index()
#         .sort_values(
#             group_cols
#         )
#     )

#     return _write_dataframe_sheet(
#         wb=wb,
#         sheet_name="По категориям",
#         df=summary,
#         report_date=report_date,
#         title="Сводка остатков по категориям",
#         subtitle=(
#             "Количество, продажи за 7 дней, "
#             "розничная и себестоимостная оценка"
#         ),
#     )


# def _apply_sheet_tab_colors(
#     wb,
# ):
#     for idx, ws in enumerate(
#         wb.worksheets
#     ):
#         ws.sheet_properties.tabColor = (
#             SHEET_TAB_COLORS[
#                 idx % len(
#                     SHEET_TAB_COLORS
#                 )
#             ]
#         )


# # ---------------------------------------------------------------------
# # Public API
# # ---------------------------------------------------------------------

# def make_stocks_excel(
#     df: pd.DataFrame,
#     report_date=None,
# ) -> bytes:
#     report_date = (
#         report_date
#         or datetime.today().date()
#     )

#     df = _prepare_stocks_df(
#         df,
#         report_date=report_date,
#     )

#     wb = Workbook()
#     default_ws = wb.active
#     wb.remove(
#         default_ws
#     )

#     sheets_info = []

#     summary_name = _build_summary_sheet(
#         wb,
#         df,
#         report_date,
#     )
#     sheets_info.append(
#         {
#             "name": summary_name,
#             "description": (
#                 "Ключевые показатели по количеству, "
#                 "розничной стоимости и себестоимости"
#             ),
#         }
#     )

#     all_name = _write_dataframe_sheet(
#         wb=wb,
#         sheet_name="Все товары",
#         df=df,
#         report_date=report_date,
#         title="Детальные остатки товаров",
#         subtitle=(
#             "Остатки, текущие розничные цены, "
#             "оборачиваемость за 7 дней и возраст товара"
#         ),
#     )
#     sheets_info.append(
#         {
#             "name": all_name,
#             "description": (
#                 "Полная детализация по всем товарам"
#             ),
#         }
#     )

#     category_name = _build_category_summary(
#         wb,
#         df,
#         report_date,
#     )

#     if category_name:
#         sheets_info.append(
#             {
#                 "name": category_name,
#                 "description": (
#                     "Сводка по категориям и брендам"
#                 ),
#             }
#         )

#     if "Бренд" in df.columns:
#         brands = (
#             df["Бренд"]
#             .fillna(
#                 "Бренд не указан"
#             )
#             .astype(str)
#             .sort_values()
#             .unique()
#             .tolist()
#         )

#         for brand in brands:
#             brand_df = df[
#                 df["Бренд"]
#                 .fillna(
#                     "Бренд не указан"
#                 )
#                 .astype(str)
#                 == brand
#             ]

#             sheet_name = _write_dataframe_sheet(
#                 wb=wb,
#                 sheet_name=f"Бренд_{brand}",
#                 df=brand_df,
#                 report_date=report_date,
#                 title=(
#                     f"Остатки товаров — {brand}"
#                 ),
#                 subtitle=(
#                     "Остатки, розничные цены, "
#                     "оборачиваемость и возраст товара"
#                 ),
#             )

#             sheets_info.append(
#                 {
#                     "name": sheet_name,
#                     "description": (
#                         f"Остатки по бренду {brand}"
#                     ),
#                 }
#             )

#     _build_toc_sheet(
#         wb,
#         sheets_info,
#         report_date,
#     )

#     _apply_sheet_tab_colors(
#         wb
#     )

#     output = BytesIO()
#     wb.save(
#         output
#     )
#     output.seek(
#         0
#     )

#     return output.read()



# gear/app/daily_sales/stocks/excel.py
from io import BytesIO
from datetime import datetime
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .styles import (
    COLORS,
    FONT_NAME,
    FONT_NAME_BOLD,
    THIN_BORDER,
    HEADER_FONT,
    BODY_FONT,
    TITLE_FONT,
    SUBTITLE_FONT,
    SMALL_MUTED_FONT,
    BUTTON_FONT,
    CENTER,
    LEFT,
)


TOC_SHEET_NAME = "Оглавление"

SHEET_TAB_COLORS = [
    "2F6656",
    "4F7F70",
    "7A9E92",
    "A33A3A",
    "8C6A3F",
    "5B6F8C",
    "6F5B8C",
]


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def _safe_sheet_name(name: str) -> str:
    name = str(name or "Без названия").strip()
    name = re.sub(r"[\\/*?:\[\]]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:31] or "Без названия"


def _fmt_date(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return pd.to_datetime(value).strftime("%d.%m.%Y")


def _first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for col in candidates:
        if col in df.columns:
            return col
    return None


def _sum_col(df: pd.DataFrame, col_name: str):
    if col_name not in df.columns:
        return 0
    return pd.to_numeric(df[col_name], errors="coerce").fillna(0).sum()


def _safe_div(num, den):
    if den is None or pd.isna(den) or den == 0:
        return None
    return num / den


# ---------------------------------------------------------------------
# Подготовка данных
# ---------------------------------------------------------------------

def _prepare_stocks_df(
    df: pd.DataFrame,
    report_date=None,
) -> pd.DataFrame:
    """
    Подготавливает детальную таблицу остатков.

    Ожидаемые поля из data.py могут называться как по-русски, так и raw-именами.

    Основная логика:
    - last_price = текущая розничная цена;
    - розничная стоимость остатка = Итого количество * текущая розничная цена;
    - оборачиваемость за 7 дней = запас в днях:
          Итого количество / (Продажи за 7 дней / 7)
    - возраст товара = дата отчёта - дата последнего прихода.

    Если data.py уже передал готовые поля:
    - "Оборачиваемость 7 дней"
    - "Возраст товара, дней"
    они не пересчитываются повторно.
    """
    df = df.copy()
    report_date = pd.to_datetime(
        report_date or datetime.today().date()
    ).normalize()

    # --------------------------------------------------------------
    # Нормализуем названия новых полей
    # --------------------------------------------------------------
    rename_map = {}

    aliases = {
        "Последняя наша розничная цена": [
            "last_price",
            "Last Price",
            "Последняя розничная цена",
            "Розничная цена",
        ],
        "Продажи за 7 дней": [
            "sales_7d",
            "sold_7d",
            "quantity_sold_7d",
            "Продано за 7 дней",
        ],
        "Дата последнего прихода": [
            "last_income_date",
            "last_receipt_date",
            "last_supply_date",
            "Последний приход",
        ],
        "Оборачиваемость 7 дней": [
            "turnover_7d",
            "turnover_days_7d",
            "Оборачиваемость за 7 дней",
        ],
        "Возраст товара, дней": [
            "product_age_days",
            "age_days",
            "Возраст, дней",
        ],
    }

    for target, candidates in aliases.items():
        if target in df.columns:
            continue

        source = _first_existing_column(df, candidates)
        if source:
            rename_map[source] = target

    if rename_map:
        df = df.rename(columns=rename_map)

    if "Дата" in df.columns:
        df = df.drop(columns=["Дата"])

    # --------------------------------------------------------------
    # Денежные поля себестоимости приходят в копейках
    # --------------------------------------------------------------
    cost_money_cols = [
        "Бух. с/с за ед.",
        "Упр. с/с за ед.",
        "Бух. с/с всего",
        "Упр. с/с всего",
    ]

    for col in cost_money_cols:
        if col in df.columns:
            df[col] = (
                pd.to_numeric(df[col], errors="coerce")
                .fillna(0)
                / 100
            )

    # last_price — розничная цена.
    # Предполагаем, что она уже приходит в рублях.
    if "Последняя наша розничная цена" in df.columns:
        df["Последняя наша розничная цена"] = (
            pd.to_numeric(
                df["Последняя наша розничная цена"],
                errors="coerce",
            )
            .fillna(0)
        )

    # --------------------------------------------------------------
    # Дельта бухгалтерской / управленческой себестоимости
    # --------------------------------------------------------------
    if {"Бух. с/с за ед.", "Упр. с/с за ед."}.issubset(df.columns):
        df["Δ с/с за ед."] = (
            df["Упр. с/с за ед."] - df["Бух. с/с за ед."]
        )

        df["Δ с/с за ед., %"] = df.apply(
            lambda x: (
                x["Δ с/с за ед."] / x["Бух. с/с за ед."]
                if x["Бух. с/с за ед."]
                else 0
            ),
            axis=1,
        )

    # --------------------------------------------------------------
    # Остатки в текущих розничных ценах
    # --------------------------------------------------------------
    if (
        "Итого количество" in df.columns
        and "Последняя наша розничная цена" in df.columns
    ):
        qty = pd.to_numeric(
            df["Итого количество"],
            errors="coerce",
        ).fillna(0)

        retail_price = pd.to_numeric(
            df["Последняя наша розничная цена"],
            errors="coerce",
        ).fillna(0)

        df["Остатки в розничных ценах"] = qty * retail_price

    # --------------------------------------------------------------
    # Контроль убыточной розничной цены
    # Если управленческая себестоимость ВЫШЕ последней нашей
    # розничной цены, товар продаётся ниже управленческой себестоимости.
    # --------------------------------------------------------------
    if {"Упр. с/с за ед.", "Последняя наша розничная цена"}.issubset(df.columns):
        man_cost = pd.to_numeric(
            df["Упр. с/с за ед."],
            errors="coerce",
        ).fillna(0)

        retail_price = pd.to_numeric(
            df["Последняя наша розничная цена"],
            errors="coerce",
        ).fillna(0)

        df["Упр. с/с выше розничной цены"] = (
            (retail_price > 0)
            & (man_cost > retail_price)
        )

        # Положительная дельта = управленческая себестоимость
        # выше розничной цены, то есть потенциально убыточная продажа.
        df["Δ упр. с/с к розничной цене"] = man_cost - retail_price

    # --------------------------------------------------------------
    # Продажи за последние 7 дней
    # --------------------------------------------------------------
    if "Продажи за 7 дней" in df.columns:
        df["Продажи за 7 дней"] = pd.to_numeric(
            df["Продажи за 7 дней"],
            errors="coerce",
        ).fillna(0)

    # --------------------------------------------------------------
    # Оборачиваемость за 7 дней в днях запаса
    #
    # Пример:
    # остаток 70 шт, за 7 дней продано 35 шт
    # средние продажи в день = 5 шт
    # оборачиваемость = 70 / 5 = 14 дней
    # --------------------------------------------------------------
    if (
        "Оборачиваемость 7 дней" not in df.columns
        and "Итого количество" in df.columns
        and "Продажи за 7 дней" in df.columns
    ):
        qty = pd.to_numeric(
            df["Итого количество"],
            errors="coerce",
        ).fillna(0)

        sold_7d = pd.to_numeric(
            df["Продажи за 7 дней"],
            errors="coerce",
        ).fillna(0)

        avg_daily_sales = sold_7d / 7

        df["Оборачиваемость 7 дней"] = [
            _safe_div(stock, daily_sales)
            for stock, daily_sales in zip(qty, avg_daily_sales)
        ]

    if "Оборачиваемость 7 дней" in df.columns:
        df["Оборачиваемость 7 дней"] = pd.to_numeric(
            df["Оборачиваемость 7 дней"],
            errors="coerce",
        )

    # --------------------------------------------------------------
    # Последний приход и возраст товара
    # --------------------------------------------------------------
    if "Дата последнего прихода" in df.columns:
        df["Дата последнего прихода"] = pd.to_datetime(
            df["Дата последнего прихода"],
            errors="coerce",
        )

    if (
        "Возраст товара, дней" not in df.columns
        and "Дата последнего прихода" in df.columns
    ):
        df["Возраст товара, дней"] = (
            report_date - df["Дата последнего прихода"].dt.normalize()
        ).dt.days

        # На случай некорректной даты прихода из будущего
        df.loc[
            df["Возраст товара, дней"] < 0,
            "Возраст товара, дней",
        ] = 0

    if "Возраст товара, дней" in df.columns:
        df["Возраст товара, дней"] = pd.to_numeric(
            df["Возраст товара, дней"],
            errors="coerce",
        )

    # --------------------------------------------------------------
    # ID как текст, чтобы Excel не переводил длинные ID в scientific
    # --------------------------------------------------------------
    for col in ["USK", "NM ID", "Chrt ID"]:
        if col in df.columns:
            df[col] = df[col].astype("string").fillna("")

    # --------------------------------------------------------------
    # Красивый и логичный порядок колонок
    # --------------------------------------------------------------
    preferred_order = [
        "USK",
        "Бренд",
        "Категория",
        "Пол",
        "Артикул",
        "Наименование",
        "Размер",

        # Остатки
        "Итого количество",
        "Остаток на складе",
        "В пути от клиента",
        "В пути к клиенту",

        # Розница
        "Последняя наша розничная цена",
        "Упр. с/с за ед.",
        "Δ упр. с/с к розничной цене",
        "Упр. с/с выше розничной цены",
        "Остатки в розничных ценах",

        # Продажи / оборачиваемость
        "Продажи за 7 дней",
        "Оборачиваемость 7 дней",

        # Возраст товара
        "Дата последнего прихода",
        "Возраст товара, дней",

        # Себестоимость
        "Бух. с/с за ед.",
        "Δ с/с за ед.",
        "Δ с/с за ед., %",
        "Бух. с/с всего",
        "Упр. с/с всего",

        # Технические ID в конце
        "NM ID",
        "Chrt ID",
    ]

    existing_order = [
        col for col in preferred_order
        if col in df.columns
    ]
    other_cols = [
        col for col in df.columns
        if col not in existing_order
    ]

    df = df[existing_order + other_cols]

    sort_cols = [
        col
        for col in [
            "Бренд",
            "Категория",
            "Наименование",
            "Размер",
        ]
        if col in df.columns
    ]

    if sort_cols:
        df = df.sort_values(
            sort_cols,
            na_position="last",
        )

    return df




def _prepare_warehouse_products_df(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Подготавливает детальные складские остатки для Excel.

    Входная детализация:
        регион
        + склад
        + nm_id
        + chrt_id

    Денежные значения из data.py приходят в копейках,
    поэтому переводим их в рубли здесь.
    """

    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()

    # --------------------------------------------------------------
    # Денежные значения: копейки -> рубли
    # --------------------------------------------------------------
    money_cols = [
        "Бух. с/с за ед.",
        "Упр. с/с за ед.",
        "Бух. стоимость остатка",
        "Упр. стоимость остатка",
    ]

    for col in money_cols:
        if col in df.columns:
            df[col] = (
                pd.to_numeric(
                    df[col],
                    errors="coerce",
                )
                .fillna(0)
                / 100
            )

    # --------------------------------------------------------------
    # Количество
    # --------------------------------------------------------------
    quantity_cols = [
        "Итого количество",
        "Остаток на складе",
        "В пути от клиента",
        "В пути к клиенту",
    ]

    for col in quantity_cols:
        if col in df.columns:
            df[col] = (
                pd.to_numeric(
                    df[col],
                    errors="coerce",
                )
                .fillna(0)
            )

    # --------------------------------------------------------------
    # Разница бухгалтерской и управленческой оценки
    # --------------------------------------------------------------
    if {
        "Бух. стоимость остатка",
        "Упр. стоимость остатка",
    }.issubset(df.columns):
        df["Δ стоимости"] = (
            df["Упр. стоимость остатка"]
            - df["Бух. стоимость остатка"]
        )

    # --------------------------------------------------------------
    # ID сохраняем как текст
    # --------------------------------------------------------------
    for col in [
        "NM ID",
        "Chrt ID",
    ]:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype("string")
                .fillna("")
            )

    # --------------------------------------------------------------
    # Порядок колонок
    # --------------------------------------------------------------
    preferred_order = [
        "Регион",
        "Склад",

        "Бренд",
        "Категория",
        "Пол",
        "Артикул",
        "Наименование",
        "Размер",

        "Итого количество",
        "Остаток на складе",
        "В пути от клиента",
        "В пути к клиенту",

        "Бух. с/с за ед.",
        "Упр. с/с за ед.",

        "Бух. стоимость остатка",
        "Упр. стоимость остатка",
        "Δ стоимости",

        "NM ID",
        "Chrt ID",
    ]

    existing_order = [
        col
        for col in preferred_order
        if col in df.columns
    ]

    other_cols = [
        col
        for col in df.columns
        if col not in existing_order
    ]

    df = df[
        existing_order + other_cols
    ]

    sort_cols = [
        col
        for col in [
            "Регион",
            "Склад",
            "Бренд",
            "Категория",
            "Наименование",
            "Размер",
        ]
        if col in df.columns
    ]

    if sort_cols:
        df = df.sort_values(
            sort_cols,
            na_position="last",
        )

    return df.reset_index(
        drop=True
    )


# ---------------------------------------------------------------------
# Excel styles
# ---------------------------------------------------------------------

def _autosize_columns(ws, min_width=10, max_width=45):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(
                row=row_idx,
                column=col_idx,
            ).value

            if value is not None:
                max_len = max(
                    max_len,
                    len(str(value)),
                )

        ws.column_dimensions[letter].width = max(
            min_width,
            min(max_len + 2, max_width),
        )


def _add_back_button(ws, last_col):
    if ws.title == TOC_SHEET_NAME:
        return

    btn_last_col = min(3, last_col)

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=btn_last_col,
    )

    cell = ws.cell(
        row=1,
        column=1,
        value="← ОГЛАВЛЕНИЕ",
    )
    cell.hyperlink = f"#'{TOC_SHEET_NAME}'!A1"
    cell.font = BUTTON_FONT
    cell.fill = PatternFill(
        "solid",
        fgColor=COLORS["success"],
    )
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24


def _style_title(
    ws,
    title,
    subtitle,
    report_date,
    last_col,
):
    ws.sheet_view.showGridLines = False

    _add_back_button(
        ws,
        last_col,
    )

    ws.merge_cells(
        start_row=3,
        start_column=1,
        end_row=3,
        end_column=last_col,
    )
    cell = ws.cell(
        row=3,
        column=1,
        value=title,
    )
    cell.font = TITLE_FONT
    cell.alignment = LEFT
    ws.row_dimensions[3].height = 30

    ws.merge_cells(
        start_row=4,
        start_column=1,
        end_row=4,
        end_column=last_col,
    )
    cell = ws.cell(
        row=4,
        column=1,
        value=(
            f"{subtitle} · "
            f"дата остатков: {_fmt_date(report_date)}"
        ),
    )
    cell.font = SUBTITLE_FONT
    cell.alignment = LEFT
    ws.row_dimensions[4].height = 22

    ws.merge_cells(
        start_row=5,
        start_column=1,
        end_row=5,
        end_column=last_col,
    )
    cell = ws.cell(
        row=5,
        column=1,
        value=(
            "Сформировано: "
            f"{datetime.now().strftime('%d.%m.%Y в %H:%M')}"
        ),
    )
    cell.font = SMALL_MUTED_FONT
    cell.alignment = LEFT
    ws.row_dimensions[5].height = 20


def _add_sheet_summary_cards(
    ws,
    df,
    start_row,
    last_col,
):
    """
    Добавляет верхние карточки показателей.

    Поддерживает как основной лист "Все товары",
    так и складскую детализацию "По складам",
    где названия стоимостных колонок отличаются.
    """

    total_qty = _sum_col(
        df,
        "Итого количество",
    )

    # --------------------------------------------------------------
    # Розничная стоимость
    # На складском листе этой колонки пока может не быть.
    # --------------------------------------------------------------
    total_retail = _sum_col(
        df,
        "Остатки в розничных ценах",
    )

    # --------------------------------------------------------------
    # Бухгалтерская стоимость
    #
    # Основной лист:
    #   "Бух. с/с всего"
    #
    # Лист "По складам":
    #   "Бух. стоимость остатка"
    # --------------------------------------------------------------
    if "Бух. с/с всего" in df.columns:
        total_buh = _sum_col(
            df,
            "Бух. с/с всего",
        )
    else:
        total_buh = _sum_col(
            df,
            "Бух. стоимость остатка",
        )

    # --------------------------------------------------------------
    # Управленческая стоимость
    #
    # Основной лист:
    #   "Упр. с/с всего"
    #
    # Лист "По складам":
    #   "Упр. стоимость остатка"
    # --------------------------------------------------------------
    if "Упр. с/с всего" in df.columns:
        total_man = _sum_col(
            df,
            "Упр. с/с всего",
        )
    else:
        total_man = _sum_col(
            df,
            "Упр. стоимость остатка",
        )

    cards = [
        (
            "Строк",
            len(df),
            "SKU",
        ),
        (
            "Количество",
            total_qty,
            "шт",
        ),
    ]

    # Розничную карточку показываем только там,
    # где действительно есть розничная стоимость.
    if "Остатки в розничных ценах" in df.columns:
        cards.append(
            (
                "Розничная стоимость",
                total_retail,
                "₽",
            )
        )

    cards.extend(
        [
            (
                "Бух. стоимость",
                total_buh,
                "₽",
            ),
            (
                "Упр. стоимость",
                total_man,
                "₽",
            ),
        ]
    )

    max_cards = min(
        len(cards),
        max(
            1,
            last_col // 2,
        ),
    )

    row = start_row
    col = 1

    for idx, (
        title,
        value,
        subtitle,
    ) in enumerate(
        cards[:max_cards]
    ):
        c1 = col + idx * 2
        c2 = c1 + 1

        ws.merge_cells(
            start_row=row,
            start_column=c1,
            end_row=row,
            end_column=c2,
        )

        title_cell = ws.cell(
            row=row,
            column=c1,
            value=title,
        )
        title_cell.font = Font(
            name=FONT_NAME_BOLD,
            size=9,
            bold=True,
            color=COLORS["muted"],
        )
        title_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_green"],
        )
        title_cell.alignment = CENTER

        ws.merge_cells(
            start_row=row + 1,
            start_column=c1,
            end_row=row + 1,
            end_column=c2,
        )

        value_cell = ws.cell(
            row=row + 1,
            column=c1,
            value=value,
        )
        value_cell.font = Font(
            name=FONT_NAME_BOLD,
            size=13,
            bold=True,
            color=COLORS["dark_green"],
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_gray"],
        )
        value_cell.alignment = CENTER

        if "стоимость" in title.lower():
            value_cell.number_format = '#,##0.00 ₽'
        else:
            value_cell.number_format = "#,##0"

        ws.merge_cells(
            start_row=row + 2,
            start_column=c1,
            end_row=row + 2,
            end_column=c2,
        )

        sub_cell = ws.cell(
            row=row + 2,
            column=c1,
            value=subtitle,
        )
        sub_cell.font = Font(
            name=FONT_NAME,
            size=8,
            color=COLORS["muted"],
        )
        sub_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_gray"],
        )
        sub_cell.alignment = CENTER

        for rr in range(
            row,
            row + 3,
        ):
            for cc in range(
                c1,
                c2 + 1,
            ):
                ws.cell(
                    rr,
                    cc,
                ).border = THIN_BORDER

    return start_row + 5

def _style_body_cell(
    cell,
    col_name,
    row_idx,
):
    cell.font = BODY_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER

    if col_name in [
        "USK",
        "NM ID",
        "Chrt ID",
    ]:
        cell.number_format = "@"
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    elif col_name in [
        "Итого количество",
        "Остаток на складе",
        "В пути от клиента",
        "В пути к клиенту",

    ]:
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["qty"],
        )
        cell.number_format = "#,##0"
        cell.alignment = Alignment(
            horizontal="right",
            vertical="center",
        )
        
        
    elif col_name in [
        "Продажи за 7 дней",
    ]:
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["qty"],
        )
        cell.number_format = "#,##0"
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    elif col_name in [
        "Последняя наша розничная цена",
        "Остатки в розничных ценах",
        "Δ упр. с/с к розничной цене",

        "Бух. с/с за ед.",
        "Упр. с/с за ед.",
        "Бух. с/с всего",
        "Упр. с/с всего",
        "Δ с/с за ед.",

        # Складская детализация
        "Бух. стоимость остатка",
        "Упр. стоимость остатка",
        "Δ стоимости",
    ]:
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["money"],
        )
        cell.number_format = '#,##0.00 ₽'
        cell.alignment = Alignment(
            horizontal="right",
            vertical="center",
        )
        
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["money"],
        )
        cell.number_format = '#,##0.00 ₽'
        cell.alignment = Alignment(
            horizontal="right",
            vertical="center",
        )

    elif col_name == "Упр. с/с выше розничной цены":
        is_loss = bool(cell.value)
        cell.value = "ДА" if is_loss else "Нет"
        cell.alignment = CENTER
        cell.font = Font(
            name=FONT_NAME_BOLD,
            size=10,
            bold=True,
            color=(
                COLORS["discount"]
                if is_loss
                else COLORS["dark_green"]
            ),
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=(
                COLORS["warning"]
                if is_loss
                else COLORS["success"]
            ),
        )

    elif col_name == "Δ с/с за ед., %":
        value = cell.value or 0

        cell.fill = PatternFill(
            "solid",
            fgColor=(
                COLORS["warning"]
                if value > 0
                else COLORS["success"]
            ),
        )
        cell.font = Font(
            name=FONT_NAME_BOLD,
            size=10,
            bold=True,
            color=(
                COLORS["discount"]
                if value > 0
                else COLORS["dark_green"]
            ),
        )
        cell.number_format = "0.00%"
        cell.alignment = Alignment(
            horizontal="right",
            vertical="center",
        )

    elif col_name == "Оборачиваемость 7 дней":
        value = cell.value

        if value is None:
            cell.value = "Нет продаж"
            cell.fill = PatternFill(
                "solid",
                fgColor=COLORS["warning"],
            )
            cell.alignment = CENTER
        else:
            cell.number_format = '0.0 "дн."'
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            # Чем больше дней запаса — тем внимательнее надо смотреть товар
            if value >= 90:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["warning"],
                )
            elif value <= 30:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["success"],
                )
            else:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["light_green"],
                )

    elif col_name == "Дата последнего прихода":
        if cell.value is not None and not pd.isna(cell.value):
            cell.value = pd.to_datetime(
                cell.value,
            ).to_pydatetime()
            cell.number_format = "dd.mm.yyyy"

        cell.alignment = CENTER
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_green"],
        )

    elif col_name == "Возраст товара, дней":
        value = cell.value

        cell.number_format = '0 "дн."'
        cell.alignment = Alignment(
            horizontal="right",
            vertical="center",
        )

        if value is not None and not pd.isna(value):
            if value >= 365:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["warning"],
                )
            elif value <= 90:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["success"],
                )
            else:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["light_green"],
                )

    elif row_idx % 2 == 0:
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_gray"],
        )


# ---------------------------------------------------------------------
# Объединение показателей уровня USK
# ---------------------------------------------------------------------

def _merge_usk_metric_cells(
    ws,
    df: pd.DataFrame,
    first_data_row: int,
):
    """
    Вертикально объединяет ячейки показателей, которые рассчитаны
    на уровне USK, для всех подряд идущих строк одного USK.

    Объединяются:
    - Продажи за 7 дней;
    - Оборачиваемость 7 дней.

    Важно:
    сами данные в DataFrame не удаляются и не зануляются.
    Это только визуальное объединение ячеек в Excel.
    """

    if df.empty or "USK" not in df.columns:
        return

    merge_columns = [
        col_name
        for col_name in [
            "Продажи за 7 дней",
            "Оборачиваемость 7 дней",
        ]
        if col_name in df.columns
    ]

    if not merge_columns:
        return

    usk_values = (
        df["USK"]
        .astype("string")
        .fillna("")
        .str.strip()
        .tolist()
    )

    start_pos = 0

    while start_pos < len(usk_values):
        current_usk = usk_values[start_pos]

        # Пустые USK не объединяем.
        if not current_usk:
            start_pos += 1
            continue

        end_pos = start_pos

        while (
            end_pos + 1 < len(usk_values)
            and usk_values[end_pos + 1] == current_usk
        ):
            end_pos += 1

        # Объединяем только если у USK больше одной строки.
        excel_start_row = first_data_row + start_pos
        excel_end_row = first_data_row + end_pos

        for col_name in merge_columns:
            col_idx = df.columns.get_loc(col_name) + 1

            # Если у USK несколько строк — объединяем.
            if end_pos > start_pos:
                ws.merge_cells(
                    start_row=excel_start_row,
                    start_column=col_idx,
                    end_row=excel_end_row,
                    end_column=col_idx,
                )

            # Центрируем и объединённые, и одиночные ячейки.
            # Поэтому продажи = 0 тоже будут строго по центру.
            cell = ws.cell(
                row=excel_start_row,
                column=col_idx,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        start_pos = end_pos + 1


# ---------------------------------------------------------------------
# Детальные листы
# ---------------------------------------------------------------------

def _write_dataframe_sheet(
    wb,
    sheet_name,
    df,
    report_date,
    title,
    subtitle,
):
    ws = wb.create_sheet(
        _safe_sheet_name(sheet_name)
    )

    df = df.copy()
    last_col = max(
        len(df.columns),
        1,
    )

    _style_title(
        ws,
        title,
        subtitle,
        report_date,
        last_col,
    )

    header_row = _add_sheet_summary_cards(
        ws=ws,
        df=df,
        start_row=7,
        last_col=last_col,
    )

    for col_idx, col_name in enumerate(
        df.columns,
        start=1,
    ):
        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=col_name,
        )
        cell.font = HEADER_FONT
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["dark_green"],
        )
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(
        df.itertuples(index=False),
        start=header_row + 1,
    ):
        ws.row_dimensions[row_idx].height = 20

        for col_idx, value in enumerate(
            row,
            start=1,
        ):
            col_name = df.columns[
                col_idx - 1
            ]

            # pandas NaN / NaT в Excel лучше писать как None
            if pd.isna(value):
                value = None

            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )

            _style_body_cell(
                cell,
                col_name,
                row_idx,
            )

        # Подсветка контроля: если управленческая себестоимость
        # выше последней нашей розничной цены, выделяем ключевые ячейки красным.
        if {
            "Упр. с/с за ед.",
            "Последняя наша розничная цена",
            "Упр. с/с выше розничной цены",
        }.issubset(df.columns):
            risk_col_idx = df.columns.get_loc("Упр. с/с выше розничной цены") + 1
            risk_value = ws.cell(row=row_idx, column=risk_col_idx).value

            if risk_value == "ДА":
                for risk_col_name in [
                    "Последняя наша розничная цена",
                    "Упр. с/с за ед.",
                    "Δ упр. с/с к розничной цене",
                    "Упр. с/с выше розничной цены",
                ]:
                    if risk_col_name not in df.columns:
                        continue

                    risk_cell = ws.cell(
                        row=row_idx,
                        column=df.columns.get_loc(risk_col_name) + 1,
                    )
                    risk_cell.fill = PatternFill(
                        "solid",
                        fgColor=COLORS["warning"],
                    )
                    risk_cell.font = Font(
                        name=FONT_NAME_BOLD,
                        size=10,
                        bold=True,
                        color=COLORS["discount"],
                    )

    # Продажи и оборачиваемость рассчитаны на уровне USK.
    # Поэтому для нескольких размерных строк одного USK
    # объединяем соответствующие ячейки по вертикали.
    _merge_usk_metric_cells(
        ws=ws,
        df=df,
        first_data_row=header_row + 1,
    )

    last_row = (
        header_row + len(df)
    )

    # Замораживаем строку заголовка и первые 3 колонки
    freeze_col = 4
    ws.freeze_panes = (
        f"{get_column_letter(freeze_col)}"
        f"{header_row + 1}"
    )

    if len(df) > 0:
        ws.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(last_col)}"
            f"{last_row}"
        )

    # Детализацию по движению WB прячем в outline-группу
    detail_cols = [
        "Остаток на складе",
        "В пути от клиента",
        "В пути к клиенту",
    ]

    detail_indexes = [
        df.columns.get_loc(col) + 1
        for col in detail_cols
        if col in df.columns
    ]

    if detail_indexes:
        first = min(
            detail_indexes
        )
        last = max(
            detail_indexes
        )

        for col_idx in range(
            first,
            last + 1,
        ):
            letter = get_column_letter(
                col_idx
            )
            ws.column_dimensions[
                letter
            ].outlineLevel = 1
            ws.column_dimensions[
                letter
            ].hidden = True

        ws.sheet_properties.outlinePr.summaryRight = False

    _autosize_columns(
        ws
    )

    widths = {
        "USK": 16,
        "Бренд": 20,
        "Категория": 22,
        "Пол": 12,
        "Артикул": 18,
        "Наименование": 42,
        "Размер": 14,

        "Итого количество": 15,
        "Остаток на складе": 15,
        "В пути от клиента": 15,
        "В пути к клиенту": 15,

        "Последняя наша розничная цена": 24,
        "Δ упр. с/с к розничной цене": 24,
        "Упр. с/с выше розничной цены": 25,
        "Остатки в розничных ценах": 24,

        "Продажи за 7 дней": 17,
        "Оборачиваемость 7 дней": 22,

        "Дата последнего прихода": 21,
        "Возраст товара, дней": 20,

        "Бух. с/с за ед.": 16,
        "Упр. с/с за ед.": 16,
        "Δ с/с за ед.": 16,
        "Δ с/с за ед., %": 16,
        "Бух. с/с всего": 17,
        "Упр. с/с всего": 17,

        "NM ID": 16,
        "Chrt ID": 16,
    }

    for col_name, width in widths.items():
        if col_name in df.columns:
            letter = get_column_letter(
                df.columns.get_loc(
                    col_name
                ) + 1
            )
            ws.column_dimensions[
                letter
            ].width = width

    return ws.title


# ---------------------------------------------------------------------
# Сводка
# ---------------------------------------------------------------------
def _build_warehouse_products_sheet(
    wb,
    df,
    report_date,
):
    """
    Создаёт единый лист с номенклатурными остатками
    в разрезе региона и склада.

    Фильтры Excel позволяют выбирать:
    - регион;
    - склад;
    - бренд;
    - категорию;
    - наименование;
    - NM ID.
    """

    if df is None or df.empty:
        return None

    df = _prepare_warehouse_products_df(
        df
    )

    if df.empty:
        return None

    sheet_name = _write_dataframe_sheet(
        wb=wb,
        sheet_name="По складам",
        df=df,
        report_date=report_date,
        title="Номенклатурные остатки по складам",
        subtitle=(
            "Фактический товар на складах: "
            "регион, склад, номенклатура, количество "
            "и стоимость по себестоимости"
        ),
    )

    ws = wb[sheet_name]

    # --------------------------------------------------------------
    # Закрепляем регион и склад
    #
    # Первые две колонки остаются видимыми при прокрутке.
    # --------------------------------------------------------------
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Регион":
                header_row = cell.row
                ws.freeze_panes = (
                    f"C{header_row + 1}"
                )
                break
        else:
            continue
        break

    # --------------------------------------------------------------
    # Ширина колонок
    # --------------------------------------------------------------
    widths = {
        "Регион": 24,
        "Склад": 30,
        "Бренд": 20,
        "Категория": 24,
        "Пол": 12,
        "Артикул": 18,
        "Наименование": 44,
        "Размер": 14,

        "Остаток на складе": 18,

        "Бух. с/с за ед.": 17,
        "Упр. с/с за ед.": 17,

        "Бух. стоимость остатка": 23,
        "Упр. стоимость остатка": 23,
        "Δ стоимости": 18,

        "NM ID": 16,
        "Chrt ID": 16,
    }

    for col_name, width in widths.items():
        if col_name not in df.columns:
            continue

        letter = get_column_letter(
            df.columns.get_loc(
                col_name
            ) + 1
        )

        ws.column_dimensions[
            letter
        ].width = width

    return sheet_name



def _build_summary_sheet(
    wb,
    df,
    report_date,
):
    ws = wb.create_sheet(
        "Сводка"
    )
    ws.sheet_view.showGridLines = False

    last_col = 10

    _add_back_button(
        ws,
        last_col,
    )

    total_qty = _sum_col(
        df,
        "Итого количество",
    )
    on_hand = _sum_col(
        df,
        "Остаток на складе",
    )
    in_way_client = _sum_col(
        df,
        "В пути к клиенту",
    )
    in_way_from = _sum_col(
        df,
        "В пути от клиента",
    )
    total_retail = _sum_col(
        df,
        "Остатки в розничных ценах",
    )
    total_buh = _sum_col(
        df,
        "Бух. с/с всего",
    )
    total_man = _sum_col(
        df,
        "Упр. с/с всего",
    )

    ws.merge_cells(
        "A3:J3"
    )
    cell = ws["A3"]
    cell.value = (
        "ОТЧЕТ ПО ОСТАТКАМ ТОВАРОВ"
    )
    cell.font = Font(
        name=FONT_NAME_BOLD,
        size=18,
        bold=True,
        color=COLORS["dark_green"],
    )
    cell.alignment = LEFT

    ws.merge_cells(
        "A4:J4"
    )
    cell = ws["A4"]
    cell.value = (
        f"Дата остатков: "
        f"{_fmt_date(report_date)}"
    )
    cell.font = Font(
        name=FONT_NAME,
        size=11,
        bold=True,
        color=COLORS["muted"],
    )

    ws.merge_cells(
        "A5:J5"
    )
    cell = ws["A5"]
    cell.value = (
        "Сформировано: "
        f"{datetime.now().strftime('%d.%m.%Y в %H:%M')}"
    )
    cell.font = SMALL_MUTED_FONT

    cards = [
        ("SKU / строк", len(df), "позиций в выгрузке"),
        ("Итого количество", total_qty, "шт"),
        ("На складе", on_hand, "шт"),
        ("В пути к клиенту", in_way_client, "шт"),
        ("В пути от клиента", in_way_from, "шт"),
        ("Розничная стоимость", total_retail, "в текущих розничных ценах"),
        ("Бух. стоимость", total_buh, "₽"),
        ("Упр. стоимость", total_man, "₽"),
    ]

    row = 7

    for idx, (
        title,
        value,
        subtitle,
    ) in enumerate(cards):
        c1 = 1 + (
            idx % 4
        ) * 2
        r1 = row + (
            idx // 4
        ) * 4

        ws.merge_cells(
            start_row=r1,
            start_column=c1,
            end_row=r1,
            end_column=c1 + 1,
        )

        title_cell = ws.cell(
            r1,
            c1,
            title,
        )
        title_cell.font = Font(
            name=FONT_NAME_BOLD,
            size=9,
            bold=True,
            color=COLORS["muted"],
        )
        title_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_green"],
        )
        title_cell.alignment = CENTER

        ws.merge_cells(
            start_row=r1 + 1,
            start_column=c1,
            end_row=r1 + 1,
            end_column=c1 + 1,
        )

        value_cell = ws.cell(
            r1 + 1,
            c1,
            value,
        )
        value_cell.font = Font(
            name=FONT_NAME_BOLD,
            size=14,
            bold=True,
            color=COLORS["dark_green"],
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_gray"],
        )
        value_cell.alignment = CENTER

        if "стоимость" in title.lower():
            value_cell.number_format = '#,##0.00 ₽'
        else:
            value_cell.number_format = "#,##0"

        ws.merge_cells(
            start_row=r1 + 2,
            start_column=c1,
            end_row=r1 + 2,
            end_column=c1 + 1,
        )

        sub_cell = ws.cell(
            r1 + 2,
            c1,
            subtitle,
        )
        sub_cell.font = Font(
            name=FONT_NAME,
            size=8,
            color=COLORS["muted"],
        )
        sub_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_gray"],
        )
        sub_cell.alignment = CENTER

        for rr in range(
            r1,
            r1 + 3,
        ):
            for cc in range(
                c1,
                c1 + 2,
            ):
                ws.cell(
                    rr,
                    cc,
                ).border = THIN_BORDER

    ws.column_dimensions[
        "A"
    ].width = 18

    for col_letter in [
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
    ]:
        ws.column_dimensions[
            col_letter
        ].width = 16

    return ws.title


# ---------------------------------------------------------------------
# Оглавление
# ---------------------------------------------------------------------

def _build_toc_sheet(
    wb,
    sheets_info,
    report_date,
):
    ws = wb.create_sheet(
        TOC_SHEET_NAME,
        0,
    )
    ws.sheet_view.showGridLines = False

    ws.merge_cells(
        "B2:D2"
    )
    cell = ws["B2"]
    cell.value = (
        "ОГЛАВЛЕНИЕ ОТЧЕТА ПО ОСТАТКАМ"
    )
    cell.font = Font(
        name=FONT_NAME_BOLD,
        size=18,
        bold=True,
        color=COLORS["dark_green"],
    )

    ws.merge_cells(
        "B3:D3"
    )
    cell = ws["B3"]
    cell.value = (
        f"Дата остатков: "
        f"{_fmt_date(report_date)}"
    )
    cell.font = Font(
        name=FONT_NAME,
        size=11,
        bold=True,
        color=COLORS["muted"],
    )

    ws.merge_cells(
        "B4:D4"
    )
    cell = ws["B4"]
    cell.value = (
        "Для перехода к листу нажмите на название раздела"
    )
    cell.font = SMALL_MUTED_FONT

    headers = [
        "№",
        "Лист",
        "Описание",
    ]
    start_row = 7

    for col_idx, header in enumerate(
        headers,
        start=2,
    ):
        cell = ws.cell(
            start_row,
            col_idx,
            header,
        )
        cell.font = HEADER_FONT
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["dark_green"],
        )
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row = (
        start_row + 1
    )

    for idx, item in enumerate(
        sheets_info,
        start=1,
    ):
        sheet_name = item[
            "name"
        ]

        values = [
            f"{idx:02d}",
            sheet_name,
            item.get(
                "description",
                "",
            ),
        ]

        for col_idx, value in enumerate(
            values,
            start=2,
        ):
            cell = ws.cell(
                row,
                col_idx,
                value,
            )
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT

            if col_idx == 3:
                cell.font = Font(
                    name=FONT_NAME_BOLD,
                    size=10,
                    bold=True,
                    color=COLORS["link"],
                )
                cell.hyperlink = (
                    f"#'{sheet_name}'!A1"
                )

            if idx % 2 == 0:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["light_gray"],
                )

        ws.row_dimensions[
            row
        ].height = 26
        row += 1

    ws.column_dimensions[
        "A"
    ].width = 3
    ws.column_dimensions[
        "B"
    ].width = 8
    ws.column_dimensions[
        "C"
    ].width = 34
    ws.column_dimensions[
        "D"
    ].width = 70


# ---------------------------------------------------------------------
# Сводка по категориям
# ---------------------------------------------------------------------

def _build_category_summary(
    wb,
    df,
    report_date,
):
    group_cols = [
        col
        for col in [
            "Категория",
            "Бренд",
        ]
        if col in df.columns
    ]

    if not group_cols:
        return None

    value_cols = [
        col
        for col in [
            "Итого количество",
            "Остаток на складе",
            "В пути от клиента",
            "В пути к клиенту",
            "Продажи за 7 дней",
            "Остатки в розничных ценах",
            "Бух. с/с всего",
            "Упр. с/с всего",
        ]
        if col in df.columns
    ]

    summary_source = df.copy()

    # Продажи рассчитаны на уровне USK и повторяются
    # в каждой размерной строке одного товара.
    # Перед агрегацией по категории / бренду учитываем
    # продажи каждого USK только один раз.
    if (
        "USK" in summary_source.columns
        and "Продажи за 7 дней" in summary_source.columns
    ):
        usk_values = (
            summary_source["USK"]
            .astype("string")
            .fillna("")
            .str.strip()
        )

        dedup_subset = group_cols + ["USK"]

        duplicate_sales_mask = (
            usk_values.ne("")
            & summary_source.duplicated(
                subset=dedup_subset,
                keep="first",
            )
        )

        summary_source.loc[
            duplicate_sales_mask,
            "Продажи за 7 дней",
        ] = 0

    summary = (
        summary_source.groupby(
            group_cols,
            dropna=False,
        )[value_cols]
        .sum()
        .reset_index()
        .sort_values(
            group_cols
        )
    )

    return _write_dataframe_sheet(
        wb=wb,
        sheet_name="По категориям",
        df=summary,
        report_date=report_date,
        title="Сводка остатков по категориям",
        subtitle=(
            "Количество, продажи за 7 дней, "
            "розничная и себестоимостная оценка"
        ),
    )


def _apply_sheet_tab_colors(
    wb,
):
    for idx, ws in enumerate(
        wb.worksheets
    ):
        ws.sheet_properties.tabColor = (
            SHEET_TAB_COLORS[
                idx % len(
                    SHEET_TAB_COLORS
                )
            ]
        )


# ---------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------

def make_stocks_excel(
    df: pd.DataFrame,
    report_date=None,
    warehouse_products_df: pd.DataFrame | None = None,
) -> bytes:
    report_date = (
        report_date
        or datetime.today().date()
    )

    df = _prepare_stocks_df(
        df,
        report_date=report_date,
    )

    wb = Workbook()
    default_ws = wb.active
    wb.remove(
        default_ws
    )

    sheets_info = []

    summary_name = _build_summary_sheet(
        wb,
        df,
        report_date,
    )
    sheets_info.append(
        {
            "name": summary_name,
            "description": (
                "Ключевые показатели по количеству, "
                "розничной стоимости и себестоимости"
            ),
        }
    )

    all_name = _write_dataframe_sheet(
        wb=wb,
        sheet_name="Все товары",
        df=df,
        report_date=report_date,
        title="Детальные остатки товаров",
        subtitle=(
            "Остатки, текущие розничные цены, "
            "оборачиваемость за 7 дней и возраст товара"
        ),
    )
    sheets_info.append(
        {
            "name": all_name,
            "description": (
                "Полная детализация по всем товарам"
            ),
        }
    )
    
    warehouse_name = _build_warehouse_products_sheet(
        wb=wb,
        df=warehouse_products_df,
        report_date=report_date,
    )

    if warehouse_name:
        sheets_info.append(
            {
                "name": warehouse_name,
                "description": (
                    "Номенклатура по регионам и складам "
                    "с количеством и стоимостью остатков"
                ),
            }
        )

    category_name = _build_category_summary(
        wb,
        df,
        report_date,
    )

    if category_name:
        sheets_info.append(
            {
                "name": category_name,
                "description": (
                    "Сводка по категориям и брендам"
                ),
            }
        )

    if "Бренд" in df.columns:
        brands = (
            df["Бренд"]
            .fillna(
                "Бренд не указан"
            )
            .astype(str)
            .sort_values()
            .unique()
            .tolist()
        )

        for brand in brands:
            brand_df = df[
                df["Бренд"]
                .fillna(
                    "Бренд не указан"
                )
                .astype(str)
                == brand
            ]

            sheet_name = _write_dataframe_sheet(
                wb=wb,
                sheet_name=f"Бренд_{brand}",
                df=brand_df,
                report_date=report_date,
                title=(
                    f"Остатки товаров — {brand}"
                ),
                subtitle=(
                    "Остатки, розничные цены, "
                    "оборачиваемость и возраст товара"
                ),
            )

            sheets_info.append(
                {
                    "name": sheet_name,
                    "description": (
                        f"Остатки по бренду {brand}"
                    ),
                }
            )

    _build_toc_sheet(
        wb,
        sheets_info,
        report_date,
    )

    _apply_sheet_tab_colors(
        wb
    )

    output = BytesIO()
    wb.save(
        output
    )
    output.seek(
        0
    )

    return output.read()