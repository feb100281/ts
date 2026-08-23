# # gear/app/daily_sales/stocks/incident_loss_export.py
# """
# Экспорт данных для оценки товарного ущерба по происшествиям на складах.

# Модуль полностью самостоятельный: не зависит от Dash и от конкретной
# реализации dashboard_data — на вход принимает уже собранные данные
# о происшествиях (тот же список `events`, который строится в
# dashboard_stock/incidents_panel.py::build_incidents_panel()).

# Публичные функции:

#     build_incident_loss_excel(events, ...)   -> (bytes, filename)
#     build_incident_cover_letter_pdf(events, ...) -> (bytes, filename)

# Формат одного элемента events (см. incidents_panel.py):

#     {
#         "date": "2026-07-22",                # дата происшествия
#         "warehouse_name": "Краснодар",
#         "incident": {
#             "type": "fire",
#             "title": "Пожар на складе",
#             "status": "Происшествие",
#             "description": "...",
#         },
#         "snapshot": {
#             "effective_date": "2026-07-21",
#             "on_hand": 1234,
#             "nm_count": 87,
#             "accounting_cost": 1500000.0,
#             "management_cost": 1380000.0,
#             "no_accounting_cost_qty": 0,
#             "no_accounting_cost_nm_count": 0,
#             "no_management_cost_qty": 0,
#             "no_management_cost_nm_count": 0,
#         },

#         # Необязательно. Постатейная детализация остатка (если она
#         # доступна) — список словарей или pd.DataFrame со столбцами:
#         #   nm_id, name, brand, qty,
#         #   accounting_unit_cost, management_unit_cost
#         # Если не передана — в лист склада попадёт только сводная
#         # часть (без построчной детализации по товарам).
#         "items": None,
#     }
# """

# from __future__ import annotations

# import io
# from datetime import datetime, date
# from pathlib import Path

# import pandas as pd

# from openpyxl import Workbook
# from openpyxl.styles import (
#     Alignment,
#     Border,
#     Font,
#     NamedStyle,
#     PatternFill,
#     Side,
# )
# from openpyxl.utils import get_column_letter
# from openpyxl.worksheet.worksheet import Worksheet

# from xml.sax.saxutils import escape as _xml_escape

# from reportlab.lib import colors
# from reportlab.lib.pagesizes import A4
# from reportlab.lib.styles import ParagraphStyle
# from reportlab.lib.units import mm
# from reportlab.pdfbase import pdfmetrics
# from reportlab.pdfbase.ttfonts import TTFont
# from reportlab.platypus import (
#     HRFlowable,
#     KeepTogether,
#     Paragraph,
#     SimpleDocTemplate,
#     Spacer,
#     Table,
#     TableStyle,
# )


# # =============================================================================
# # ФИРМЕННАЯ ПАЛИТРА (единая для Excel и PDF, соответствует dashboard)
# # =============================================================================

# TEXT_DARK = "18352F"
# MUTED = "60746D"
# BORDER = "D6DFDB"

# ACCENT_GREEN = "315E52"
# ACCENT_RED = "A43E3E"

# FILL_HEADER = "18352F"       # тёмная плашка заголовков таблиц
# FILL_SUBHEADER = "F7F9F8"    # светлая подложка
# FILL_STRIPE = "F7F9F8"       # чередование строк
# FILL_INCIDENT = "FFF8F8"     # плашка происшествия (красноватая)
# FILL_TOTAL = "EAF0ED"        # плашка итогов

# FONT_NAME = "Helvetica Light"

# # Палитра письма (PDF) — повторяет образец-шаблон пользователя:
# # тёплый тёмный текст + бордовый акцент, отдельно от зелёной палитры
# # дашборда/Excel выше (ACCENT_RED там тоже #A43E3E — специально
# # оставлен тем же числом, чтобы акцентный цвет совпадал в Excel и PDF).
# LETTER_ACCENT = "A43E3E"
# LETTER_TEXT = "2A2020"
# LETTER_MUTED = "7B6A6A"
# LETTER_CARD_BG = "FDF6F6"
# LETTER_BORDER = "E6D6D6"


# # =============================================================================
# # EXCEL: базовые стили
# # =============================================================================

# def _thin_border(color: str = BORDER) -> Border:
#     side = Side(
#         style="thin",
#         color=color,
#     )
#     return Border(
#         left=side,
#         right=side,
#         top=side,
#         bottom=side,
#     )


# def _set_col_widths(
#     ws: Worksheet,
#     widths: dict[str, float],
# ) -> None:
#     for col, width in widths.items():
#         ws.column_dimensions[col].width = width


# def _write_title_block(
#     ws: Worksheet,
#     *,
#     row: int,
#     title: str,
#     subtitle: str | None,
#     ncols: int,
# ) -> int:
#     """Пишет заголовочный блок (название отчёта) и возвращает следующую строку."""

#     ws.merge_cells(
#         start_row=row,
#         start_column=1,
#         end_row=row,
#         end_column=ncols,
#     )
#     cell = ws.cell(row=row, column=1, value=title)
#     cell.font = Font(name=FONT_NAME, size=15, bold=True, color=TEXT_DARK)
#     cell.alignment = Alignment(vertical="center")
#     ws.row_dimensions[row].height = 26
#     row += 1

#     if subtitle:
#         ws.merge_cells(
#             start_row=row,
#             start_column=1,
#             end_row=row,
#             end_column=ncols,
#         )
#         cell = ws.cell(row=row, column=1, value=subtitle)
#         cell.font = Font(name=FONT_NAME, size=10, color=MUTED)
#         row += 1

#     return row + 1


# def _write_table_header(
#     ws: Worksheet,
#     *,
#     row: int,
#     headers: list[str],
#     start_col: int = 1,
# ) -> None:
#     for offset, header in enumerate(headers):
#         col = start_col + offset
#         cell = ws.cell(row=row, column=col, value=header)
#         cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
#         cell.fill = PatternFill(
#             fill_type="solid",
#             fgColor=FILL_HEADER,
#         )
#         cell.alignment = Alignment(
#             vertical="center",
#             horizontal="center",
#             wrap_text=True,
#         )
#         cell.border = _thin_border()
#     ws.row_dimensions[row].height = 30


# def _sanitize_sheet_title(
#     raw_name: str,
#     used: set[str],
# ) -> str:
#     """
#     Excel: максимум 31 символ, запрещены : \\ / ? * [ ].
#     Гарантирует уникальность имени листа.
#     """

#     forbidden = set(':\\/?*[]')

#     cleaned = "".join(
#         ch if ch not in forbidden else " "
#         for ch in str(raw_name or "Склад")
#     ).strip()

#     if not cleaned:
#         cleaned = "Склад"

#     base = cleaned[:31]

#     candidate = base
#     suffix = 2

#     while candidate.lower() in used:
#         tail = f" ({suffix})"
#         candidate = base[: 31 - len(tail)] + tail
#         suffix += 1

#     used.add(candidate.lower())

#     return candidate


# def _fmt_ru_date(value) -> str:
#     if not value:
#         return "нет данных"
#     try:
#         return pd.to_datetime(value).strftime("%d.%m.%Y")
#     except Exception:
#         return str(value)


# # =============================================================================
# # EXCEL: лист "Сводка"
# # =============================================================================

# SUMMARY_HEADERS = [
#     "№",
#     "Склад",
#     "Происшествие",
#     "Дата происшествия",
#     "Остатки на дату",
#     "Физ. остаток, шт",
#     "Товаров, NM ID",
#     "Бухгалтерская с/с, ₽",
#     "Управленческая с/с, ₽",
#     "Без бух. с/с, шт",
#     "Без упр. с/с, шт",
# ]

# SUMMARY_COL_WIDTHS = {
#     "A": 5,
#     "B": 30,
#     "C": 20,
#     "D": 16,
#     "E": 16,
#     "F": 16,
#     "G": 14,
#     "H": 20,
#     "I": 20,
#     "J": 14,
#     "K": 14,
# }


# HYPERLINK_COLOR = "1155CC"


# def _build_summary_sheet(
#     wb: Workbook,
#     sorted_events: list[dict],
#     warehouse_refs: dict[int, dict],
#     warehouse_sheet_titles: dict[int, str],
#     *,
#     generated_label: str,
# ) -> None:
#     ws = wb.create_sheet("Сводка", 0)

#     ncols = len(SUMMARY_HEADERS)

#     _set_col_widths(ws, SUMMARY_COL_WIDTHS)

#     row = _write_title_block(
#         ws,
#         row=1,
#         title="Отчёт по товарным остаткам для оценки ущерба",
#         subtitle=(
#             f"Сформировано: {generated_label}  ·  "
#             f"Происшествий в отчёте: {len(sorted_events)}  ·  "
#             "Оценка по физическому остатку на конец дня, "
#             "предшествующего происшествию. Товары в пути не учтены."
#         ),
#         ncols=ncols,
#     )

#     header_row = row
#     _write_table_header(ws, row=header_row, headers=SUMMARY_HEADERS)

#     first_data_row = header_row + 1
#     r = first_data_row

#     for idx, item in enumerate(sorted_events, start=1):
#         incident = item.get("incident") or {}
#         snapshot = item.get("snapshot") or {}
#         refs = warehouse_refs.get(id(item), {})

#         warehouse_name = item.get("warehouse_name", "")
#         sheet_title = warehouse_sheet_titles.get(id(item))

#         ws.cell(row=r, column=1, value=idx)

#         warehouse_cell = ws.cell(row=r, column=2, value=warehouse_name)

#         # Кликабельный переход на лист склада. Внутренняя ссылка —
#         # это просто адрес вида "#'Имя листа'!A1", Excel сам
#         # прокручивает на нужный лист и ячейку.
#         if sheet_title:
#             warehouse_cell.hyperlink = f"#'{sheet_title}'!A1"
#             warehouse_cell.font = Font(
#                 name=FONT_NAME,
#                 size=10,
#                 color=HYPERLINK_COLOR,
#                 underline="single",
#             )

#         ws.cell(
#             row=r,
#             column=3,
#             value=incident.get("title", "Происшествие"),
#         )
#         ws.cell(
#             row=r,
#             column=4,
#             value=_fmt_ru_date(item.get("date")),
#         )
#         ws.cell(
#             row=r,
#             column=5,
#             value=_fmt_ru_date(snapshot.get("effective_date")),
#         )

#         # Числовые значения — формулой со ссылкой на лист склада,
#         # если он был построен, иначе — прямым значением.
#         def _num_cell(col: int, key: str, default=0):
#             sheet_ref = refs.get(key)
#             if sheet_ref:
#                 ws.cell(row=r, column=col, value=f"={sheet_ref}")
#             else:
#                 ws.cell(row=r, column=col, value=snapshot.get(key, default) or default)

#         _num_cell(6, "on_hand")
#         _num_cell(7, "nm_count")
#         _num_cell(8, "accounting_cost")
#         _num_cell(9, "management_cost")
#         _num_cell(10, "no_accounting_cost_qty")
#         _num_cell(11, "no_management_cost_qty")

#         stripe = (idx % 2 == 0)
#         for col in range(1, ncols + 1):
#             cell = ws.cell(row=r, column=col)
#             cell.border = _thin_border()
#             # Колонка 2 ("Склад") уже получила шрифт гиперссылки выше —
#             # не перезаписываем его обычным цветом.
#             if not (col == 2 and sheet_title):
#                 cell.font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
#             if stripe:
#                 cell.fill = PatternFill(fill_type="solid", fgColor=FILL_STRIPE)
#             if col == 1:
#                 cell.alignment = Alignment(horizontal="center")
#             if col in (6, 7, 8, 9, 10, 11):
#                 cell.alignment = Alignment(horizontal="right")
#                 cell.number_format = (
#                     '#,##0" ₽";[Red]-#,##0" ₽";"–"'
#                     if col in (8, 9)
#                     else '#,##0;[Red]-#,##0;"–"'
#                 )

#         r += 1

#     last_data_row = r - 1

#     # ------------------------------------------------------------------ #
#     # Итоговая строка
#     # ------------------------------------------------------------------ #

#     ws.cell(row=r, column=2, value="ИТОГО")
#     ws.cell(row=r, column=2).font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)

#     for col in (6, 7, 8, 9, 10, 11):
#         col_letter = get_column_letter(col)
#         cell = ws.cell(
#             row=r,
#             column=col,
#             value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})",
#         )
#         cell.number_format = (
#             '#,##0" ₽";[Red]-#,##0" ₽";"–"'
#             if col in (8, 9)
#             else '#,##0;[Red]-#,##0;"–"'
#         )
#         cell.alignment = Alignment(horizontal="right")

#     for col in range(1, ncols + 1):
#         cell = ws.cell(row=r, column=col)
#         cell.font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
#         cell.fill = PatternFill(fill_type="solid", fgColor=FILL_TOTAL)
#         cell.border = _thin_border()

#     ws.row_dimensions[r].height = 20

#     # Замораживаем и строки-заголовки сверху, и колонки №/Склад слева —
#     # при прокрутке длинной таблицы вправо/вниз название склада
#     # и номер строки остаются на экране.
#     ws.freeze_panes = ws.cell(row=first_data_row, column=3).coordinate
#     ws.auto_filter.ref = (
#         f"A{header_row}:{get_column_letter(ncols)}{last_data_row}"
#     )
#     ws.sheet_view.showGridLines = False


# # =============================================================================
# # EXCEL: лист склада
# # =============================================================================

# ITEMS_HEADERS = [
#     "№",
#     "Артикул WB (NM ID)",
#     "Наименование",
#     "Бренд",
#     "Размер",
#     "Кол-во, шт",
#     "Цена, бух. с/с, ₽",
#     "Сумма, бух. с/с, ₽",
#     "Цена, упр. с/с, ₽",
#     "Сумма, упр. с/с, ₽",
# ]

# ITEMS_COL_WIDTHS = {
#     "A": 5,
#     "B": 16,
#     "C": 40,
#     "D": 18,
#     "E": 10,
#     "F": 12,
#     "G": 16,
#     "H": 16,
#     "I": 16,
#     "J": 16,
# }


# def _normalize_items(items) -> pd.DataFrame | None:
#     if items is None:
#         return None

#     if isinstance(items, pd.DataFrame):
#         df = items.copy()
#     else:
#         df = pd.DataFrame(list(items))

#     if df.empty:
#         return None

#     for col in (
#         "nm_id",
#         "name",
#         "brand",
#         "size",
#         "qty",
#         "accounting_unit_cost",
#         "management_unit_cost",
#     ):
#         if col not in df.columns:
#             df[col] = None

#     return df


# def _write_warehouse_sheet(
#     wb: Workbook,
#     item: dict,
#     used_titles: set[str],
# ) -> tuple[str, dict[str, str]]:
#     """
#     Строит лист склада.

#     Возвращает (имя_листа, refs), где refs — адреса ключевых ячеек
#     (для формул на листе "Сводка"), например:

#         {"on_hand": "'Краснодар'!$D$8", ...}
#     """

#     warehouse_name = item.get("warehouse_name", "Склад")
#     incident = item.get("incident") or {}
#     snapshot = item.get("snapshot") or {}

#     sheet_title = _sanitize_sheet_title(warehouse_name, used_titles)
#     ws = wb.create_sheet(sheet_title)

#     items_df = _normalize_items(item.get("items"))

#     ncols = 10
#     _set_col_widths(ws, ITEMS_COL_WIDTHS)
#     ws.sheet_view.showGridLines = False

#     # ------------------------------------------------------------------ #
#     # Заголовок
#     # ------------------------------------------------------------------ #

#     row = _write_title_block(
#         ws,
#         row=1,
#         title=warehouse_name,
#         subtitle=None,
#         ncols=ncols,
#     )

#     # Кликабельная ссылка назад на "Сводку" — правый край, отдельной
#     # строкой над плашкой происшествия.
#     back_cell = ws.cell(row=row, column=ncols, value="← Сводка")
#     back_cell.hyperlink = "#'Сводка'!A1"
#     back_cell.font = Font(
#         name=FONT_NAME,
#         size=9,
#         color=HYPERLINK_COLOR,
#         underline="single",
#     )
#     back_cell.alignment = Alignment(horizontal="right")
#     row += 1

#     ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
#     badge = ws.cell(
#         row=row,
#         column=1,
#         value=(
#             f"{incident.get('title', 'Происшествие')}  ·  "
#             f"{_fmt_ru_date(item.get('date'))}  ·  "
#             f"{incident.get('status', 'Происшествие')}"
#         ),
#     )
#     badge.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_RED)
#     badge.fill = PatternFill(fill_type="solid", fgColor=FILL_INCIDENT)
#     for c in range(1, ncols + 1):
#         ws.cell(row=row, column=c).fill = PatternFill(fill_type="solid", fgColor=FILL_INCIDENT)
#         ws.cell(row=row, column=c).border = _thin_border()
#     ws.row_dimensions[row].height = 20
#     row += 2

#     description = incident.get("description", "")
#     if description:
#         ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
#         cell = ws.cell(row=row, column=1, value=description)
#         cell.font = Font(name=FONT_NAME, size=9, italic=True, color=MUTED)
#         cell.alignment = Alignment(wrap_text=True, vertical="top")
#         ws.row_dimensions[row].height = 28
#         row += 2

#     # ------------------------------------------------------------------ #
#     # Блок ключевых показателей (label / value)
#     # ------------------------------------------------------------------ #

#     kpi_rows = [
#         ("on_hand", "Физический остаток на складе, шт", snapshot.get("on_hand", 0), '#,##0;[Red]-#,##0;"–"'),
#         ("nm_count", "Количество товаров, NM ID", snapshot.get("nm_count", 0), '#,##0;[Red]-#,##0;"–"'),
#         ("accounting_cost", "Бухгалтерская себестоимость, ₽", snapshot.get("accounting_cost", 0), '#,##0" ₽";[Red]-#,##0" ₽";"–"'),
#         ("management_cost", "Управленческая себестоимость, ₽", snapshot.get("management_cost", 0), '#,##0" ₽";[Red]-#,##0" ₽";"–"'),
#         ("no_accounting_cost_qty", "Без бухгалтерской с/с, шт", snapshot.get("no_accounting_cost_qty", 0), '#,##0;[Red]-#,##0;"–"'),
#         ("no_management_cost_qty", "Без управленческой с/с, шт", snapshot.get("no_management_cost_qty", 0), '#,##0;[Red]-#,##0;"–"'),
#     ]

#     ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
#     ws.cell(row=row, column=1, value="Дата снимка остатков:")
#     ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=10, color=MUTED)
#     ws.cell(row=row, column=4, value=_fmt_ru_date(snapshot.get("effective_date")))
#     ws.cell(row=row, column=4).font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
#     row += 1

#     refs: dict[str, str] = {}

#     kpi_start_row = row

#     for key, label, value, number_format in kpi_rows:
#         ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
#         label_cell = ws.cell(row=row, column=1, value=label)
#         label_cell.font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
#         label_cell.fill = PatternFill(fill_type="solid", fgColor=FILL_SUBHEADER)

#         value_cell = ws.cell(row=row, column=4, value=value or 0)
#         value_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=TEXT_DARK)
#         value_cell.number_format = number_format
#         value_cell.alignment = Alignment(horizontal="right")
#         value_cell.fill = PatternFill(fill_type="solid", fgColor=FILL_SUBHEADER)

#         for c in range(1, 5):
#             ws.cell(row=row, column=c).border = _thin_border()

#         refs[key] = f"'{sheet_title}'!${get_column_letter(4)}${row}"
#         row += 1

#     row += 1

#     # ------------------------------------------------------------------ #
#     # Постатейная детализация (если передана)
#     # ------------------------------------------------------------------ #

#     if items_df is not None and not items_df.empty:
#         header_row = row
#         _write_table_header(ws, row=header_row, headers=ITEMS_HEADERS)
#         r = header_row + 1
#         first_item_row = r

#         for idx, (_, line) in enumerate(items_df.iterrows(), start=1):
#             ws.cell(row=r, column=1, value=idx)
#             ws.cell(row=r, column=2, value=line.get("nm_id"))
#             ws.cell(row=r, column=3, value=line.get("name"))
#             ws.cell(row=r, column=4, value=line.get("brand"))
#             ws.cell(row=r, column=5, value=line.get("size"))

#             qty = float(line.get("qty") or 0)
#             acc_unit = float(line.get("accounting_unit_cost") or 0)
#             mgmt_unit = float(line.get("management_unit_cost") or 0)

#             ws.cell(row=r, column=6, value=qty)
#             ws.cell(row=r, column=7, value=acc_unit)
#             ws.cell(
#                 row=r,
#                 column=8,
#                 value=f"=F{r}*G{r}",
#             )
#             ws.cell(row=r, column=9, value=mgmt_unit)
#             ws.cell(
#                 row=r,
#                 column=10,
#                 value=f"=F{r}*I{r}",
#             )

#             stripe = (idx % 2 == 0)
#             for col in range(1, ncols + 1):
#                 cell = ws.cell(row=r, column=col)
#                 cell.border = _thin_border()
#                 cell.font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
#                 if stripe:
#                     cell.fill = PatternFill(fill_type="solid", fgColor=FILL_STRIPE)
#                 if col in (6,):
#                     cell.number_format = '#,##0;[Red]-#,##0;"–"'
#                     cell.alignment = Alignment(horizontal="right")
#                 if col in (7, 8, 9, 10):
#                     cell.number_format = '#,##0.00" ₽";[Red]-#,##0.00" ₽";"–"'
#                     cell.alignment = Alignment(horizontal="right")

#             r += 1

#         last_item_row = r - 1

#         ws.cell(row=r, column=5, value="ИТОГО:")
#         ws.cell(row=r, column=5).font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
#         ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")

#         ws.cell(row=r, column=6, value=f"=SUM(F{first_item_row}:F{last_item_row})")
#         ws.cell(row=r, column=8, value=f"=SUM(H{first_item_row}:H{last_item_row})")
#         ws.cell(row=r, column=10, value=f"=SUM(J{first_item_row}:J{last_item_row})")

#         for col in (6, 8, 10):
#             fmt = '#,##0;"–"' if col == 6 else '#,##0.00" ₽";"–"'
#             cell = ws.cell(row=r, column=col)
#             cell.number_format = fmt
#             cell.alignment = Alignment(horizontal="right")

#         for col in range(1, ncols + 1):
#             cell = ws.cell(row=r, column=col)
#             cell.font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
#             cell.fill = PatternFill(fill_type="solid", fgColor=FILL_TOTAL)
#             cell.border = _thin_border()

#         # Замораживаем строки-заголовки сверху И колонки №/NM ID/
#         # Наименование слева — при прокрутке длинного списка товаров
#         # вниз и вправо видно, к какой позиции относится строка.
#         ws.freeze_panes = ws.cell(row=first_item_row, column=4).coordinate
#         ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_item_row}"

#         row = r + 2
#     else:
#         ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
#         note = ws.cell(
#             row=row,
#             column=1,
#             value=(
#                 "Постатейная детализация по товарам недоступна — "
#                 "приведена только сводная оценка выше."
#             ),
#         )
#         note.font = Font(name=FONT_NAME, size=9, italic=True, color=MUTED)
#         row += 2

#         # Без построчной детализации сам лист короткий, но заголовок
#         # (название склада + плашка происшествия) всё равно закрепляем.
#         ws.freeze_panes = "A4"

#     # ------------------------------------------------------------------ #
#     # Примечание по методологии
#     # ------------------------------------------------------------------ #

#     ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
#     footnote = ws.cell(
#         row=row,
#         column=1,
#         value=(
#             "Оценка рассчитана по товару, физически находившемуся на складе "
#             "на конец дня, предшествующего происшествию. Товары в пути в расчёт "
#             "не включены. Позиции без определённой себестоимости не включены "
#             "в соответствующую стоимостную оценку."
#         ),
#     )
#     footnote.font = Font(name=FONT_NAME, size=8, italic=True, color=MUTED)
#     footnote.alignment = Alignment(wrap_text=True, vertical="top")
#     ws.row_dimensions[row].height = 26

#     return sheet_title, refs


# # =============================================================================
# # EXCEL: точка входа
# # =============================================================================

# def build_incident_loss_excel(
#     events: list[dict],
#     *,
#     generated_at: datetime | None = None,
# ) -> tuple[bytes, str]:
#     """
#     Строит книгу Excel: лист "Сводка" + отдельный лист на каждый склад.

#     events — см. docstring модуля.

#     Возвращает (bytes, filename), готовые для dcc.send_bytes.
#     """

#     if not events:
#         raise ValueError("Список происшествий пуст — нечего экспортировать.")

#     generated_at = generated_at or datetime.now()

#     sorted_events = sorted(
#         events,
#         key=lambda x: (x.get("date", ""), x.get("warehouse_name", "")),
#         reverse=True,
#     )

#     wb = Workbook()
#     wb.remove(wb.active)

#     used_titles: set[str] = set()
#     warehouse_refs: dict[int, dict] = {}
#     warehouse_sheet_titles: dict[int, str] = {}

#     for item in sorted_events:
#         sheet_title, refs = _write_warehouse_sheet(wb, item, used_titles)
#         warehouse_refs[id(item)] = refs
#         warehouse_sheet_titles[id(item)] = sheet_title

#     _build_summary_sheet(
#         wb,
#         sorted_events,
#         warehouse_refs,
#         warehouse_sheet_titles,
#         generated_label=generated_at.strftime("%d.%m.%Y %H:%M"),
#     )

#     # "Сводка" должна остаться первым (активным) листом
#     wb.active = 0

#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     filename = (
#         f"Оценка_ущерба_склады_{generated_at.strftime('%Y-%m-%d')}.xlsx"
#     )

#     return buffer.getvalue(), filename


# # =============================================================================
# # PDF: сопроводительное письмо
# # =============================================================================

# _FONTS_REGISTERED = False

# # Стандартные 14 шрифтов reportlab (Helvetica/Times/Courier) физически
# # не содержат кириллицу — это ограничение формата PDF, а не прихоть:
# # base-14 шрифты включают только латиницу. Поэтому нужен настоящий TTF
# # с кириллицей, и берём его из уже установленных в системе шрифтов —
# # никаких дополнительных папок в репозитории не требуется.
# #
# # Для каждого начертания — список стандартных путей на разных ОС,
# # пробуем по очереди и берём первый найденный:
# #   1) Arial из macOS (стоит на любом Mac из коробки);
# #   2) Arial из Windows;
# #   3) Arial из пакета msttcorefonts (Linux, если ставили);
# #   4) Liberation Sans — метрический аналог Arial, часто уже стоит
# #      на серверных Linux-дистрибутивах;
# #   5) DejaVu Sans — тоже часто предустановлен на Linux.

# _FONT_CANDIDATES: dict[str, list[str]] = {
#     "Arial": [
#         "/System/Library/Fonts/Supplemental/Arial.ttf",
#         "/Library/Fonts/Arial.ttf",
#         "C:/Windows/Fonts/arial.ttf",
#         "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
#         "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
#         "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
#     ],
#     "Arial-Bold": [
#         "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
#         "/Library/Fonts/Arial Bold.ttf",
#         "C:/Windows/Fonts/arialbd.ttf",
#         "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold.ttf",
#         "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
#         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
#     ],
#     "Arial-Italic": [
#         "/System/Library/Fonts/Supplemental/Arial Italic.ttf",
#         "/Library/Fonts/Arial Italic.ttf",
#         "C:/Windows/Fonts/ariali.ttf",
#         "/usr/share/fonts/truetype/msttcorefonts/Arial_Italic.ttf",
#         "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
#         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
#     ],
#     "Arial-BoldItalic": [
#         "/System/Library/Fonts/Supplemental/Arial Bold Italic.ttf",
#         "/Library/Fonts/Arial Bold Italic.ttf",
#         "C:/Windows/Fonts/arialbi.ttf",
#         "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold_Italic.ttf",
#         "/usr/share/fonts/truetype/liberation/LiberationSans-BoldItalic.ttf",
#         "/usr/share/fonts/truetype/dejavu/DejaVuSans-BoldOblique.ttf",
#     ],
# }


# def _find_font_file(candidates: list[str]) -> str | None:
#     for path_str in candidates:
#         if Path(path_str).is_file():
#             return path_str
#     return None


# def _register_pdf_fonts() -> None:
#     """
#     Регистрирует первый найденный в системе шрифт с кириллицей под
#     именами Arial/Arial-Bold/... — чтобы весь остальной код мог просто
#     использовать fontName="Arial".

#     Если ни один стандартный путь не подошёл — кидает понятную ошибку
#     с инструкцией, что установить, а не глухой TTFError из недр
#     reportlab.
#     """

#     global _FONTS_REGISTERED

#     if _FONTS_REGISTERED:
#         return

#     missing = []

#     for font_name, candidates in _FONT_CANDIDATES.items():
#         path = _find_font_file(candidates)

#         if path is None:
#             missing.append(font_name)
#             continue

#         pdfmetrics.registerFont(TTFont(font_name, path))

#     if missing:
#         raise RuntimeError(
#             "Не найден шрифт с кириллицей для PDF (начертания: "
#             + ", ".join(sorted(set(missing)))
#             + "). На Mac он должен быть по умолчанию "
#             "(/System/Library/Fonts/Supplemental/Arial*.ttf) — "
#             "проверьте, что файл существует. На Linux-сервере "
#             "поставьте пакет с Arial или Liberation Sans, например: "
#             "apt-get install ttf-mscorefonts-installer "
#             "(или fonts-liberation)."
#         )

#     _FONTS_REGISTERED = True


# def _pdf_styles() -> dict[str, ParagraphStyle]:
#     text = colors.HexColor(f"#{LETTER_TEXT}")
#     muted = colors.HexColor(f"#{LETTER_MUTED}")
#     accent = colors.HexColor(f"#{LETTER_ACCENT}")

#     return {
#         "eyebrow": ParagraphStyle(
#             "eyebrow",
#             fontName="Arial-Bold",
#             fontSize=8.5,
#             textColor=accent,
#             leading=11,
#             spaceAfter=3,
#         ),
#         "title": ParagraphStyle(
#             "title",
#             fontName="Arial-Bold",
#             fontSize=19,
#             textColor=text,
#             leading=23,
#             spaceAfter=5,
#         ),
#         "subtitle": ParagraphStyle(
#             "subtitle",
#             fontName="Arial",
#             fontSize=9.5,
#             textColor=muted,
#             leading=13,
#         ),
#         "meta_right": ParagraphStyle(
#             "meta_right",
#             fontName="Arial",
#             fontSize=9.5,
#             textColor=text,
#             leading=13,
#             alignment=2,  # right
#         ),
#         "salutation": ParagraphStyle(
#             "salutation",
#             fontName="Arial-Bold",
#             fontSize=11,
#             textColor=text,
#             spaceBefore=4,
#             spaceAfter=10,
#         ),
#         "body": ParagraphStyle(
#             "body",
#             fontName="Arial",
#             fontSize=10,
#             textColor=text,
#             leading=15,
#             spaceAfter=10,
#             alignment=4,  # justify
#         ),
#         "section": ParagraphStyle(
#             "section",
#             fontName="Arial-Bold",
#             fontSize=9.5,
#             textColor=accent,
#             leading=13,
#             spaceBefore=16,
#             spaceAfter=8,
#         ),
#         "kpi_label": ParagraphStyle(
#             "kpi_label",
#             fontName="Arial-Bold",
#             fontSize=7.5,
#             textColor=muted,
#             leading=10,
#         ),
#         "kpi_value": ParagraphStyle(
#             "kpi_value",
#             fontName="Arial-Bold",
#             fontSize=15,
#             textColor=accent,
#             leading=19,
#             spaceBefore=3,
#         ),
#         "bullet_body": ParagraphStyle(
#             "bullet_body",
#             fontName="Arial",
#             fontSize=10,
#             textColor=text,
#             leading=14.5,
#             leftIndent=14,
#             firstLineIndent=-14,
#             spaceAfter=8,
#         ),
#         "closing": ParagraphStyle(
#             "closing",
#             fontName="Arial",
#             fontSize=10,
#             textColor=text,
#             leading=15,
#             spaceBefore=8,
#             alignment=4,
#         ),
#         "sign_name": ParagraphStyle(
#             "sign_name",
#             fontName="Arial-Bold",
#             fontSize=11,
#             textColor=text,
#             spaceBefore=2,
#         ),
#         "sign_email": ParagraphStyle(
#             "sign_email",
#             fontName="Arial",
#             fontSize=9.5,
#             textColor=muted,
#             spaceBefore=3,
#         ),
#         "table_header": ParagraphStyle(
#             "table_header",
#             fontName="Arial-Bold",
#             fontSize=8.5,
#             textColor=colors.white,
#             leading=11,
#         ),
#         "table_cell": ParagraphStyle(
#             "table_cell",
#             fontName="Arial",
#             fontSize=9.5,
#             textColor=text,
#             leading=12,
#         ),
#         "table_cell_muted": ParagraphStyle(
#             "table_cell_muted",
#             fontName="Arial",
#             fontSize=9.5,
#             textColor=muted,
#             leading=12,
#         ),
#     }


# _RU_MONTHS = [
#     "января", "февраля", "марта", "апреля", "мая", "июня",
#     "июля", "августа", "сентября", "октября", "ноября", "декабря",
# ]


# def _fmt_ru_date_long(dt: datetime) -> str:
#     return f"{dt.day} {_RU_MONTHS[dt.month - 1]} {dt.year} г."


# def _fmt_ru_date_short(value) -> str:
#     """08.06 без года — для компактной плашки периода."""
#     if not value:
#         return ""
#     try:
#         return pd.to_datetime(value).strftime("%d.%m")
#     except Exception:
#         return str(value)


# def _file_word(count: int) -> str:
#     """1 файл, 2 файла, 5 файлов, 21 файл, 23 файла, 27 файлов."""

#     count = abs(int(count or 0))
#     last_two = count % 100
#     last_one = count % 10

#     if 11 <= last_two <= 14:
#         return "файлов"
#     if last_one == 1:
#         return "файл"
#     if last_one in {2, 3, 4}:
#         return "файла"
#     return "файлов"


# def _short_warehouse_label(warehouse_name: str) -> str:
#     """
#     Электросталь -> Электросталь
#     Симферополь, Молодежненское -> Симферополь
#     Красный Бор (Питер) WB -> Красный
#     Санкт-Петербург Уткина Заводь -> Санкт-Петербург

#     Правило: первое слово названия, без хвостовой пунктуации.
#     """

#     name = str(warehouse_name or "").strip()
#     if not name:
#         return "Склад"

#     first_token = name.split()[0]

#     return first_token.rstrip(",;:")


# def build_incident_cover_letter_pdf(
#     events: list[dict],
#     *,
#     letter_title: str = "Реестр пожаров на складах Wildberries",
#     author_short_name: str = "Войтенко Д. В.",
#     author_name: str = "Дарья Войтенко",
#     author_email: str = "daria031288d@gmail.com",
#     closing_text: str = (
#         ""
  
#     ),
#     generated_at: datetime | None = None,
# ) -> tuple[bytes, str]:
#     """
#     Строит служебное письмо (PDF) — реестр происшествий по образцу
#     пользовательского шаблона: эйбрау + заголовок, плашки-метрики,
#     таблица реестра происшествий, методика оценки, подпись.

#     Приложения (файлы/листы Excel) в письме не прикладываются —
#     это только сам реестр; Excel-книга с остатками скачивается
#     отдельной кнопкой (build_incident_loss_excel).
#     """

#     if not events:
#         raise ValueError("Список происшествий пуст — нечего экспортировать.")

#     _register_pdf_fonts()
#     styles = _pdf_styles()

#     generated_at = generated_at or datetime.now()

#     # Реестр в письме идёт в ХРОНОЛОГИЧЕСКОМ порядке (старые сверху) —
#     # так же, как в образце (№1 — самое раннее происшествие).
#     chronological_events = sorted(
#         events,
#         key=lambda x: (x.get("date", ""), x.get("warehouse_name", "")),
#     )

#     warehouse_count = len(
#         {e.get("warehouse_name", "") for e in events}
#     )

#     dates = [
#         pd.to_datetime(e.get("date"))
#         for e in events
#         if e.get("date")
#     ]

#     period_short = ""
#     period_full = ""

#     if dates:
#         min_date = min(dates)
#         max_date = max(dates)
#         period_short = (
#             f"{min_date.strftime('%d.%m')} — {max_date.strftime('%d.%m')}"
#         )
#         period_full = (
#             f"{min_date.strftime('%d.%m.%Y')} — "
#             f"{max_date.strftime('%d.%m.%Y')}"
#         )

#     accent = colors.HexColor(f"#{LETTER_ACCENT}")
#     card_bg = colors.HexColor(f"#{LETTER_CARD_BG}")
#     border_c = colors.HexColor(f"#{LETTER_BORDER}")

#     buffer = io.BytesIO()

#     doc = SimpleDocTemplate(
#         buffer,
#         pagesize=A4,
#         leftMargin=20 * mm,
#         rightMargin=20 * mm,
#         topMargin=18 * mm,
#         bottomMargin=18 * mm,
#         title=letter_title,
#     )

#     story = []

#     # ------------------------------------------------------------------ #
#     # Шапка: эйбрау + заголовок + подзаголовок (слева),
#     # автор + дата (справа)
#     # ------------------------------------------------------------------ #

#     left_col = [
#         Paragraph(_xml_escape(letter_title), styles["title"]),
#         Paragraph(
#             (
#                 f"Период {period_full} · оценка товарного остатка "
#                 "на конец дня, предшествующего происшествию"
#             ),
#             styles["subtitle"],
#         ),
#     ]

#     right_col = [
#         Paragraph(_xml_escape(author_short_name), styles["meta_right"]),
#         Paragraph(_fmt_ru_date_long(generated_at), styles["meta_right"]),
#     ]

#     header_table = Table(
#         [[left_col, right_col]],
#         colWidths=[125 * mm, 45 * mm],
#     )
#     header_table.setStyle(
#         TableStyle(
#             [
#                 ("VALIGN", (0, 0), (-1, -1), "TOP"),
#                 ("LEFTPADDING", (0, 0), (-1, -1), 0),
#                 ("RIGHTPADDING", (0, 0), (-1, -1), 0),
#                 ("TOPPADDING", (0, 0), (-1, -1), 0),
#                 ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
#             ]
#         )
#     )
#     story.append(header_table)

#     story.append(Spacer(1, 5 * mm))
#     story.append(
#         HRFlowable(
#             width="100%",
#             thickness=1.2,
#             color=accent,
#             spaceAfter=12,
#         )
#     )

#     # ------------------------------------------------------------------ #
#     # Приветствие и вводный абзац
#     # ------------------------------------------------------------------ #

#     story.append(Paragraph("Добрый день!", styles["salutation"]))

#     period_from = dates and min(dates).strftime("%d.%m.%Y") or ""
#     period_to = dates and max(dates).strftime("%d.%m.%Y") or ""

#     intro = (
#         "Направляю сводный реестр пожаров на складах Wildberries, "
#         f"зафиксированных в период с {period_from} по {period_to}. "
#         f"Всего затронуто <font color=\"#{LETTER_ACCENT}\"><b>"
#         f"{warehouse_count}</b></font> складов. "
#         "Прилагаю выгрузки товарных остатков на "
#         "конец дня, предшествующего дате пожара — "
#         f"<font color=\"#{LETTER_ACCENT}\"><b>в расчёт включён только "
#         "физический остаток на складе, товары в пути не "
#         "учитывались.</b></font>"
#     )

#     story.append(Paragraph(intro, styles["body"]))

#     # ------------------------------------------------------------------ #
#     # Реестр происшествий
#     # ------------------------------------------------------------------ #

#     story.append(Paragraph("РЕЕСТР ПРОИСШЕСТВИЙ", styles["section"]))

#     def _cell(text, style_name="table_cell"):
#         return Paragraph(_xml_escape(str(text)), styles[style_name])

#     header_row = [
#         _cell("№", "table_header"),
#         _cell("СКЛАД", "table_header"),
#         _cell("ДАТА ПОЖАРА", "table_header"),
#         _cell("ОСТАТКИ НА ДАТУ", "table_header"),
#     ]

#     reg_rows = [header_row]

#     for idx, item in enumerate(chronological_events, start=1):
#         snapshot = item.get("snapshot") or {}

#         reg_rows.append(
#             [
#                 _cell(idx, "table_cell_muted"),
#                 _cell(item.get("warehouse_name", "")),
#                 _cell(_fmt_ru_date(item.get("date"))),
#                 _cell(_fmt_ru_date(snapshot.get("effective_date"))),
#             ]
#         )

#     col_widths = [10 * mm, 82 * mm, 36 * mm, 42 * mm]

#     reg_table = Table(reg_rows, colWidths=col_widths, repeatRows=1)

#     reg_style = [
#         ("BACKGROUND", (0, 0), (-1, 0), accent),
#         ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
#         ("ALIGN", (0, 0), (0, -1), "CENTER"),
#         ("TOPPADDING", (0, 0), (-1, -1), 8),
#         ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
#         ("LEFTPADDING", (0, 0), (-1, -1), 8),
#         ("RIGHTPADDING", (0, 0), (-1, -1), 8),
#         ("LINEBELOW", (0, 1), (-1, -1), 0.6, border_c),
#     ]
#     reg_table.setStyle(TableStyle(reg_style))

#     story.append(reg_table)

#     # ------------------------------------------------------------------ #
#     # Методика оценки
#     # ------------------------------------------------------------------ #

#     story.append(Paragraph("МЕТОДИКА ОЦЕНКИ", styles["section"]))

#     methodology_points = [
#         (
#             "В оценку включён только физический товарный остаток на "
#             "складе (quantity). Товары в пути не учитываются."
#         ),
#         (
#             "Используется снимок остатков на конец календарного дня, "
#             "предшествующего дате происшествия."
#         ),
#         (
#             "Позиции без определённой себестоимости в стоимостную "
#             "оценку не включены."
#         ),
#     ]

#     for point in methodology_points:
#         story.append(
#             Paragraph(
#                 f'<font color="#{LETTER_ACCENT}">●</font>&nbsp;&nbsp;{point}',
#                 styles["bullet_body"],
#             )
#         )

#     # ------------------------------------------------------------------ #
#     # Заключительный абзац
#     # ------------------------------------------------------------------ #

#     story.append(Paragraph(closing_text, styles["closing"]))

#     story.append(Spacer(1, 10 * mm))

#     # ------------------------------------------------------------------ #
#     # Подпись
#     # ------------------------------------------------------------------ #

#     story.append(
#         HRFlowable(
#             width="100%",
#             thickness=0.6,
#             color=border_c,
#             spaceAfter=8,
#         )
#     )
#     story.append(Paragraph(_xml_escape(author_name), styles["sign_name"]))
#     story.append(Paragraph(_xml_escape(author_email), styles["sign_email"]))

#     doc.build(story)

#     buffer.seek(0)

#     filename = (
#         f"Сопроводительное_письмо_{generated_at.strftime('%Y-%m-%d')}.pdf"
#     )

#     return buffer.getvalue(), filename


# gear/app/daily_sales/stocks/incident_loss_export.py
"""
Экспорт данных для оценки товарного ущерба по происшествиям на складах.

Модуль полностью самостоятельный: не зависит от Dash и от конкретной
реализации dashboard_data — на вход принимает уже собранные данные
о происшествиях (тот же список `events`, который строится в
dashboard_stock/incidents_panel.py::build_incidents_panel()).

Публичные функции:

    build_incident_loss_excel(events, ...)   -> (bytes, filename)
    build_incident_cover_letter_pdf(events, ...) -> (bytes, filename)

Формат одного элемента events (см. incidents_panel.py):

    {
        "date": "2026-07-22",                # дата происшествия
        "warehouse_name": "Краснодар",
        "incident": {
            "type": "fire",
            "title": "Пожар на складе",
            "status": "Происшествие",
            "description": "...",
        },
        "snapshot": {
            "effective_date": "2026-07-21",
            "on_hand": 1234,
            "nm_count": 87,
            "accounting_cost": 1500000.0,
            "management_cost": 1380000.0,
            "no_accounting_cost_qty": 0,
            "no_accounting_cost_nm_count": 0,
            "no_management_cost_qty": 0,
            "no_management_cost_nm_count": 0,
        },

        # Необязательно. Постатейная детализация остатка (если она
        # доступна) — список словарей или pd.DataFrame со столбцами:
        #   nm_id, name, brand, qty,
        #   accounting_unit_cost, management_unit_cost
        # Если не передана — в лист склада попадёт только сводная
        # часть (без построчной детализации по товарам).
        "items": None,
    }
"""

from __future__ import annotations

import io
from datetime import datetime, date
from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from xml.sax.saxutils import escape as _xml_escape

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# =============================================================================
# ФИРМЕННАЯ ПАЛИТРА (единая для Excel и PDF, соответствует dashboard)
# =============================================================================

TEXT_DARK = "18352F"
MUTED = "60746D"
BORDER = "D6DFDB"

ACCENT_GREEN = "315E52"
ACCENT_RED = "A43E3E"

FILL_HEADER = "18352F"       # тёмная плашка заголовков таблиц
FILL_SUBHEADER = "F7F9F8"    # светлая подложка
FILL_STRIPE = "F7F9F8"       # чередование строк
FILL_INCIDENT = "FFF8F8"     # плашка происшествия (красноватая)
FILL_TOTAL = "EAF0ED"        # плашка итогов
FILL_WARNING = "FFF4D6"      # позиции без определённой себестоимости

FONT_NAME = "Helvetica Light"

# Палитра письма (PDF) — повторяет образец-шаблон пользователя:
# тёплый тёмный текст + бордовый акцент, отдельно от зелёной палитры
# дашборда/Excel выше (ACCENT_RED там тоже #A43E3E — специально
# оставлен тем же числом, чтобы акцентный цвет совпадал в Excel и PDF).
LETTER_ACCENT = "A43E3E"
LETTER_TEXT = "2A2020"
LETTER_MUTED = "7B6A6A"
LETTER_CARD_BG = "FDF6F6"
LETTER_BORDER = "E6D6D6"


# =============================================================================
# EXCEL: базовые стили
# =============================================================================

def _thin_border(color: str = BORDER) -> Border:
    side = Side(
        style="thin",
        color=color,
    )
    return Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )


def _set_col_widths(
    ws: Worksheet,
    widths: dict[str, float],
) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_title_block(
    ws: Worksheet,
    *,
    row: int,
    title: str,
    subtitle: str | None,
    ncols: int,
) -> int:
    """Пишет заголовочный блок (название отчёта) и возвращает следующую строку."""

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=ncols,
    )
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name=FONT_NAME, size=15, bold=True, color=TEXT_DARK)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 26
    row += 1

    if subtitle:
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=ncols,
        )
        cell = ws.cell(row=row, column=1, value=subtitle)
        cell.font = Font(name=FONT_NAME, size=10, color=MUTED)
        row += 1

    return row + 1


def _write_table_header(
    ws: Worksheet,
    *,
    row: int,
    headers: list[str],
    start_col: int = 1,
) -> None:
    for offset, header in enumerate(headers):
        col = start_col + offset
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=FILL_HEADER,
        )
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center",
            wrap_text=True,
        )
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 30


def _sanitize_sheet_title(
    raw_name: str,
    used: set[str],
) -> str:
    """
    Excel: максимум 31 символ, запрещены : \\ / ? * [ ].
    Гарантирует уникальность имени листа.
    """

    forbidden = set(':\\/?*[]')

    cleaned = "".join(
        ch if ch not in forbidden else " "
        for ch in str(raw_name or "Склад")
    ).strip()

    if not cleaned:
        cleaned = "Склад"

    base = cleaned[:31]

    candidate = base
    suffix = 2

    while candidate.lower() in used:
        tail = f" ({suffix})"
        candidate = base[: 31 - len(tail)] + tail
        suffix += 1

    used.add(candidate.lower())

    return candidate


def _fmt_ru_date(value) -> str:
    if not value:
        return "нет данных"
    try:
        return pd.to_datetime(value).strftime("%d.%m.%Y")
    except Exception:
        return str(value)


# =============================================================================
# EXCEL: лист "Сводка"
# =============================================================================

SUMMARY_HEADERS = [
    "№",
    "Склад",
    "Происшествие",
    "Дата происшествия",
    "Остатки на дату",
    "Физ. остаток, шт",
    "Товаров, NM ID",
    "Бухгалтерская с/с, ₽",
    "Управленческая с/с, ₽",
    "Без бух. с/с, шт",
    "Без упр. с/с, шт",
]

SUMMARY_COL_WIDTHS = {
    "A": 5,
    "B": 30,
    "C": 20,
    "D": 16,
    "E": 16,
    "F": 16,
    "G": 14,
    "H": 20,
    "I": 20,
    "J": 14,
    "K": 14,
}


HYPERLINK_COLOR = "1155CC"


def _build_summary_sheet(
    wb: Workbook,
    sorted_events: list[dict],
    warehouse_refs: dict[int, dict],
    warehouse_sheet_titles: dict[int, str],
    *,
    generated_label: str,
) -> None:
    ws = wb.create_sheet("Сводка", 0)

    ncols = len(SUMMARY_HEADERS)

    _set_col_widths(ws, SUMMARY_COL_WIDTHS)

    row = _write_title_block(
        ws,
        row=1,
        title="Отчёт по товарным остаткам для оценки ущерба",
        subtitle=(
            f"Сформировано: {generated_label}  ·  "
            f"Происшествий в отчёте: {len(sorted_events)}  ·  "
            "Оценка по физическому остатку на конец дня, "
            "предшествующего происшествию. Товары в пути не учтены."
        ),
        ncols=ncols,
    )

    header_row = row
    _write_table_header(ws, row=header_row, headers=SUMMARY_HEADERS)

    first_data_row = header_row + 1
    r = first_data_row

    for idx, item in enumerate(sorted_events, start=1):
        incident = item.get("incident") or {}
        snapshot = item.get("snapshot") or {}
        refs = warehouse_refs.get(id(item), {})

        warehouse_name = item.get("warehouse_name", "")
        sheet_title = warehouse_sheet_titles.get(id(item))

        ws.cell(row=r, column=1, value=idx)

        warehouse_cell = ws.cell(row=r, column=2, value=warehouse_name)

        # Кликабельный переход на лист склада. Внутренняя ссылка —
        # это просто адрес вида "#'Имя листа'!A1", Excel сам
        # прокручивает на нужный лист и ячейку.
        if sheet_title:
            warehouse_cell.hyperlink = f"#'{sheet_title}'!A1"
            warehouse_cell.font = Font(
                name=FONT_NAME,
                size=10,
                color=HYPERLINK_COLOR,
                underline="single",
            )

        ws.cell(
            row=r,
            column=3,
            value=incident.get("title", "Происшествие"),
        )
        ws.cell(
            row=r,
            column=4,
            value=_fmt_ru_date(item.get("date")),
        )
        ws.cell(
            row=r,
            column=5,
            value=_fmt_ru_date(snapshot.get("effective_date")),
        )

        # Числовые значения — формулой со ссылкой на лист склада,
        # если он был построен, иначе — прямым значением.
        def _num_cell(col: int, key: str, default=0):
            sheet_ref = refs.get(key)
            if sheet_ref:
                ws.cell(row=r, column=col, value=f"={sheet_ref}")
            else:
                ws.cell(row=r, column=col, value=snapshot.get(key, default) or default)

        _num_cell(6, "on_hand")
        _num_cell(7, "nm_count")
        _num_cell(8, "accounting_cost")
        _num_cell(9, "management_cost")
        _num_cell(10, "no_accounting_cost_qty")
        _num_cell(11, "no_management_cost_qty")

        stripe = (idx % 2 == 0)
        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = _thin_border()
            # Колонка 2 ("Склад") уже получила шрифт гиперссылки выше —
            # не перезаписываем его обычным цветом.
            if not (col == 2 and sheet_title):
                cell.font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
            if stripe:
                cell.fill = PatternFill(fill_type="solid", fgColor=FILL_STRIPE)
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            if col in (6, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = (
                    '#,##0" ₽";[Red]-#,##0" ₽";"–"'
                    if col in (8, 9)
                    else '#,##0;[Red]-#,##0;"–"'
                )

        r += 1

    last_data_row = r - 1

    # ------------------------------------------------------------------ #
    # Итоговая строка
    # ------------------------------------------------------------------ #

    ws.cell(row=r, column=2, value="ИТОГО")
    ws.cell(row=r, column=2).font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)

    for col in (6, 7, 8, 9, 10, 11):
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=r,
            column=col,
            value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})",
        )
        cell.number_format = (
            '#,##0" ₽";[Red]-#,##0" ₽";"–"'
            if col in (8, 9)
            else '#,##0;[Red]-#,##0;"–"'
        )
        cell.alignment = Alignment(horizontal="right")

    for col in range(1, ncols + 1):
        cell = ws.cell(row=r, column=col)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
        cell.fill = PatternFill(fill_type="solid", fgColor=FILL_TOTAL)
        cell.border = _thin_border()

    ws.row_dimensions[r].height = 20

    # Замораживаем и строки-заголовки сверху, и колонки №/Склад слева —
    # при прокрутке длинной таблицы вправо/вниз название склада
    # и номер строки остаются на экране.
    ws.freeze_panes = ws.cell(row=first_data_row, column=3).coordinate
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(ncols)}{last_data_row}"
    )
    ws.sheet_view.showGridLines = False


# =============================================================================
# EXCEL: лист склада
# =============================================================================

ITEMS_HEADERS = [
    "№",
    "Артикул WB (NM ID)",
    "Наименование",
    "Бренд",
    "Размер",
    "Кол-во, шт",
    "Цена, бух. с/с, ₽",
    "Сумма, бух. с/с, ₽",
    "Цена, упр. с/с, ₽",
    "Сумма, упр. с/с, ₽",
    "Без с/с",
]

ITEMS_COL_WIDTHS = {
    "A": 5,
    "B": 16,
    "C": 40,
    "D": 18,
    "E": 10,
    "F": 12,
    "G": 16,
    "H": 16,
    "I": 16,
    "J": 16,
    "K": 14,
}


def _normalize_items(items) -> pd.DataFrame | None:
    if items is None:
        return None

    if isinstance(items, pd.DataFrame):
        df = items.copy()
    else:
        df = pd.DataFrame(list(items))

    if df.empty:
        return None

    for col in (
        "nm_id",
        "name",
        "brand",
        "size",
        "qty",
        "accounting_unit_cost",
        "management_unit_cost",
    ):
        if col not in df.columns:
            df[col] = None

    return df


def _write_warehouse_sheet(
    wb: Workbook,
    item: dict,
    used_titles: set[str],
) -> tuple[str, dict[str, str]]:
    """
    Строит лист склада.

    Возвращает (имя_листа, refs), где refs — адреса ключевых ячеек
    (для формул на листе "Сводка"), например:

        {"on_hand": "'Краснодар'!$D$8", ...}
    """

    warehouse_name = item.get("warehouse_name", "Склад")
    incident = item.get("incident") or {}
    snapshot = item.get("snapshot") or {}

    sheet_title = _sanitize_sheet_title(warehouse_name, used_titles)
    ws = wb.create_sheet(sheet_title)

    items_df = _normalize_items(item.get("items"))

    ncols = 11
    _set_col_widths(ws, ITEMS_COL_WIDTHS)
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------ #
    # Заголовок
    # ------------------------------------------------------------------ #

    row = _write_title_block(
        ws,
        row=1,
        title=warehouse_name,
        subtitle=None,
        ncols=ncols,
    )

    # Кликабельная ссылка назад на "Сводку" — правый край, отдельной
    # строкой над плашкой происшествия.
    back_cell = ws.cell(row=row, column=ncols, value="← Сводка")
    back_cell.hyperlink = "#'Сводка'!A1"
    back_cell.font = Font(
        name=FONT_NAME,
        size=9,
        color=HYPERLINK_COLOR,
        underline="single",
    )
    back_cell.alignment = Alignment(horizontal="right")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    badge = ws.cell(
        row=row,
        column=1,
        value=(
            f"{incident.get('title', 'Происшествие')}  ·  "
            f"{_fmt_ru_date(item.get('date'))}  ·  "
            f"{incident.get('status', 'Происшествие')}"
        ),
    )
    badge.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_RED)
    badge.fill = PatternFill(fill_type="solid", fgColor=FILL_INCIDENT)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(fill_type="solid", fgColor=FILL_INCIDENT)
        ws.cell(row=row, column=c).border = _thin_border()
    ws.row_dimensions[row].height = 20
    row += 2

    description = incident.get("description", "")
    if description:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=description)
        cell.font = Font(name=FONT_NAME, size=9, italic=True, color=MUTED)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28
        row += 2

    # ------------------------------------------------------------------ #
    # Блок ключевых показателей (label / value)
    # ------------------------------------------------------------------ #

    kpi_rows = [
        ("on_hand", "Физический остаток на складе, шт", snapshot.get("on_hand", 0), '#,##0;[Red]-#,##0;"–"'),
        ("nm_count", "Количество товаров, NM ID", snapshot.get("nm_count", 0), '#,##0;[Red]-#,##0;"–"'),
        ("accounting_cost", "Бухгалтерская себестоимость, ₽", snapshot.get("accounting_cost", 0), '#,##0" ₽";[Red]-#,##0" ₽";"–"'),
        ("management_cost", "Управленческая себестоимость, ₽", snapshot.get("management_cost", 0), '#,##0" ₽";[Red]-#,##0" ₽";"–"'),
        ("no_accounting_cost_qty", "Без бухгалтерской с/с, шт", snapshot.get("no_accounting_cost_qty", 0), '#,##0;[Red]-#,##0;"–"'),
        ("no_management_cost_qty", "Без управленческой с/с, шт", snapshot.get("no_management_cost_qty", 0), '#,##0;[Red]-#,##0;"–"'),
    ]

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="Дата снимка остатков:")
    ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=10, color=MUTED)
    ws.cell(row=row, column=4, value=_fmt_ru_date(snapshot.get("effective_date")))
    ws.cell(row=row, column=4).font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
    row += 1

    refs: dict[str, str] = {}

    kpi_start_row = row

    for key, label, value, number_format in kpi_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
        label_cell.fill = PatternFill(fill_type="solid", fgColor=FILL_SUBHEADER)

        value_cell = ws.cell(row=row, column=4, value=value or 0)
        value_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=TEXT_DARK)
        value_cell.number_format = number_format
        value_cell.alignment = Alignment(horizontal="right")
        value_cell.fill = PatternFill(fill_type="solid", fgColor=FILL_SUBHEADER)

        for c in range(1, 5):
            ws.cell(row=row, column=c).border = _thin_border()

        refs[key] = f"'{sheet_title}'!${get_column_letter(4)}${row}"
        row += 1

    row += 1

    # ------------------------------------------------------------------ #
    # Постатейная детализация (если передана)
    # ------------------------------------------------------------------ #

    if items_df is not None and not items_df.empty:
        header_row = row
        _write_table_header(ws, row=header_row, headers=ITEMS_HEADERS)
        r = header_row + 1
        first_item_row = r

        for idx, (_, line) in enumerate(items_df.iterrows(), start=1):
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=line.get("nm_id"))
            ws.cell(row=r, column=3, value=line.get("name"))
            ws.cell(row=r, column=4, value=line.get("brand"))
            ws.cell(row=r, column=5, value=line.get("size"))

            qty = float(line.get("qty") or 0)

            # "Без с/с" ставится по факту нуля/отсутствия цены —
            # бухгалтерская и управленческая проверяются НЕЗАВИСИМО
            # друг от друга, флаг ставится даже если только одна
            # из двух не определена.
            acc_raw = line.get("accounting_unit_cost")
            mgmt_raw = line.get("management_unit_cost")

            acc_unit = 0.0 if pd.isna(acc_raw) else float(acc_raw)
            mgmt_unit = 0.0 if pd.isna(mgmt_raw) else float(mgmt_raw)

            missing_acc = acc_unit == 0
            missing_mgmt = mgmt_unit == 0

            ws.cell(row=r, column=6, value=qty)
            ws.cell(row=r, column=7, value=acc_unit)
            ws.cell(
                row=r,
                column=8,
                value=f"=F{r}*G{r}",
            )
            ws.cell(row=r, column=9, value=mgmt_unit)
            ws.cell(
                row=r,
                column=10,
                value=f"=F{r}*I{r}",
            )

            missing_parts = []
            if missing_acc:
                missing_parts.append("Бух.")
            if missing_mgmt:
                missing_parts.append("Упр.")

            no_cost_cell = ws.cell(
                row=r,
                column=11,
                value=(", ".join(missing_parts) or None),
            )
            no_cost_cell.alignment = Alignment(horizontal="center")

            has_warning = missing_acc or missing_mgmt

            stripe = (idx % 2 == 0)
            for col in range(1, ncols + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = _thin_border()
                cell.font = Font(
                    name=FONT_NAME,
                    size=9,
                    color=ACCENT_RED if (has_warning and col == 11) else TEXT_DARK,
                    bold=(has_warning and col == 11),
                )
                if has_warning:
                    cell.fill = PatternFill(fill_type="solid", fgColor=FILL_WARNING)
                elif stripe:
                    cell.fill = PatternFill(fill_type="solid", fgColor=FILL_STRIPE)
                if col in (6,):
                    cell.number_format = '#,##0;[Red]-#,##0;"–"'
                    cell.alignment = Alignment(horizontal="right")
                if col in (7, 8, 9, 10):
                    cell.number_format = '#,##0.00" ₽";[Red]-#,##0.00" ₽";"–"'
                    cell.alignment = Alignment(horizontal="right")

            r += 1

        last_item_row = r - 1

        ws.cell(row=r, column=5, value="ИТОГО:")
        ws.cell(row=r, column=5).font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")

        ws.cell(row=r, column=6, value=f"=SUM(F{first_item_row}:F{last_item_row})")
        ws.cell(row=r, column=8, value=f"=SUM(H{first_item_row}:H{last_item_row})")
        ws.cell(row=r, column=10, value=f"=SUM(J{first_item_row}:J{last_item_row})")

        for col in (6, 8, 10):
            fmt = '#,##0;"–"' if col == 6 else '#,##0.00" ₽";"–"'
            cell = ws.cell(row=r, column=col)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")

        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_DARK)
            cell.fill = PatternFill(fill_type="solid", fgColor=FILL_TOTAL)
            cell.border = _thin_border()

        # Счётчик позиций без определённой себестоимости — сколько
        # строк выше получили пометку в колонке "Без с/с". Пишем
        # ПОСЛЕ общего цикла стилизации итоговой строки, иначе он
        # перезаписал бы этот акцентный (красный) шрифт обычным.
        no_cost_total = ws.cell(
            row=r,
            column=11,
            value=(
                f'=COUNTIF(K{first_item_row}:K{last_item_row},"<>")'
                ' & " шт"'
            ),
        )
        no_cost_total.alignment = Alignment(horizontal="center")
        no_cost_total.font = Font(name=FONT_NAME, size=9, bold=True, color=ACCENT_RED)
        no_cost_total.fill = PatternFill(fill_type="solid", fgColor=FILL_TOTAL)

        # Замораживаем строки-заголовки сверху И колонки №/NM ID/
        # Наименование слева — при прокрутке длинного списка товаров
        # вниз и вправо видно, к какой позиции относится строка.
        ws.freeze_panes = ws.cell(row=first_item_row, column=4).coordinate
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_item_row}"

        row = r + 2
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        note = ws.cell(
            row=row,
            column=1,
            value=(
                "Постатейная детализация по товарам недоступна — "
                "приведена только сводная оценка выше."
            ),
        )
        note.font = Font(name=FONT_NAME, size=9, italic=True, color=MUTED)
        row += 2

        # Без построчной детализации сам лист короткий, но заголовок
        # (название склада + плашка происшествия) всё равно закрепляем.
        ws.freeze_panes = "A4"

    # ------------------------------------------------------------------ #
    # Примечание по методологии
    # ------------------------------------------------------------------ #

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    footnote = ws.cell(
        row=row,
        column=1,
        value=(
            "Оценка рассчитана по товару, физически находившемуся на складе "
            "на конец дня, предшествующего происшествию. Товары в пути в расчёт "
            "не включены. Позиции без определённой себестоимости не включены "
            "в соответствующую стоимостную оценку."
        ),
    )
    footnote.font = Font(name=FONT_NAME, size=8, italic=True, color=MUTED)
    footnote.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 26

    return sheet_title, refs


# =============================================================================
# EXCEL: точка входа
# =============================================================================

def build_incident_loss_excel(
    events: list[dict],
    *,
    generated_at: datetime | None = None,
) -> tuple[bytes, str]:
    """
    Строит книгу Excel: лист "Сводка" + отдельный лист на каждый склад.

    events — см. docstring модуля.

    Возвращает (bytes, filename), готовые для dcc.send_bytes.
    """

    if not events:
        raise ValueError("Список происшествий пуст — нечего экспортировать.")

    generated_at = generated_at or datetime.now()

    sorted_events = sorted(
        events,
        key=lambda x: (x.get("date", ""), x.get("warehouse_name", "")),
        reverse=True,
    )

    wb = Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    warehouse_refs: dict[int, dict] = {}
    warehouse_sheet_titles: dict[int, str] = {}

    for item in sorted_events:
        sheet_title, refs = _write_warehouse_sheet(wb, item, used_titles)
        warehouse_refs[id(item)] = refs
        warehouse_sheet_titles[id(item)] = sheet_title

    _build_summary_sheet(
        wb,
        sorted_events,
        warehouse_refs,
        warehouse_sheet_titles,
        generated_label=generated_at.strftime("%d.%m.%Y %H:%M"),
    )

    # "Сводка" должна остаться первым (активным) листом
    wb.active = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"Оценка_ущерба_склады_{generated_at.strftime('%Y-%m-%d')}.xlsx"
    )

    return buffer.getvalue(), filename


# =============================================================================
# PDF: сопроводительное письмо
# =============================================================================

_FONTS_REGISTERED = False

# Стандартные 14 шрифтов reportlab (Helvetica/Times/Courier) физически
# не содержат кириллицу — это ограничение формата PDF, а не прихоть:
# base-14 шрифты включают только латиницу. Поэтому нужен настоящий TTF
# с кириллицей, и берём его из уже установленных в системе шрифтов —
# никаких дополнительных папок в репозитории не требуется.
#
# Для каждого начертания — список стандартных путей на разных ОС,
# пробуем по очереди и берём первый найденный:
#   1) Arial из macOS (стоит на любом Mac из коробки);
#   2) Arial из Windows;
#   3) Arial из пакета msttcorefonts (Linux, если ставили);
#   4) Liberation Sans — метрический аналог Arial, часто уже стоит
#      на серверных Linux-дистрибутивах;
#   5) DejaVu Sans — тоже часто предустановлен на Linux.

_FONT_CANDIDATES: dict[str, list[str]] = {
    "Arial": [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ],
    "Arial-Bold": [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/Library/Fonts/Arial Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ],
    "Arial-Italic": [
        "/System/Library/Fonts/Supplemental/Arial Italic.ttf",
        "/Library/Fonts/Arial Italic.ttf",
        "C:/Windows/Fonts/ariali.ttf",
        "/usr/share/fonts/truetype/msttcorefonts/Arial_Italic.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
    ],
    "Arial-BoldItalic": [
        "/System/Library/Fonts/Supplemental/Arial Bold Italic.ttf",
        "/Library/Fonts/Arial Bold Italic.ttf",
        "C:/Windows/Fonts/arialbi.ttf",
        "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold_Italic.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-BoldItalic.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-BoldOblique.ttf",
    ],
}


def _find_font_file(candidates: list[str]) -> str | None:
    for path_str in candidates:
        if Path(path_str).is_file():
            return path_str
    return None


def _register_pdf_fonts() -> None:
    """
    Регистрирует первый найденный в системе шрифт с кириллицей под
    именами Arial/Arial-Bold/... — чтобы весь остальной код мог просто
    использовать fontName="Arial".

    Если ни один стандартный путь не подошёл — кидает понятную ошибку
    с инструкцией, что установить, а не глухой TTFError из недр
    reportlab.
    """

    global _FONTS_REGISTERED

    if _FONTS_REGISTERED:
        return

    missing = []

    for font_name, candidates in _FONT_CANDIDATES.items():
        path = _find_font_file(candidates)

        if path is None:
            missing.append(font_name)
            continue

        pdfmetrics.registerFont(TTFont(font_name, path))

    if missing:
        raise RuntimeError(
            "Не найден шрифт с кириллицей для PDF (начертания: "
            + ", ".join(sorted(set(missing)))
            + "). На Mac он должен быть по умолчанию "
            "(/System/Library/Fonts/Supplemental/Arial*.ttf) — "
            "проверьте, что файл существует. На Linux-сервере "
            "поставьте пакет с Arial или Liberation Sans, например: "
            "apt-get install ttf-mscorefonts-installer "
            "(или fonts-liberation)."
        )

    _FONTS_REGISTERED = True


def _pdf_styles() -> dict[str, ParagraphStyle]:
    text = colors.HexColor(f"#{LETTER_TEXT}")
    muted = colors.HexColor(f"#{LETTER_MUTED}")
    accent = colors.HexColor(f"#{LETTER_ACCENT}")

    return {
        "eyebrow": ParagraphStyle(
            "eyebrow",
            fontName="Arial-Bold",
            fontSize=8.5,
            textColor=accent,
            leading=11,
            spaceAfter=3,
        ),
        "title": ParagraphStyle(
            "title",
            fontName="Arial-Bold",
            fontSize=19,
            textColor=text,
            leading=23,
            spaceAfter=5,
        ),
        "subtitle": ParagraphStyle(
            "subtitle",
            fontName="Arial",
            fontSize=9.5,
            textColor=muted,
            leading=13,
        ),
        "meta_right": ParagraphStyle(
            "meta_right",
            fontName="Arial",
            fontSize=9.5,
            textColor=text,
            leading=13,
            alignment=2,  # right
        ),
        "salutation": ParagraphStyle(
            "salutation",
            fontName="Arial-Bold",
            fontSize=11,
            textColor=text,
            spaceBefore=4,
            spaceAfter=10,
        ),
        "body": ParagraphStyle(
            "body",
            fontName="Arial",
            fontSize=10,
            textColor=text,
            leading=15,
            spaceAfter=10,
            alignment=4,  # justify
        ),
        "section": ParagraphStyle(
            "section",
            fontName="Arial-Bold",
            fontSize=9.5,
            textColor=accent,
            leading=13,
            spaceBefore=16,
            spaceAfter=8,
        ),
        "kpi_label": ParagraphStyle(
            "kpi_label",
            fontName="Arial-Bold",
            fontSize=7.5,
            textColor=muted,
            leading=10,
        ),
        "kpi_value": ParagraphStyle(
            "kpi_value",
            fontName="Arial-Bold",
            fontSize=15,
            textColor=accent,
            leading=19,
            spaceBefore=3,
        ),
        "bullet_body": ParagraphStyle(
            "bullet_body",
            fontName="Arial",
            fontSize=10,
            textColor=text,
            leading=14.5,
            leftIndent=14,
            firstLineIndent=-14,
            spaceAfter=8,
        ),
        "closing": ParagraphStyle(
            "closing",
            fontName="Arial",
            fontSize=10,
            textColor=text,
            leading=15,
            spaceBefore=8,
            alignment=4,
        ),
        "sign_name": ParagraphStyle(
            "sign_name",
            fontName="Arial-Bold",
            fontSize=11,
            textColor=text,
            spaceBefore=2,
        ),
        "sign_email": ParagraphStyle(
            "sign_email",
            fontName="Arial",
            fontSize=9.5,
            textColor=muted,
            spaceBefore=3,
        ),
        "table_header": ParagraphStyle(
            "table_header",
            fontName="Arial-Bold",
            fontSize=8.5,
            textColor=colors.white,
            leading=11,
        ),
        "table_cell": ParagraphStyle(
            "table_cell",
            fontName="Arial",
            fontSize=9.5,
            textColor=text,
            leading=12,
        ),
        "table_cell_muted": ParagraphStyle(
            "table_cell_muted",
            fontName="Arial",
            fontSize=9.5,
            textColor=muted,
            leading=12,
        ),
    }


_RU_MONTHS = [
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]


def _fmt_ru_date_long(dt: datetime) -> str:
    return f"{dt.day} {_RU_MONTHS[dt.month - 1]} {dt.year} г."


def _fmt_ru_date_short(value) -> str:
    """08.06 без года — для компактной плашки периода."""
    if not value:
        return ""
    try:
        return pd.to_datetime(value).strftime("%d.%m")
    except Exception:
        return str(value)


def _file_word(count: int) -> str:
    """1 файл, 2 файла, 5 файлов, 21 файл, 23 файла, 27 файлов."""

    count = abs(int(count or 0))
    last_two = count % 100
    last_one = count % 10

    if 11 <= last_two <= 14:
        return "файлов"
    if last_one == 1:
        return "файл"
    if last_one in {2, 3, 4}:
        return "файла"
    return "файлов"


def _short_warehouse_label(warehouse_name: str) -> str:
    """
    Электросталь -> Электросталь
    Симферополь, Молодежненское -> Симферополь
    Красный Бор (Питер) WB -> Красный
    Санкт-Петербург Уткина Заводь -> Санкт-Петербург

    Правило: первое слово названия, без хвостовой пунктуации.
    """

    name = str(warehouse_name or "").strip()
    if not name:
        return "Склад"

    first_token = name.split()[0]

    return first_token.rstrip(",;:")


def build_incident_cover_letter_pdf(
    events: list[dict],
    *,
    letter_title: str = "Реестр пожаров на складах Wildberries",
    author_short_name: str = "Войтенко Д. В.",
    author_name: str = "Дарья Войтенко",
    author_email: str = "daria031288d@gmail.com",
    closing_text: str = (
        ""
  
    ),
    generated_at: datetime | None = None,
) -> tuple[bytes, str]:
    """
    Строит служебное письмо (PDF) — реестр происшествий по образцу
    пользовательского шаблона: эйбрау + заголовок, плашки-метрики,
    таблица реестра происшествий, методика оценки, подпись.

    Приложения (файлы/листы Excel) в письме не прикладываются —
    это только сам реестр; Excel-книга с остатками скачивается
    отдельной кнопкой (build_incident_loss_excel).
    """

    if not events:
        raise ValueError("Список происшествий пуст — нечего экспортировать.")

    _register_pdf_fonts()
    styles = _pdf_styles()

    generated_at = generated_at or datetime.now()

    # Реестр в письме идёт в ХРОНОЛОГИЧЕСКОМ порядке (старые сверху) —
    # так же, как в образце (№1 — самое раннее происшествие).
    chronological_events = sorted(
        events,
        key=lambda x: (x.get("date", ""), x.get("warehouse_name", "")),
    )

    warehouse_count = len(
        {e.get("warehouse_name", "") for e in events}
    )

    dates = [
        pd.to_datetime(e.get("date"))
        for e in events
        if e.get("date")
    ]

    period_short = ""
    period_full = ""

    if dates:
        min_date = min(dates)
        max_date = max(dates)
        period_short = (
            f"{min_date.strftime('%d.%m')} — {max_date.strftime('%d.%m')}"
        )
        period_full = (
            f"{min_date.strftime('%d.%m.%Y')} — "
            f"{max_date.strftime('%d.%m.%Y')}"
        )

    accent = colors.HexColor(f"#{LETTER_ACCENT}")
    card_bg = colors.HexColor(f"#{LETTER_CARD_BG}")
    border_c = colors.HexColor(f"#{LETTER_BORDER}")

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=letter_title,
    )

    story = []

    # ------------------------------------------------------------------ #
    # Шапка: эйбрау + заголовок + подзаголовок (слева),
    # автор + дата (справа)
    # ------------------------------------------------------------------ #

    left_col = [
        Paragraph(_xml_escape(letter_title), styles["title"]),
        Paragraph(
            (
                f"Период {period_full} · оценка товарного остатка "
                "на конец дня, предшествующего происшествию"
            ),
            styles["subtitle"],
        ),
    ]

    right_col = [
        Paragraph(_xml_escape(author_short_name), styles["meta_right"]),
        Paragraph(_fmt_ru_date_long(generated_at), styles["meta_right"]),
    ]

    header_table = Table(
        [[left_col, right_col]],
        colWidths=[125 * mm, 45 * mm],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_table)

    story.append(Spacer(1, 5 * mm))
    story.append(
        HRFlowable(
            width="100%",
            thickness=1.2,
            color=accent,
            spaceAfter=12,
        )
    )

    # ------------------------------------------------------------------ #
    # Приветствие и вводный абзац
    # ------------------------------------------------------------------ #

    story.append(Paragraph("Добрый день!", styles["salutation"]))

    period_from = dates and min(dates).strftime("%d.%m.%Y") or ""
    period_to = dates and max(dates).strftime("%d.%m.%Y") or ""

    intro = (
        "Направляю сводный реестр пожаров на складах Wildberries, "
        f"зафиксированных в период с {period_from} по {period_to}. "
        f"Всего затронуто <font color=\"#{LETTER_ACCENT}\"><b>"
        f"{warehouse_count}</b></font> складов. "
        "Прилагаю выгрузки товарных остатков на "
        "конец дня, предшествующего дате пожара — "
        f"<font color=\"#{LETTER_ACCENT}\"><b>в расчёт включён только "
        "физический остаток на складе, товары в пути не "
        "учитывались.</b></font>"
    )

    story.append(Paragraph(intro, styles["body"]))

    # ------------------------------------------------------------------ #
    # Реестр происшествий
    # ------------------------------------------------------------------ #

    story.append(Paragraph("РЕЕСТР ПРОИСШЕСТВИЙ", styles["section"]))

    def _cell(text, style_name="table_cell"):
        return Paragraph(_xml_escape(str(text)), styles[style_name])

    header_row = [
        _cell("№", "table_header"),
        _cell("СКЛАД", "table_header"),
        _cell("ДАТА ПОЖАРА", "table_header"),
        _cell("ОСТАТКИ НА ДАТУ", "table_header"),
    ]

    reg_rows = [header_row]

    for idx, item in enumerate(chronological_events, start=1):
        snapshot = item.get("snapshot") or {}

        reg_rows.append(
            [
                _cell(idx, "table_cell_muted"),
                _cell(item.get("warehouse_name", "")),
                _cell(_fmt_ru_date(item.get("date"))),
                _cell(_fmt_ru_date(snapshot.get("effective_date"))),
            ]
        )

    col_widths = [10 * mm, 82 * mm, 36 * mm, 42 * mm]

    reg_table = Table(reg_rows, colWidths=col_widths, repeatRows=1)

    reg_style = [
        ("BACKGROUND", (0, 0), (-1, 0), accent),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 1), (-1, -1), 0.6, border_c),
    ]
    reg_table.setStyle(TableStyle(reg_style))

    story.append(reg_table)

    # ------------------------------------------------------------------ #
    # Методика оценки
    # ------------------------------------------------------------------ #

    story.append(Paragraph("МЕТОДИКА ОЦЕНКИ", styles["section"]))

    methodology_points = [
        (
            "В оценку включён только физический товарный остаток на "
            "складе (quantity). Товары в пути не учитываются."
        ),
        (
            "Используется снимок остатков на конец календарного дня, "
            "предшествующего дате происшествия."
        ),
        (
            "Позиции без определённой себестоимости в стоимостную "
            "оценку не включены."
        ),
    ]

    for point in methodology_points:
        story.append(
            Paragraph(
                f'<font color="#{LETTER_ACCENT}">●</font>&nbsp;&nbsp;{point}',
                styles["bullet_body"],
            )
        )

    # ------------------------------------------------------------------ #
    # Заключительный абзац
    # ------------------------------------------------------------------ #

    story.append(Paragraph(closing_text, styles["closing"]))

    story.append(Spacer(1, 10 * mm))

    # ------------------------------------------------------------------ #
    # Подпись
    # ------------------------------------------------------------------ #

    story.append(
        HRFlowable(
            width="100%",
            thickness=0.6,
            color=border_c,
            spaceAfter=8,
        )
    )
    story.append(Paragraph(_xml_escape(author_name), styles["sign_name"]))
    story.append(Paragraph(_xml_escape(author_email), styles["sign_email"]))

    doc.build(story)

    buffer.seek(0)

    filename = (
        f"Сопроводительное_письмо_{generated_at.strftime('%Y-%m-%d')}.pdf"
    )

    return buffer.getvalue(), filename