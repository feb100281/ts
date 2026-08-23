from __future__ import annotations

from io import BytesIO

import pandas as pd

from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)

from openpyxl.utils import (
    get_column_letter,
)


# =========================================================
# Названия колонок
# =========================================================

COLUMN_NAMES = {
    "name": "Наименование",

    "revenue_vat": "Выручка с НДС",

    "vat_amount": "НДС",

    "revenue_vatless": (
        "Выручка без НДС"
    ),

    "revenue_share_pct": (
        "Доля выручки, %"
    ),

    "net_comission": (
        "Комиссия WB"
    ),

    "commission_pct": (
        "Комиссия WB, %"
    ),

    "cogs_book": (
        "Себестоимость бух."
    ),

    "gross_profit_book": (
        "Прибыль бух."
    ),

    "margin_book_pct": (
        "Маржа бух., %"
    ),

    "cogs_man": (
        "Себестоимость упр."
    ),

    "gross_profit_man": (
        "Прибыль упр."
    ),

    "margin_man_pct": (
        "Маржа упр., %"
    ),

    "net_qty": (
        "Количество нетто"
    ),

    "average_revenue": (
        "Средняя выручка / ед."
    ),

    "products_count": (
        "Товаров"
    ),

    "no_book_cost": (
        "Без бух. себестоимости"
    ),

    "no_man_cost": (
        "Без упр. себестоимости"
    ),
}


EXPORT_COLUMNS = list(
    COLUMN_NAMES.keys()
)


# =========================================================
# Excel
# =========================================================

def build_revenue_excel(
    rows: list[dict],
    sheet_name: str,
) -> bytes:
    """
    Формирует Excel-файл таблицы структуры выручки.

    Все денежные и процентные значения:
    0 знаков после запятой.
    """

    df = pd.DataFrame(
        rows or []
    )

    if df.empty:
        df = pd.DataFrame(
            columns=EXPORT_COLUMNS
        )

    available_columns = [
        column
        for column in EXPORT_COLUMNS
        if column in df.columns
    ]

    df = df[
        available_columns
    ].copy()

    df = df.rename(
        columns=COLUMN_NAMES
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,

        engine="openpyxl",

    ) as writer:

        df.to_excel(
            writer,

            sheet_name=sheet_name,

            index=False,
        )

        workbook = writer.book

        worksheet = workbook[
            sheet_name
        ]

        # ---------------------------------------------
        # Вид листа
        # ---------------------------------------------

        worksheet.sheet_view.showGridLines = (
            False
        )

        worksheet.freeze_panes = "B2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        # ---------------------------------------------
        # Заголовки
        # ---------------------------------------------

        header_fill = PatternFill(
            fill_type="solid",

            fgColor="EAF2F8",
        )

        header_font = Font(
            bold=True,
        )

        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(
                vertical="center",
            )

        worksheet.row_dimensions[
            1
        ].height = 24

        # ---------------------------------------------
        # Форматы чисел
        # ---------------------------------------------

        percent_headers = {
            "Доля выручки, %",
            "Комиссия WB, %",
            "Маржа бух., %",
            "Маржа упр., %",
        }

        numeric_headers = {
            "Выручка с НДС",
            "НДС",
            "Выручка без НДС",
            "Комиссия WB",
            "Себестоимость бух.",
            "Прибыль бух.",
            "Себестоимость упр.",
            "Прибыль упр.",
            "Количество нетто",
            "Средняя выручка / ед.",
            "Товаров",
            "Без бух. себестоимости",
            "Без упр. себестоимости",
        }

        for column_idx, cell in enumerate(
            worksheet[1],
            start=1,
        ):

            header = cell.value

            if (
                header in percent_headers
                or header in numeric_headers
            ):

                for row_idx in range(
                    2,
                    worksheet.max_row + 1,
                ):

                    worksheet.cell(
                        row=row_idx,
                        column=column_idx,
                    ).number_format = (
                        '# ##0'
                    )

        # ---------------------------------------------
        # Ширина колонок
        # ---------------------------------------------

        for column_idx, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):

            max_length = 0

            for cell in column_cells:

                value = (
                    ""
                    if cell.value is None
                    else str(
                        cell.value
                    )
                )

                max_length = max(
                    max_length,
                    len(value),
                )

            width = min(
                max(
                    max_length + 2,
                    12,
                ),
                35,
            )

            worksheet.column_dimensions[
                get_column_letter(
                    column_idx
                )
            ].width = width

        # Первая колонка шире.
        worksheet.column_dimensions[
            "A"
        ].width = 30

    output.seek(0)

    return output.getvalue()