# gear/app/daily_sales/excel_styles.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


COLORS = {
    "dark_green": "2F6656",
    "light_gray": "F7F7F7",
    "border_gray": "D9D9D9",
    "white": "FFFFFF",
    "text": "050505",
    "warning": "FDECEC",
    "success": "EDF7F3",
    "discount": "A33A3A",
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "none": PatternFill(fill_type=None),
    "warning": PatternFill("solid", fgColor=COLORS["warning"]),
    "success": PatternFill("solid", fgColor=COLORS["success"]),
}

FONTS = {
    "header": Font(
        name="Roboto Light",
        size=10,
        bold=True,
        color=COLORS["white"],
    ),
    "normal": Font(
        name="Roboto Light",
        size=10,
        color=COLORS["text"],
    ),
    "bold": Font(
    name="Roboto Light",
    size=10,
    bold=True,
    color=COLORS["text"],
    ),

    "discount": Font(
        name="Roboto Light",
        size=10,
        color=COLORS["discount"],
    ),
}

thin = Side(style="thin", color=COLORS["border_gray"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "right": Alignment(horizontal="right", vertical="center"),
}

FORMATS = {
    "decimal": "#,##0.00",
    "date": "dd.mm.yyyy",
}



COLUMN_WIDTHS = {
    "Дата": 12,
    'Дата остатка': 15,
    "USK": 15,
    "NM ID": 15,
    "Наименование": 35,
    "Бренд": 20,
    "Категория": 22,
    "Пол": 12,

    "Остаток всего": 16,
    "На WB": 14,
    "FBS": 14,
    "В пути": 14,

    "Выручка": 16,
    "WB реализовал": 16,
    "Фин результат WB": 18,
}


def apply_excel_style(
    ws,
    freeze_panes="B2",
    numeric_columns=None,
):
    numeric_columns = numeric_columns or set()

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_panes

    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.fill = FILLS["header"]
        cell.font = FONTS["header"]
        cell.border = BORDERS["thin"]
        cell.alignment = ALIGNMENTS["center"]

    for row_idx in range(2, max_row + 1):
        is_alt = row_idx % 2 == 0

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header = ws.cell(row=1, column=col_idx).value

            cell.border = BORDERS["thin"]

            if header in {
                "Выручка",
                "WB реализовал",
                "Наименование",
                "Фин результат WB",
  
            }:
                cell.fill = FILLS["success"]

            elif header in {
                "Q без себест.",
                "Нет на складе",
                "Нет прихода",
            }:
                cell.fill = FILLS["warning"]

            else:
                cell.fill = FILLS["alt"] if is_alt else FILLS["none"]

            if header in {
                "USK",
                "Дата",
            }:
                cell.font = FONTS["bold"]

            elif header == "WB дисконт":
                cell.font = FONTS["discount"]

            else:
                cell.font = FONTS["normal"]

            if col_idx in numeric_columns:
                cell.number_format = FORMATS["decimal"]
                cell.alignment = ALIGNMENTS["right"]
            else:
                cell.alignment = ALIGNMENTS["left"]

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        header = column_cells[0].value

        # Если для колонки задана фиксированная ширина
        if header in COLUMN_WIDTHS:
            ws.column_dimensions[column_letter].width = COLUMN_WIDTHS[header]
            continue

        # Для остальных — автоматическая ширина
        max_len = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[column_letter].width = min(
            max_len + 2,
            38,
        )