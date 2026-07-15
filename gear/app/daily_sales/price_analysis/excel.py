# gear/app/daily_sales/price_analysis/excel.py
from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import CV_RANK_ORDER
from .styles import (
    BODY_FONT,
    BUTTON_FONT,
    CENTER,
    COLORS,
    FONT_NAME,
    FONT_NAME_BOLD,
    HEADER_FONT,
    LEFT,
    RIGHT,
    SMALL_MUTED_FONT,
    SUBTITLE_FONT,
    THIN_BORDER,
    TITLE_FONT,
)


TOC_SHEET_NAME = "Оглавление"

SHEET_TAB_COLORS = [
    "2F6656",
    "B45309",
    "A33A3A",
    "9A6700",
    "3B6B8F",
    "6B7280",
]


def _safe_sheet_name(name: str) -> str:
    name = str(name or "Без названия").strip()
    name = re.sub(r"[\\/*?:\[\]]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:31] or "Без названия"


def _fmt_date(value) -> str:
    if value is None:
        return "не определена"
    return pd.to_datetime(value).strftime("%d.%m.%Y")


# def _prepare_df(df: pd.DataFrame) -> pd.DataFrame:
#     work = df.copy()
    


#     for col in ["nm_id"]:
#         if col in work.columns:
#             work[col] = work[col].astype("string").fillna("")

#     date_cols = ["Первая дата УПД", "Последняя дата УПД", "Дата УПД"]
#     for col in date_cols:
#         if col in work.columns:
#             work[col] = pd.to_datetime(work[col], errors="coerce").dt.date

#     preferred = [
#         "nm_id",
#         "Наименование",
#         "Бренд",
#         "Категория",
#         "Первая дата УПД",
#         "Последняя дата УПД",
#         "Кол-во записей",
#         "Кол-во, шт",
#         "Кол-во УПД",

#         "Ранг CV, бух",
#         "Коэффициент вариации, %, бух",
#         "Медиана цены, бух",
#         "Средняя цена, бух",
#         "Ст. отклонение, руб., бух",
#         "Мин. цена, бух",
#         "Макс. цена, бух",
#         "Диапазон цены, бух",
#         "Кол-во разных цен, бух",
#         "Макс. отклонение от медианы, %, бух",
#         "Мин. отклонение от медианы, %, бух",

#         "Ранг CV, упр",
#         "Коэффициент вариации, %, упр",
#         "Медиана цены, упр",
#         "Средняя цена, упр",
#         "Ст. отклонение, руб., упр",
#         "Мин. цена, упр",
#         "Макс. цена, упр",
#         "Диапазон цены, упр",
#         "Кол-во разных цен, упр",
#         "Макс. отклонение от медианы, %, упр",
#         "Мин. отклонение от медианы, %, упр",

#         "Δ медианы упр-бух, руб.",
#         "Δ медианы упр-бух, %",
#         "История цен, бух",
#         "История цен, упр",
#     ]

#     existing = [col for col in preferred if col in work.columns]
#     other = [col for col in work.columns if col not in existing]

#     return work[existing + other]




def _fmt_date(value) -> str:
    if value is None:
        return "не определена"
    return pd.to_datetime(value).strftime("%d.%m.%Y")


def _prepare_df(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()

    if "nm_id" in work.columns:
        work["nm_id"] = (
            pd.to_numeric(work["nm_id"], errors="coerce")
            .astype("Int64")
            .astype("string")
            .fillna("")
        )

    date_cols = [
        "Первая дата УПД",
        "Последняя дата УПД",
        "Дата УПД",
    ]

    for col in date_cols:
        if col in work.columns:
            work[col] = pd.to_datetime(
                work[col],
                errors="coerce",
            ).dt.date

    preferred = [
        "nm_id",
        "Наименование",
        "Бренд",
        "Категория",
        "Первая дата УПД",
        "Последняя дата УПД",
        "Дата УПД",
        "Кол-во записей",
        "Кол-во, шт",
        "Кол-во УПД",

        "Ранг CV, бух",
        "Коэффициент вариации, %, бух",
        "Медиана цены, бух",
        "Средняя цена, бух",
        "Ст. отклонение, руб., бух",
        "Мин. цена, бух",
        "Макс. цена, бух",
        "Диапазон цены, бух",
        "Кол-во разных цен, бух",
        "Макс. отклонение от медианы, %, бух",
        "Мин. отклонение от медианы, %, бух",

        "Ранг CV, упр",
        "Коэффициент вариации, %, упр",
        "Медиана цены, упр",
        "Средняя цена, упр",
        "Ст. отклонение, руб., упр",
        "Мин. цена, упр",
        "Макс. цена, упр",
        "Диапазон цены, упр",
        "Кол-во разных цен, упр",
        "Макс. отклонение от медианы, %, упр",
        "Мин. отклонение от медианы, %, упр",

        "Δ медианы упр-бух, руб.",
        "Δ медианы упр-бух, %",
        "История цен, бух",
        "История цен, упр",
    ]

    existing = [
        col
        for col in preferred
        if col in work.columns
    ]

    other = [
        col
        for col in work.columns
        if col not in existing
    ]

    return work[existing + other]


def _autosize_columns(ws, min_width=10, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, min(ws.max_row, 250) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        ws.column_dimensions[letter].width = max(
            min_width,
            min(max_len + 2, max_width),
        )


def _add_back_button(ws, last_col):
    if ws.title == TOC_SHEET_NAME:
        return

    end_col = min(last_col, 3)
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_col,
    )

    cell = ws.cell(1, 1, "← ОГЛАВЛЕНИЕ")
    cell.hyperlink = f"#'{TOC_SHEET_NAME}'!A1"
    cell.font = BUTTON_FONT
    cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 24


def _style_title(
    ws,
    title,
    subtitle,
    start_date,
    end_date,
    last_col,
):
    ws.sheet_view.showGridLines = False
    _add_back_button(ws, last_col)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws.cell(3, 1, title)
    ws.cell(3, 1).font = TITLE_FONT
    ws.cell(3, 1).alignment = LEFT
    ws.row_dimensions[3].height = 30

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    ws.cell(
        4,
        1,
        f"{subtitle} · период: {_fmt_date(start_date)} — {_fmt_date(end_date)}",
    )
    ws.cell(4, 1).font = SUBTITLE_FONT
    ws.cell(4, 1).alignment = LEFT

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=last_col)
    ws.cell(
        5,
        1,
        f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
    )
    ws.cell(5, 1).font = SMALL_MUTED_FONT
    ws.cell(5, 1).alignment = LEFT


def _rank_fill(rank):
    mapping = {
        "0. Одна цена": COLORS["light_gray"],
        "1. До 25%": COLORS["very_light_green"],
        "2. От 25% до 50%": COLORS["light_blue"],
        "3. От 50% до 75%": COLORS["light_yellow"],
        "4. 75% и выше": COLORS["light_red"],
    }
    return mapping.get(rank, COLORS["white"])


def _style_body_cell(cell, col_name, row_idx):
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = LEFT

    if col_name in ["nm_id", "ID УПД"]:
        cell.number_format = "@"
        cell.alignment = LEFT

    elif "дата" in col_name.lower():
        cell.number_format = "DD.MM.YYYY"
        cell.alignment = CENTER

    elif col_name.startswith("Ранг CV"):
        cell.fill = PatternFill("solid", fgColor=_rank_fill(cell.value))
        cell.font = Font(
            name=FONT_NAME_BOLD,
            size=9,
            bold=True,
            color=COLORS["text"],
        )
        cell.alignment = CENTER

    elif "%," in col_name or col_name.endswith(", %"):
        cell.number_format = '0.00"%"'
        cell.alignment = RIGHT

    elif any(token in col_name for token in [
        "цена",
        "Медиана",
        "Средняя",
        "отклонение, руб.",
        "Диапазон",
        "Δ медианы упр-бух, руб.",
    ]):
        cell.number_format = '#,##0.00" ₽"'
        cell.alignment = RIGHT

    elif any(token in col_name for token in [
        "Кол-во",
        "Количество",
    ]):
        cell.number_format = "#,##0"
        cell.alignment = RIGHT

    elif row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=COLORS["light_gray"])


def _write_dataframe_sheet(
    wb,
    sheet_name,
    df,
    start_date,
    end_date,
    title,
    subtitle,
):
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    work = df.copy()
    last_col = max(len(work.columns), 1)

    _style_title(
        ws,
        title,
        subtitle,
        start_date,
        end_date,
        last_col,
    )

    header_row = 7

    for col_idx, col_name in enumerate(work.columns, start=1):
        cell = ws.cell(header_row, col_idx, col_name)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(
        work.itertuples(index=False),
        start=header_row + 1,
    ):
        ws.row_dimensions[row_idx].height = 20

        for col_idx, value in enumerate(row, start=1):
            col_name = work.columns[col_idx - 1]
            cell = ws.cell(row_idx, col_idx, value)
            _style_body_cell(cell, col_name, row_idx)

    last_row = header_row + len(work)
    ws.freeze_panes = f"D{header_row + 1}"

    if len(work):
        ws.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(last_col)}{last_row}"
        )

    _autosize_columns(ws)

    custom_widths = {
        "nm_id": 16,
        "Наименование": 42,
        "Бренд": 20,
        "Категория": 24,
        "Первая дата УПД": 14,
        "Последняя дата УПД": 14,
        "Ранг CV, бух": 21,
        "Ранг CV, упр": 21,
        "История цен, бух": 55,
        "История цен, упр": 55,
        "Поставщик": 35, 
    }

    for col_name, width in custom_widths.items():
        if col_name in work.columns:
            letter = get_column_letter(work.columns.get_loc(col_name) + 1)
            ws.column_dimensions[letter].width = width

    for col_name in [
        "Коэффициент вариации, %, бух",
        "Коэффициент вариации, %, упр",
    ]:
        if col_name in work.columns and len(work):
            col_idx = work.columns.get_loc(col_name) + 1
            letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{letter}{header_row + 1}:{letter}{last_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color="E7F1ED",
                    mid_type="num",
                    mid_value=50,
                    mid_color="FFF6D8",
                    end_type="num",
                    end_value=100,
                    end_color="FDECEC",
                ),
            )

    return ws.title


def _build_summary_sheet(
    wb,
    df,
    start_date,
    end_date,
):
    ws = wb.create_sheet("Сводка")
    ws.sheet_view.showGridLines = False
    _add_back_button(ws, 10)

    ws.merge_cells("A3:J3")
    ws["A3"] = "АНАЛИЗ СЕБЕСТОИМОСТИ"
    ws["A3"].font = Font(
        name=FONT_NAME_BOLD,
        size=19,
        bold=True,
        color=COLORS["dark_green"],
    )

    ws.merge_cells("A4:J4")
    ws["A4"] = (
        f"Период УПД: {_fmt_date(start_date)} — {_fmt_date(end_date)}"
    )
    ws["A4"].font = SUBTITLE_FONT

    ws.merge_cells("A5:J5")
    ws["A5"] = (
        f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}"
    )
    ws["A5"].font = SMALL_MUTED_FONT

    total_products = len(df)
    total_documents = pd.to_numeric(
        df.get("Кол-во УПД", 0),
        errors="coerce",
    ).fillna(0).sum()

    critical_buh = (
        df.get("Ранг CV, бух", pd.Series(dtype=str))
        .eq("4. 75% и выше")
        .sum()
    )
    critical_upr = (
        df.get("Ранг CV, упр", pd.Series(dtype=str))
        .eq("4. 75% и выше")
        .sum()
    )
    one_price_buh = (
        df.get("Ранг CV, бух", pd.Series(dtype=str))
        .eq("0. Одна цена")
        .sum()
    )
    one_price_upr = (
        df.get("Ранг CV, упр", pd.Series(dtype=str))
        .eq("0. Одна цена")
        .sum()
    )

    cards = [
        ("Товаров", total_products, "NM ID"),
        ("УПД", total_documents, "документов"),
        ("Критический CV, бух", critical_buh, "75% и выше"),
        ("Критический CV, упр", critical_upr, "75% и выше"),
        ("Одна цена, бух", one_price_buh, "нет динамики"),
        ("Одна цена, упр", one_price_upr, "нет динамики"),
    ]

    for idx, (title, value, subtitle) in enumerate(cards):
        row = 7 + (idx // 3) * 4
        col = 1 + (idx % 3) * 3

        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + 1,
        )
        title_cell = ws.cell(row, col, title)
        title_cell.font = Font(
            name=FONT_NAME_BOLD,
            size=9,
            bold=True,
            color=COLORS["gray"],
        )
        title_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_green"],
        )
        title_cell.alignment = CENTER

        ws.merge_cells(
            start_row=row + 1,
            start_column=col,
            end_row=row + 1,
            end_column=col + 1,
        )
        value_cell = ws.cell(row + 1, col, value)
        value_cell.font = Font(
            name=FONT_NAME_BOLD,
            size=15,
            bold=True,
            color=(
                COLORS["red"]
                if "Критический" in title
                else COLORS["dark_green"]
            ),
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=(
                COLORS["light_red"]
                if "Критический" in title
                else COLORS["light_gray"]
            ),
        )
        value_cell.alignment = CENTER
        value_cell.number_format = "#,##0"

        ws.merge_cells(
            start_row=row + 2,
            start_column=col,
            end_row=row + 2,
            end_column=col + 1,
        )
        sub_cell = ws.cell(row + 2, col, subtitle)
        sub_cell.font = Font(
            name=FONT_NAME,
            size=8,
            color=COLORS["gray"],
        )
        sub_cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["light_gray"],
        )
        sub_cell.alignment = CENTER

        for rr in range(row, row + 3):
            for cc in range(col, col + 2):
                ws.cell(rr, cc).border = THIN_BORDER

    rank_start = 17
    ws.cell(rank_start, 1, "Распределение по рангам")
    ws.cell(rank_start, 1).font = Font(
        name=FONT_NAME_BOLD,
        size=12,
        bold=True,
        color=COLORS["dark_green"],
    )

    headers = ["Ранг", "Бухгалтерская", "Управленческая"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(rank_start + 2, idx, header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["dark_green"],
        )
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_offset, rank in enumerate(CV_RANK_ORDER, start=1):
        row = rank_start + 2 + row_offset
        values = [
            rank,
            int(df.get("Ранг CV, бух", pd.Series(dtype=str)).eq(rank).sum()),
            int(df.get("Ранг CV, упр", pd.Series(dtype=str)).eq(rank).sum()),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx > 1 else LEFT
            cell.font = BODY_FONT
            cell.fill = PatternFill("solid", fgColor=_rank_fill(rank))

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions["A"].width = 24

    return ws.title


def _build_toc_sheet(wb, sheets_info, start_date, end_date):
    ws = wb.create_sheet(TOC_SHEET_NAME, 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:D2")
    ws["B2"] = "ОГЛАВЛЕНИЕ — АНАЛИЗ СЕБЕСТОИМОСТИ"
    ws["B2"].font = Font(
        name=FONT_NAME_BOLD,
        size=18,
        bold=True,
        color=COLORS["dark_green"],
    )

    ws.merge_cells("B3:D3")
    ws["B3"] = (
        f"Период УПД: {_fmt_date(start_date)} — {_fmt_date(end_date)}"
    )
    ws["B3"].font = SUBTITLE_FONT

    ws.merge_cells("B4:D4")
    ws["B4"] = "Нажмите на название листа для перехода"
    ws["B4"].font = SMALL_MUTED_FONT

    start_row = 7
    headers = ["№", "Лист", "Описание"]

    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(
            "solid",
            fgColor=COLORS["dark_green"],
        )
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for idx, item in enumerate(sheets_info, start=1):
        row = start_row + idx
        values = [
            f"{idx:02d}",
            item["name"],
            item["description"],
        ]

        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
            cell.font = BODY_FONT

            if col_idx == 3:
                cell.hyperlink = f"#'{item['name']}'!A1"
                cell.font = Font(
                    name=FONT_NAME_BOLD,
                    size=10,
                    bold=True,
                    color=COLORS["link"],
                )

            if idx % 2 == 0:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLORS["light_gray"],
                )

        ws.row_dimensions[row].height = 26

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 72


def _apply_sheet_tab_colors(wb):
    for idx, ws in enumerate(wb.worksheets):
        ws.sheet_properties.tabColor = (
            SHEET_TAB_COLORS[idx % len(SHEET_TAB_COLORS)]
        )


def make_price_analysis_excel(
    analysis_df: pd.DataFrame,
    history_df: pd.DataFrame,
    start_date=None,
    end_date=None,
) -> bytes:
    analysis_df = _prepare_df(analysis_df)
    history_df = _prepare_df(history_df)

    wb = Workbook()
    wb.remove(wb.active)

    sheets_info = []



    all_name = _write_dataframe_sheet(
        wb,
        "Все товары",
        analysis_df,
        start_date,
        end_date,
        "Детальный анализ себестоимости",
        "Бухгалтерская и управленческая себестоимость по товарам",
    )
    sheets_info.append({
        "name": all_name,
        "description": "Полный анализ по всем NM ID",
    })

    

    
    if not history_df.empty:
        history_name = _write_dataframe_sheet(
            wb,
            "История цен",
            history_df,
            start_date,
            end_date,
            "История себестоимости по УПД",
            "Детализация по датам и документам",
        )
        sheets_info.append({
            "name": history_name,
            "description": "Построчная история бухгалтерской и управленческой цены",
        })

    _build_toc_sheet(
        wb,
        sheets_info,
        start_date,
        end_date,
    )
    _apply_sheet_tab_colors(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()
