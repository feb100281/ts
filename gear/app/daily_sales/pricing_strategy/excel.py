# gear/app/daily_sales/pricing_strategy/excel.py
"""
EXCEL ПО ЦЕНОВЫМ РЕШЕНИЯМ.

Файл делается не «чтобы выгрузить данные», а чтобы его можно
было открыть на совещании и работать прямо в нём:

    — понятные русские заголовки вместо технических имён полей;
    — денежные и процентные форматы, а не голые числа;
    — цветом подсвечено то, что требует решения;
    — закреплённая шапка и автофильтр на каждом листе;
    — отдельный лист с методикой — тем же текстом, что в
      приложении, чтобы не пересказывать её на словах.

Порядок листов повторяет порядок разговора: сначала итог,
потом проблема, потом весь ассортимент, потом детали.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .config import TARGET_MARGIN_PCT
from .methodology import methodology_rows


# ============================================================
# ОФОРМЛЕНИЕ
# ============================================================

FONT_NAME = "Helvetica Light"
FONT_SIZE = 10

COLOR_HEADER = "1E293B"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_BORDER = "D8DEE6"
COLOR_ZEBRA = "F8FAFC"
COLOR_TITLE = "0F172A"

FILL_DANGER = "FEE2E2"
FILL_WARNING = "FEF3C7"
FILL_GOOD = "DCFCE7"
FILL_ACCENT = "EEF2FF"

TEXT_DANGER = "B91C1C"
TEXT_WARNING = "92400E"
TEXT_GOOD = "166534"

MONEY_FORMAT = '# ##0 ₽;[Red]-# ##0 ₽;"—"'
MONEY_2_FORMAT = '# ##0.00 ₽;[Red]-# ##0.00 ₽;"—"'
INT_FORMAT = '# ##0;[Red]-# ##0;"—"'
PCT_FORMAT = '0.0" %";[Red]-0.0" %";"—"'
SIGNED_PCT_FORMAT = '+0.0" %";-0.0" %";"0 %"'
RATIO_FORMAT = '0.00;[Red]-0.00;"—"'

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ============================================================
# ОПИСАНИЕ КОЛОНОК
#
# Один словарь на всё: и заголовок, и формат, и ширина.
# Добавили поле в анализ — дописали строку здесь, и оно
# появилось во всех выгрузках сразу.
# ============================================================

def _c(field, title, fmt=None, width=16):
    return {
        "field": field,
        "title": title,
        "format": fmt,
        "width": width,
    }


DECISION_COLUMNS = [
    _c("status_ru", "Что делать", None, 16),
    _c("priority", "Приоритет", INT_FORMAT, 11),
    _c("nm_id", "NM ID", None, 13),
    _c("sa_name", "Артикул", None, 16),
    _c("brand", "Бренд", None, 18),
    _c("category", "Категория", None, 22),
    _c("title", "Наименование", None, 38),

    _c("current_effective_price", "Текущая цена", MONEY_FORMAT, 15),
    _c("buyer_price_30d", "Цена покупателя 30д", MONEY_FORMAT, 18),
    _c("wb_discount_pct_30d", "Скидка WB 30д", PCT_FORMAT, 14),

    _c("breakeven_price", "Минимальная цена", MONEY_FORMAT, 17),
    _c(
        "target_margin_price",
        f"Цена под маржу {TARGET_MARGIN_PCT:.0f}%",
        MONEY_FORMAT,
        19,
    ),
    _c("price_headroom_pct", "Запас по скидке", PCT_FORMAT, 15),

    _c("action_price", "Цена к установке", MONEY_FORMAT, 17),
    _c("action_change_pct", "Изменение цены", SIGNED_PCT_FORMAT, 15),

    _c("unit_cogs", "Упр. с/с единицы", MONEY_FORMAT, 17),
    _c("unit_acc_cost", "Бух. с/с единицы", MONEY_FORMAT, 17),
    _c("ratios_source", "Источник коэффициентов", None, 22),
    _c("cost_source", "Источник себестоимости", None, 20),

    _c("total_stock", "Остаток всего", INT_FORMAT, 14),
    _c("wb_stock", "WB", INT_FORMAT, 10),
    _c("fbs_stock", "FBS", INT_FORMAT, 10),
    _c("in_transit", "В пути", INT_FORMAT, 11),
    _c("days_of_stock", "Запас, дн.", INT_FORMAT, 12),
    _c("stock_age_days", "Возраст, дн.", INT_FORMAT, 13),

    _c("sales_qty_7d", "Продажи 7д", INT_FORMAT, 12),
    _c("sales_qty_30d", "Продажи 30д", INT_FORMAT, 13),
    _c("sales_speed_trend_pct", "Тренд скорости", PCT_FORMAT, 14),

    _c("margin_man_30d", "Маржа 30д", MONEY_FORMAT, 14),
    _c("margin_pct_30d", "Маржа %", PCT_FORMAT, 11),
    _c("stock_at_risk_value", "Убыток на остатке", MONEY_FORMAT, 18),
    _c("margin_upside_day", "Потенциал ₽/день", MONEY_FORMAT, 17),

    _c("elasticity", "Эластичность", RATIO_FORMAT, 13),
    _c("elasticity_confidence", "Надёжность", None, 18),
    _c("reason", "Почему", None, 80),
]


PORTFOLIO_COLUMNS = [
    _c("brand", "Бренд", None, 20),
    _c("category", "Категория", None, 24),
    _c("products", "Артикулов", INT_FORMAT, 12),
    _c("action_products", "Требуют действия", INT_FORMAT, 17),
    _c("below_breakeven", "Ниже минимальной цены", INT_FORMAT, 21),
    _c("stock_at_risk_value", "Убыток на остатке", MONEY_FORMAT, 18),
    _c("stock_units", "Остаток всего", INT_FORMAT, 14),
    _c("wb_stock", "WB", INT_FORMAT, 10),
    _c("fbs_stock", "FBS", INT_FORMAT, 10),
    _c("in_transit", "В пути", INT_FORMAT, 11),
    _c("sales_30d", "Продажи 30д", INT_FORMAT, 13),
    _c("stock_days", "Запас, дн.", INT_FORMAT, 12),
    _c("current_margin_30d", "Маржа 30д", MONEY_FORMAT, 15),
    _c("margin_upside_day", "Потенциал ₽/день", MONEY_FORMAT, 17),
]


SCENARIO_COLUMNS = [
    _c("nm_id", "NM ID", None, 13),
    _c("brand", "Бренд", None, 18),
    _c("category", "Категория", None, 22),
    _c("price_change_pct", "Изменение цены", SIGNED_PCT_FORMAT, 15),
    _c("seller_price", "Наша цена", MONEY_FORMAT, 14),
    _c("buyer_price", "Цена покупателя", MONEY_FORMAT, 16),
    _c("projected_daily_qty", "Продаж в день", RATIO_FORMAT, 14),
    _c("projected_qty", "Продажи 30д, прогноз", INT_FORMAT, 19),
    _c("projected_margin", "Маржа 30д, прогноз", MONEY_FORMAT, 18),
    _c("projected_margin_pct", "Маржа %, прогноз", PCT_FORMAT, 16),
    _c("projected_stock_days", "Запас, дн.", INT_FORMAT, 12),
]


HISTORY_COLUMNS = [
    _c("nm_id", "NM ID", None, 13),
    _c("date_from", "Дата", None, 12),
    _c("sales_qty", "Продажи", INT_FORMAT, 11),
    _c("returns_qty", "Возвраты", INT_FORMAT, 11),
    _c("seller_price", "Наша цена", MONEY_2_FORMAT, 14),
    _c("buyer_price", "Цена покупателя", MONEY_2_FORMAT, 16),
    _c("amount_vatless", "Выручка без НДС", MONEY_FORMAT, 16),
    _c("cogs_man", "Себестоимость", MONEY_FORMAT, 15),
    _c("net_comission", "Комиссия WB", MONEY_FORMAT, 14),
    _c("margin_man", "Маржа", MONEY_FORMAT, 13),
]


STATUS_RU = {
    "LOSS": "Убыток",
    "CLEARANCE": "Распродажа",
    "REDUCE": "Снизить цену",
    "RAISE": "Повысить цену",
    "TEST": "Тест цены",
    "HOLD": "Оставить цену",
}


# ============================================================
# ПОДГОТОВКА ДАННЫХ
# ============================================================

def to_frame(rows) -> pd.DataFrame:
    if rows is None:
        return pd.DataFrame()

    if isinstance(rows, pd.DataFrame):
        return rows.copy()

    return pd.DataFrame(rows or [])


def prepare_decisions(payload) -> pd.DataFrame:
    """Таблица решений с человеческим статусом."""

    frame = to_frame(payload.get("recommendations"))

    if frame.empty:
        return frame

    frame = frame.copy()

    frame["status_ru"] = (
        frame.get("status", pd.Series(dtype=object))
        .map(STATUS_RU)
        .fillna(frame.get("status"))
    )

    return frame


def loss_frame(payload) -> pd.DataFrame:
    """
    Только то, что горит: цена ниже точки безубыточности.
    Отсортировано по сумме потерь, а не по проценту —
    сначала разбираем то, где больше денег.
    """

    frame = prepare_decisions(payload)

    if frame.empty or "below_breakeven" not in frame.columns:
        return pd.DataFrame()

    mask = frame["below_breakeven"].fillna(False).astype(bool)

    frame = frame[mask].copy()

    if frame.empty:
        return frame

    frame["stock_at_risk_value"] = pd.to_numeric(
        frame.get("stock_at_risk_value"),
        errors="coerce",
    ).fillna(0)

    return frame.sort_values(
        "stock_at_risk_value",
        ascending=False,
    )


# ============================================================
# ЗАПИСЬ ЛИСТА
# ============================================================

# Поле, по которому закрепляем колонки (freeze panes) на
# листах с товарными таблицами: «Наименование» и всё левее
# него остаются на месте при прокрутке вправо. На листах, где
# такого поля нет (портфель, сценарии, история), закрепляется
# только шапка — как и раньше.
FREEZE_COLUMN_FIELD = "title"


def _new_sheet(workbook, name):
    """
    Создаёт лист и сразу убирает сетку ячеек (gridlines) —
    так таблица со своими границами и подсветкой читается
    чище, без лишних серых линий вокруг.
    """

    ws = workbook.create_sheet(name)
    ws.sheet_view.showGridLines = False

    return ws


def _write_title(ws, title, subtitle, width):

    cell = ws.cell(1, 1, title)
    cell.font = Font(
        name=FONT_NAME,
        size=FONT_SIZE,
        bold=True,
        color=COLOR_TITLE,
    )

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max(width, 1),
    )

    cell = ws.cell(2, 1, subtitle)
    cell.font = Font(
        name=FONT_NAME,
        size=FONT_SIZE,
        italic=True,
        color="64748B",
    )

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=max(width, 1),
    )

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def write_sheet(
    ws,
    frame: pd.DataFrame,
    columns: list[dict],
    *,
    title: str,
    subtitle: str = "",
    highlight: bool = True,
    styled_rows: bool = True,
    max_rows: int | None = None,
):
    """
    Пишет таблицу с шапкой, форматами и подсветкой.

    highlight=True включает смысловую подсветку: убыток —
    красным, пограничные значения — жёлтым, рост — зелёным.

    styled_rows=False отключает оформление КАЖДОЙ ячейки —
    шрифт, границы, чередование строк. Нужно для больших
    листов: на десятках тысяч строк пооблачное оформление
    занимает больше времени, чем сама выгрузка, и запрос
    успевает отвалиться. Числовые форматы остаются.
    """

    frame = to_frame(frame)

    available = [
        column
        for column in columns
        if column["field"] in frame.columns
    ]

    if frame.empty or not available:
        _write_title(ws, title, subtitle or "Нет данных", 6)

        ws.cell(4, 1, "Нет данных для этого раздела.").font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            italic=True,
            color="94A3B8",
        )
        return

    truncated = False

    total_rows = len(frame)

    if max_rows and total_rows > max_rows:
        frame = frame.head(max_rows)
        truncated = True

    _write_title(ws, title, subtitle, len(available))

    header_row = 4

    for index, column in enumerate(available, start=1):

        cell = ws.cell(
            header_row,
            index,
            column["title"],
        )

        cell.font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            bold=True,
            color=COLOR_HEADER_TEXT,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = BORDER

        ws.column_dimensions[
            get_column_letter(index)
        ].width = column["width"]

    ws.row_dimensions[header_row].height = 34

    fields = [column["field"] for column in available]

    values = frame[fields].where(
        pd.notna(frame[fields]),
        None,
    )

    for row_index, record in enumerate(
        values.to_dict("records"),
        start=header_row + 1,
    ):

        zebra = (row_index - header_row) % 2 == 0

        for col_index, column in enumerate(
            available,
            start=1,
        ):

            value = record.get(column["field"])

            if (
                value is not None
                and hasattr(value, "isoformat")
                and not isinstance(value, str)
            ):
                try:
                    value = value.isoformat()
                except Exception:
                    pass

            if isinstance(value, bool):
                value = "да" if value else ""

            cell = ws.cell(row_index, col_index, value)

            if column["format"]:
                cell.number_format = column["format"]

            if not styled_rows:
                continue

            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                color="0F172A",
            )

            cell.border = BORDER

            if column["format"]:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=(column["width"] > 30),
                )

            if zebra:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=COLOR_ZEBRA,
                )

            if highlight:
                _highlight(cell, column["field"], value)

    # Закрепляем шапку и, если на листе есть колонка
    # «Наименование», всё, что левее неё включительно — чтобы
    # название товара не убегало при прокрутке вправо.
    freeze_column = 1

    for index, column in enumerate(available, start=1):
        if column["field"] == FREEZE_COLUMN_FIELD:
            freeze_column = index + 1
            break

    ws.freeze_panes = ws.cell(
        header_row + 1,
        freeze_column,
    ).coordinate

    ws.auto_filter.ref = (
        f"A{header_row}:"
        f"{get_column_letter(len(available))}"
        f"{header_row + len(frame)}"
    )

    if truncated:
        note_row = header_row + len(frame) + 2

        ws.cell(
            note_row,
            1,
            (
                f"Показаны {max_rows:,} строк из "
                f"{total_rows:,}: лист ограничен, чтобы файл "
                f"собирался за разумное время. Полные данные "
                f"есть в приложении."
            ).replace(",", " "),
        ).font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            italic=True,
            color=TEXT_WARNING,
        )


def _highlight(cell, field, value):
    """Смысловая подсветка отдельных колонок."""

    if value is None:
        return

    try:
        number = float(value)
    except (TypeError, ValueError):
        number = None

    if field == "status_ru":

        if value == "Убыток":
            cell.fill = PatternFill("solid", fgColor=FILL_DANGER)
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=TEXT_DANGER,
            )

        elif value == "Распродажа":
            cell.fill = PatternFill("solid", fgColor=FILL_WARNING)
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=TEXT_WARNING,
            )

        elif value == "Повысить цену":
            cell.fill = PatternFill("solid", fgColor=FILL_GOOD)
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=TEXT_GOOD,
            )

        return

    if number is None:
        return

    if field == "price_headroom_pct":

        if number < 0:
            cell.fill = PatternFill("solid", fgColor=FILL_DANGER)
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=TEXT_DANGER,
            )

        elif number < 10:
            cell.fill = PatternFill("solid", fgColor=FILL_WARNING)

        return

    if field in ("stock_at_risk_value", "below_breakeven"):

        if number > 0:
            cell.fill = PatternFill("solid", fgColor=FILL_DANGER)
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=TEXT_DANGER,
            )

        return

    if field in ("breakeven_price", "action_price"):
        cell.fill = PatternFill("solid", fgColor=FILL_ACCENT)
        cell.font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            bold=True,
            color="1E293B",
        )
        return

    if field in ("margin_man_30d", "margin_pct_30d"):

        if number < 0:
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=TEXT_DANGER,
            )

        return

    if field == "days_of_stock":

        if number >= 365:
            cell.fill = PatternFill("solid", fgColor=FILL_DANGER)

        elif number >= 180:
            cell.fill = PatternFill("solid", fgColor=FILL_WARNING)


# ============================================================
# ЛИСТ СВОДКИ
# ============================================================

def _write_summary(ws, payload):

    summary = payload.get("summary") or {}

    _write_title(
        ws,
        "Ценовые решения по товарному остатку",
        (
            f"Дата анализа: {payload.get('report_date', '—')} · "
            f"остатки WB: {payload.get('wb_date', '—')} · "
            f"FBS: {payload.get('fbs_date', '—')}"
        ),
        4,
    )

    rows = [
        ("Что в выборке", None, None),
        (
            "Артикулов с остатком",
            summary.get("products"),
            INT_FORMAT,
        ),
        (
            "Общий остаток, шт",
            summary.get("stock_units"),
            INT_FORMAT,
        ),
        (
            "  в том числе WB",
            summary.get("wb_stock_units"),
            INT_FORMAT,
        ),
        (
            "  в том числе FBS",
            summary.get("fbs_stock_units"),
            INT_FORMAT,
        ),
        (
            "  в том числе в пути",
            summary.get("in_transit_units"),
            INT_FORMAT,
        ),
        (None, None, None),

        ("Где мы теряем", None, None),
        (
            "Продаём ниже минимальной цены, артикулов",
            summary.get("below_breakeven_products"),
            INT_FORMAT,
        ),
        (
            "Потенциальный убыток на остатке",
            summary.get("stock_at_risk_value"),
            MONEY_FORMAT,
        ),
        (
            "На грани: запас по скидке меньше 10%",
            summary.get("margin_at_risk_products"),
            INT_FORMAT,
        ),
        (
            "Без себестоимости: минимальная цена не посчитана",
            summary.get("no_cost_products"),
            INT_FORMAT,
        ),
        (None, None, None),

        ("Что делать", None, None),
        (
            "Требуют действия, артикулов",
            summary.get("action_products"),
            INT_FORMAT,
        ),
        (
            "Кандидаты на распродажу",
            summary.get("clearance_products"),
            INT_FORMAT,
        ),
        (
            "Можно повышать цену",
            summary.get("raise_products"),
            INT_FORMAT,
        ),
        (
            "Модельный потенциал маржи, ₽ в день",
            summary.get("margin_upside_day"),
            MONEY_FORMAT,
        ),
    ]

    row_index = 4

    for label, value, fmt in rows:

        if label is None:
            row_index += 1
            continue

        cell = ws.cell(row_index, 1, label)

        if value is None:
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=COLOR_HEADER,
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=FILL_ACCENT,
            )

            ws.merge_cells(
                start_row=row_index,
                start_column=1,
                end_row=row_index,
                end_column=2,
            )

        else:
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE)

            value_cell = ws.cell(row_index, 2, value)

            value_cell.number_format = fmt or INT_FORMAT

            value_cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
            )

            value_cell.alignment = Alignment(horizontal="right")

        row_index += 1

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22

    note = (
        "Маржа в отчёте — выручка без НДС минус себестоимость "
        "минус комиссия WB. Логистика, реклама, хранение, "
        "приёмка и штрафы в неё не входят. Подробности — на "
        "листе «Методика»."
    )

    ws.cell(row_index + 1, 1, note).font = Font(
        name=FONT_NAME,
        size=FONT_SIZE,
        italic=True,
        color="64748B",
    )


# ============================================================
# ЛИСТ МЕТОДИКИ
# ============================================================

def _write_methodology(ws):

    _write_title(
        ws,
        "Методика расчёта",
        "Тот же текст, что в приложении по кнопке «Как это считается»",
        3,
    )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 96

    row_index = 4

    for section, term, text in methodology_rows():

        if section:
            cell = ws.cell(row_index, 1, section)
            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
                color=COLOR_HEADER,
            )

            ws.cell(row_index, 3, text).font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                italic=True,
                color="475569",
            )

        elif term:
            ws.cell(row_index, 2, term).font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
                bold=True,
            )

            ws.cell(row_index, 3, text).font = Font(
                name=FONT_NAME,
                size=FONT_SIZE,
            )

        for column in (1, 2, 3):
            ws.cell(row_index, column).alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

        row_index += 1


# ============================================================
# ГЛАВНАЯ СБОРКА
# ============================================================

def build_pricing_excel(payload) -> bytes:
    """Полный отчёт: сводка, решения, убытки, портфель, детали."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    decisions = prepare_decisions(payload)

    _write_summary(
        _new_sheet(workbook, "Сводка"),
        payload,
    )

    write_sheet(
        _new_sheet(workbook, "Ценовые решения"),
        decisions,
        DECISION_COLUMNS,
        title="Ценовые решения по каждому артикулу",
        subtitle=(
            "Минимальная цена — граница убытка. Цена к установке "
            "никогда не ниже минимальной, кроме распродажи."
        ),
    )

    write_sheet(
        _new_sheet(workbook, "Продаём в убыток"),
        loss_frame(payload),
        DECISION_COLUMNS,
        title="Товары, которые продаются ниже минимальной цены",
        subtitle=(
            "Отсортировано по сумме потерь на остатке: сверху то, "
            "где больше денег."
        ),
    )

    write_sheet(
        _new_sheet(workbook, "Бренды и категории"),
        to_frame(payload.get("portfolio")),
        PORTFOLIO_COLUMNS,
        title="Срез по брендам и категориям",
        subtitle="Где сосредоточен запас и где сосредоточены потери",
    )

    write_sheet(
        _new_sheet(workbook, "Сценарии"),
        to_frame(payload.get("scenarios")),
        SCENARIO_COLUMNS,
        title="Сценарии изменения цены",
        subtitle=(
            "Прогноз продаж и маржи на 30 дней при разных "
            "изменениях нашей цены"
        ),
        highlight=False,
        styled_rows=False,
        max_rows=30_000,
    )

    history = to_frame(payload.get("history"))

    # Сортируем от свежих дат к старым: если строк окажется
    # больше потолка, в файл попадут последние дни, а не
    # первые артикулы по алфавиту.
    if not history.empty and "date_from" in history.columns:
        history = history.sort_values(
            ["date_from", "nm_id"],
            ascending=[False, True],
        )

    write_sheet(
        _new_sheet(workbook, "История"),
        history,
        HISTORY_COLUMNS,
        title="История продаж, цен и маржи по дням",
        subtitle=(
            "Исходные данные, на которых построена модель. "
            "Свежие даты сверху."
        ),
        highlight=False,
        styled_rows=False,
        max_rows=30_000,
    )

    _write_methodology(
        _new_sheet(workbook, "Методика"),
    )

    for name, color in (
        ("Сводка", "1E293B"),
        ("Ценовые решения", "4F46E5"),
        ("Продаём в убыток", "B91C1C"),
        ("Бренды и категории", "0F766E"),
        ("Сценарии", "0369A1"),
        ("История", "64748B"),
        ("Методика", "CA8A04"),
    ):
        if name in workbook.sheetnames:
            workbook[name].sheet_properties.tabColor = color

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return stream.getvalue()


def build_loss_excel(payload) -> bytes:
    """Короткий файл к действию: только убыточные позиции."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    frame = loss_frame(payload)

    write_sheet(
        _new_sheet(workbook, "Продаём в убыток"),
        frame,
        DECISION_COLUMNS,
        title="Товары, которые продаются ниже минимальной цены",
        subtitle=(
            f"Дата анализа: {payload.get('report_date', '—')}. "
            "Каждая проданная единица уменьшает прибыль."
        ),
    )

    _write_methodology(
        _new_sheet(workbook, "Методика"),
    )

    workbook["Продаём в убыток"].sheet_properties.tabColor = "B91C1C"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return stream.getvalue()


def build_product_excel(payload, nm_id) -> bytes:
    """Разбор одного товара: решение, сценарии, история."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    nm_id = str(nm_id)

    decisions = prepare_decisions(payload)

    if not decisions.empty:
        decisions = decisions[
            decisions["nm_id"].astype(str) == nm_id
        ]

    scenarios = to_frame(payload.get("scenarios"))

    if not scenarios.empty:
        scenarios = scenarios[
            scenarios["nm_id"].astype(str) == nm_id
        ]

    history = to_frame(payload.get("history"))

    if not history.empty:
        history = history[
            history["nm_id"].astype(str) == nm_id
        ]

    title = nm_id

    if not decisions.empty:
        title = (
            f"{decisions.iloc[0].get('brand', '')} · "
            f"{decisions.iloc[0].get('title', '')}"
        )

    write_sheet(
        _new_sheet(workbook, "Решение"),
        decisions,
        DECISION_COLUMNS,
        title=f"NM ID {nm_id} · {title}",
        subtitle="Итоговое ценовое решение и его основание",
    )

    write_sheet(
        _new_sheet(workbook, "Сценарии"),
        scenarios,
        SCENARIO_COLUMNS,
        title="Сценарии изменения цены",
        subtitle="Прогноз на 30 дней",
        highlight=False,
    )

    write_sheet(
        _new_sheet(workbook, "История"),
        history,
        HISTORY_COLUMNS,
        title="История продаж и цен",
        subtitle="Данные, на которых построена модель",
        highlight=False,
    )

    _write_methodology(
        _new_sheet(workbook, "Методика"),
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return stream.getvalue()
