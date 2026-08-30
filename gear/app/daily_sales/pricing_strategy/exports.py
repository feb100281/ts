# gear/app/daily_sales/pricing_strategy/exports.py
"""
ФАЙЛЫ, КОТОРЫЕ РЕАЛЬНО НУЖНЫ.

Не «выгрузка всего в Excel», а четыре разных документа под
четыре разные задачи:

    1. Ценовые решения (xlsx)   — разобрать и обсудить
    2. Продаём в убыток (xlsx)  — короткий список к действию
    3. Новые цены (csv)         — взять и загрузить на WB
    4. Пакет (zip)              — отдать целиком, с методикой

CSV сознательно сделан минимальным: nmID и цена. Любые лишние
колонки в таком файле мешают загрузке, а всё остальное всегда
можно посмотреть в Excel.
"""

from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd

from .excel import (
    build_loss_excel,
    build_pricing_excel,
    build_product_excel,
    loss_frame,
    prepare_decisions,
)
from .methodology import methodology_text


# Меньше этого изменения цену не трогаем: суета вокруг
# половины процента дороже, чем эффект от неё.
MIN_MEANINGFUL_CHANGE_PCT = 1.0


def _stamp(payload) -> str:
    report_date = str(
        payload.get("report_date")
        or datetime.now().date()
    )

    return report_date[:10]


# ============================================================
# CSV ДЛЯ ЗАГРУЗКИ ЦЕН
# ============================================================

def price_change_frame(payload) -> pd.DataFrame:
    """
    Артикулы, у которых цену действительно нужно поменять.

    Отбор:
        — цена к установке посчитана;
        — изменение не меньше 1%;
        — товар не в статусе «Оставить цену».
    """

    frame = prepare_decisions(payload)

    if frame.empty:
        return pd.DataFrame()

    if "action_price" not in frame.columns:
        return pd.DataFrame()

    work = frame.copy()

    work["action_price"] = pd.to_numeric(
        work["action_price"],
        errors="coerce",
    )

    work["action_change_pct"] = pd.to_numeric(
        work.get("action_change_pct"),
        errors="coerce",
    )

    mask = (
        work["action_price"].notna()
        & (work["action_price"] > 0)
        & work["action_change_pct"].notna()
        & (
            work["action_change_pct"].abs()
            >= MIN_MEANINGFUL_CHANGE_PCT
        )
    )

    work = work[mask].copy()

    if work.empty:
        return work

    # Цена на WB задаётся целыми рублями.
    work["price_to_set"] = (
        work["action_price"]
        .round(0)
        .astype("Int64")
    )

    return work.sort_values(
        "priority",
        ascending=False,
    )


def build_wb_price_csv(payload) -> bytes:
    """
    Минимальный файл для загрузки цен: nmID и цена.

    Кодировка UTF-8 с BOM — иначе Excel на Windows
    показывает кракозябры, и файл считают битым.
    Разделитель «;» — привычный для русской локали.
    """

    frame = price_change_frame(payload)

    buffer = StringIO()

    writer = csv.writer(
        buffer,
        delimiter=";",
        lineterminator="\r\n",
        quoting=csv.QUOTE_MINIMAL,
    )

    writer.writerow(["nmID", "price"])

    if not frame.empty:
        for _, row in frame.iterrows():

            price = row.get("price_to_set")

            if pd.isna(price):
                continue

            writer.writerow(
                [
                    str(row.get("nm_id")),
                    int(price),
                ]
            )

    return (
        "﻿"
        + buffer.getvalue()
    ).encode("utf-8")


def build_price_change_excel(payload) -> bytes:
    """
    Тот же список цен, но с контекстом: старая цена,
    новая цена, минимальная цена и причина.

    Нужен тому, кто цены утверждает, а не загружает.
    """

    from openpyxl import Workbook

    from .excel import (
        DECISION_COLUMNS,
        _write_methodology,
        write_sheet,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Новые цены"),
        price_change_frame(payload),
        DECISION_COLUMNS,
        title="Цены к установке",
        subtitle=(
            "Только артикулы, где изменение цены не меньше 1%. "
            "Цена к установке не опускается ниже минимальной, "
            "кроме статуса «Распродажа»."
        ),
    )

    _write_methodology(
        workbook.create_sheet("Методика"),
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return stream.getvalue()


# ============================================================
# ZIP
# ============================================================

def build_zip(payload) -> bytes:
    """Всё вместе: три документа, CSV и README с методикой."""

    stamp = _stamp(payload)

    folder = f"pricing_strategy_{stamp}"

    decisions = prepare_decisions(payload)

    losses = loss_frame(payload)

    changes = price_change_frame(payload)

    summary = payload.get("summary") or {}

    readme = (
        methodology_text()
        + "\n\n"
        + "СОСТАВ ПАКЕТА\n"
        + "=" * 60
        + "\n\n"
        + f"Дата анализа: {stamp}\n"
        + f"Остатки WB: {payload.get('wb_date', '—')}\n"
        + f"Остатки FBS: {payload.get('fbs_date', '—')}\n\n"
        + f"Артикулов в анализе: {summary.get('products', 0)}\n"
        + (
            "Продаётся ниже минимальной цены: "
            f"{summary.get('below_breakeven_products', 0)}\n"
        )
        + (
            "Потенциальный убыток на остатке: "
            f"{summary.get('stock_at_risk_value', 0):,.0f} ₽\n\n"
        ).replace(",", " ")
        + "1. pricing_decisions.xlsx — полный отчёт\n"
        + f"   строк: {len(decisions)}\n"
        + "2. pricing_loss.xlsx — продаём в убыток\n"
        + f"   строк: {len(losses)}\n"
        + "3. pricing_new_prices.xlsx — цены к установке\n"
        + f"   строк: {len(changes)}\n"
        + "4. wb_prices.csv — nmID и цена для загрузки\n"
        + f"   строк: {len(changes)}\n"
    )

    buffer = BytesIO()

    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:

        archive.writestr(
            f"{folder}/pricing_decisions_{stamp}.xlsx",
            build_pricing_excel(payload),
        )

        archive.writestr(
            f"{folder}/pricing_loss_{stamp}.xlsx",
            build_loss_excel(payload),
        )

        archive.writestr(
            f"{folder}/pricing_new_prices_{stamp}.xlsx",
            build_price_change_excel(payload),
        )

        archive.writestr(
            f"{folder}/wb_prices_{stamp}.csv",
            build_wb_price_csv(payload),
        )

        archive.writestr(
            f"{folder}/README.txt",
            readme.encode("utf-8"),
        )

    buffer.seek(0)

    return buffer.read()


# ============================================================
# ИМЕНА ФАЙЛОВ
#
# Держим в одном месте: имя файла — это тоже интерфейс.
# Через месяц в папке «Загрузки» должно быть понятно,
# что это за файл и на какую дату.
# ============================================================

def filename(kind: str, payload, nm_id=None) -> str:

    stamp = _stamp(payload)

    names = {
        "excel": f"Ценовые_решения_{stamp}.xlsx",
        "loss": f"Продаём_в_убыток_{stamp}.xlsx",
        "prices": f"Новые_цены_{stamp}.xlsx",
        "csv": f"wb_prices_{stamp}.csv",
        "zip": f"Управление_ценами_{stamp}.zip",
        "product": f"Товар_{nm_id}_{stamp}.xlsx",
    }

    return names.get(kind, f"pricing_{stamp}.xlsx")


__all__ = [
    "build_loss_excel",
    "build_price_change_excel",
    "build_pricing_excel",
    "build_product_excel",
    "build_wb_price_csv",
    "build_zip",
    "filename",
    "price_change_frame",
]
