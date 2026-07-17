# gear/app/costs_control/modal.py
from __future__ import annotations

from io import BytesIO
from typing import Any

import dash_ag_grid as dag
import dash_mantine_components as dmc
import pandas as pd
from dash import Input, Output, State, dcc, html, no_update
from dash_iconify import DashIconify
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import COLORS
from .data import get_price_history_data
from .ids import (
    CHART_PRODUCT_CLIPBOARD_ID,
    CHART_PRODUCT_MODAL_ID,
    CHART_PRODUCT_NAME_ID,
    CHART_PRODUCT_NM_ID,
    CHART_PRODUCT_UPD_CSV_BTN_ID,
    CHART_PRODUCT_UPD_CSV_DOWNLOAD_ID,
    CHART_PRODUCT_UPD_EXCEL_BTN_ID,
    CHART_PRODUCT_UPD_EXCEL_DOWNLOAD_ID,
    CHART_PRODUCT_UPD_GRID_ID,
    MEDIAN_DEVIATION_CHART_ID,
    MEDIAN_DEVIATION_CLICK_STORE_ID,
    TOP_CV_CHART_ID,
    TOP_CV_CLICK_STORE_ID,
)


# ---------------------------------------------------------------------
# Цвета модального окна
# ---------------------------------------------------------------------

GREEN = COLORS.get("green", "#2F6656")
DARK_GREEN = COLORS.get("dark_green", "#245245")
BORDER = COLORS.get("border", "#D9DEE2")
MUTED = COLORS.get("muted", "#667085")
RED = COLORS.get("red", "#D92D20")
LIGHT_RED = COLORS.get("light_red", "#FFF1F0")
YELLOW = COLORS.get("yellow", "#B54708")
LIGHT_YELLOW = COLORS.get("light_yellow", "#FFF7E8")
VERY_LIGHT_GREEN = COLORS.get("very_light_green", "#EFF8F4")
SOFT_GRAY = "#F8FAFB"
TEXT = "#1F2937"


# ---------------------------------------------------------------------
# Форматирование данных
# ---------------------------------------------------------------------


def _normalise_nm_id(value: Any) -> str:
    """Приводит NM ID к строке без окончания .0."""

    if value is None:
        return ""

    text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    return text



def _extract_product_from_chart(click_data) -> tuple[str, str]:
    """
    Получает наименование и NM ID из clickData графика.

    customdata[0] — наименование;
    customdata[1] — NM ID.
    """

    if not click_data:
        return "", ""

    points = click_data.get("points") or []

    if not points:
        return "", ""

    customdata = points[0].get("customdata") or []

    if len(customdata) < 2:
        return "", ""

    product_name = str(customdata[0] or "").strip()
    nm_id = _normalise_nm_id(customdata[1])

    return product_name, nm_id



def _first_existing_column(
    df: pd.DataFrame,
    candidates: list[str],
) -> str | None:
    """Возвращает первую найденную колонку из списка."""

    for column in candidates:
        if column in df.columns:
            return column

    return None



def _safe_text(value: Any) -> str:
    """Безопасно преобразует значение в строку."""

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    return text



def _safe_number(value: Any) -> float | None:
    """Безопасно преобразует значение в число."""

    if value is None:
        return None

    try:
        numeric = pd.to_numeric(value, errors="coerce")
    except (TypeError, ValueError):
        return None

    if pd.isna(numeric):
        return None

    return float(numeric)



def _format_number(value: float | None, decimals: int = 2) -> str:
    """Форматирует число с пробелами и десятичной запятой."""

    if value is None:
        return ""

    formatted = f"{value:,.{decimals}f}"
    return formatted.replace(",", " ").replace(".", ",")


def _format_quantity(value: float | None) -> str:
    """Форматирует количество без лишних нулей."""

    if value is None:
        return ""

    if float(value).is_integer():
        return f"{int(value):,}".replace(",", " ")

    return _format_number(value, 2)


def _format_money(value: float | None) -> str:
    """Форматирует денежное значение."""

    if value is None:
        return ""

    return f"{_format_number(value, 2)} ₽"


def _format_percent(value: float | None) -> str:
    """Форматирует процент с двумя знаками."""

    if value is None:
        return ""

    sign = "+" if value > 0 else ""
    return f"{sign}{_format_number(value, 2)} %"


def _safe_date(value: Any) -> str:
    """Форматирует дату как ДД.ММ.ГГГГ."""

    if value is None:
        return ""

    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        return ""

    return parsed.strftime("%d.%m.%Y")


# ---------------------------------------------------------------------
# История УПД
# ---------------------------------------------------------------------


def _build_product_upd_rows(nm_id: str) -> list[dict]:
    """Получает историю УПД товара и приводит её к структуре AG Grid."""

    if not nm_id:
        return []

    history_df = get_price_history_data().copy()

    if history_df.empty:
        return []

    nm_id_column = _first_existing_column(
        history_df,
        ["nm_id", "NM ID", "nmId"],
    )

    if not nm_id_column:
        return []

    history_nm_ids = (
        history_df[nm_id_column]
        .astype(str)
        .map(_normalise_nm_id)
    )

    product_history = (
        history_df.loc[history_nm_ids == nm_id]
        .copy()
        .reset_index(drop=True)
    )

    if product_history.empty:
        return []

    document_column = _first_existing_column(
        product_history,
        [
            "Номер УПД",
            "Номер",
            "УПД",
            "upd_number",
            "upd_document_number",
            "upd_document_id",
        ],
    )

    date_column = _first_existing_column(
        product_history,
        ["Дата УПД", "Дата", "upd_date", "date"],
    )

    supplier_column = _first_existing_column(
        product_history,
        ["Поставщик", "Контрагент", "supplier", "supplier_name"],
    )

    quantity_column = _first_existing_column(
        product_history,
        [
            "Кол-во, шт",
            "Количество, шт",
            "Количество",
            "Кол-во",
            "upd_qty",
            "quantity",
        ],
    )

    accounting_price_column = _first_existing_column(
        product_history,
        [
            "Цена, бух",
            "Цена бух",
            "Бухгалтерская цена",
            "Себестоимость, бух",
            "Бух. себестоимость",
            "upd_price_vatless",
        ],
    )

    management_price_column = _first_existing_column(
        product_history,
        [
            "Цена, упр",
            "Цена упр",
            "Управленческая цена",
            "Себестоимость, упр",
            "Упр. себестоимость",
            "man_cost_per_unit",
        ],
    )

    if date_column:
        product_history["__sort_date"] = pd.to_datetime(
            product_history[date_column],
            errors="coerce",
        )
        product_history = product_history.sort_values(
            "__sort_date",
            ascending=False,
            na_position="last",
        )

    rows: list[dict] = []

    for _, row in product_history.iterrows():
        accounting_price = (
            _safe_number(row.get(accounting_price_column))
            if accounting_price_column
            else None
        )

        management_price = (
            _safe_number(row.get(management_price_column))
            if management_price_column
            else None
        )

        price_delta = None
        price_delta_pct = None

        if accounting_price is not None and management_price is not None:
            price_delta = management_price - accounting_price

            if accounting_price != 0:
                price_delta_pct = price_delta / accounting_price * 100

        rows.append(
            {
                "upd_number": (
                    _safe_text(row.get(document_column))
                    if document_column
                    else ""
                ),
                "upd_date": (
                    _safe_date(row.get(date_column))
                    if date_column
                    else ""
                ),
                "supplier": (
                    _safe_text(row.get(supplier_column))
                    if supplier_column
                    else ""
                ),
                "quantity": (
                    _safe_number(row.get(quantity_column))
                    if quantity_column
                    else None
                ),
                "quantity_display": _format_quantity(
                    _safe_number(row.get(quantity_column))
                    if quantity_column
                    else None
                ),
                "accounting_price": accounting_price,
                "accounting_price_display": _format_money(accounting_price),
                "management_price": management_price,
                "management_price_display": _format_money(management_price),
                "price_delta": price_delta,
                "price_delta_display": _format_money(price_delta),
                "price_delta_pct": price_delta_pct,
                "price_delta_pct_display": _format_percent(price_delta_pct),
            }
        )

    return rows


# ---------------------------------------------------------------------
# AG Grid
# ---------------------------------------------------------------------

def build_product_upd_grid():
    """Таблица истории закупочных документов выбранного товара."""

    delta_rubles_style = {
        "styleConditions": [
            {
                "condition": "data.price_delta > 0",
                "style": {
                    "backgroundColor": LIGHT_RED,
                    "color": RED,
                    "fontWeight": "700",
                },
            },
            {
                "condition": "data.price_delta < 0",
                "style": {
                    "backgroundColor": VERY_LIGHT_GREEN,
                    "color": DARK_GREEN,
                    "fontWeight": "700",
                },
            },
        ],
        "defaultStyle": {
            "color": MUTED,
        },
    }

    delta_percent_style = {
        "styleConditions": [
            {
                "condition": "Math.abs(data.price_delta_pct) >= 25",
                "style": {
                    "backgroundColor": LIGHT_RED,
                    "color": RED,
                    "fontWeight": "700",
                },
            },
            {
                "condition": (
                    "Math.abs(data.price_delta_pct) >= 10 "
                    "&& Math.abs(data.price_delta_pct) < 25"
                ),
                "style": {
                    "backgroundColor": LIGHT_YELLOW,
                    "color": YELLOW,
                    "fontWeight": "700",
                },
            },
            {
                "condition": (
                    "Math.abs(data.price_delta_pct) > 0.01 "
                    "&& Math.abs(data.price_delta_pct) < 10"
                ),
                "style": {
                    "backgroundColor": VERY_LIGHT_GREEN,
                    "color": DARK_GREEN,
                    "fontWeight": "600",
                },
            },
        ],
        "defaultStyle": {
            "color": MUTED,
        },
    }

    return dag.AgGrid(
        id=CHART_PRODUCT_UPD_GRID_ID,
        rowData=[],
        columnDefs=[
            {
                "headerName": "№ УПД",
                "field": "upd_number",
                "minWidth": 120,
                "maxWidth": 170,
                "flex": 0.9,
                "filter": "agTextColumnFilter",
                "pinned": "left",
                "cellStyle": {
                    "fontWeight": "600",
                    "backgroundColor": SOFT_GRAY,
                },
            },
            {
                "headerName": "Дата УПД",
                "field": "upd_date",
                "minWidth": 115,
                "maxWidth": 140,
                "flex": 0.75,
                "filter": "agTextColumnFilter",
                "cellStyle": {
                    "backgroundColor": SOFT_GRAY,
                    "borderRight": f"1px solid {BORDER}",
                },
            },
            {
                "headerName": "Поставщик",
                "field": "supplier",
                "minWidth": 260,
                "flex": 2.4,
                "filter": "agTextColumnFilter",
                "tooltipField": "supplier",
            },
            {
                "headerName": "Количество, шт",
                "field": "quantity_display",
                "minWidth": 135,
                "maxWidth": 165,
                "flex": 0.9,
            },
            {
                "headerName": "Цена, бух",
                "field": "accounting_price_display",
                "minWidth": 135,
                "maxWidth": 175,
                "flex": 0.95,
                "cellStyle": {
                    "backgroundColor": "#F5F3FF",
                    "fontWeight": "600",
                },
            },
            {
                "headerName": "Цена, упр",
                "field": "management_price_display",
                "minWidth": 135,
                "maxWidth": 175,
                "flex": 0.95,
                "cellStyle": {
                    "backgroundColor": "#EEF2FF",
                    "fontWeight": "600",
                },
            },
            {
                "headerName": "Отклонение, ₽",
                "field": "price_delta_display",
                "minWidth": 145,
                "maxWidth": 180,
                "flex": 0.95,
                "cellStyle": delta_rubles_style,
            },
            {
                "headerName": "Отклонение, %",
                "field": "price_delta_pct_display",
                "minWidth": 145,
                "maxWidth": 175,
                "flex": 0.9,
                "cellStyle": delta_percent_style,
            },
        ],
        defaultColDef={
            "sortable": True,
            "resizable": True,
            "filter": True,
            "editable": False,
            "floatingFilter": False,
            "suppressHeaderMenuButton": False,
            "wrapHeaderText": False,
            "cellStyle": {
                "fontSize": "12px",
                "lineHeight": "1.25",
            },
        },
        dashGridOptions={
            "animateRows": False,
            "rowHeight": 38,
            "headerHeight": 42,
            "pagination": True,
            "paginationPageSize": 15,
            "paginationPageSizeSelector": [15, 30, 50, 100],
            "domLayout": "normal",
            "enableCellTextSelection": True,
            "ensureDomOrder": True,
            "suppressCellFocus": False,
            "tooltipShowDelay": 250,
            "autoSizeStrategy": {
                "type": "fitGridWidth",
                "defaultMinWidth": 110,
            },
            "rowClassRules": {
                "cost-row-critical": (
                    "Math.abs(data.price_delta_pct || 0) >= 25"
                ),
                "cost-row-warning": (
                    "Math.abs(data.price_delta_pct || 0) >= 10 "
                    "&& Math.abs(data.price_delta_pct || 0) < 25"
                ),
            },
            "overlayNoRowsTemplate": (
                "<span style='color:#667085;font-size:12px;'>"
                "История УПД по товару не найдена"
                "</span>"
            ),
        },
        style={
            "height": "535px",
            "width": "100%",
        },
        className="ag-theme-quartz costs-product-history-grid",
    )


# ---------------------------------------------------------------------
# Компоненты модального окна
# ---------------------------------------------------------------------


def _modal_title():
    return dmc.Group(
        gap=10,
        wrap="nowrap",
        children=[
            html.Div(
                style={
                    "display": "flex",
                    "alignItems": "center",
                    "justifyContent": "center",
                    "width": "32px",
                    "height": "32px",
                    "backgroundColor": VERY_LIGHT_GREEN,
                    "border": f"1px solid {BORDER}",
                },
                children=DashIconify(
                    icon="solar:document-text-linear",
                    width=18,
                    height=18,
                    color=GREEN,
                ),
            ),
            html.Div(
                children=[
                    dmc.Text(
                        "История закупочной цены",
                        size="sm",
                        fw=700,
                        c=TEXT,
                    ),
                    dmc.Text(
                        "Документы УПД по выбранному товару",
                        size="xs",
                        c="dimmed",
                        mt=1,
                    ),
                ],
            ),
        ],
    )



def _product_information_block():
    return html.Div(
        style={
            "display": "grid",
            "gridTemplateColumns": "minmax(360px, 1fr) 310px",
            "gap": "22px",
            "alignItems": "stretch",
            "padding": "16px",
            "backgroundColor": SOFT_GRAY,
            "border": f"1px solid {BORDER}",
            "marginBottom": "16px",
        },
        children=[
            html.Div(
                children=[
                    dmc.Group(
                        gap=7,
                        wrap="nowrap",
                        mb=7,
                        children=[
                            DashIconify(
                                icon="solar:tag-linear",
                                width=15,
                                color=MUTED,
                            ),
                            dmc.Text(
                                "Наименование товара",
                                size="xs",
                                fw=600,
                                c="dimmed",
                            ),
                        ],
                    ),
                    dmc.Text(
                        id=CHART_PRODUCT_NAME_ID,
                        children="—",
                        size="sm",
                        fw=700,
                        c=TEXT,
                        style={
                            "lineHeight": "21px",
                            "wordBreak": "break-word",
                        },
                    ),
                ],
            ),
            html.Div(
                style={
                    "paddingLeft": "20px",
                    "borderLeft": f"1px solid {BORDER}",
                },
                children=[
                    dmc.Group(
                        gap=7,
                        wrap="nowrap",
                        mb=7,
                        children=[
                            DashIconify(
                                icon="solar:hashtag-linear",
                                width=15,
                                color=MUTED,
                            ),
                            dmc.Text(
                                "Артикул WB · NM ID",
                                size="xs",
                                fw=600,
                                c="dimmed",
                            ),
                        ],
                    ),
                    html.Div(
                        style={
                            "display": "flex",
                            "alignItems": "center",
                            "gap": "8px",
                        },
                        children=[
                            dmc.TextInput(
                                id=CHART_PRODUCT_NM_ID,
                                value="",
                                readOnly=True,
                                radius=0,
                                size="sm",
                                style={"flex": 1},
                                styles={
                                    "input": {
                                        "height": "38px",
                                        "fontFamily": "Inter, Arial, sans-serif",
                                        "fontSize": "14px",
                                        "fontWeight": 700,
                                        "color": TEXT,
                                        "cursor": "text",
                                        "userSelect": "text",
                                        "backgroundColor": "#FFFFFF",
                                        "borderColor": BORDER,
                                    },
                                },
                            ),
                            dmc.Tooltip(
                                label="Копировать NM ID",
                                position="top",
                                withArrow=True,
                                children=dcc.Clipboard(
                                    id=CHART_PRODUCT_CLIPBOARD_ID,
                                    content="",
                                    title="Копировать NM ID",
                                    style={
                                        "display": "inline-flex",
                                        "alignItems": "center",
                                        "justifyContent": "center",
                                        "width": "38px",
                                        "height": "38px",
                                        "flex": "0 0 38px",
                                        "cursor": "pointer",
                                        "fontSize": "18px",
                                        "color": GREEN,
                                        "border": f"1px solid {BORDER}",
                                        "backgroundColor": "#FFFFFF",
                                    },
                                ),
                            ),
                        ],
                    ),
                ],
            ),
        ],
    )



def _table_header():
    return html.Div(
        style={
            "display": "flex",
            "alignItems": "center",
            "justifyContent": "space-between",
            "gap": "16px",
            "marginBottom": "10px",
        },
        children=[
            html.Div(
                children=[
                    dmc.Group(
                        gap=8,
                        wrap="nowrap",
                        children=[
                            DashIconify(
                                icon="solar:bill-list-linear",
                                width=17,
                                color=GREEN,
                            ),
                            dmc.Text(
                                "История документов УПД",
                                size="sm",
                                fw=700,
                                c=TEXT,
                            ),
                        ],
                    ),
                    dmc.Text(
                        (
                            "Цены показаны без НДС. Отклонение рассчитано "
                            "как управленческая цена минус бухгалтерская."
                        ),
                        size="xs",
                        c="dimmed",
                        mt=3,
                    ),
                ],
            ),
            dmc.Group(
                gap=8,
                wrap="nowrap",
                children=[
                    dmc.Button(
                        "Скачать CSV",
                        id=CHART_PRODUCT_UPD_CSV_BTN_ID,
                        variant="default",
                        radius=0,
                        size="xs",
                        leftSection=DashIconify(
                            icon="solar:file-text-linear",
                            width=14,
                        ),
                        styles={
                            "root": {
                                "height": "34px",
                                "fontWeight": 600,
                                "borderColor": BORDER,
                            },
                        },
                    ),
                    dmc.Button(
                        "Скачать Excel",
                        id=CHART_PRODUCT_UPD_EXCEL_BTN_ID,
                        variant="light",
                        color="#087f5b",
                        radius=0,
                        size="xs",
                        leftSection=DashIconify(
                            icon="solar:file-download-linear",
                            width=14,
                        ),
                        styles={
                            "root": {
                                "height": "34px",
                                "fontWeight": 600,
                            },
                        },
                    ),
                ],
            ),
        ],
    )



def _legend():
    items = [
        ("≥ 25%", LIGHT_RED, RED, "Критическое отклонение"),
        ("10–25%", LIGHT_YELLOW, YELLOW, "Требует внимания"),
        ("< 10%", VERY_LIGHT_GREEN, DARK_GREEN, "Небольшое отклонение"),
    ]

    return dmc.Group(
        gap=16,
        wrap="wrap",
        mt=10,
        children=[
            dmc.Group(
                gap=6,
                wrap="nowrap",
                children=[
                    html.Span(
                        style={
                            "display": "inline-block",
                            "width": "10px",
                            "height": "10px",
                            "backgroundColor": background,
                            "border": f"1px solid {color}",
                        }
                    ),
                    dmc.Text(
                        f"{label} — {description}",
                        size="xs",
                        c="dimmed",
                    ),
                ],
            )
            for label, background, color, description in items
        ],
    )



def build_chart_product_modal():
    """Модальное окно с товаром и историей его закупочной цены."""

    return dmc.Modal(
        id=CHART_PRODUCT_MODAL_ID,
        opened=False,
        centered=True,
        size="92%",
        radius=0,
        withCloseButton=True,
        closeOnClickOutside=True,
        closeOnEscape=True,
        title=_modal_title(),
        styles={
            "header": {
                "padding": "14px 18px",
                "borderBottom": f"1px solid {BORDER}",
            },
            "body": {
                "padding": "16px 18px 18px",
            },
            "content": {
                "maxWidth": "1500px",
            },
            "close": {
                "borderRadius": 0,
            },
        },
        children=[
            _product_information_block(),
            _table_header(),
            build_product_upd_grid(),
            _legend(),
        ],
    )



def build_chart_product_modal_components():
    """Store-компоненты, загрузки и модальное окно."""

    return html.Div(
        children=[
            dcc.Store(
                id=TOP_CV_CLICK_STORE_ID,
                storage_type="memory",
            ),
            dcc.Store(
                id=MEDIAN_DEVIATION_CLICK_STORE_ID,
                storage_type="memory",
            ),
            dcc.Download(
                id=CHART_PRODUCT_UPD_CSV_DOWNLOAD_ID,
            ),
            dcc.Download(
                id=CHART_PRODUCT_UPD_EXCEL_DOWNLOAD_ID,
            ),
            build_chart_product_modal(),
        ],
    )


# ---------------------------------------------------------------------
# Экспорт
# ---------------------------------------------------------------------


def _prepare_upd_export_dataframe(rows) -> pd.DataFrame:
    """Формирует DataFrame для выгрузки истории УПД."""

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    column_mapping = {
        "upd_number": "Номер УПД",
        "upd_date": "Дата УПД",
        "supplier": "Поставщик",
        "quantity": "Количество, шт",
        "accounting_price": "Цена, бух",
        "management_price": "Цена, упр",
        "price_delta": "Отклонение, ₽",
        "price_delta_pct": "Отклонение, %",
    }

    existing_columns = [
        column
        for column in column_mapping
        if column in df.columns
    ]

    return df[existing_columns].rename(columns=column_mapping)



def _build_excel_bytes(export_df: pd.DataFrame) -> bytes:
    """Создаёт аккуратно оформленный Excel-файл в памяти."""

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(
            writer,
            sheet_name="История УПД",
            index=False,
        )

        worksheet = writer.sheets["История УПД"]
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="2F6656",
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
            name="Arial",
            size=10,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.row_dimensions[1].height = 24

        column_widths = {
            "A": 18,
            "B": 14,
            "C": 34,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 18,
            "H": 18,
        }

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

        money_columns = {5, 6, 7}
        numeric_columns = {4}
        percent_columns = {8}

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal=(
                        "right"
                        if cell.column in money_columns | numeric_columns | percent_columns
                        else "left"
                    ),
                )

                if cell.column in money_columns:
                    cell.number_format = '#,##0.00 [$₽-ru-RU]'
                elif cell.column in numeric_columns:
                    cell.number_format = '#,##0.00'
                elif cell.column in percent_columns:
                    cell.number_format = '0.0" %"'

        for column_index in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].bestFit = False

    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------
# Callbacks
# ---------------------------------------------------------------------


def register_chart_product_modal_callbacks(app):
    """Регистрирует callbacks модального окна."""

    @app.callback(
        Output(TOP_CV_CLICK_STORE_ID, "data"),
        Input(TOP_CV_CHART_ID, "clickData"),
        prevent_initial_call=True,
    )
    def save_top_cv_chart_click(click_data):
        product_name, nm_id = _extract_product_from_chart(click_data)

        if not nm_id:
            return no_update

        return {
            "nm_id": nm_id,
            "name": product_name or "Без наименования",
        }

    @app.callback(
        Output(MEDIAN_DEVIATION_CLICK_STORE_ID, "data"),
        Input(MEDIAN_DEVIATION_CHART_ID, "clickData"),
        prevent_initial_call=True,
    )
    def save_median_deviation_chart_click(click_data):
        product_name, nm_id = _extract_product_from_chart(click_data)

        if not nm_id:
            return no_update

        return {
            "nm_id": nm_id,
            "name": product_name or "Без наименования",
        }

    @app.callback(
        Output(CHART_PRODUCT_MODAL_ID, "opened"),
        Output(CHART_PRODUCT_NAME_ID, "children"),
        Output(CHART_PRODUCT_NM_ID, "value"),
        Output(CHART_PRODUCT_CLIPBOARD_ID, "content"),
        Output(CHART_PRODUCT_UPD_GRID_ID, "rowData"),
        Input(TOP_CV_CLICK_STORE_ID, "modified_timestamp"),
        Input(MEDIAN_DEVIATION_CLICK_STORE_ID, "modified_timestamp"),
        State(TOP_CV_CLICK_STORE_ID, "data"),
        State(MEDIAN_DEVIATION_CLICK_STORE_ID, "data"),
        prevent_initial_call=True,
    )
    def open_chart_product_modal(
        top_cv_timestamp,
        median_timestamp,
        top_cv_data,
        median_data,
    ):
        top_cv_timestamp = (
            top_cv_timestamp
            if top_cv_timestamp is not None
            else -1
        )
        median_timestamp = (
            median_timestamp
            if median_timestamp is not None
            else -1
        )

        selected_product = (
            median_data
            if median_timestamp > top_cv_timestamp
            else top_cv_data
        )

        if not selected_product:
            return (no_update,) * 5

        nm_id = _normalise_nm_id(selected_product.get("nm_id"))
        product_name = str(
            selected_product.get("name") or "Без наименования"
        ).strip()

        if not nm_id:
            return (no_update,) * 5

        upd_rows = _build_product_upd_rows(nm_id)

        return (
            True,
            product_name,
            nm_id,
            nm_id,
            upd_rows,
        )

    @app.callback(
        Output(CHART_PRODUCT_UPD_CSV_DOWNLOAD_ID, "data"),
        Input(CHART_PRODUCT_UPD_CSV_BTN_ID, "n_clicks"),
        State(CHART_PRODUCT_UPD_GRID_ID, "rowData"),
        State(CHART_PRODUCT_NM_ID, "value"),
        prevent_initial_call=True,
    )
    def download_product_upd_csv(n_clicks, rows, nm_id):
        if not n_clicks or not rows:
            return no_update

        export_df = _prepare_upd_export_dataframe(rows)

        if export_df.empty:
            return no_update

        safe_nm_id = _normalise_nm_id(nm_id) or "product"

        return dcc.send_data_frame(
            export_df.to_csv,
            filename=f"upd_history_{safe_nm_id}.csv",
            index=False,
            sep=";",
            encoding="utf-8-sig",
            decimal=",",
        )

    @app.callback(
        Output(CHART_PRODUCT_UPD_EXCEL_DOWNLOAD_ID, "data"),
        Input(CHART_PRODUCT_UPD_EXCEL_BTN_ID, "n_clicks"),
        State(CHART_PRODUCT_UPD_GRID_ID, "rowData"),
        State(CHART_PRODUCT_NM_ID, "value"),
        prevent_initial_call=True,
    )
    def download_product_upd_excel(n_clicks, rows, nm_id):
        if not n_clicks or not rows:
            return no_update

        export_df = _prepare_upd_export_dataframe(rows)

        if export_df.empty:
            return no_update

        safe_nm_id = _normalise_nm_id(nm_id) or "product"
        excel_bytes = _build_excel_bytes(export_df)

        return dcc.send_bytes(
            excel_bytes,
            filename=f"upd_history_{safe_nm_id}.xlsx",
        )