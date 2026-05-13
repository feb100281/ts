# # ЭТО НОВЫЕ ИМПОРТЫ
# import pandas as pd

# from pathlib import Path


# source_folder = Path('/Users/daria/Desktop/ТРЕНДСЕТТЕР/Бух данные/upds/all_stock_upd')

# # source_folder = Path('/Users/pavelustenko/Downloads/2024')
# base_path = Path.cwd()
# upd_path = base_path / "data" / "all_stock_upd"
# upd_path.mkdir(parents=True, exist_ok=True)

# UPD_LIST = {

#     "УПД HM-047 от 12 марта 2026 г.xlsx": {
#         "number": "HM-047",
#         "date": "2026-03-12",
#         "columns": "B:N",
#         "header_row": 8,
#         "nrows":3566,
#         "dropna_col": "№",
#         "supplier": "ЗАО «Арминвест»",
#         "sheet": "1. Счет-Фактура"
#     },
    
    
    
           
# }


# for filename, config in UPD_LIST.items():

#     df = pd.read_excel(
#     source_folder / filename,
#     sheet_name=config["sheet"],
#     skiprows=config["header_row"],
#     usecols=config["columns"],
#     nrows=config["nrows"],
#     dtype=str,
#     )
#     df.columns = df.columns.astype(str).str.strip()
#     df = df.dropna(subset=[config["dropna_col"]])
#     df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
#     print(df.columns)
#     df['file_name'] = filename
#     df['number'] = config['number']
#     df['date']=config['date']
#     df['supplier']=config['supplier']
    
#     output_file = upd_path / f"{Path(filename).stem}.parquet"
#     df.to_parquet(output_file, index=False)

#     print(f"Saved: {output_file} | rows: {len(df)}")




# utils/upd_import/upd_import_new.py

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


source_folder = Path('/Users/daria/Desktop/ТРЕНДСЕТТЕР/Бух данные/upds/all_stock_upd')

base_path = Path.cwd()
upd_path = base_path / "data" / "all_stock_upd"
upd_path.mkdir(parents=True, exist_ok=True)


UPD_LIST = {
    "УПД HM-047 от 12 марта 2026 г.xlsx": {
        "number": "HM-047",
        "date": "2026-03-12",
        "columns": "B:N",
        "header_row": 17,
        "nrows": 3566,
        "dropna_col": "№",
        "supplier": "ЗАО «Арминвест»",
        "sheet": "1. Счет-Фактура"
    },
}


def find_existing_filepath(source_folder, filename):
    filepath = source_folder / filename

    if filepath.exists():
        return filepath

    # если в списке указано .xlsx, а реально лежит .xls
    if filepath.suffix.lower() == ".xlsx":
        xls_path = filepath.with_suffix(".xls")
        if xls_path.exists():
            return xls_path

    return filepath


def convert_xls_to_xlsx(filepath):
    filepath = Path(filepath)

    if filepath.suffix.lower() != ".xls":
        return filepath

    new_path = filepath.with_suffix(".xlsx")

    if new_path.exists():
        print(f"XLSX уже есть: {new_path.name}")
        return new_path

    sheets = pd.read_excel(
        filepath,
        sheet_name=None,
        header=None,
        dtype=str,
        engine="xlrd"
    )

    with pd.ExcelWriter(new_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False
            )

    print(f"Конвертирован: {filepath.name} -> {new_path.name}")

    return new_path


def fill_merged_cells_xlsx(filepath, sheet_name):
    """
    Разъединяет объединенные ячейки и заполняет весь диапазон
    значением из верхней левой ячейки.
    """

    filepath = Path(filepath)

    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    merged_ranges = list(ws.merged_cells.ranges)

    if not merged_ranges:
        return filepath

    for merged_range in merged_ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row

        value = ws.cell(row=min_row, column=min_col).value

        ws.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = value

    normalized_path = filepath.with_name(filepath.stem + "_normalized.xlsx")
    wb.save(normalized_path)

    print(f"Объединенные ячейки обработаны: {normalized_path.name}")

    return normalized_path


def to_number(s):
    return (
        s.astype(str)
        .str.replace("\xa0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
    )


def find_amount_column(df):
    for col in df.columns:
        col_lower = str(col).lower()

        if (
            "стоимость" in col_lower
            and "налог" in col_lower
            and "всего" in col_lower
        ):
            return col

        if "стоимость" in col_lower and "с ндс" in col_lower:
            return col

    return None


for filename, config in UPD_LIST.items():

    filepath = find_existing_filepath(source_folder, filename)

    if not filepath.exists():
        print(f"Файл не найден: {filename}")
        continue

    filepath = convert_xls_to_xlsx(filepath)

    filepath = fill_merged_cells_xlsx(
        filepath=filepath,
        sheet_name=config["sheet"]
    )

    df = pd.read_excel(
        filepath,
        sheet_name=config["sheet"],
        skiprows=config["header_row"] - 1,
        usecols=config["columns"],
        nrows=config["nrows"],
        dtype=str,
        engine="openpyxl"
    )

    df.columns = (
        df.columns
        .astype(str)
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

    if config["dropna_col"] not in df.columns:
        print(f"Нет колонки {config['dropna_col']} в файле: {filename}")
        print("Колонки в файле:", df.columns.tolist())
        continue

    df = df.dropna(subset=[config["dropna_col"]])

    amount_col = find_amount_column(df)

    if amount_col:
        total_amount = to_number(df[amount_col]).fillna(0).sum()
    else:
        total_amount = 0
        print(f"Не найдена колонка суммы в файле: {filename}")
        print("Колонки в файле:", df.columns.tolist())

    df["file_name"] = filename
    df["number"] = config["number"]
    df["date"] = config["date"]
    df["supplier"] = config["supplier"]
    df["итого_сумма_по_файлу"] = f"{total_amount:.2f}"
    df = df.astype(str)

    output_file = upd_path / f"{Path(filename).stem}.parquet"
    df.to_parquet(output_file, index=False)

    print(
        f"Готово: {filename} | "
        f"строк: {len(df)} | "
        f"сумма: {total_amount:,.2f}"
    )