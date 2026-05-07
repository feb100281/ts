# # utils/upd_issues/builder.py
# import io
# import zipfile
# from io import BytesIO
# from openpyxl import Workbook
# from .sheets.error_sheet import ErrorSheet
# from .sheets.toc_sheet import TOCSheet
# from .queries import UpdIssuesQueries
# from .pdf_exporter import build_upd_issues_pdf_response


# class UpdIssuesReportGenerator:
#     def __init__(self):
#         self.queries = UpdIssuesQueries()
    
#     def generate(self) -> BytesIO:
#         files_df = self.queries.get_files_list()
        
#         if files_df.empty:
#             raise ValueError("Нет данных о косяках в УПД")
        
#         # Получаем общую статистику для сопроводилки
#         summary_stats = self.queries.get_summary_stats()
        
#         zip_buffer = BytesIO()
        
#         with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
#             # Генерируем PDF-сопроводилку
#             pdf_response = build_upd_issues_pdf_response(files_df, summary_stats)
#             zip_file.writestr("Сопроводительное_письмо.pdf", pdf_response.content)
            
#             # Для каждого УПД создаем отдельный Excel файл
#             for _, file_row in files_df.iterrows():
#                 full_name = file_row['full_name']
#                 supplier = file_row.get('supplier', '—')
                
#                 file_stats = self.queries.get_summary_stats(full_name=full_name)
#                 file_stats['total_positions'] = int(file_row['total_positions'])
#                 file_stats['name_mismatch'] = int(file_row['name_mismatch'])
#                 file_stats['article_mismatch'] = int(file_row.get('article_mismatch', 0))
#                 file_stats['size_mismatch'] = int(file_row['size_mismatch'])
#                 file_stats['vat_mismatch'] = int(file_row['vat_mismatch'])
#                 file_stats['cert_issues'] = int(file_row['cert_issues'])
#                 file_stats['total_issues'] = int(file_row['total_issues'])
                
#                 document_totals = self.queries.get_document_totals(full_name=full_name)
#                 df = self.queries.get_all_issues(full_name=full_name)
                
#                 wb = Workbook()
                
#                 if "Sheet" in wb.sheetnames:
#                     wb.remove(wb["Sheet"])
                
#                 toc = TOCSheet(wb, full_name, supplier, file_stats, document_totals)
#                 toc.build()
                
#                 name_errors = df[df['name_match'] == False].copy()
#                 article_errors = df[df['match_article'] == False].copy()
#                 size_errors = df[df['size_match'] == False].copy()
#                 vat_errors = df[df['match_vats'] == False].copy()
#                 cert_errors = df[df['cert_match'] == False].copy()
                
#                 if not name_errors.empty:
#                     sheet = ErrorSheet(wb, "Названия")
#                     sheet.build_for_file(name_errors, "name", full_name, supplier)
                
#                 if not article_errors.empty:
#                     sheet = ErrorSheet(wb, "Артикулы")
#                     sheet.build_for_file(article_errors, "article", full_name, supplier)
                
#                 if not size_errors.empty:
#                     sheet = ErrorSheet(wb, "Размеры")
#                     sheet.build_for_file(size_errors, "size", full_name, supplier)
                
#                 if not vat_errors.empty:
#                     sheet = ErrorSheet(wb, "НДС")
#                     sheet.build_for_file(vat_errors, "vat", full_name, supplier)
                
#                 if not cert_errors.empty:
#                     sheet = ErrorSheet(wb, "Сертификаты")
#                     sheet.build_for_file(cert_errors, "cert", full_name, supplier)
                
#                 excel_buffer = BytesIO()
#                 wb.save(excel_buffer)
#                 excel_buffer.seek(0)
                
#                 safe_filename = self._safe_filename(full_name)
#                 filename = f"{safe_filename}.xlsx"
#                 zip_file.writestr(filename, excel_buffer.getvalue())
        
#         zip_buffer.seek(0)
#         return zip_buffer
    
#     def _safe_filename(self, filename: str) -> str:
#         import re
#         safe = re.sub(r'[<>:"/\\|?*]', '_', filename)
#         if len(safe) > 100:
#             safe = safe[:100]
#         return safe



# utils/upd_issues/builder.py
import io
import zipfile
from io import BytesIO
from openpyxl import Workbook
from .sheets.error_sheet import ErrorSheet
from .sheets.toc_sheet import TOCSheet
from .queries import UpdIssuesQueries
from .pdf_exporter import build_upd_issues_pdf_response


class UpdIssuesReportGenerator:
    def __init__(self):
        self.queries = UpdIssuesQueries()
    
    def generate(self) -> BytesIO:
        files_df = self.queries.get_files_list()
        
        if files_df.empty:
            raise ValueError("Нет данных о косяках в УПД")
        
        # Получаем общую статистику для сопроводилки
        summary_stats = self.queries.get_summary_stats()
        
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Генерируем PDF-сопроводилку
            pdf_response = build_upd_issues_pdf_response(files_df, summary_stats)
            zip_file.writestr("Сопроводительное_письмо.pdf", pdf_response.content)
            
            # Для каждого УПД создаем отдельный Excel файл
            for _, file_row in files_df.iterrows():
                full_name = file_row['full_name']
                supplier = file_row.get('supplier', '—')
                
                file_stats = self.queries.get_summary_stats(full_name=full_name)
                file_stats['total_positions'] = int(file_row['total_positions'])  # количество позиций
                file_stats['name_mismatch'] = int(file_row['name_mismatch'])
                file_stats['article_mismatch'] = int(file_row.get('article_mismatch', 0))
                file_stats['size_mismatch'] = int(file_row['size_mismatch'])
                file_stats['vat_mismatch'] = int(file_row['vat_mismatch'])
                file_stats['cert_issues'] = int(file_row['cert_issues'])
                file_stats['total_issues'] = int(file_row['total_issues'])
                
                document_totals = self.queries.get_document_totals(full_name=full_name)
                df = self.queries.get_all_issues(full_name=full_name)
                
                wb = Workbook()
                
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                
                toc = TOCSheet(wb, full_name, supplier, file_stats, document_totals)
                toc.build()
                
                name_errors = df[df['name_match'] == False].copy()
                article_errors = df[df['match_article'] == False].copy()
                size_errors = df[df['size_match'] == False].copy()
                vat_errors = df[df['match_vats'] == False].copy()
                cert_errors = df[df['cert_match'] == False].copy()
                
                if not name_errors.empty:
                    sheet = ErrorSheet(wb, "Названия")
                    sheet.build_for_file(name_errors, "name", full_name, supplier)
                
                if not article_errors.empty:
                    sheet = ErrorSheet(wb, "Артикулы")
                    sheet.build_for_file(article_errors, "article", full_name, supplier)
                
                if not size_errors.empty:
                    sheet = ErrorSheet(wb, "Размеры")
                    sheet.build_for_file(size_errors, "size", full_name, supplier)
                
                if not vat_errors.empty:
                    sheet = ErrorSheet(wb, "НДС")
                    sheet.build_for_file(vat_errors, "vat", full_name, supplier)
                
                if not cert_errors.empty:
                    sheet = ErrorSheet(wb, "Сертификаты")
                    sheet.build_for_file(cert_errors, "cert", full_name, supplier)
                
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                safe_filename = self._safe_filename(full_name)
                filename = f"{safe_filename}.xlsx"
                zip_file.writestr(filename, excel_buffer.getvalue())
        
        zip_buffer.seek(0)
        return zip_buffer
    
    def _safe_filename(self, filename: str) -> str:
        import re
        safe = re.sub(r'[<>:"/\\|?*]', '_', filename)
        if len(safe) > 100:
            safe = safe[:100]
        return safe