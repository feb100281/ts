# upd_issues/components/tables.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles.theme import COLORS, FONTS, BORDERS, ALIGNMENTS, FILLS


class TableComponent:
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw(self, start_row, headers, data_rows, start_col=2, 
             highlight_errors=None, column_widths=None):
        """Рисует таблицу с подсветкой ошибок"""
        row = start_row
        start_col = start_col
        
        # Заголовки
        for col_idx, header in enumerate(headers):
            cell = self.ws.cell(row=row, column=start_col + col_idx, value=header)
            cell.font = FONTS["header"]
            cell.fill = FILLS["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 35
        row += 1
        
        # Данные
        for idx, data_row in enumerate(data_rows):
            fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]
            
            for col_idx, value in enumerate(data_row):
                cell = self.ws.cell(row=row, column=start_col + col_idx, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                cell.font = FONTS["normal"]
                
                # Подсветка ошибок
                if highlight_errors and col_idx in highlight_errors:
                    if isinstance(value, (int, float)):
                        cell.font = FONTS["error"]
                        cell.fill = FILLS["error"]
                
                # Выравнивание
                if col_idx == 0:
                    cell.alignment = ALIGNMENTS["left"]
                elif isinstance(value, (int, float)):
                    cell.alignment = ALIGNMENTS["right"]
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = ALIGNMENTS["center"]
            
            self.ws.row_dimensions[row].height = 25
            row += 1
        
        # Настройка ширины колонок
        if column_widths:
            for letter, width in column_widths.items():
                self.ws.column_dimensions[letter].width = width
        
        return row


def create_table(worksheet):
    return TableComponent(worksheet)