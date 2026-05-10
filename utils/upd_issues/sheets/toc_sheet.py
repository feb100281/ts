# utils/upd_issues/sheets/toc_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS, FILLS


class TOCSheet(BaseSheet):
    def __init__(self, workbook, full_name: str, supplier: str, stats: dict, totals: dict):
        super().__init__(workbook, "Справка")
        self.full_name = full_name
        self.supplier = supplier
        self.stats = stats
        self.totals = totals

    def build(self):
        row = 1
        table_start_col = 2
        table_end_col = 5

        # ============================================================
        # ЗАГОЛОВОЧНЫЙ БЛОК
        # ============================================================
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        title_cell = self.ws.cell(row=row, column=table_start_col, value="СПРАВКА ПО ДОКУМЕНТУ")
        title_cell.font = Font(name="Roboto", size=18, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 35
        row += 1

        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        subtitle_cell = self.ws.cell(row=row, column=table_start_col, value="Детальный отчет по ошибкам в УПД")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 22
        row += 1

        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        date_cell = self.ws.cell(row=row, column=table_start_col, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 1

        doc_text = f"ФАЙЛ: {self.full_name}"
        if self.supplier and self.supplier != '—':
            doc_text += f" | ПОСТАВЩИК: {self.supplier}"
        
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        doc_cell = self.ws.cell(row=row, column=table_start_col, value=doc_text)
        doc_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        doc_cell.alignment = Alignment(horizontal="left", vertical="center")
        doc_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        self.ws.row_dimensions[row].height = 28
        row += 2

        # ============================================================
        # ДЕКОРАТИВНЫЙ РАЗДЕЛИТЕЛЬ
        # ============================================================
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
            cell.border = Border(bottom=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 4
        row += 2

        # ============================================================
        # БЛОК СТАТИСТИКИ ПО ДОКУМЕНТУ
        # ============================================================
        stats_title = self.ws.cell(row=row, column=table_start_col, value="СТАТИСТИКА ПО ДОКУМЕНТУ")
        stats_title.font = Font(name="Roboto", size=13, bold=True, color=COLORS["dark_green"])
        stats_title.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        self.ws.row_dimensions[row].height = 28
        row += 1

        light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        stats_items = [
            ("Всего позиций в документе:", self._format_number(self.stats.get('total_positions', 0))),
            ("Количество товаров (ед.):", self._format_number(self.totals.get('total_qty', 0))),
            ("Сумма без НДС:", self._format_currency(self.totals.get('total_amount_vatless', 0))),
            ("Сумма НДС:", self._format_currency(self.totals.get('total_vat_amount', 0)))
        ]

        for label, value in stats_items:
            cell = self.ws.cell(row=row, column=table_start_col)
            cell.value = f"{label} {value}"
            cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
            cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
            self.ws.row_dimensions[row].height = 22
            
            for c in range(table_start_col, table_end_col + 1):
                self.ws.cell(row=row, column=c).fill = light_gray_fill
            row += 1

        row += 1

        total_cell = self.ws.cell(row=row, column=table_start_col)
        total_cell.value = f"ОБЩАЯ СУММА С НДС: {self._format_currency(self.totals.get('total_amount_vatadd', 0))}"
        total_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
        total_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        self.ws.row_dimensions[row].height = 28

        for c in range(table_start_col, table_end_col + 1):
            self.ws.cell(row=row, column=c).fill = light_gray_fill

        row += 2

        # ============================================================
        # БЛОК НАВИГАЦИИ
        # ============================================================
        nav_cell = self.ws.cell(row=row, column=table_start_col, value="ЛИСТЫ С ОШИБКАМИ")
        nav_cell.font = Font(name="Roboto", size=13, bold=True, color=COLORS["dark_green"])
        nav_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        self.ws.row_dimensions[row].height = 28
        row += 1

        # ============================================================
        # ТАБЛИЦА НАВИГАЦИИ
        # ============================================================
        headers = ["№", "ТИП ОШИБКИ", "КОЛИЧЕСТВО", "ПЕРЕЙТИ"]

        for col_idx, header in enumerate(headers, start=table_start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
            cell.border = Border(
                left=Side(style="thin", color=COLORS["border_gray"]),
                right=Side(style="thin", color=COLORS["border_gray"]),
                top=Side(style="thin", color=COLORS["border_gray"]),
                bottom=Side(style="medium", color=COLORS["dark_green"])
            )

        self.ws.row_dimensions[row].height = 32
        row += 1

        sheets = []
        if self.stats.get('name_mismatch', 0) > 0:
            sheets.append(("Названия", "Несоответствие названий", self.stats['name_mismatch']))
        if self.stats.get('article_mismatch', 0) > 0:
            sheets.append(("Артикулы", "Несоответствие артикулов", self.stats['article_mismatch']))
        if self.stats.get('size_mismatch', 0) > 0:
            sheets.append(("Размеры", "Несоответствие размеров", self.stats['size_mismatch']))
        if self.stats.get('vat_mismatch', 0) > 0:
            sheets.append(("НДС", "Несоответствие ставки НДС", self.stats['vat_mismatch']))
        if self.stats.get('cert_issues', 0) > 0:
            sheets.append(("Сертификаты", "Проблемы с сертификатами", self.stats['cert_issues']))

        link_color = COLORS["dark_green"]

        for idx, (sheet_name, sheet_title, count) in enumerate(sheets):
            num_cell = self.ws.cell(row=row, column=table_start_col, value=idx + 1)
            num_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
            num_cell.alignment = Alignment(horizontal="center", vertical="center")
            num_cell.border = BORDERS["thin"]
            num_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")

            type_cell = self.ws.cell(row=row, column=table_start_col + 1, value=sheet_title)
            type_cell.font = Font(name="Roboto", size=10, bold=True, color=link_color)
            type_cell.alignment = Alignment(horizontal="left", vertical="center")
            type_cell.border = BORDERS["thin"]
            type_cell.hyperlink = f"#'{sheet_name}'!A1"

            count_cell = self.ws.cell(row=row, column=table_start_col + 2, value=count)
            count_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["error_red"])
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.border = BORDERS["thin"]
            count_cell.fill = PatternFill(start_color=COLORS["error_bg"], end_color=COLORS["error_bg"], fill_type="solid")

            link_cell = self.ws.cell(row=row, column=table_start_col + 3, value="Открыть")
            link_cell.font = Font(name="Roboto", size=9, bold=True, color=link_color, underline="single")
            link_cell.alignment = Alignment(horizontal="center", vertical="center")
            link_cell.border = BORDERS["thin"]
            link_cell.hyperlink = f"#'{sheet_name}'!A1"

            if idx % 2 == 0:
                row_fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")
            else:
                row_fill = PatternFill(start_color=COLORS["white"], end_color=COLORS["white"], fill_type="solid")

            type_cell.fill = row_fill
            link_cell.fill = row_fill

            self.ws.row_dimensions[row].height = 28
            row += 1

        if not sheets:
            no_errors_cell = self.ws.cell(row=row, column=table_start_col, value="Ошибок не найдено")
            no_errors_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
            no_errors_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_start_col + 3)
            row += 1

        for col in range(table_start_col, table_start_col + 4):
            cell = self.ws.cell(row=row, column=col)
            cell.border = Border(bottom=Side(style="medium", color=COLORS["dark_green"]))
        self.ws.row_dimensions[row].height = 8
        row += 1

        # ============================================================
        # ИТОГОВАЯ СТРОКА
        # ============================================================
        if sheets:
            total_sheets_cell = self.ws.cell(row=row, column=table_start_col, value=f"Всего листов с ошибками: {len(sheets)}")
            total_sheets_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["white"])
            total_sheets_cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
            total_sheets_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_start_col + 3)
            self.ws.row_dimensions[row].height = 32
            row += 2

        # ============================================================
        # НИЖНЯЯ ДЕКОРАТИВНАЯ ПОЛОСА
        # ============================================================
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
            cell.border = Border(top=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 6

        # ============================================================
        # НАСТРОЙКА ШИРИНЫ КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 6
        self.ws.column_dimensions["C"].width = 38
        self.ws.column_dimensions["D"].width = 14
        self.ws.column_dimensions["E"].width = 10

        self.ws.sheet_view.showGridLines = False

    def _format_number(self, value):
        if value is None:
            return "0"
        try:
            return f"{int(value):,}".replace(",", " ")
        except (ValueError, TypeError):
            return str(value)

    def _format_currency(self, value):
        if value is None:
            return "0,00"
        try:
            formatted = f"{float(value):,.2f}"
            return formatted.replace(",", " ").replace(".", ",")
        except (ValueError, TypeError):
            return str(value)