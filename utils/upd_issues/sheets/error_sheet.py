# utils/upd_issues/sheets/error_sheet.py
from openpyxl.styles import Font, Alignment, PatternFill
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, ALIGNMENTS


class ErrorSheet(BaseSheet):
    """Лист с ошибками определенного типа для конкретного УПД"""
    
    def __init__(self, workbook, sheet_name):
        super().__init__(workbook, sheet_name)
    
    def build_for_file(self, df, error_type, full_name: str, supplier: str = '—'):
        """Создает лист с ошибками"""
        row = 1
        start_col = 1
        
        # Кнопка возврата к справке
        back_cell = self.ws.cell(row=row, column=start_col, value="← К СПРАВКЕ")
        back_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        back_cell.alignment = ALIGNMENTS["left"]
        back_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        back_cell.border = self._thin_border()
        back_cell.hyperlink = "#'Справка'!A1"
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 2)
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # Заголовок
        title = self._get_title(error_type)
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 10)
        title_cell = self.ws.cell(row=row, column=start_col, value=title)
        title_cell.font = Font(name="Roboto", size=14, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 28
        row += 1
        
        # Информация о документе
        info_text = f"Документ: {full_name} | Поставщик: {supplier} | Найдено ошибок: {len(df)}"
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 10)
        info_cell = self.ws.cell(row=row, column=start_col, value=info_text)
        info_cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
        info_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 22
        row += 2
        
        # Подготовка данных
        headers, data_rows, highlight_cols = self._prepare_data(df, error_type)
        
        # Заголовки таблицы
        for col_idx, header in enumerate(headers, start=start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
            cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
            cell.alignment = ALIGNMENTS["center"]
            cell.border = self._thin_border()
        
        self.ws.row_dimensions[row].height = 28
        row += 1
        
        start_data_row = row
        
        # Данные
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=start_col):
                cell = self.ws.cell(row=row, column=col_idx, value=value)
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
                cell.alignment = self._get_alignment(col_idx, headers)
                cell.border = self._thin_border()
            
            # Подсветка только нужных колонок (не всей строки!)
            for col_idx in highlight_cols:
                abs_col_idx = start_col + col_idx
                self.ws.cell(row=row, column=abs_col_idx).fill = PatternFill(
                    start_color=COLORS["error_bg"], end_color=COLORS["error_bg"], fill_type="solid"
                )
            
            self.ws.row_dimensions[row].height = 22
            row += 1
        
        # Настройка ширины колонок
        self._set_column_widths(error_type)
        
        # Фильтр - ПРАВИЛЬНО на заголовки таблицы (строкой выше)
        last_col_letter = self._col_letter(start_col + len(headers) - 1)
        header_row = start_data_row - 1  # строка с заголовками
        self.ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{row - 1}'
        
        # Заморозка на строке с заголовками
        self.ws.freeze_panes = f'G{header_row + 1}'
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False
    
    def _get_title(self, error_type):
        titles = {
            "name": "НЕСООТВЕТСТВИЕ НАЗВАНИЙ",
            "article": "НЕСООТВЕТСТВИЕ АРТИКУЛОВ",
            "size": "НЕСООТВЕТСТВИЕ РАЗМЕРОВ",
            "vat": "НЕСООТВЕТСТВИЕ СТАВКИ НДС",
            "cert": "ПРОБЛЕМЫ С СЕРТИФИКАТАМИ"
        }
        return titles.get(error_type, "ОШИБКИ")
    
    def _prepare_data(self, df, error_type):
        """Подготовка данных с числами как числами, а не строками"""
        if error_type == "name":
            headers = [
                "ID", "Артикул УПД", "Артикул WB", "Бренд",
                "Название УПД", "Название WB",
                "Кол-во", "Цена без НДС", "Сумма без НДС", "Сумма НДС", "Итого с НДС"
            ]
            # Индексы колонок для подсветки (начиная с 0 для списка data_row)
            highlight_cols = [4, 5]  # Название УПД, Название WB
            data_rows = []
            for _, item in df.iterrows():
                data_rows.append([
                    item['id'],
                    item['upd_sa_name'],
                    item['sa_pid'],
                    item['brand'],
                    str(item['upd_title'])[:100],
                    str(item['cards_titles'])[:100],
                    self._get_numeric_value(item['upd_qty']),      # число
                    self._get_numeric_value(item['upd_price_vatless']),  # число
                    self._get_numeric_value(item['upd_amount_vatless']), # число
                    self._get_numeric_value(item['upd_vat_amount']),      # число
                    self._get_numeric_value(item['upd_amount_vatadd']),   # число
                ])
        
        elif error_type == "article":
            headers = [
                "ID", "Артикул УПД", "Артикул WB", "Бренд",
                "Название товара",
                "Кол-во", "Цена без НДС", "Сумма без НДС", "Сумма НДС", "Итого с НДС"
            ]
            highlight_cols = [1, 2]  # Артикул УПД, Артикул WB
            data_rows = []
            for _, item in df.iterrows():
                data_rows.append([
                    item['id'],
                    item['upd_sa_name'],
                    item['sa_pid'],
                    item['brand'],
                    str(item['upd_title'])[:80],
                    self._get_numeric_value(item['upd_qty']),
                    self._get_numeric_value(item['upd_price_vatless']),
                    self._get_numeric_value(item['upd_amount_vatless']),
                    self._get_numeric_value(item['upd_vat_amount']),
                    self._get_numeric_value(item['upd_amount_vatadd']),
                ])
        
        elif error_type == "size":
            headers = [
                "ID", "Артикул УПД", "Артикул WB", "Бренд",
                "Размер в УПД", "Доступные размеры WB", "Название УПД", 
                "Кол-во", "Цена без НДС", "Сумма без НДС", "Сумма НДС", "Итого с НДС"
            ]
            highlight_cols = [4, 5,]  # Название УПД, Размер в УПД, Доступные размеры WB
            data_rows = []
            for _, item in df.iterrows():
                data_rows.append([
                    item['id'],
                    item['upd_sa_name'],
                    item['sa_pid'],
                    item['brand'], 
                    item['upd_size'],
                    item['available_sizes'],
                    str(item['upd_title'])[:80], 
                    self._get_numeric_value(item['upd_qty']),
                    self._get_numeric_value(item['upd_price_vatless']),
                    self._get_numeric_value(item['upd_amount_vatless']),
                    self._get_numeric_value(item['upd_vat_amount']),
                    self._get_numeric_value(item['upd_amount_vatadd']),
                ])
        
        elif error_type == "vat":
            headers = [
                "ID", "Артикул УПД", "Артикул WB", "Бренд",
                "Название товара", "НДС УПД", "НДС WB",
                "Кол-во", "Цена без НДС", "Сумма без НДС", "Сумма НДС", "Итого с НДС"
            ]
            highlight_cols = [5, 6]  # НДС УПД, НДС WB
            data_rows = []
            for _, item in df.iterrows():
                data_rows.append([
                    item['id'],
                    item['upd_sa_name'],
                    item['sa_pid'],
                    item['brand'],
                    str(item['upd_title'])[:80],
                    item['upd_vat_rate'],
                    item['card_vat_rate'],
                    self._get_numeric_value(item['upd_qty']),
                    self._get_numeric_value(item['upd_price_vatless']),
                    self._get_numeric_value(item['upd_amount_vatless']),
                    self._get_numeric_value(item['upd_vat_amount']),
                    self._get_numeric_value(item['upd_amount_vatadd']),
                ])
        
        else:  # cert
            headers = [
                "ID", "Артикул УПД", "Артикул WB", "Бренд",
                "Название товара", "Дата окончания сертификата", "Статус сертификата",
                "Кол-во", "Цена без НДС", "Сумма без НДС", "Сумма НДС", "Итого с НДС"
            ]
            highlight_cols = [5, 6]  # Дата, Статус
            data_rows = []
            for _, item in df.iterrows():
                status = item['cert_status']
                if status == 'expired':
                    status_display = 'Истек'
                elif status == 'expiring_soon':
                    status_display = 'Истекает скоро'
                elif status == 'ok':
                    status_display = 'Действителен'
                else:
                    status_display = str(status) if status != '—' else 'Нет данных'
                
                data_rows.append([
                    item['id'],
                    item['upd_sa_name'],
                    item['sa_pid'],
                    item['brand'],
                    str(item['upd_title'])[:80],
                    item['cert_end_date'],
                    status_display,
                    self._get_numeric_value(item['upd_qty']),
                    self._get_numeric_value(item['upd_price_vatless']),
                    self._get_numeric_value(item['upd_amount_vatless']),
                    self._get_numeric_value(item['upd_vat_amount']),
                    self._get_numeric_value(item['upd_amount_vatadd']),
                ])
        
        return headers, data_rows, highlight_cols
    
    def _get_numeric_value(self, value):
        """
        Возвращает числовое значение для Excel (чтобы форматировалось как число)
        """
        if value is None or value == '—' or str(value) == 'nan':
            return None  # Пустая ячейка
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    
    def _set_column_widths(self, error_type):
        widths = {
            'A': 8, 'B': 14, 'C': 14, 'D': 14,
            'E': 35, 'F': 35, 'G': 10, 'H': 10,
            'I': 14, 'J': 14, 'K': 14, 'L': 14
        }
        
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width
        
        if error_type == "article":
            self.ws.column_dimensions['E'].width = 50
            self.ws.column_dimensions['F'].width = 10
        elif error_type == "size":
            self.ws.column_dimensions['E'].width = 20  # Название УПД
            self.ws.column_dimensions['F'].width = 20  # Размер в УПД
            self.ws.column_dimensions['G'].width = 40  # Доступные размеры WB
        elif error_type == "vat":
            self.ws.column_dimensions['E'].width = 30
            self.ws.column_dimensions['F'].width = 10
            self.ws.column_dimensions['G'].width = 10
        elif error_type == "cert":
            self.ws.column_dimensions['E'].width = 30
            self.ws.column_dimensions['F'].width = 25
            self.ws.column_dimensions['G'].width = 25
    
    def _thin_border(self):
        from openpyxl.styles import Border, Side
        thin = Side(style="thin", color=COLORS["border_gray"])
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def _get_alignment(self, col_idx, headers):
        col_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
        # Числовые колонки выравниваем вправо
        if col_name in ["ID", "Кол-во", "Цена без НДС", "Сумма без НДС", "Сумма НДС", "Итого с НДС"]:
            return ALIGNMENTS["right"]
        elif col_name in ["НДС УПД", "НДС WB"]:
            return ALIGNMENTS["center"]
        return ALIGNMENTS["left"]
    
    def _col_letter(self, col_num):
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result