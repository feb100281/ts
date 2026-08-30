import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def style_worksheet(ws, currency_columns=None, percent_columns=None):
    currency_columns = currency_columns or []
    percent_columns = percent_columns or []

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if cell.column in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            if cell.column in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def export_to_excel(daily_sheet, monthly_sheet, yearly_sheet, analysis_sheet, settings_sheet, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        daily_sheet.to_excel(writer, sheet_name="По дням", index=False)
        monthly_sheet.to_excel(writer, sheet_name="По месяцам", index=False)
        yearly_sheet.to_excel(writer, sheet_name="По годам", index=False)
        analysis_sheet.to_excel(writer, sheet_name="Анализ", index=False)
        settings_sheet.to_excel(writer, sheet_name="Настройки", index=False)

        wb = writer.book

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            headers = [cell.value for cell in ws[1]]

            currency_cols = []
            percent_cols = []

            for idx, h in enumerate(headers, start=1):
                if h and ("%" in str(h)):
                    percent_cols.append(idx)
                if h in [
                    "Факт", "Прогноз", "Нижняя граница прогноза", "Верхняя граница прогноза",
                    "Тренд", "Итог", "Отклонение факт-прогноз", "Значение",
                    "Изменение к пред. месяцу, ₽", "Изменение к тому же месяцу прошлого года, ₽",
                    "Изменение к пред. году, ₽"
                ]:
                    currency_cols.append(idx)

            style_worksheet(ws, currency_columns=currency_cols, percent_columns=percent_cols)