# # utils/forecast_revenue.py

# import psycopg
# from psycopg.rows import dict_row
# import pandas as pd
# import numpy as np
# from pprint import pprint
# from psycopg import Connection
# from prophet import Prophet

# def connect_db():
#     return psycopg.connect(
#         dbname="ts_db",  # DB_NAME
#         user="ts_user",  # DB_USER
#         password="Dec8108079",  # DB_PASSWORD
#         host="127.0.0.1",  # DB_HOST
#         port="5433",  # DB_PORT
#         connect_timeout=10,
#     )


# def get_forecast_data(conn, start_date, end_date):
#     SQL = """
#     SELECT
#     date_from as ds,
#     sum(dt-cr) as y
#     from gl.fact
#     where acc_id = 46 and subconto_id in (52,87)
#     and date_from > %s and date_from <= %s
#     group by date_from
#     order by 1  
#     """
    
#     with conn.cursor(row_factory=dict_row) as cur:
#         cur.execute(SQL,(start_date,end_date))        
#         rows = cur.fetchall()
    
#     df = pd.DataFrame(rows)
    
#     if df.empty:
#         return df

#     df["ds"] = pd.to_datetime(df["ds"])
#     df["y"] = pd.to_numeric(df["y"], errors="coerce")
#     df = df.dropna(subset=["ds", "y"]).sort_values("ds")

#     return df

# def revenue_prophet_forecast(conn, start_date, end_date, forecast_period, freq="D"):
    
#     #Получаем данные для погноза
#     data = get_forecast_data(conn, start_date, end_date)
#     model = Prophet(
#         yearly_seasonality=True,
#         weekly_seasonality=True,
#         daily_seasonality=False,
#     )
#     periods = (pd.to_datetime(forecast_period) - pd.to_datetime(end_date)).days

#     model.fit(data)

#     future = model.make_future_dataframe(periods=periods, freq=freq)
#     forecast = model.predict(future)

#     return data, forecast, model


# def main():
#     conn = connect_db()
    
#     START_DATE = '2024-01-31'
#     END_DATE = '2026-03-15'
#     FORECAST_DATE = '2026-12-31'
    
#     data, forecast, model = revenue_prophet_forecast(conn, START_DATE, END_DATE, FORECAST_DATE)
    
    
#     pprint(forecast)
    


# if __name__ == "__main__":
#     main()




# # utils/forecast_revenue.py

# import psycopg
# from psycopg.rows import dict_row
# import pandas as pd
# from prophet import Prophet


# def connect_db():
#     return psycopg.connect(
#         dbname="ts_db",
#         user="ts_user",
#         password="Dec8108079",
#         host="127.0.0.1",
#         port="5433",
#         connect_timeout=10,
#     )


# def get_forecast_data(conn, start_date, end_date):
#     sql = """
#     SELECT
#         date_from AS ds,
#         SUM(dt - cr) AS y
#     FROM gl.fact
#     WHERE acc_id = 46
#       AND subconto_id IN (52, 87)
#       AND date_from > %s
#       AND date_from <= %s
#     GROUP BY date_from
#     ORDER BY 1
#     """

#     with conn.cursor(row_factory=dict_row) as cur:
#         cur.execute(sql, (start_date, end_date))
#         rows = cur.fetchall()

#     df = pd.DataFrame(rows)

#     if df.empty:
#         return df

#     df["ds"] = pd.to_datetime(df["ds"])
#     df["y"] = pd.to_numeric(df["y"], errors="coerce") / 100
#     df = df.dropna(subset=["ds", "y"]).sort_values("ds")

#     return df


# def revenue_prophet_forecast(conn, start_date, end_date, forecast_date, freq="D"):
#     data = get_forecast_data(conn, start_date, end_date)

#     if data.empty:
#         raise ValueError("Нет данных для построения прогноза.")

#     model = Prophet(
#         yearly_seasonality=True,
#         weekly_seasonality=True,
#         daily_seasonality=False,
#     )

#     periods = (pd.to_datetime(forecast_date) - pd.to_datetime(end_date)).days
#     if periods < 0:
#         raise ValueError("forecast_date должен быть позже end_date.")

#     model.fit(data)

#     future = model.make_future_dataframe(periods=periods, freq=freq)
#     forecast = model.predict(future)

#     return data, forecast, model


# def prepare_daily_sheet(data, forecast, end_date):
#     end_date = pd.to_datetime(end_date)

#     # Берем только нужные колонки из прогноза
#     daily = forecast[["ds", "yhat", "yhat_lower", "yhat_upper"]].copy()

#     # Подтягиваем факт
#     daily = daily.merge(
#         data[["ds", "y"]],
#         on="ds",
#         how="left"
#     )

#     # Тип строки: факт / прогноз
#     daily["type"] = daily["ds"].apply(
#         lambda x: "Факт" if x <= end_date else "Прогноз"
#     )

#     # Для удобства можно сделать отдельную итоговую колонку:
#     # до end_date показываем факт, после end_date — прогноз
#     daily["value"] = daily["y"]
#     daily.loc[daily["ds"] > end_date, "value"] = daily.loc[daily["ds"] > end_date, "yhat"]

#     # Формат даты
#     daily["ds"] = pd.to_datetime(daily["ds"]).dt.date

#     # Переименование колонок
#     daily = daily.rename(columns={
#         "ds": "Дата",
#         "y": "Факт",
#         "yhat": "Прогноз",
#         "yhat_lower": "Нижняя граница прогноза",
#         "yhat_upper": "Верхняя граница прогноза",
#         "type": "Тип",
#         "value": "Итог"
#     })

#     return daily


# def prepare_monthly_sheet(daily_sheet):
#     monthly = daily_sheet.copy()
#     monthly["Дата"] = pd.to_datetime(monthly["Дата"])
#     monthly["Месяц"] = monthly["Дата"].dt.to_period("M").dt.to_timestamp()

#     monthly = (
#         monthly.groupby("Месяц", as_index=False)
#         .agg({
#             "Факт": "sum",
#             "Прогноз": "sum",
#             "Нижняя граница прогноза": "sum",
#             "Верхняя граница прогноза": "sum",
#             "Итог": "sum"
#         })
#     )

#     monthly["Месяц"] = monthly["Месяц"].dt.strftime("%Y-%m")

#     return monthly


# def export_forecast_to_excel(data, forecast, end_date, file_name="forecast_revenue.xlsx"):
#     daily_sheet = prepare_daily_sheet(data, forecast, end_date)
#     monthly_sheet = prepare_monthly_sheet(daily_sheet)

#     with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
#         daily_sheet.to_excel(writer, sheet_name="По дням", index=False)
#         monthly_sheet.to_excel(writer, sheet_name="По месяцам", index=False)

#     return file_name


# def main():
#     conn = connect_db()

#     try:
#         start_date = "2024-01-31"
#         end_date = "2026-03-15"
#         forecast_date = "2026-12-31"

#         data, forecast, model = revenue_prophet_forecast(
#             conn=conn,
#             start_date=start_date,
#             end_date=end_date,
#             forecast_date=forecast_date,
#             freq="D"
#         )

#         file_name = export_forecast_to_excel(
#             data=data,
#             forecast=forecast,
#             end_date=end_date,
#             file_name="forecast_revenue.xlsx"
#         )

#         print(f"Файл успешно сохранен: {file_name}")

#     finally:
#         conn.close()


# if __name__ == "__main__":
#     main()








# utils/forecast_revenue.py

import psycopg
from psycopg.rows import dict_row
import pandas as pd
from prophet import Prophet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =========================
# НАСТРОЙКИ
# =========================

DB_CONFIG = {
    "dbname": "ts_db",
    "user": "ts_user",
    "password": "Dec8108079",
    "host": "127.0.0.1",
    "port": "5433",
    "connect_timeout": 10,
}

START_DATE = "2024-01-31"
END_DATE = "2026-03-15"
FORECAST_DATE = "2026-12-31"
OUTPUT_FILE = "forecast_revenue.xlsx"

# Основные параметры Prophet, которыми удобно управлять
PROPHET_PARAMS = {
    # Тип сезонности:
    # "additive" — если сезонные колебания примерно одинаковые по размеру
    # "multiplicative" — если при росте выручки растет и амплитуда сезонности
    "seasonality_mode": "additive",

    # Насколько модель чувствительна к изменениям тренда.
    # Больше значение = модель быстрее подстраивается под переломы тренда.
    # Часто удобно тестировать: 0.03 / 0.05 / 0.1 / 0.2
    "changepoint_prior_scale": 0.08,

    # Насколько выраженными можно делать сезонности.
    # Часто удобно тестировать: 5 / 10 / 15 / 20
    "seasonality_prior_scale": 10.0,

    # Для праздников, если потом будешь добавлять holidays
    "holidays_prior_scale": 10.0,

    # Интервал неопределенности прогноза
    "interval_width": 0.8,

    # Сезонности
    "yearly_seasonality": True,
    "weekly_seasonality": True,
    "daily_seasonality": False,

    # Доп. месячная сезонность
    "add_monthly_seasonality": True,
    "monthly_period": 30.5,
    "monthly_fourier_order": 5,
}

# Нужна ли агрегация по пропущенным дням:
# если в некоторые даты нет проводок, то для финансового ряда часто полезно
# заполнять такие дни нулями, чтобы модель видела непрерывный ряд.
FILL_MISSING_DATES = True


# =========================
# ПОДКЛЮЧЕНИЕ К БД
# =========================

def connect_db():
    return psycopg.connect(**DB_CONFIG)


# =========================
# ПОЛУЧЕНИЕ ДАННЫХ
# =========================

def get_forecast_data(conn, start_date, end_date, fill_missing_dates=True):
    sql = """
    SELECT
        date_from AS ds,
        SUM(dt - cr) AS y
    FROM gl.fact
    WHERE acc_id = 46
      AND subconto_id IN (52, 87)
      AND date_from > %s
      AND date_from <= %s
    GROUP BY date_from
    ORDER BY 1
    """

    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute(sql, (start_date, end_date))
        rows = cur.fetchall()

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df["ds"] = pd.to_datetime(df["ds"])
    # Было в копейках -> переводим в рубли
    df["y"] = (pd.to_numeric(df["y"], errors="coerce") / 100).round(2)
    df = df.dropna(subset=["ds", "y"]).sort_values("ds")

    if fill_missing_dates:
        full_range = pd.date_range(df["ds"].min(), df["ds"].max(), freq="D")
        df = (
            df.set_index("ds")
              .reindex(full_range)
              .rename_axis("ds")
              .reset_index()
        )
        df["y"] = df["y"].fillna(0)

    return df


# =========================
# PROPHET
# =========================

def build_prophet_model(params: dict) -> Prophet:
    model = Prophet(
        yearly_seasonality=params["yearly_seasonality"],
        weekly_seasonality=params["weekly_seasonality"],
        daily_seasonality=params["daily_seasonality"],
        seasonality_mode=params["seasonality_mode"],
        changepoint_prior_scale=params["changepoint_prior_scale"],
        seasonality_prior_scale=params["seasonality_prior_scale"],
        holidays_prior_scale=params["holidays_prior_scale"],
        interval_width=params["interval_width"],
    )

    # Дополнительная месячная сезонность
    if params.get("add_monthly_seasonality", False):
        model.add_seasonality(
            name="monthly",
            period=params.get("monthly_period", 30.5),
            fourier_order=params.get("monthly_fourier_order", 5)
        )

    return model


def revenue_prophet_forecast(conn, start_date, end_date, forecast_date, freq="D", prophet_params=None):
    if prophet_params is None:
        prophet_params = PROPHET_PARAMS

    data = get_forecast_data(
        conn=conn,
        start_date=start_date,
        end_date=end_date,
        fill_missing_dates=FILL_MISSING_DATES
    )

    if data.empty:
        raise ValueError("Нет данных для построения прогноза.")

    periods = (pd.to_datetime(forecast_date) - pd.to_datetime(end_date)).days
    if periods < 0:
        raise ValueError("forecast_date должен быть позже end_date.")

    model = build_prophet_model(prophet_params)
    model.fit(data)

    future = model.make_future_dataframe(periods=periods, freq=freq)
    forecast = model.predict(future)

    return data, forecast, model


# =========================
# ПОДГОТОВКА ЛИСТОВ
# =========================

def prepare_daily_sheet(data, forecast, end_date):
    end_date = pd.to_datetime(end_date)

    daily = forecast[["ds", "trend", "yhat", "yhat_lower", "yhat_upper"]].copy()
    daily = daily.merge(data[["ds", "y"]], on="ds", how="left")

    daily["Тип"] = daily["ds"].apply(lambda x: "Факт" if x <= end_date else "Прогноз")
    daily["Итог"] = daily["y"]
    daily.loc[daily["ds"] > end_date, "Итог"] = daily.loc[daily["ds"] > end_date, "yhat"]

    daily["Отклонение факт-прогноз"] = daily["y"] - daily["yhat"]
    daily["Отклонение %"] = None
    mask_fact = daily["Тип"] == "Факт"
    daily.loc[mask_fact, "Отклонение %"] = (
        daily.loc[mask_fact, "Отклонение факт-прогноз"] /
        daily.loc[mask_fact, "yhat"].replace(0, pd.NA)
    )

    daily["Дата"] = pd.to_datetime(daily["ds"]).dt.date

    daily = daily.rename(columns={
        "y": "Факт",
        "yhat": "Прогноз",
        "yhat_lower": "Нижняя граница прогноза",
        "yhat_upper": "Верхняя граница прогноза",
        "trend": "Тренд",
    })

    daily = daily[
        [
            "Дата",
            "Тип",
            "Факт",
            "Прогноз",
            "Нижняя граница прогноза",
            "Верхняя граница прогноза",
            "Тренд",
            "Итог",
            "Отклонение факт-прогноз",
            "Отклонение %",
        ]
    ]

    return daily


def prepare_monthly_sheet(daily_sheet):
    monthly = daily_sheet.copy()
    monthly["Дата"] = pd.to_datetime(monthly["Дата"])
    monthly["Месяц"] = monthly["Дата"].dt.to_period("M").dt.to_timestamp()

    monthly = (
        monthly.groupby("Месяц", as_index=False)
        .agg({
            "Факт": "sum",
            "Прогноз": "sum",
            "Нижняя граница прогноза": "sum",
            "Верхняя граница прогноза": "sum",
            "Тренд": "sum",
            "Итог": "sum",
        })
    )

    monthly["Изменение к пред. месяцу, ₽"] = monthly["Итог"].diff()
    monthly["Изменение к пред. месяцу, %"] = monthly["Итог"].pct_change()

    monthly["Изменение к тому же месяцу прошлого года, ₽"] = monthly["Итог"] - monthly["Итог"].shift(12)
    monthly["Изменение к тому же месяцу прошлого года, %"] = monthly["Итог"] / monthly["Итог"].shift(12) - 1

    monthly["Год"] = monthly["Месяц"].dt.year
    monthly["Месяц_номер"] = monthly["Месяц"].dt.month
    monthly["Месяц"] = monthly["Месяц"].dt.strftime("%Y-%m")

    monthly = monthly[
        [
            "Месяц",
            "Год",
            "Месяц_номер",
            "Факт",
            "Прогноз",
            "Нижняя граница прогноза",
            "Верхняя граница прогноза",
            "Тренд",
            "Итог",
            "Изменение к пред. месяцу, ₽",
            "Изменение к пред. месяцу, %",
            "Изменение к тому же месяцу прошлого года, ₽",
            "Изменение к тому же месяцу прошлого года, %",
        ]
    ]

    return monthly


def prepare_yearly_sheet(monthly_sheet):
    yearly = monthly_sheet.copy()

    yearly = (
        yearly.groupby("Год", as_index=False)
        .agg({
            "Факт": "sum",
            "Прогноз": "sum",
            "Нижняя граница прогноза": "sum",
            "Верхняя граница прогноза": "sum",
            "Тренд": "sum",
            "Итог": "sum",
        })
    )

    yearly["Изменение к пред. году, ₽"] = yearly["Итог"].diff()
    yearly["Изменение к пред. году, %"] = yearly["Итог"].pct_change()

    return yearly


def prepare_analysis_sheet(monthly_sheet, yearly_sheet, end_date):
    end_date = pd.to_datetime(end_date)

    analysis_rows = []

    # Годовой анализ
    for _, row in yearly_sheet.iterrows():
        year = int(row["Год"])
        total = row["Итог"]
        delta_abs = row["Изменение к пред. году, ₽"]
        delta_pct = row["Изменение к пред. году, %"]

        analysis_rows.append({
            "Блок": "Годовой анализ",
            "Показатель": f"Выручка {year}",
            "Значение": total,
            "Комментарий": f"Итоговая выручка за {year} год"
        })

        if pd.notna(delta_abs):
            analysis_rows.append({
                "Блок": "Годовой анализ",
                "Показатель": f"Изменение {year} к предыдущему году, ₽",
                "Значение": delta_abs,
                "Комментарий": f"Абсолютное изменение выручки {year} года к предыдущему году"
            })
            analysis_rows.append({
                "Блок": "Годовой анализ",
                "Показатель": f"Изменение {year} к предыдущему году, %",
                "Значение": delta_pct,
                "Комментарий": f"Темп роста / снижения выручки {year} года к предыдущему году"
            })

    # Последние 12 месяцев
    temp = monthly_sheet.copy()
    temp["Месяц_dt"] = pd.to_datetime(temp["Месяц"] + "-01")
    last_12 = temp.sort_values("Месяц_dt").tail(12)

    if not last_12.empty:
        analysis_rows.append({
            "Блок": "Последние 12 месяцев",
            "Показатель": "Выручка за последние 12 месяцев",
            "Значение": last_12["Итог"].sum(),
            "Комментарий": "Сумма по последним 12 месяцам"
        })

        analysis_rows.append({
            "Блок": "Последние 12 месяцев",
            "Показатель": "Среднемесячная выручка за последние 12 месяцев",
            "Значение": last_12["Итог"].mean(),
            "Комментарий": "Среднее значение выручки по последним 12 месяцам"
        })

        analysis_rows.append({
            "Блок": "Последние 12 месяцев",
            "Показатель": "Лучший месяц за последние 12 месяцев",
            "Значение": last_12["Итог"].max(),
            "Комментарий": last_12.loc[last_12["Итог"].idxmax(), "Месяц"]
        })

        analysis_rows.append({
            "Блок": "Последние 12 месяцев",
            "Показатель": "Худший месяц за последние 12 месяцев",
            "Значение": last_12["Итог"].min(),
            "Комментарий": last_12.loc[last_12["Итог"].idxmin(), "Месяц"]
        })

    # Анализ прогноза
    forecast_only = temp[temp["Месяц_dt"] > end_date.replace(day=1)]
    if not forecast_only.empty:
        analysis_rows.append({
            "Блок": "Прогноз",
            "Показатель": "Среднемесячный прогноз после даты отсечения",
            "Значение": forecast_only["Итог"].mean(),
            "Комментарий": f"Средний месячный прогноз после {end_date.date()}"
        })

        analysis_rows.append({
            "Блок": "Прогноз",
            "Показатель": "Максимальный прогнозный месяц",
            "Значение": forecast_only["Итог"].max(),
            "Комментарий": forecast_only.loc[forecast_only["Итог"].idxmax(), "Месяц"]
        })

        analysis_rows.append({
            "Блок": "Прогноз",
            "Показатель": "Минимальный прогнозный месяц",
            "Значение": forecast_only["Итог"].min(),
            "Комментарий": forecast_only.loc[forecast_only["Итог"].idxmin(), "Месяц"]
        })

    analysis_df = pd.DataFrame(analysis_rows)

    return analysis_df


def prepare_settings_sheet():
    rows = [
        {"Параметр": "START_DATE", "Значение": START_DATE, "Комментарий": "Начало исторического периода"},
        {"Параметр": "END_DATE", "Значение": END_DATE, "Комментарий": "Последняя фактическая дата"},
        {"Параметр": "FORECAST_DATE", "Значение": FORECAST_DATE, "Комментарий": "Дата окончания прогноза"},
        {"Параметр": "FILL_MISSING_DATES", "Значение": str(FILL_MISSING_DATES), "Комментарий": "Заполнять пропуски по дням нулями"},
    ]

    for k, v in PROPHET_PARAMS.items():
        rows.append({
            "Параметр": f"PROPHET_PARAMS.{k}",
            "Значение": str(v),
            "Комментарий": "Параметр модели Prophet"
        })

    return pd.DataFrame(rows)


# =========================
# ФОРМАТИРОВАНИЕ EXCEL
# =========================

def style_worksheet(ws, currency_columns=None, percent_columns=None):
    if currency_columns is None:
        currency_columns = []
    if percent_columns is None:
        percent_columns = []

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if cell.column in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            if cell.column in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    # Автоширина
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


# =========================
# ВЫГРУЗКА В EXCEL
# =========================

def export_forecast_to_excel(
    data,
    forecast,
    end_date,
    file_name="forecast_revenue.xlsx"
):
    daily_sheet = prepare_daily_sheet(data, forecast, end_date)
    monthly_sheet = prepare_monthly_sheet(daily_sheet)
    yearly_sheet = prepare_yearly_sheet(monthly_sheet)
    analysis_sheet = prepare_analysis_sheet(monthly_sheet, yearly_sheet, end_date)
    settings_sheet = prepare_settings_sheet()

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        daily_sheet.to_excel(writer, sheet_name="По дням", index=False)
        monthly_sheet.to_excel(writer, sheet_name="По месяцам", index=False)
        yearly_sheet.to_excel(writer, sheet_name="По годам", index=False)
        analysis_sheet.to_excel(writer, sheet_name="Анализ", index=False)
        settings_sheet.to_excel(writer, sheet_name="Настройки", index=False)

        wb = writer.book

        # По дням
        ws = wb["По дням"]
        headers = [cell.value for cell in ws[1]]
        currency_cols = [headers.index(x) + 1 for x in headers if x in [
            "Факт",
            "Прогноз",
            "Нижняя граница прогноза",
            "Верхняя граница прогноза",
            "Тренд",
            "Итог",
            "Отклонение факт-прогноз",
        ]]
        percent_cols = [headers.index(x) + 1 for x in headers if x in ["Отклонение %"]]
        style_worksheet(ws, currency_columns=currency_cols, percent_columns=percent_cols)

        # По месяцам
        ws = wb["По месяцам"]
        headers = [cell.value for cell in ws[1]]
        currency_cols = [headers.index(x) + 1 for x in headers if x in [
            "Факт",
            "Прогноз",
            "Нижняя граница прогноза",
            "Верхняя граница прогноза",
            "Тренд",
            "Итог",
            "Изменение к пред. месяцу, ₽",
            "Изменение к тому же месяцу прошлого года, ₽",
        ]]
        percent_cols = [headers.index(x) + 1 for x in headers if x in [
            "Изменение к пред. месяцу, %",
            "Изменение к тому же месяцу прошлого года, %",
        ]]
        style_worksheet(ws, currency_columns=currency_cols, percent_columns=percent_cols)

        # По годам
        ws = wb["По годам"]
        headers = [cell.value for cell in ws[1]]
        currency_cols = [headers.index(x) + 1 for x in headers if x in [
            "Факт",
            "Прогноз",
            "Нижняя граница прогноза",
            "Верхняя граница прогноза",
            "Тренд",
            "Итог",
            "Изменение к пред. году, ₽",
        ]]
        percent_cols = [headers.index(x) + 1 for x in headers if x in [
            "Изменение к пред. году, %",
        ]]
        style_worksheet(ws, currency_columns=currency_cols, percent_columns=percent_cols)

        # Анализ
        ws = wb["Анализ"]
        headers = [cell.value for cell in ws[1]]
        currency_cols = [headers.index("Значение") + 1] if "Значение" in headers else []
        style_worksheet(ws, currency_columns=currency_cols, percent_columns=[])

        # Настройки
        ws = wb["Настройки"]
        style_worksheet(ws, currency_columns=[], percent_columns=[])

    return file_name


# =========================
# MAIN
# =========================

def main():
    conn = connect_db()

    try:
        data, forecast, model = revenue_prophet_forecast(
            conn=conn,
            start_date=START_DATE,
            end_date=END_DATE,
            forecast_date=FORECAST_DATE,
            freq="D",
            prophet_params=PROPHET_PARAMS,
        )

        file_name = export_forecast_to_excel(
            data=data,
            forecast=forecast,
            end_date=END_DATE,
            file_name=OUTPUT_FILE
        )

        print(f"Файл успешно сохранен: {file_name}")

    finally:
        conn.close()


if __name__ == "__main__":
    main()