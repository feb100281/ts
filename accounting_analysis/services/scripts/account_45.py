# accounting_analysis/services/scripts/account_45.py


from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from accounting_analysis.services.reports.account_45_report import build_account_45_report


def clean_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    lines = [line for line in lines if line]

    if not lines:
        return None

    return "\n".join(lines)


def clean_one_line(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    return text.replace("\n", " | ")


def to_number(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text == "":
        return None

    text = text.replace("\u00A0", "").replace(" ", "")

    if text.count(",") > 0 and text.count(".") > 0:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")

    if text in {"", "-"}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def is_total_row(name: str | None) -> bool:
    if not name:
        return False
    return str(name).strip().lower() == "итого"


def is_account_code(name: str | None) -> bool:
    if not name:
        return False
    return bool(re.fullmatch(r"\d{2}(?:\.\d+)?", str(name).strip()))


def is_service_header(name: str | None) -> bool:
    if not name:
        return False
    return str(name).strip() in {
        "Счет",
        "Контрагенты",
        "Номенклатура, Артикул",
    }


def unmerge_and_fill_sheet(ws) -> None:
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row

        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def find_header_row(ws) -> int:
    for row_idx in range(1, ws.max_row + 1):
        a = clean_text(ws.cell(row=row_idx, column=1).value)
        b = clean_text(ws.cell(row=row_idx, column=2).value)

        if a == "Счет" and b and "Показа" in b:
            return row_idx

    raise ValueError("Не удалось найти строку заголовков ОСВ (Счет / Показатели).")


def extract_report_meta(ws) -> dict[str, Any]:
    company = clean_one_line(ws["A1"].value)
    report_title = clean_one_line(ws["A2"].value)
    selection = clean_one_line(ws["A4"].value)

    if report_title:
        report_title = report_title.replace("Оборотно-сальдовая ведомость", "ОСВ").strip()

    if selection:
        selection = selection.replace("Отбор:", "").strip()
        selection = selection.replace("Контрагенты", "").strip(" :")

    return {
        "company": company,
        "report_title": report_title,
        "selection": selection,
    }


def open_excel_sheet(input_file: Path):
    if not input_file.exists():
        raise FileNotFoundError(f"Файл не найден: {input_file}")

    suffix = input_file.suffix.lower()

    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(
            f"Неподдерживаемый формат файла: {suffix}. "
            f"Нужен Excel-файл .xlsx/.xlsm/.xltx/.xltm"
        )

    try:
        wb = load_workbook(input_file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return wb, ws
    except InvalidFileException as e:
        raise ValueError(
            f"Файл не удалось открыть как Excel: {input_file}\n"
            f"Причина: {e}\n"
            f"Проверь, что файл действительно сохранен как .xlsx"
        ) from e


def read_sheet_to_raw_dataframe(ws) -> pd.DataFrame:
    header_row = find_header_row(ws)
    data_start_row = header_row + 3

    rows = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        rows.append(
            {
                "excel_row": row_idx,
                "name_raw": ws.cell(row=row_idx, column=1).value,
                "indicator_raw": ws.cell(row=row_idx, column=2).value,
                "opening_debit_raw": ws.cell(row=row_idx, column=3).value,
                "opening_credit_raw": ws.cell(row=row_idx, column=4).value,
                "turnover_debit_raw": ws.cell(row=row_idx, column=5).value,
                "turnover_credit_raw": ws.cell(row=row_idx, column=6).value,
                "ending_debit_raw": ws.cell(row=row_idx, column=7).value,
                "ending_credit_raw": ws.cell(row=row_idx, column=8).value,
                "indent": ws.cell(row=row_idx, column=1).alignment.indent or 0,
            }
        )

    df = pd.DataFrame(rows)

    df["name"] = df["name_raw"].apply(clean_one_line)
    df["indicator"] = df["indicator_raw"].apply(clean_one_line)

    numeric_pairs = [
        ("opening_debit_raw", "opening_debit"),
        ("opening_credit_raw", "opening_credit"),
        ("turnover_debit_raw", "turnover_debit"),
        ("turnover_credit_raw", "turnover_credit"),
        ("ending_debit_raw", "ending_debit"),
        ("ending_credit_raw", "ending_credit"),
    ]

    for src, dst in numeric_pairs:
        df[dst] = df[src].apply(to_number)

    value_cols = [
        "opening_debit",
        "opening_credit",
        "turnover_debit",
        "turnover_credit",
        "ending_debit",
        "ending_credit",
    ]

    df = df[
        df["name"].notna()
        | df["indicator"].notna()
        | df[value_cols].notna().any(axis=1)
    ].copy()

    df = df.reset_index(drop=True)
    return df


def fill_name_for_qty_rows(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()

    df["name_filled"] = None
    last_name = None

    for idx in df.index:
        current_name = df.at[idx, "name"]
        indicator = df.at[idx, "indicator"]

        if current_name and not is_total_row(current_name):
            last_name = current_name
            df.at[idx, "name_filled"] = current_name
        else:
            if indicator in {"БУ", "Кол."}:
                df.at[idx, "name_filled"] = last_name
            else:
                df.at[idx, "name_filled"] = current_name

    return df


def get_item_names_by_deepest_named_level(raw_df: pd.DataFrame) -> set[str]:
    named_df = raw_df[
        raw_df["name"].notna()
        & (~raw_df["name"].apply(is_total_row))
        & (~raw_df["name"].apply(is_service_header))
        & (~raw_df["name"].apply(is_account_code))
    ].copy()

    if named_df.empty:
        raise ValueError("Не удалось определить уровень товарных строк.")

    max_indent = int(named_df["indent"].max())

    item_names = set(
        named_df.loc[named_df["indent"] == max_indent, "name"].dropna().tolist()
    )

    if not item_names:
        raise ValueError("Не удалось выделить товарные строки по максимальному indent.")

    return item_names


def build_item_dataframe(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = fill_name_for_qty_rows(raw_df)
    item_names = get_item_names_by_deepest_named_level(raw_df)

    df = df[~df["name_filled"].apply(is_total_row)].copy()
    df = df[~df["name_filled"].apply(is_service_header)].copy()
    df = df[~df["name_filled"].apply(is_account_code)].copy()

    df = df[df["indicator"].isin(["БУ", "Кол."])].copy()
    df = df[df["name_filled"].isin(item_names)].copy()

    value_cols = [
        "opening_debit",
        "opening_credit",
        "turnover_debit",
        "turnover_credit",
        "ending_debit",
        "ending_credit",
    ]
    df = df[df[value_cols].notna().any(axis=1)].copy()

    grouped_records = []

    for item_name, grp in df.groupby("name_filled", sort=False):
        rec: dict[str, Any] = {
            "item_name": item_name,
            "excel_rows": ", ".join(str(x) for x in grp["excel_row"].tolist()),
        }

        bu = grp[grp["indicator"] == "БУ"].head(1)
        qty = grp[grp["indicator"] == "Кол."].head(1)

        rec["opening_amount_debit"] = bu["opening_debit"].iloc[0] if not bu.empty else None
        rec["opening_amount_credit"] = bu["opening_credit"].iloc[0] if not bu.empty else None
        rec["turnover_amount_debit"] = bu["turnover_debit"].iloc[0] if not bu.empty else None
        rec["turnover_amount_credit"] = bu["turnover_credit"].iloc[0] if not bu.empty else None
        rec["ending_amount_debit"] = bu["ending_debit"].iloc[0] if not bu.empty else None
        rec["ending_amount_credit"] = bu["ending_credit"].iloc[0] if not bu.empty else None

        rec["opening_qty_debit"] = qty["opening_debit"].iloc[0] if not qty.empty else None
        rec["opening_qty_credit"] = qty["opening_credit"].iloc[0] if not qty.empty else None
        rec["turnover_qty_debit"] = qty["turnover_debit"].iloc[0] if not qty.empty else None
        rec["turnover_qty_credit"] = qty["turnover_credit"].iloc[0] if not qty.empty else None
        rec["ending_qty_debit"] = qty["ending_debit"].iloc[0] if not qty.empty else None
        rec["ending_qty_credit"] = qty["ending_credit"].iloc[0] if not qty.empty else None

        grouped_records.append(rec)

    items_df = pd.DataFrame(grouped_records)

    if items_df.empty:
        raise ValueError("Не удалось собрать номенклатуру. Проверь структуру выгрузки.")

    numeric_cols = [
        "opening_amount_debit",
        "opening_amount_credit",
        "turnover_amount_debit",
        "turnover_amount_credit",
        "ending_amount_debit",
        "ending_amount_credit",
        "opening_qty_debit",
        "opening_qty_credit",
        "turnover_qty_debit",
        "turnover_qty_credit",
        "ending_qty_debit",
        "ending_qty_credit",
    ]

    for col in numeric_cols:
        items_df[col] = pd.to_numeric(items_df[col], errors="coerce")

    items_df["opening_qty"] = items_df["opening_qty_debit"]
    items_df["opening_amount"] = items_df["opening_amount_debit"]

    items_df["ending_qty"] = items_df["ending_qty_debit"]
    items_df["ending_amount"] = items_df["ending_amount_debit"]

    items_df["sold_qty"] = items_df["turnover_qty_credit"].fillna(0)
    items_df["sold_amount"] = items_df["turnover_amount_credit"].fillna(0)

    items_df["avg_sale_price"] = items_df.apply(
        lambda row: row["sold_amount"] / row["sold_qty"]
        if pd.notna(row["sold_qty"]) and row["sold_qty"] != 0
        else None,
        axis=1,
    )

    items_df["is_negative_ending_qty"] = items_df["ending_qty"].fillna(0) < 0
    items_df["is_negative_ending_amount"] = items_df["ending_amount"].fillna(0) < 0


    
    items_df["qty_exists_amount_missing_at_end"] = (
        (items_df["ending_qty"].fillna(0) > 0)
        & (items_df["ending_amount"].fillna(0) == 0)
    )

    items_df["amount_exists_qty_missing_at_end"] = (
        (items_df["ending_amount"].fillna(0) != 0)
        & (items_df["ending_qty"].fillna(0) == 0)
    )

    items_df["article_candidate"] = items_df["item_name"].str.extract(r"(\d{6,})", expand=False)

    return items_df



def build_summary(items_df: pd.DataFrame, meta: dict[str, Any]) -> pd.DataFrame:
    negative_qty_df = items_df[items_df["is_negative_ending_qty"]].copy()
    negative_amount_df = items_df[items_df["is_negative_ending_amount"]].copy()
    qty_no_amount_df = items_df[items_df["qty_exists_amount_missing_at_end"]].copy()
    amount_no_qty_df = items_df[items_df["amount_exists_qty_missing_at_end"]].copy()

    total_sold_qty = items_df["sold_qty"].fillna(0).sum()
    total_sold_amount = items_df["sold_amount"].fillna(0).sum()
    avg_sale_price_total = total_sold_amount / total_sold_qty if total_sold_qty else None

    negative_qty_abs_sum = abs(negative_qty_df["ending_qty"].fillna(0).sum())

    summary_rows = [
        ("Компания", meta.get("company")),
        ("ОСВ", meta.get("report_title")),
        ("Контрагент", meta.get("selection")),
        ("Полная детализация", "Все позиции"),
        ("Количество номенклатурных позиций", len(items_df)),
        ("Позиции с отрицательным количеством на конец", len(negative_qty_df)),
        ("Сумма отрицательных остатков в штуках (abs)", negative_qty_abs_sum),
        ("Позиции с отрицательной суммой на конец", len(negative_amount_df)),
        ("Позиции, где есть количество, но нет суммы на конец", len(qty_no_amount_df)),
        ("Позиции, где есть сумма, но нет количества на конец", len(amount_no_qty_df)),
        ("Продано за период, шт", total_sold_qty),
        ("Продано за период, сумма", total_sold_amount),
        ("Средняя цена продажи за 1 единицу", avg_sale_price_total),
        ("Остаток на начало, всего шт", items_df["opening_qty"].fillna(0).sum()),
        ("Остаток на начало, всего сумма", items_df["opening_amount"].fillna(0).sum()),
        ("Остаток на конец, всего шт", items_df["ending_qty"].fillna(0).sum()),
        ("Остаток на конец, всего сумма", items_df["ending_amount"].fillna(0).sum()),
    ]

    return pd.DataFrame(summary_rows, columns=["Показатель", "Значение"])


def format_output_frames(
    items_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    negative_output = items_df[items_df["is_negative_ending_qty"]].copy()
    qty_no_amount_output = items_df[items_df["qty_exists_amount_missing_at_end"]].copy()
    negative_amount_output = items_df[items_df["is_negative_ending_amount"]].copy()
    amount_no_qty_output = items_df[items_df["amount_exists_qty_missing_at_end"]].copy()

    items_output = items_df[
        [
            "item_name",
            "article_candidate",
            "opening_qty",
            "opening_amount",
            "sold_qty",
            "sold_amount",
            "avg_sale_price",
            "ending_qty",
            "ending_amount",
            "is_negative_ending_qty",
            "is_negative_ending_amount",
            "qty_exists_amount_missing_at_end",
            "amount_exists_qty_missing_at_end",
            "excel_rows",
        ]
    ].copy()

    negative_output = negative_output[
        ["item_name", "article_candidate", "ending_qty", "ending_amount", "excel_rows"]
    ].copy()

    qty_no_amount_output = qty_no_amount_output[
        ["item_name", "article_candidate", "ending_qty", "ending_amount", "excel_rows"]
    ].copy()

    negative_amount_output = negative_amount_output[
        ["item_name", "article_candidate", "ending_qty", "ending_amount", "excel_rows"]
    ].copy()

    amount_no_qty_output = amount_no_qty_output[
        ["item_name", "article_candidate", "ending_qty", "ending_amount", "excel_rows"]
    ].copy()

    negative_output = negative_output.sort_values(by="ending_qty", ascending=True)
    qty_no_amount_output = qty_no_amount_output.sort_values(by="ending_qty", ascending=True)
    negative_amount_output = negative_amount_output.sort_values(by="ending_amount", ascending=True)
    amount_no_qty_output = amount_no_qty_output.sort_values(by="ending_amount", ascending=False)

    return (
        items_output,
        negative_output,
        qty_no_amount_output,
        negative_amount_output,
        amount_no_qty_output,
    )


def build_control_sheet(raw_df: pd.DataFrame, items_df: pd.DataFrame) -> pd.DataFrame:
    df = fill_name_for_qty_rows(raw_df).copy()
    item_names = get_item_names_by_deepest_named_level(raw_df)

    df["is_item_name"] = df["name_filled"].isin(item_names)
    df["used_in_items"] = (
        df["name_filled"].isin(set(items_df["item_name"].tolist()))
        & df["indicator"].isin(["БУ", "Кол."])
    )

    return df[
        [
            "excel_row",
            "name",
            "name_filled",
            "indicator",
            "indent",
            "opening_debit",
            "opening_credit",
            "turnover_debit",
            "turnover_credit",
            "ending_debit",
            "ending_credit",
            "is_item_name",
            "used_in_items",
        ]
    ].copy()


def analyze_osv_45(input_file: Path, output_file: Path) -> None:
    _, ws = open_excel_sheet(input_file)

    meta = extract_report_meta(ws)

    unmerge_and_fill_sheet(ws)
    raw_df = read_sheet_to_raw_dataframe(ws)
    items_df = build_item_dataframe(raw_df)
    summary_df = build_summary(items_df, meta)

    (
        items_output,
        negative_output,
        qty_no_amount_output,
        negative_amount_output,
        amount_no_qty_output,
    ) = format_output_frames(items_df)


    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        build_account_45_report(
            writer=writer,
            summary_df=summary_df,
            items_output=items_output,
            negative_output=negative_output,
            qty_no_amount_output=qty_no_amount_output,
            negative_amount_output=negative_amount_output,
            amount_no_qty_output=amount_no_qty_output,
        )


def run_account_45(input_file_path: str) -> str:
    input_path = Path(input_file_path)
    output_path = input_path.with_name(input_path.stem + "_analysis.xlsx")

    analyze_osv_45(input_path, output_path)

    return str(output_path)