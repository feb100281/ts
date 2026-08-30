# accounting_analysis/services/styles/theme.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED",
    "tab_light_green": "CFE5DC",
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "back_fill": "E7F1ED",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "red_soft": "FDECEC",
    "red_text": "7A2E2E",
    "yellow_soft": "FFF7DB",
    "yellow_text": "8A6D1D",
    "conclusion_fill": "F4F8F6",
    "conclusion_border": "9DB7AC",
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "back": PatternFill("solid", fgColor=COLORS["back_fill"]),
    "danger": PatternFill("solid", fgColor=COLORS["red_soft"]),
    "warning": PatternFill("solid", fgColor=COLORS["yellow_soft"]),
    "none": PatternFill(fill_type=None),
    "conclusion": PatternFill("solid", fgColor=COLORS["conclusion_fill"]),
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=10, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "danger": Font(name="Roboto", size=10, color=COLORS["red_text"]),
    "warning": Font(name="Roboto", size=10, bold=True, color=COLORS["yellow_text"]),
    "conclusion_title": Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"]),
    "conclusion_text": Font(name="Roboto", size=10, color=COLORS["black"]),
    "danger_bold": Font(name="Roboto", size=10, bold=True, color=COLORS["red_text"]),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium_dark = Side(style="medium", color=COLORS["dark_green"])
dashed_green = Side(style="dashed", color=COLORS["dark_green"])
conclusion_side = Side(style="thin", color=COLORS["conclusion_border"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium_dark),
    "top_bottom_medium": Border(top=medium_dark, bottom=medium_dark),
    "top_bottom_dashed": Border(top=dashed_green, bottom=dashed_green),
    "none": Border(),
    "conclusion_box": Border(left=conclusion_side, right=conclusion_side, top=conclusion_side, bottom=conclusion_side,
    ),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
}

FORMATS = {
    "money": '#,##0.00;(#,##0.00)',
    "int": '#,##0',
    "qty": '#,##0.000',
    "date": 'dd.mm.yyyy',
    "text": '@',
}