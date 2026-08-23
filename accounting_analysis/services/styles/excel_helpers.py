# accounting_analysis/services/styles/excel_helpers.py
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import CellIsRule
from copy import copy
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from .theme import FILLS, FONTS, BORDERS, ALIGNMENTS, FORMATS, COLORS


def hide_grid_and_freeze(ws, freeze_cell="A1"):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_cell


def set_column_widths(ws, widths: dict[str, float]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_heights(ws, heights: dict[int, float]):
    for row_idx, height in heights.items():
        ws.row_dimensions[row_idx].height = height


def draw_sheet_header(ws, title, subtitle="", note="", line_to_col=8):
    ws["A2"] = title
    ws["A3"] = subtitle
    ws["A4"] = note

    ws["A2"].font = FONTS["title"]
    ws["A3"].font = FONTS["subtitle"]
    ws["A4"].font = FONTS["subtitle"]

    ws["A2"].alignment = ALIGNMENTS["left"]
    ws["A3"].alignment = ALIGNMENTS["left"]
    ws["A4"].alignment = ALIGNMENTS["left"]

    for col in range(1, line_to_col + 1):
        ws.cell(row=6, column=col).border = BORDERS["bottom_medium"]


def draw_section_title(ws, row, col_start, col_end, title):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILLS["section"]
        cell.border = BORDERS["bottom_thin"]
        if col == col_start:
            cell.value = title
            cell.font = FONTS["section"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.value = None


def draw_table_header(ws, row, headers, start_col=1, wrap=True):
    alignment = ALIGNMENTS["center_wrap"] if wrap else ALIGNMENTS["center"]

    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i, value=header)
        cell.fill = FILLS["header"]
        cell.font = FONTS["header_white"]
        cell.alignment = alignment
        cell.border = BORDERS["thin"]


def style_data_row(ws, row, values, start_col=1, number_formats=None):
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = FONTS["normal"]
        cell.alignment = ALIGNMENTS["left"] if i == start_col else ALIGNMENTS["right"]
        cell.border = BORDERS["thin"]
        cell.fill = FILLS["alt"] if row % 2 == 0 else FILLS["none"]

        if number_formats and i in number_formats:
            cell.number_format = number_formats[i]


def style_total_row(ws, row, values, start_col=1, number_formats=None):
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = FONTS["total"]
        cell.alignment = ALIGNMENTS["left"] if i == start_col else ALIGNMENTS["right"]
        cell.border = BORDERS["top_bottom_medium"]
        cell.fill = FILLS["total"]

        if number_formats and i in number_formats:
            cell.number_format = number_formats[i]


def autosize_by_content(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column

        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if "\n" in value:
                    value = max(value.split("\n"), key=len)
                max_len = max(max_len, len(value))
            except Exception:
                pass

        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_negative_highlight(ws, start_row, end_row, col_idx):
    red_fill = FILLS["danger"]
    red_font = FONTS["danger"]

    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.fill = red_fill
            cell.font = red_font


def apply_zero_warning(ws, start_row, end_row, qty_col_idx, amount_col_idx):
    for row in range(start_row, end_row + 1):
        qty_cell = ws.cell(row=row, column=qty_col_idx)
        amount_cell = ws.cell(row=row, column=amount_col_idx)

        qty = qty_cell.value or 0
        amount = amount_cell.value or 0

        if qty != 0 and amount == 0:
            qty_cell.fill = FILLS["warning"]
            amount_cell.fill = FILLS["warning"]
            qty_cell.font = FONTS["warning"]
            amount_cell.font = FONTS["warning"]

        if amount != 0 and qty == 0:
            qty_cell.fill = FILLS["warning"]
            amount_cell.fill = FILLS["warning"]
            qty_cell.font = FONTS["warning"]
            amount_cell.font = FONTS["warning"]


def add_filter(ws, header_row, last_col):
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{header_row}"


def thin_border_box(ws, row_start, row_end, col_start, col_end):
    thin = Side(style="thin", color=COLORS["border_gray"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = border
            
            
            
def add_back_to_summary_link(ws, row=1, col=1, target_sheet="Summary"):
    cell = ws.cell(row=row, column=col, value="← Вернуться к Summary")
    cell.hyperlink = f"#{target_sheet}!A1"
    cell.font = FONTS["bold"]
    cell.alignment = ALIGNMENTS["left"]
    cell.fill = FILLS["back"]
    cell.border = BORDERS["thin"]

    # чтобы кнопка выглядела аккуратнее
    if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < 20:
        ws.row_dimensions[row].height = 20


def set_tab_color(ws, color: str):
    ws.sheet_properties.tabColor = color
    
    

def insert_section_row(ws, row, title, col_start=1, col_end=3):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)

        cell.fill = FILLS["section"]
        cell.border = BORDERS["top_bottom_dashed"]
       

        if col == col_start:
            cell.value = title
            cell.font = FONTS["section"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.value = None

    ws.row_dimensions[row].height = 22
    
    
    
def draw_conclusion_block(
    ws,
    start_row: int,
    col_start: int,
    col_end: int,
    title: str,
    conclusions: list[dict],
) -> int:
    # Заголовок блока
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = FILLS["section"]
        cell.border = BORDERS["conclusion_box"]

        if col == col_start:
            cell.value = title
            cell.font = FONTS["conclusion_title"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.value = None

    ws.row_dimensions[start_row].height = 22

    current_row = start_row + 1

    normal_inline = InlineFont(
        rFont="Roboto",
        sz=10,
        color=COLORS["black"],
    )

    accent_inline = InlineFont(
        rFont="Roboto",
        sz=10,
        b=True,
        color=COLORS["red_text"],
    )

    for item in conclusions:
        # merge на всю ширину блока
        ws.merge_cells(
            start_row=current_row,
            start_column=col_start,
            end_row=current_row,
            end_column=col_end,
        )

        cell = ws.cell(row=current_row, column=col_start)
        cell.fill = FILLS["conclusion"]
        cell.border = BORDERS["conclusion_box"]
        cell.alignment = ALIGNMENTS["left_wrap"]

        # т.к. после merge border/fill нужно продублировать по всем ячейкам диапазона
        for col in range(col_start, col_end + 1):
            merged_cell = ws.cell(row=current_row, column=col)
            merged_cell.fill = FILLS["conclusion"]
            merged_cell.border = BORDERS["conclusion_box"]

        if isinstance(item, str):
            cell.value = f"• {item}"
            cell.font = FONTS["conclusion_text"]

        elif item.get("type") == "rich":
            prefix = item.get("prefix", "")
            highlight = item.get("highlight", "")
            suffix = item.get("suffix", "")

            rich_value = CellRichText(
                "• ",
                TextBlock(normal_inline, prefix),
                TextBlock(accent_inline, highlight),
                TextBlock(normal_inline, suffix),
            )
            cell.value = rich_value
            cell.font = FONTS["conclusion_text"]

        else:
            text = item.get("text", "")
            cell.value = f"• {text}"
            cell.font = FONTS["conclusion_text"]

        ws.row_dimensions[current_row].height = 42
        current_row += 1

    return current_row