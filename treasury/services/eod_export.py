from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule


from treasury.models import BankStatements, CfData


def export_eod_xlsx(request):
    raw = request.GET.get("in_period_date")
    if not raw:
        return HttpResponse("Не задан параметр in_period_date", status=400)

    try:
        selected_date = datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Некорректная дата", status=400)

    # те же фильтры, что в changelist_view
    bss = (
        BankStatements.objects
        .filter(start__lte=selected_date, finish__gte=selected_date)
        .select_related("owner", "ba", "ba__bank")
    )

    owner_id = request.GET.get("owner__id__exact")
    ba_id = request.GET.get("ba__id__exact")
    if owner_id:
        bss = bss.filter(owner_id=owner_id)
    if ba_id:
        bss = bss.filter(ba_id=ba_id)

    bs_ids = list(bss.values_list("id", flat=True))
    if not bs_ids:
        return HttpResponse("Нет выписок, покрывающих дату", status=404)

    qs = (
        CfData.objects
        .filter(bs_id__in=bs_ids, date__lte=selected_date)
        .select_related("cp_final", "cp", "ba", "bs", "cfitem", "contract", "bs__owner", "ba__bank")
        .order_by("date", "id")
    )

    # отдельный набор: строго выбранный день
    qs_day = qs.filter(date=selected_date)

    # --- helpers ---
    def cp_name(obj):
        if obj.cp_final_id and obj.cp_final:
            return str(obj.cp_final)
        if obj.cp_id and obj.cp:
            return str(obj.cp)
        if obj.cp_bs_name:
            return obj.cp_bs_name
        return "—"

    def cf_name(obj):
        return str(obj.cfitem) if getattr(obj, "cfitem_id", None) else "—"

    def bank_name(obj):
        if obj.ba_id and obj.ba and getattr(obj.ba, "bank", None):
            return (getattr(obj.ba.bank, "name", None) or "").strip()
        return ""

    def rs(obj):
        if obj.ba_id and obj.ba:
            return (getattr(obj.ba, "account", None) or "").strip()
        return ""

    def owner_name(obj):
        if obj.bs_id and obj.bs and getattr(obj.bs, "owner", None):
            return str(obj.bs.owner)
        return ""

    def dt_val(obj):
        return obj.dt or Decimal("0.00")

    def cr_val(obj):
        return obj.cr or Decimal("0.00")

    def net_amount(obj):
        return dt_val(obj) - cr_val(obj)

    def flow(obj):
        dt = dt_val(obj)
        cr = cr_val(obj)
        if dt > 0:
            return "Дт"
        if cr > 0:
            return "Кт"
        return "—"

    def amount_signed(obj):
        # для pivot: поступления +, списания -
        dt = dt_val(obj)
        cr = cr_val(obj)
        if dt > 0:
            return dt
        if cr > 0:
            return -cr
        return Decimal("0.00")

    def bs_period(obj):
        if obj.bs_id and obj.bs and obj.bs.start and obj.bs.finish:
            return f"{obj.bs.start:%d.%m.%Y}-{obj.bs.finish:%d.%m.%Y}"
        return ""

    # --- агрегации + QC списки (по всем до даты) ---
    by_cp = {}
    by_cf = {}

    qc_no_cf = []
    qc_no_contract = []
    qc_no_cp_final = []
    qc_zero_amount = []

    for r in qs:
        dt = dt_val(r)
        cr = cr_val(r)

        name = cp_name(r)
        acc = by_cp.get(name, {"dt": Decimal("0.00"), "cr": Decimal("0.00"), "cnt": 0})
        acc["dt"] += dt
        acc["cr"] += cr
        acc["cnt"] += 1
        by_cp[name] = acc

        cf = cf_name(r)
        acc2 = by_cf.get(cf, {"dt": Decimal("0.00"), "cr": Decimal("0.00"), "cnt": 0})
        acc2["dt"] += dt
        acc2["cr"] += cr
        acc2["cnt"] += 1
        by_cf[cf] = acc2

        if not getattr(r, "cfitem_id", None):
            qc_no_cf.append(r)
        if not getattr(r, "contract_id", None):
            qc_no_contract.append(r)
        if not getattr(r, "cp_final_id", None):
            qc_no_cp_final.append(r)
        if dt == 0 and cr == 0:
            qc_zero_amount.append(r)

    # --- агрегации по выбранному дню (для быстрых summary при желании) ---
    by_cp_day = {}
    by_cf_day = {}
    for r in qs_day:
        dt = dt_val(r)
        cr = cr_val(r)

        name = cp_name(r)
        acc = by_cp_day.get(name, {"dt": Decimal("0.00"), "cr": Decimal("0.00"), "cnt": 0})
        acc["dt"] += dt
        acc["cr"] += cr
        acc["cnt"] += 1
        by_cp_day[name] = acc

        cf = cf_name(r)
        acc2 = by_cf_day.get(cf, {"dt": Decimal("0.00"), "cr": Decimal("0.00"), "cnt": 0})
        acc2["dt"] += dt
        acc2["cr"] += cr
        acc2["cnt"] += 1
        by_cf_day[cf] = acc2

    # ----------------- Excel: стиль (спокойная палитра) -----------------
    wb = Workbook()

    # Спокойная “офисная” палитра
    fill_title = PatternFill("solid", fgColor="2B6CB0")  # заголовок листа
    fill_header = PatternFill("solid", fgColor="E2E8F0")  # шапка таблицы
    fill_total = PatternFill("solid", fgColor="F8FAFC")  # итоги
    fill_soft = PatternFill("solid", fgColor="F1F5F9")  # мягкая подсветка
    fill_zebra = PatternFill("solid", fgColor="F9FAFB")  # зебра
    fill_warn = PatternFill("solid", fgColor="FEF3C7")  # warning
    fill_bad = PatternFill("solid", fgColor="FEE2E2")  # negative

    f_header = Font(bold=True, color="111827")
    f_title = Font(bold=True, size=14, color="FFFFFF")
    f_bold = Font(bold=True)
    f_dim = Font(color="6B7280")

    a_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    a_right = Alignment(horizontal="right", vertical="center")

    thin = Side(style="thin", color="CBD5E1")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cols_width(sheet, widths):
        for idx, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = w

    def autosize(sheet, min_w=10, max_w=80):
        for col in sheet.columns:
            length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                length = max(length, len(v))
            sheet.column_dimensions[col_letter].width = max(min_w, min(max_w, length + 2))

    def apply_sheet_prefs(sheet, freeze_cell="B3"):
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = freeze_cell

    def apply_title(sheet, title, ncols):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        cell = sheet.cell(row=1, column=1, value=title)
        cell.font = f_title
        cell.fill = fill_title
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 24

    def apply_table_header(sheet, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = sheet.cell(row=row_idx, column=c)
            cell.font = f_header
            cell.fill = fill_header
            cell.alignment = a_center
            cell.border = border_thin
        sheet.row_dimensions[row_idx].height = 24

    def apply_borders(sheet, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                sheet.cell(row=r, column=c).border = border_thin

    def apply_zebra(sheet, start_row, end_row, ncols):
        for r in range(start_row, end_row + 1):
            if (r - start_row) % 2 == 1:
                for c in range(1, ncols + 1):
                    sheet.cell(row=r, column=c).fill = fill_zebra

    def fmt_money_range(sheet, min_row, max_row, cols):
        for r in range(min_row, max_row + 1):
            for c in cols:
                cell = sheet.cell(row=r, column=c)
                cell.number_format = "#,##0.00"
                cell.alignment = a_right

    def fmt_int_range(sheet, min_row, max_row, cols):
        for r in range(min_row, max_row + 1):
            for c in cols:
                cell = sheet.cell(row=r, column=c)
                cell.number_format = "0"
                cell.alignment = a_right

    def add_totals_row(sheet, start_row, end_row, ncols, money_cols):
        if end_row < start_row:
            return end_row
        total_row = end_row + 1
        sheet.cell(row=total_row, column=1, value="Итого").font = f_bold
        for c in range(1, ncols + 1):
            sheet.cell(row=total_row, column=c).fill = fill_total
            sheet.cell(row=total_row, column=c).border = border_thin
        for c in money_cols:
            col_letter = get_column_letter(c)
            cell = sheet.cell(
                row=total_row,
                column=c,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            )
            cell.number_format = "#,##0.00"
            cell.alignment = a_right
            cell.font = f_bold
        return total_row

    # --- оглавление / ссылки ---
    def add_back_to_toc(sheet, ncols, toc_title="Оглавление"):
        # Ссылка в правом верхнем углу (A1 уже занято заголовком)
        # Делаем кликабельную ячейку в последней колонке второй строки.
        cell = sheet.cell(row=2, column=ncols, value="⇦ Оглавление")
        cell.font = Font(color="1D4ED8", underline="single", bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.hyperlink = f"#'{toc_title}'!A1"

    def add_internal_link(cell, sheet_title, target="A1"):
        cell.hyperlink = f"#'{sheet_title}'!{target}"
        cell.font = Font(color="1D4ED8", underline="single")
        cell.alignment = a_left

    # ----------------- Лист 0: Оглавление -----------------
    ws_toc = wb.active
    ws_toc.title = "Оглавление"
    ws_toc.sheet_view.showGridLines = False

    toc_headers = ["Раздел", "Описание"]
    apply_title(ws_toc, f"EOD отчет на {selected_date:%d.%m.%Y} — оглавление", ncols=len(toc_headers))
    ws_toc.append(toc_headers)
    apply_table_header(ws_toc, row_idx=2, ncols=len(toc_headers))
    ws_toc.row_dimensions[2].height = 22

    toc_start = 3

    def toc_add_row(sheet_title, description):
        r = ws_toc.max_row + 1
        c1 = ws_toc.cell(row=r, column=1, value=sheet_title)
        add_internal_link(c1, sheet_title, target="A1")
        c2 = ws_toc.cell(row=r, column=2, value=description)
        c2.alignment = a_left
        for c in (c1, c2):
            c.border = border_thin

    # ----------------- Лист 1: Контрагенты (до даты) -----------------
    ws = wb.create_sheet("Расчеты с контрагентами")
    headers1 = ["Контрагент", "Поступление (Дт)", "Списание (Кт)", "Сальдо", "Операций"]

    apply_title(ws, f"EOD отчет на {selected_date:%d.%m.%Y}", ncols=len(headers1))
    ws.append(headers1)
    apply_table_header(ws, row_idx=2, ncols=len(headers1))
    add_back_to_toc(ws, ncols=len(headers1))

    start_data_row = 3
    for name, a in sorted(by_cp.items(), key=lambda x: (-(x[1]["dt"] + x[1]["cr"]), x[0])):
        dt = a["dt"]
        cr = a["cr"]
        net = dt - cr
        ws.append([name, dt, cr, net, a["cnt"]])

    end_data_row = ws.max_row
    fmt_money_range(ws, start_data_row, end_data_row, cols=[2, 3, 4])
    fmt_int_range(ws, start_data_row, end_data_row, cols=[5])

    apply_zebra(ws, start_data_row, end_data_row, ncols=len(headers1))

    ws.conditional_formatting.add(
        f"D{start_data_row}:D{end_data_row}",
        FormulaRule(formula=[f"D{start_data_row}<0"], fill=fill_bad)
    )

    total_row = add_totals_row(ws, start_data_row, end_data_row, ncols=len(headers1), money_cols=[2, 3, 4])
    apply_borders(ws, 2, total_row, 1, len(headers1))

    set_cols_width(ws, [42, 18, 18, 16, 10])
    autosize(ws, min_w=10, max_w=60)
    apply_sheet_prefs(ws, freeze_cell="B3")

    # ----------------- Лист 2: Cash Flow (до даты) -----------------
    ws2 = wb.create_sheet("Cash Flow")
    headers2 = ["Статья CF", "Поступление (Дт)", "Списание (Кт)", "Сальдо", "Операций"]

    apply_title(ws2, f"Cash Flow на {selected_date:%d.%m.%Y}", ncols=len(headers2))
    ws2.append(headers2)
    apply_table_header(ws2, row_idx=2, ncols=len(headers2))
    add_back_to_toc(ws2, ncols=len(headers2))

    start2 = 3
    for name, a in sorted(by_cf.items(), key=lambda x: (-(x[1]["dt"] + x[1]["cr"]), x[0])):
        dt = a["dt"]
        cr = a["cr"]
        net = dt - cr
        ws2.append([name, dt, cr, net, a["cnt"]])

    end2 = ws2.max_row
    fmt_money_range(ws2, start2, end2, cols=[2, 3, 4])
    fmt_int_range(ws2, start2, end2, cols=[5])

    apply_zebra(ws2, start2, end2, ncols=len(headers2))

    ws2.conditional_formatting.add(
        f"D{start2}:D{end2}",
        FormulaRule(formula=[f"D{start2}<0"], fill=fill_bad)
    )

    total_row2 = add_totals_row(ws2, start2, end2, ncols=len(headers2), money_cols=[2, 3, 4])
    apply_borders(ws2, 2, total_row2, 1, len(headers2))

    set_cols_width(ws2, [38, 18, 18, 16, 10])
    autosize(ws2, min_w=10, max_w=60)
    apply_sheet_prefs(ws2, freeze_cell="B3")

    # ----------------- Новые листы: выбранный день (детализация с группировкой) -----------------
    def build_day_grouped_sheet(
        sheet_title,
        title_text,
        rows_iterable,
        group_label,
        group_key_func,
        columns_mode="cp",  # "cp" or "cf"
    ):
        """
        Делает лист с операциями за выбранный день, сгруппированными по контрагенту или по статье CF.
        Внутри группы: строки операций + подытог.
        В конце: общий итог.
        """
        wsx = wb.create_sheet(sheet_title)

        headers = [
            "Дата", "Док №", "Документ",
            "Контрагент", "Договор", "Статья CF",
            "Дт", "Кт", "Нетто",
            "Компания", "Банк", "р/сч",
            "Назначение", "Выписка",
        ]
        apply_title(wsx, title_text, ncols=len(headers))
        wsx.append(headers)
        apply_table_header(wsx, row_idx=2, ncols=len(headers))
        add_back_to_toc(wsx, ncols=len(headers))

        start_row = 3

        rows = list(rows_iterable)
        if columns_mode == "cp":
            rows.sort(key=lambda r: (cp_name(r), r.id))
        else:
            rows.sort(key=lambda r: (cf_name(r), r.id))

        def write_group_header(row_idx, group_name):
            wsx.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = wsx.cell(row=row_idx, column=1, value=f"{group_label}: {group_name}")
            cell.font = Font(bold=True, color="111827")
            cell.fill = fill_soft
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_thin
            wsx.row_dimensions[row_idx].height = 20
            # рамки по всей строке
            for c in range(1, len(headers) + 1):
                wsx.cell(row=row_idx, column=c).border = border_thin
                wsx.cell(row=row_idx, column=c).fill = fill_soft

        def write_group_total(row_idx, dt_sum, cr_sum, net_sum, cnt):
            # подытог группы
            wsx.cell(row=row_idx, column=1, value="Итого по группе").font = f_bold
            wsx.cell(row=row_idx, column=6, value="").font = f_dim

            wsx.cell(row=row_idx, column=7, value=dt_sum).number_format = "#,##0.00"
            wsx.cell(row=row_idx, column=8, value=cr_sum).number_format = "#,##0.00"
            wsx.cell(row=row_idx, column=9, value=net_sum).number_format = "#,##0.00"

            wsx.cell(row=row_idx, column=4, value=f"Операций: {cnt}").font = f_dim

            for c in range(1, len(headers) + 1):
                cell = wsx.cell(row=row_idx, column=c)
                cell.fill = fill_total
                cell.border = border_thin
                if c in (7, 8, 9):
                    cell.alignment = a_right
                else:
                    cell.alignment = a_left

        def write_tx(row_idx, r):
            contract_txt = str(r.contract) if getattr(r, "contract_id", None) else "—"
            cf_txt = str(r.cfitem) if getattr(r, "cfitem_id", None) else "—"
            dt = dt_val(r)
            cr = cr_val(r)
            net = dt - cr

            wsx.append([
                r.date.strftime("%d.%m.%Y") if r.date else "",
                r.doc_numner or "",
                r.doc_type or "",
                cp_name(r),
                contract_txt,
                cf_txt,
                dt,
                cr,
                net,
                owner_name(r),
                bank_name(r),
                rs(r),
                (r.temp or "").strip(),
                bs_period(r),
            ])

            # формат + границы сразу
            for c in range(1, len(headers) + 1):
                cell = wsx.cell(row=row_idx, column=c)
                cell.border = border_thin
                if c in (7, 8, 9):
                    cell.number_format = "#,##0.00"
                    cell.alignment = a_right
                elif c in (13,):
                    cell.alignment = a_left
                else:
                    cell.alignment = a_left

            return dt, cr, net

        cur = start_row
        if not rows:
            wsx.cell(row=cur, column=1, value="Нет операций за выбранный день").font = f_dim
            wsx.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(headers))
            wsx.cell(row=cur, column=1).alignment = a_left
            apply_borders(wsx, 2, cur, 1, len(headers))
            set_cols_width(wsx, [12, 14, 16, 30, 22, 22, 14, 14, 14, 22, 20, 24, 70, 18])
            autosize(wsx, min_w=10, max_w=90)
            apply_sheet_prefs(wsx, freeze_cell="B3")
            return wsx

        grp_current = None
        grp_dt = Decimal("0.00")
        grp_cr = Decimal("0.00")
        grp_net = Decimal("0.00")
        grp_cnt = 0

        grand_dt = Decimal("0.00")
        grand_cr = Decimal("0.00")
        grand_net = Decimal("0.00")
        grand_cnt = 0

        for r in rows:
            g = group_key_func(r)
            if grp_current is None:
                grp_current = g
                write_group_header(cur, grp_current)
                cur += 1

            if g != grp_current:
                # закрываем предыдущую группу
                write_group_total(cur, grp_dt, grp_cr, grp_net, grp_cnt)
                cur += 1

                # новая группа
                grp_current = g
                grp_dt = Decimal("0.00")
                grp_cr = Decimal("0.00")
                grp_net = Decimal("0.00")
                grp_cnt = 0

                write_group_header(cur, grp_current)
                cur += 1

            dt, cr, net = write_tx(cur, r)
            grp_dt += dt
            grp_cr += cr
            grp_net += net
            grp_cnt += 1

            grand_dt += dt
            grand_cr += cr
            grand_net += net
            grand_cnt += 1

            cur += 1

        # закрываем последнюю группу
        write_group_total(cur, grp_dt, grp_cr, grp_net, grp_cnt)
        cur += 1

        # общий итог
        wsx.cell(row=cur, column=1, value="ОБЩИЙ ИТОГ").font = Font(bold=True, color="111827")
        wsx.cell(row=cur, column=7, value=grand_dt).number_format = "#,##0.00"
        wsx.cell(row=cur, column=8, value=grand_cr).number_format = "#,##0.00"
        wsx.cell(row=cur, column=9, value=grand_net).number_format = "#,##0.00"
        wsx.cell(row=cur, column=4, value=f"Операций: {grand_cnt}").font = f_dim

        for c in range(1, len(headers) + 1):
            cell = wsx.cell(row=cur, column=c)
            cell.fill = fill_header  # чуть заметнее
            cell.border = border_thin
            cell.alignment = a_right if c in (7, 8, 9) else a_left
        cur += 1

        # зебра только по строкам операций (между group_header и group_total не красим)
        # проще: применим зебру ко всему диапазону данных, но исключим строки с заливками fill_soft/fill_total/fill_header
        # (Оставим как есть: уже хорошо читается.)

        # подсветка отрицательного нетто
        data_first = start_row
        data_last = wsx.max_row
        wsx.conditional_formatting.add(
            f"I{data_first}:I{data_last}",
            FormulaRule(formula=[f"I{data_first}<0"], fill=fill_bad)
        )

        wsx.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{wsx.max_row}"

        set_cols_width(wsx, [12, 14, 16, 30, 22, 22, 14, 14, 14, 22, 20, 24, 70, 18])
        autosize(wsx, min_w=10, max_w=90)
        apply_sheet_prefs(wsx, freeze_cell="B3")
        return wsx

    # Лист: выбранный день, группировка по контрагентам
    ws_day_cp = build_day_grouped_sheet(
        sheet_title="День — контрагенты",
        title_text=f"Операции за {selected_date:%d.%m.%Y} (группировка по контрагентам)",
        rows_iterable=qs_day,
        group_label="Контрагент",
        group_key_func=lambda r: cp_name(r),
        columns_mode="cp",
    )

    # Лист: выбранный день, группировка по Cash Flow
    ws_day_cf = build_day_grouped_sheet(
        sheet_title="День — Cash Flow",
        title_text=f"Операции за {selected_date:%d.%m.%Y} (группировка по Cash Flow)",
        rows_iterable=qs_day,
        group_label="Статья CF",
        group_key_func=lambda r: cf_name(r),
        columns_mode="cf",
    )

    # ----------------- Лист: Без статьи CF -----------------
    ws3 = wb.create_sheet("Без статьи CF")
    headers3 = ["Дата", "Док №", "Контрагент", "Поток", "Сумма", "Банк", "р/сч", "Назначение"]

    apply_title(ws3, f"Операции без статьи CF (до {selected_date:%d.%m.%Y})", ncols=len(headers3))
    ws3.append(headers3)
    apply_table_header(ws3, row_idx=2, ncols=len(headers3))
    add_back_to_toc(ws3, ncols=len(headers3))

    start3 = 3
    for r in qc_no_cf:
        dt = dt_val(r)
        cr = cr_val(r)
        fl = "Дт" if dt > 0 else ("Кт" if cr > 0 else "—")
        amount = dt if dt > 0 else cr
        ws3.append([
            r.date.strftime("%d.%m.%Y") if r.date else "",
            r.doc_numner or "",
            cp_name(r),
            fl,
            amount,
            bank_name(r),
            rs(r),
            (r.temp or "").strip(),
        ])

    end3 = ws3.max_row
    fmt_money_range(ws3, start3, end3, cols=[5])

    for row in ws3.iter_rows(min_row=3, max_row=end3, min_col=1, max_col=len(headers3)):
        for cell in row:
            cell.fill = fill_warn
            cell.border = border_thin
            if cell.column in (1, 2, 3, 6, 7, 8):
                cell.alignment = a_left

    ws3.auto_filter.ref = f"A2:{get_column_letter(len(headers3))}{end3}"

    set_cols_width(ws3, [12, 14, 30, 8, 14, 22, 24, 60])
    autosize(ws3, min_w=10, max_w=80)
    apply_sheet_prefs(ws3, freeze_cell="B3")

    # ----------------- Лист: Все операции -----------------
    ws4 = wb.create_sheet("Все операции")
    headers4 = [
        "Дата", "Док №", "Документ", "Контрагент",
        "Договор", "Статья CF",
        "Дт", "Кт", "Нетто",
        "Компания", "Банк", "р/сч",
        "Назначение", "Выписка"
    ]

    apply_title(ws4, f"Все операции (до {selected_date:%d.%m.%Y})", ncols=len(headers4))
    ws4.append(headers4)
    apply_table_header(ws4, row_idx=2, ncols=len(headers4))
    add_back_to_toc(ws4, ncols=len(headers4))

    start4 = 3
    for r in qs:
        contract_txt = str(r.contract) if getattr(r, "contract_id", None) else "—"
        cf_txt = str(r.cfitem) if getattr(r, "cfitem_id", None) else "—"
        dt = dt_val(r)
        cr = cr_val(r)
        net = dt - cr

        ws4.append([
            r.date.strftime("%d.%m.%Y") if r.date else "",
            r.doc_numner or "",
            r.doc_type or "",
            cp_name(r),
            contract_txt,
            cf_txt,
            dt,
            cr,
            net,
            owner_name(r),
            bank_name(r),
            rs(r),
            (r.temp or "").strip(),
            bs_period(r),
        ])

    end4 = ws4.max_row
    fmt_money_range(ws4, start4, end4, cols=[7, 8, 9])

    apply_zebra(ws4, start4, end4, ncols=len(headers4))

    ws4.conditional_formatting.add(
        f"F{start4}:F{end4}",
        FormulaRule(formula=[f'F{start4}="—"'], fill=fill_warn)
    )
    ws4.conditional_formatting.add(
        f"E{start4}:E{end4}",
        FormulaRule(formula=[f'E{start4}="—"'], fill=fill_soft)
    )
    ws4.conditional_formatting.add(
        f"I{start4}:I{end4}",
        FormulaRule(formula=[f"I{start4}<0"], fill=fill_bad)
    )

    total_row4 = add_totals_row(ws4, start4, end4, ncols=len(headers4), money_cols=[7, 8, 9])
    apply_borders(ws4, 2, total_row4, 1, len(headers4))

    ws4.auto_filter.ref = f"A2:{get_column_letter(len(headers4))}{end4}"

    set_cols_width(ws4, [12, 14, 16, 30, 22, 22, 14, 14, 14, 22, 20, 24, 70, 18])
    autosize(ws4, min_w=10, max_w=90)
    apply_sheet_prefs(ws4, freeze_cell="B3")

    # ----------------- Лист: Pivot-ready -----------------
    wsp = wb.create_sheet("Pivot Data")
    headersp = [
        "Дата", "Месяц", "Год",
        "Компания",
        "Контрагент",
        "Статья CF",
        "Договор",
        "Поток",
        "Сумма (signed)",
        "Док №",
        "Документ",
        "Банк",
        "р/сч",
        "Выписка",
        "Назначение",
    ]

    apply_title(wsp, f"Pivot-ready данные (до {selected_date:%d.%m.%Y})", ncols=len(headersp))
    wsp.append(headersp)
    apply_table_header(wsp, row_idx=2, ncols=len(headersp))
    add_back_to_toc(wsp, ncols=len(headersp))

    startp = 3
    for r in qs:
        d = r.date
        month = d.strftime("%Y-%m") if d else ""
        year = d.year if d else ""

        wsp.append([
            d.strftime("%d.%m.%Y") if d else "",
            month,
            year,
            owner_name(r),
            cp_name(r),
            (str(r.cfitem) if getattr(r, "cfitem_id", None) else ""),
            (str(r.contract) if getattr(r, "contract_id", None) else ""),
            flow(r),
            amount_signed(r),
            r.doc_numner or "",
            r.doc_type or "",
            bank_name(r),
            rs(r),
            bs_period(r),
            (r.temp or "").strip(),
        ])

    endp = wsp.max_row
    fmt_money_range(wsp, startp, endp, cols=[9])
    apply_zebra(wsp, startp, endp, ncols=len(headersp))
    wsp.auto_filter.ref = f"A2:{get_column_letter(len(headersp))}{endp}"

    wsp.conditional_formatting.add(
        f"F{startp}:F{endp}",
        FormulaRule(formula=[f'F{startp}=""'], fill=fill_warn)
    )
    wsp.conditional_formatting.add(
        f"G{startp}:G{endp}",
        FormulaRule(formula=[f'G{startp}=""'], fill=fill_soft)
    )

    apply_borders(wsp, 2, endp, 1, len(headersp))

    set_cols_width(wsp, [12, 10, 8, 22, 30, 22, 22, 8, 16, 14, 16, 18, 24, 18, 80])
    autosize(wsp, min_w=8, max_w=90)
    apply_sheet_prefs(wsp, freeze_cell="B3")

    # ----------------- Лист: Контроль качества -----------------
    wsq = wb.create_sheet("Контроль качества")
    headersq = ["Проверка", "Количество", "Комментарий"]

    apply_title(wsq, f"Контроль качества (до {selected_date:%d.%m.%Y})", ncols=len(headersq))
    wsq.append(headersq)
    apply_table_header(wsq, row_idx=2, ncols=len(headersq))
    add_back_to_toc(wsq, ncols=len(headersq))

    total_ops = qs.count()

    qc_rows = [
        ("Всего операций", total_ops, "Все строки CfData до выбранной даты"),
        ("Без статьи CF", len(qc_no_cf), "Нужно разнести по статьям Cash Flow"),
        ("Без договора", len(qc_no_contract), "Проверь привязку договора (если требуется)"),
        ("Без финального контрагента", len(qc_no_cp_final), "Проверь сопоставление cp_final"),
        ("Нулевые суммы", len(qc_zero_amount), "Строки с dt=0 и cr=0 (возможно ошибка парсинга)"),
    ]

    startq = 3
    for name, cnt, comment in qc_rows:
        wsq.append([name, cnt, comment])

    endq = wsq.max_row
    fmt_int_range(wsq, startq, endq, cols=[2])

    for r in range(startq + 1, endq + 1):
        val = wsq.cell(row=r, column=2).value or 0
        if val > 0:
            for c in range(1, len(headersq) + 1):
                wsq.cell(row=r, column=c).fill = fill_warn

    apply_zebra(wsq, startq, endq, ncols=len(headersq))
    apply_borders(wsq, 2, endq, 1, len(headersq))

    cur = endq + 2
    wsq.cell(row=cur, column=1, value="Примеры (ТОП 50)").font = f_bold
    cur += 1

    example_headers = ["Дата", "Док №", "Контрагент", "Договор", "Статья CF", "Сумма (dt-cr)", "Назначение", "Банк", "р/сч"]
    wsq.append(example_headers)
    apply_table_header(wsq, row_idx=cur, ncols=len(example_headers))

    def append_examples(rows, limit=50):
        for r in rows[:limit]:
            wsq.append([
                r.date.strftime("%d.%m.%Y") if r.date else "",
                r.doc_numner or "",
                cp_name(r),
                (str(r.contract) if getattr(r, "contract_id", None) else "—"),
                (str(r.cfitem) if getattr(r, "cfitem_id", None) else "—"),
                net_amount(r),
                (r.temp or "").strip(),
                bank_name(r),
                rs(r),
            ])

    append_examples(qc_no_cf, limit=50)
    append_examples(qc_no_contract, limit=50)

    # формат денег в колонке "Сумма (dt-cr)" -> F
    for cell in wsq["F"][2:]:
        cell.number_format = "#,##0.00"
        cell.alignment = a_right

    last_row_q = wsq.max_row
    wsq.auto_filter.ref = f"A{cur}:{get_column_letter(len(example_headers))}{last_row_q}"

    apply_zebra(wsq, cur + 1, last_row_q, ncols=len(example_headers))
    apply_borders(wsq, cur, last_row_q, 1, len(example_headers))

    set_cols_width(wsq, [28, 12, 30, 22, 22, 16, 70, 18, 24])
    autosize(wsq, min_w=10, max_w=90)
    apply_sheet_prefs(wsq, freeze_cell="B3")

    # ----------------- Заполняем оглавление (кликабельные ссылки) -----------------
    toc_add_row("Расчеты с контрагентами", "Свод по контрагентам (до выбранной даты)")
    toc_add_row("Cash Flow", "Свод по статьям Cash Flow (до выбранной даты)")
    toc_add_row("День — контрагенты", "Операции строго за выбранный день, группировка по контрагентам + подытоги")
    toc_add_row("День — Cash Flow", "Операции строго за выбранный день, группировка по статьям CF + подытоги")
    toc_add_row("Без статьи CF", "QC: операции без статьи Cash Flow")
    toc_add_row("Все операции", "Все операции (до выбранной даты), построчно")
    toc_add_row("Pivot Data", "Pivot-ready набор данных")
    toc_add_row("Контроль качества", "QC-метрики и примеры строк")

    # оформление оглавления
    end_toc = ws_toc.max_row
    apply_zebra(ws_toc, toc_start, end_toc, ncols=2)
    apply_borders(ws_toc, 2, end_toc, 1, 2)
    set_cols_width(ws_toc, [34, 70])
    autosize(ws_toc, min_w=12, max_w=90)
    ws_toc.freeze_panes = "A3"

    # ----------------- ответ -----------------
    filename = f"EOD_{selected_date:%Y-%m-%d}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


