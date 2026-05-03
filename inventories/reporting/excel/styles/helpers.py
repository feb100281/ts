# inventories/reporting/excel/styles/helpers.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .theme import COLORS, FONTS, BORDERS, ALIGNMENTS, FORMATS, FILLS


def set_column_widths(ws, widths: dict):
    """Установка ширины колонок"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_heights(ws, heights: dict):
    """Установка высоты строк"""
    for row_idx, height in heights.items():
        ws.row_dimensions[row_idx].height = height


def draw_sheet_header(ws, title, subtitle="", note=""):
    """Рисование заголовка листа"""
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A3"] = note
    
    ws["A1"].font = FONTS["title"]
    ws["A2"].font = FONTS["subtitle"]
    ws["A3"].font = FONTS["subtitle"]
    
    ws["A1"].alignment = ALIGNMENTS["left"]
    ws["A2"].alignment = ALIGNMENTS["left"]
    ws["A3"].alignment = ALIGNMENTS["left"]


def draw_table_header(ws, row, headers, start_col=1, fill_key="header", font_key="header_white"):
    """Рисование заголовка таблицы"""
    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i)
        cell.value = header
        cell.fill = FILLS[fill_key]
        cell.font = FONTS[font_key]
        cell.alignment = ALIGNMENTS["center_wrap"]
        cell.border = BORDERS["thin"]


def style_data_row(ws, row, values, start_col=1, zebra=False, is_numeric_cols=None):
    """Стилизация строки данных"""
    fill = FILLS.get("alt", PatternFill(fill_type=None)) if zebra else PatternFill(fill_type=None)
    
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.fill = fill
        cell.border = BORDERS["thin"]
        cell.font = FONTS["normal"]
        
        # Проверяем, числовое ли поле
        is_numeric = is_numeric_cols and i in is_numeric_cols
        if is_numeric or isinstance(value, (int, float)):
            cell.number_format = FORMATS.get("number", '#,##0')
            cell.alignment = ALIGNMENTS["right"]
        else:
            cell.alignment = ALIGNMENTS["left"]


def style_total_row(ws, row, values, start_col=1, is_numeric_cols=None):
    """Стилизация строки итогов"""
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.fill = FILLS["total"]
        cell.border = BORDERS["bottom_medium"]
        cell.font = FONTS["total"]
        
        is_numeric = is_numeric_cols and i in is_numeric_cols
        if is_numeric or isinstance(value, (int, float)):
            cell.number_format = FORMATS.get("number", '#,##0')
            cell.alignment = ALIGNMENTS["right"]
        else:
            cell.alignment = ALIGNMENTS["left"]


def apply_money(cell):
    """Применить формат денег"""
    cell.number_format = '#,##0.00'
    cell.alignment = ALIGNMENTS["right"]


def get_delta_fill(value, default_fill):
    """Получить цвет для отклонений"""
    try:
        value = float(value or 0)
    except (TypeError, ValueError):
        return default_fill
    
    if value < 0:
        return FILLS.get("delta_red", default_fill)
    elif value > 0:
        return FILLS.get("delta_green", default_fill)
    return default_fill


def add_auto_filter(ws, start_row, start_col, end_row, end_col):
    """Добавить автофильтр"""
    from openpyxl.utils import get_column_letter
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    ws.auto_filter.ref = f"{start_letter}{start_row}:{end_letter}{end_row}"


def draw_toc_button(ws, cell="A1", text="← Оглавление", target_sheet="TOC"):
    """Рисует кнопку возврата к оглавлению"""
    cell_obj = ws[cell]
    cell_obj.value = text
    cell_obj.hyperlink = f"#'{target_sheet}'!A1"
    cell_obj.font = Font(name="Roboto", size=9, bold=True, color=COLORS["back_text_green"], underline="single")
    cell_obj.alignment = ALIGNMENTS["left"]
    cell_obj.fill = FILLS.get("back", PatternFill(fill_type=None))
    
    # Добавляем границы
    from openpyxl.styles import Border, Side
    thin_side = Side(style="thin", color=COLORS["border_gray"])
    cell_obj.border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )