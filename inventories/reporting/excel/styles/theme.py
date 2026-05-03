# inventories/reporting/excel/styles/theme.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Цвета для отчета по остаткам
COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED",
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "back_text_green": "1F5E4E",
    "negative_brown": "7A4E4E",
    "delta_red": "FBEAEA",
    "delta_green": "EAF6EE",
    "warehouse_blue": "4A90E2",
    "stock_highlight": "FFF3E0",
    "expired_bg": "FFEBEE",      # Светло-красный фон
    "expired_text": "D32F2F",    # Темно-красный текст
    "expiring_bg": "FFF3E0",     # Светло-оранжевый фон
    "expiring_text": "FF9800",   # Оранжевый текст
     
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "warehouse": PatternFill("solid", fgColor=COLORS["warehouse_blue"]),
    "stock_highlight": PatternFill("solid", fgColor=COLORS["stock_highlight"]),
    "none": PatternFill(fill_type=None),
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=12, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=11, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "warehouse": Font(name="Roboto", size=11, bold=True, color=COLORS["white"]),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium = Side(style="medium", color=COLORS["dark_green"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium),
    "none": Border(),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
}

FORMATS = {
    "number": '#,##0',
    "decimal": '#,##0.00',
    "date": 'dd.mm.yyyy',
    "datetime": 'dd.mm.yyyy hh:mm',
}