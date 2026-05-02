from openpyxl.styles import Font, Alignment
from ..styles.theme import COLORS


class SheetTitle:
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw(self, row, title, subtitle="", date_text="", start_col=2, end_col=16):
        """Рисует заголовок листа"""
        # Заголовок
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        title_cell = self.ws.cell(row=row, column=start_col, value=title)
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 35
        row += 1
        
        # Подзаголовок
        if subtitle:
            self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            sub_cell = self.ws.cell(row=row, column=start_col, value=subtitle)
            sub_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
            sub_cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.row_dimensions[row].height = 22
            row += 1
        
        # Дата
        if date_text:
            self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            date_cell = self.ws.cell(row=row, column=start_col, value=date_text)
            date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
            date_cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.row_dimensions[row].height = 20
            row += 1
        
        row += 1
        return row


def create_sheet_title(worksheet):
    """Создает экземпляр SheetTitle"""
    return SheetTitle(worksheet)