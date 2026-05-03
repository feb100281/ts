from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..styles.theme import COLORS, ALIGNMENTS, BORDERS, FILLS


class TableComponent:
    """Класс для отрисовки таблиц"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw(self, start_row, headers, data_rows, start_col=2, number_format='#,##0', 
             highlight_cols=None, column_widths=None):
        """Рисует таблицу"""
        row = start_row
        
        # Заголовки
        for col_idx, header in enumerate(headers):
            cell = self.ws.cell(row=row, column=start_col + col_idx, value=header)
            cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
            cell.fill = FILLS["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 35
        row += 1
        
        # Данные
        for idx, data_row in enumerate(data_rows):
            fill = FILLS["alt"] if idx % 2 == 1 else FILLS["none"]
            
            for col_idx, value in enumerate(data_row):
                cell = self.ws.cell(row=row, column=start_col + col_idx, value=value)
                cell.fill = fill
                cell.border = BORDERS["thin"]
                
                if col_idx == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = number_format
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
                
                if highlight_cols and col_idx in highlight_cols:
                    if isinstance(value, (int, float)) and value > 0:
                        cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS.get("warning_red", "C62828"))
            
            self.ws.row_dimensions[row].height = 22
            row += 1
        
        # Настройка ширины колонок
        if column_widths:
            for col_letter, width in column_widths.items():
                self.ws.column_dimensions[col_letter].width = width
        
        return row


def create_table(worksheet):
    """Создает экземпляр TableComponent"""
    return TableComponent(worksheet)