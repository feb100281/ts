# inventories/reporting/excel/sheets/brand_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS
from ..components import create_kpi_cards, create_table, Footnote, create_sheet_title


class BrandSheet(BaseSheet):
    """Лист с анализом остатков по брендам"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    def build(self, df, stats, report_date):
        row = 1

        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.hyperlink = "#'TOC'!A1"

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2

        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        title_cell = self.ws.cell(row=row, column=2, value="АНАЛИЗ ОСТАТКОВ ПО БРЕНДАМ")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        subtitle_cell = self.ws.cell(row=row, column=2, value=f"Дата остатков: {report_date_formatted}")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 2

        row1_cards = [
            {
                'title': 'ВСЕГО БРЕНДОВ',
                'value': self._format_number(len(df)),
                'subtitle': 'брендов с остатками',
                'color': COLORS["dark_green"],
                'width': 2
            },
            {
                'title': 'ВСЕГО ТОВАРОВ',
                'value': self._format_number(stats['total_products']),
                'subtitle': 'уникальных карточек',
                'color': COLORS["dark_green"],
                'width': 2
            },
            {
                'title': 'ОБЩИЙ ОСТАТОК',
                'value': self._format_number(df['итого'].sum()),
                'subtitle': 'на складах и в пути',
                'color': COLORS["dark_green"],
                'width': 2
            },
        ]

        row = self.kpi.draw_row(row, row1_cards)
        row += 2

        headers = [
            'Бренд',
            'Товаров',
            'На складе, шт',
            'В пути, шт',
            'Итого, шт',
            'Доля остатков, %'
        ]

        data_rows = []
        for _, row_data in df.iterrows():
            data_rows.append([
                row_data['бренд'],
                row_data['товаров'],
                row_data['на_складе'],
                row_data['в_пути'],
                row_data['итого'],
                row_data['доля_остатков_проц'],
            ])

        column_widths = {
            'A': 5,
            'B': 24,
            'C': 14,
            'D': 18,
            'E': 16,
            'F': 16,
            'G': 18,
        }

        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format='#,##0',
            highlight_cols=None,
            column_widths=column_widths
        )

        for r in range(row - len(data_rows), row):
            self.ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
            for col in [3, 4, 5, 6]:
                self.ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center")
                self.ws.cell(row=r, column=col).number_format = '#,##0'

            self.ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
            self.ws.cell(row=r, column=7).number_format = '0.00'
        
        
            # ПОДСВЕТКА КОЛОНКИ "Итого, шт" (колонка F, индекс 6)
            total_cell = self.ws.cell(row=r, column=6)
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.number_format = '#,##0'
            total_cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS["dark_green"],
            )
            total_cell.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid",
            )
            
            # ПОДСВЕТКА КОЛОНКИ "Доля остатков, %" (колонка G, индекс 7)
            share_cell = self.ws.cell(row=r, column=7)
            share_cell.alignment = Alignment(horizontal="center", vertical="center")
            share_cell.number_format = '0.00'
            share_cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS["dark_green"],
            )
            share_cell.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid",
            )

        self.ws.auto_filter.ref = f'B{row - len(data_rows) - 1}:G{row - 2}'

        header_row = row - len(data_rows) - 1
        self.ws.freeze_panes = f'C{header_row + 1}'
        self.ws.sheet_view.showGridLines = False