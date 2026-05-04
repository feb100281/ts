# # inventories/reporting/excel/sheets/certificate_sheet.py

# from datetime import datetime

# import pandas as pd
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# from .base_sheet import BaseSheet
# from ..styles.theme import COLORS
# from ..components import create_kpi_cards, create_table, Footnote, create_sheet_title


# class CertificateSheet(BaseSheet):
#     """Лист с сертификатами: просроченные, истекающие и отсутствующие"""

#     def __init__(self, workbook, sheet_number):
#         super().__init__(workbook, sheet_number)
       
#         self.sheet_title = create_sheet_title(self.ws)
#         self.kpi = create_kpi_cards(self.ws)
#         self.table = create_table(self.ws)
#         self.footnote = Footnote(self.ws)

#     @staticmethod
#     def _safe_text(value, default=""):
#         if pd.isna(value):
#             return default
#         value = str(value).strip()
#         return value if value else default

#     @staticmethod
#     def _safe_number(value, default=0):
#         if pd.isna(value):
#             return default
#         return value

#     def _format_number(self, value):
#         if value is None or pd.isna(value):
#             return "0"
#         try:
#             return f"{int(value):,}".replace(",", " ")
#         except Exception:
#             return str(value)

#     def _apply_status_style(self, cell):
#         value = str(cell.value).upper() if cell.value else ""

#         if "НЕТ СЕРТИФИКАТА" in value:
#             cell.fill = PatternFill(
#                 start_color=COLORS["delta_red"],
#                 end_color=COLORS["delta_red"],
#                 fill_type="solid",
#             )
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=COLORS["negative_brown"],
#             )

#         elif "ПРОСРОЧЕН" in value:
#             cell.fill = PatternFill(
#                 start_color=COLORS["expired_bg"],
#                 end_color=COLORS["expired_bg"],
#                 fill_type="solid",
#             )
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=COLORS["expired_text"],
#             )

#         else:
#             cell.fill = PatternFill(
#                 start_color=COLORS["expiring_bg"],
#                 end_color=COLORS["expiring_bg"],
#                 fill_type="solid",
#             )
#             cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=COLORS["expiring_text"],
#             )

#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center",
#             wrap_text=True,
#         )

#     def build(self, df, report_date):
#         row = 1

#         btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
#         btn_cell.font = Font(
#             name="Roboto",
#             size=9,
#             bold=True,
#             color=COLORS["dark_green"],
#         )
#         btn_cell.alignment = Alignment(horizontal="left", vertical="center")
#         btn_cell.fill = PatternFill(
#             start_color=COLORS["light_green"],
#             end_color=COLORS["light_green"],
#             fill_type="solid",
#         )
#         btn_cell.border = Border(
#             left=Side(style="thin", color=COLORS["border_gray"]),
#             right=Side(style="thin", color=COLORS["border_gray"]),
#             top=Side(style="thin", color=COLORS["border_gray"]),
#             bottom=Side(style="thin", color=COLORS["border_gray"]),
#         )
#         btn_cell.hyperlink = "#'TOC'!A1"

#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
#         self.ws.row_dimensions[row].height = 24
#         row += 2

#         report_date_formatted = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d.%m.%Y")

#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
#         title_cell = self.ws.cell(row=row, column=2, value="СЕРТИФИКАТЫ СООТВЕТСТВИЯ")
#         title_cell.font = Font(
#             name="Roboto",
#             size=16,
#             bold=True,
#             color=COLORS["dark_green"],
#         )
#         title_cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.row_dimensions[row].height = 32
#         row += 1

#         expired_count = (
#             len(df[df["статус_сертификата"] == "Просрочен"])
#             if not df.empty and "статус_сертификата" in df.columns
#             else 0
#         )

#         expiring_30_count = (
#             len(df[df["статус_сертификата"] == "Истекает в ближайшие 30 дней"])
#             if not df.empty and "статус_сертификата" in df.columns
#             else 0
#         )

#         no_cert_count = (
#             len(df[df["статус_сертификата"] == "Нет сертификата"])
#             if not df.empty and "статус_сертификата" in df.columns
#             else 0
#         )

#         expired_units = (
#             pd.to_numeric(
#                 df.loc[df["статус_сертификата"] == "Просрочен", "количество"],
#                 errors="coerce",
#             ).fillna(0).sum()
#             if not df.empty and {"статус_сертификата", "количество"}.issubset(df.columns)
#             else 0
#         )

#         no_cert_units = (
#             pd.to_numeric(
#                 df.loc[df["статус_сертификата"] == "Нет сертификата", "количество"],
#                 errors="coerce",
#             ).fillna(0).sum()
#             if not df.empty and {"статус_сертификата", "количество"}.issubset(df.columns)
#             else 0
#         )

#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
#         subtitle_cell = self.ws.cell(
#             row=row,
#             column=2,
#             value=(
#                 f"Дата остатков: {report_date_formatted} | "
#                 f"Просрочено: {expired_count} | "
#                 f"Истекает в ближайшие 30 дней: {expiring_30_count} | "
#                 f"Без сертификата: {no_cert_count}"
#             ),
#         )
#         subtitle_cell.font = Font(
#             name="Roboto",
#             size=11,
#             color=COLORS["text_gray"],
#         )
#         subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.row_dimensions[row].height = 24
#         row += 1

#         self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
#         date_cell = self.ws.cell(
#             row=row,
#             column=2,
#             value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
#         )
#         date_cell.font = Font(
#             name="Roboto",
#             size=9,
#             italic=True,
#             color=COLORS["text_gray"],
#         )
#         date_cell.alignment = Alignment(horizontal="left", vertical="center")
#         self.ws.row_dimensions[row].height = 20
#         row += 2

#         total_products = len(df["nm_id"].unique()) if not df.empty and "nm_id" in df.columns else 0

#         row1_cards = [
#             {
#                 "title": "ТОВАРОВ С ПРОБЛЕМАМИ",
#                 "value": self._format_number(total_products),
#                 "subtitle": "артикулы требуют внимания",
#                 "color": COLORS["dark_green"],
#                 "width": 2,
#             },
#             {
#                 "title": "ПРОСРОЧЕННЫЕ",
#                 "value": self._format_number(expired_count),
#                 "subtitle": "артикулы",
#                 "color": COLORS["dark_green"],
#                 "width": 2,
#             },
  
#             {
#                 "title": "БЕЗ СЕРТИФИКАТА",
#                 "value": self._format_number(no_cert_count),
#                 "subtitle": "артикулы",
#                 "color": COLORS["dark_green"],
#                 "width": 2,
#             },
#             {
#                 "title": "ШТ. С ПРОСРОЧЕННЫМ",
#                 "value": self._format_number(expired_units),
#                 "subtitle": "единиц на остатке",
#                 "color": COLORS["dark_green"],
#                 "width": 2,
#             },
#             {
#                 "title": "ШТ. БЕЗ СЕРТИФИКАТА",
#                 "value": self._format_number(no_cert_units),
#                 "subtitle": "единиц на остатке",
#                 "color": COLORS["dark_green"],
#                 "width": 2,
#             },
#         ]

#         row = self.kpi.draw_row(row, row1_cards)
#         row += 2

#         if df.empty:
#             self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
#             message_cell = self.ws.cell(
#                 row=row,
#                 column=2,
#                 value="✅ Нет товаров с просроченными, истекающими или отсутствующими сертификатами",
#             )
#             message_cell.font = Font(
#                 name="Roboto",
#                 size=14,
#                 bold=True,
#                 color=COLORS["dark_green"],
#             )
#             message_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center",
#                 wrap_text=True,
#             )
#             message_cell.fill = PatternFill(
#                 start_color=COLORS["light_green"],
#                 end_color=COLORS["light_green"],
#                 fill_type="solid",
#             )
#             message_cell.border = Border(
#                 left=Side(style="thin", color=COLORS["border_gray"]),
#                 right=Side(style="thin", color=COLORS["border_gray"]),
#                 top=Side(style="thin", color=COLORS["border_gray"]),
#                 bottom=Side(style="thin", color=COLORS["border_gray"]),
#             )
#             self.ws.row_dimensions[row].height = 60
#             self.ws.sheet_view.showGridLines = False
#             return

#         headers = [
#             "ID карточки WB",
#             "Бренд",
#             "Артикул продавца",
#             "Категория",
#             "Пол",
#             "Наименование",
#             "Размеры / остатки",
#             "Кол-во, шт",
#             "Дата окончания",
#             "Статус",
#         ]

#         data_rows = []

#         for _, row_data in df.iterrows():
#             nm_id = self._safe_text(row_data.get("nm_id"))
#             article = self._safe_text(row_data.get("артикул"))
#             name = self._safe_text(row_data.get("наименование"), "не указано")
#             size = self._safe_text(row_data.get("размер"), "не указан")
#             status = self._safe_text(row_data.get("статус_сертификата"))

#             if status == "Нет сертификата":
#                 status_text = "НЕТ СЕРТИФИКАТА"
#             elif status == "Просрочен":
#                 status_text = "ПРОСРОЧЕН"
#             else:
#                 status_text = "ИСТЕКАЕТ"

#             data_rows.append(
#                 [
#                     f"'{nm_id}",
#                     self._safe_text(row_data.get("бренд"), "не указан"),
#                     f"'{article}",
#                     self._safe_text(row_data.get("категория"), "не указана"),
#                     self._safe_text(row_data.get("пол"), "не указан"),
#                     name[:80],
#                     size,
#                     self._safe_number(row_data.get("количество"), 0),
#                     row_data.get("дата_окончания_сертификата") or "-",
#                     status_text,
#                 ]
#             )

#         column_widths = {
#             "A": 5,
#             "B": 15,
#             "C": 18,
#             "D": 18,
#             "E": 18,
#             "F": 14,
#             "G": 38,
#             "H": 36,
#             "I": 14,
#             "J": 18,
#             "K": 20,
#         }

#         row = self.table.draw(
#             start_row=row,
#             headers=headers,
#             data_rows=data_rows,
#             start_col=2,
#             number_format="#,##0",
#             highlight_cols=None,
#             column_widths=column_widths,
#         )

#         first_data_row = row - len(data_rows)
#         last_data_row = row - 1
#         header_row = first_data_row - 1

#         for r in range(first_data_row, row):
#             for col in range(2, 9):
#                 self.ws.cell(row=r, column=col).alignment = Alignment(
#                     horizontal="left",
#                     vertical="center",
#                     wrap_text=True,
#                 )

#             for col in range(9, 12):
#                 self.ws.cell(row=r, column=col).alignment = Alignment(
#                     horizontal="center",
#                     vertical="center",
#                     wrap_text=True,
#                 )

#             qty_cell = self.ws.cell(row=r, column=9)
#             qty_cell.number_format = "#,##0"
#             qty_cell.font = Font(
#                 name="Roboto",
#                 size=9,
#                 bold=True,
#                 color=COLORS["dark_green"],
#             )
#             qty_cell.fill = PatternFill(
#                 start_color=COLORS["light_green"],
#                 end_color=COLORS["light_green"],
#                 fill_type="solid",
#             )

#             date_cell = self.ws.cell(row=r, column=10)
#             if str(date_cell.value).lower() == "нет сертификата":
#                 date_cell.fill = PatternFill(
#                     start_color=COLORS["delta_red"],
#                     end_color=COLORS["delta_red"],
#                     fill_type="solid",
#                 )
#                 date_cell.font = Font(
#                     name="Roboto",
#                     size=9,
#                     bold=True,
#                     color=COLORS["negative_brown"],
#                 )
#             else:
#                 date_cell.number_format = "dd.mm.yyyy"

#             status_cell = self.ws.cell(row=r, column=11)
#             self._apply_status_style(status_cell)

#         self.ws.auto_filter.ref = f"B{header_row}:K{last_data_row}"
#         self.ws.freeze_panes = f"G{header_row + 1}"
#         self.ws.sheet_view.showGridLines = False

#         self.footnote.draw(
#             row=row + 1,
#             text=(
#                 "ПРОСРОЧЕН — дата окончания сертификата меньше даты отчета. "
#                 "ИСТЕКАЕТ — сертификат истекает в ближайшие 30 дней. "
#                 "НЕТ СЕРТИФИКАТА — у товара есть положительный остаток, но дата окончания сертификата не указана. "
#                 "Размеры сгруппированы по артикулу продавца и показаны списком с количеством остатков."
#             ),
#         )



# inventories/reporting/excel/sheets/certificate_sheet.py

from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .base_sheet import BaseSheet
from ..styles.theme import COLORS
from ..components import create_kpi_cards, create_table, Footnote, create_sheet_title


class CertificateSheet(BaseSheet):
    """Лист с сертификатами: просроченные, истекающие и отсутствующие"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)

        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    @staticmethod
    def _is_empty(value):
        """
        Безопасная проверка на пустое значение.
        Нужна, чтобы не было ошибки:
        The truth value of an array with more than one element is ambiguous.
        """
        if value is None:
            return True

        if isinstance(value, (list, tuple, set, np.ndarray)):
            return len(value) == 0

        try:
            return bool(pd.isna(value))
        except Exception:
            return False

    @staticmethod
    def _safe_text(value, default=""):
        if CertificateSheet._is_empty(value):
            return default

        if isinstance(value, (list, tuple, set, np.ndarray)):
            value = ", ".join(map(str, value))

        value = str(value).strip()
        return value if value else default

    @staticmethod
    def _safe_number(value, default=0):
        if CertificateSheet._is_empty(value):
            return default

        try:
            return float(value)
        except Exception:
            return default

    def _format_number(self, value):
        if self._is_empty(value):
            return "0"

        try:
            return f"{int(value):,}".replace(",", " ")
        except Exception:
            return str(value)

    def _apply_status_style(self, cell):
        value = str(cell.value).upper() if cell.value else ""

        if "НЕТ СЕРТИФИКАТА" in value:
            cell.fill = PatternFill(
                start_color=COLORS["delta_red"],
                end_color=COLORS["delta_red"],
                fill_type="solid",
            )
            cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS["negative_brown"],
            )

        elif "ПРОСРОЧЕН" in value:
            cell.fill = PatternFill(
                start_color=COLORS["expired_bg"],
                end_color=COLORS["expired_bg"],
                fill_type="solid",
            )
            cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS["expired_text"],
            )

        else:
            cell.fill = PatternFill(
                start_color=COLORS["expiring_bg"],
                end_color=COLORS["expiring_bg"],
                fill_type="solid",
            )
            cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS["expiring_text"],
            )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    def build(self, df, report_date):
        row = 1

        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(
            name="Roboto",
            size=9,
            bold=True,
            color=COLORS["dark_green"],
        )
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(
            start_color=COLORS["light_green"],
            end_color=COLORS["light_green"],
            fill_type="solid",
        )
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"]),
        )
        btn_cell.hyperlink = "#'TOC'!A1"

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2

        report_date_formatted = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d.%m.%Y")

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        title_cell = self.ws.cell(row=row, column=2, value="СЕРТИФИКАТЫ СООТВЕТСТВИЯ")
        title_cell.font = Font(
            name="Roboto",
            size=16,
            bold=True,
            color=COLORS["dark_green"],
        )
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1

        expired_count = (
            len(df[df["статус_сертификата"] == "Просрочен"])
            if not df.empty and "статус_сертификата" in df.columns
            else 0
        )

        expiring_30_count = (
            len(df[df["статус_сертификата"] == "Истекает в ближайшие 30 дней"])
            if not df.empty and "статус_сертификата" in df.columns
            else 0
        )

        no_cert_count = (
            len(df[df["статус_сертификата"] == "Нет сертификата"])
            if not df.empty and "статус_сертификата" in df.columns
            else 0
        )

        expired_units = (
            pd.to_numeric(
                df.loc[df["статус_сертификата"] == "Просрочен", "количество"],
                errors="coerce",
            ).fillna(0).sum()
            if not df.empty and {"статус_сертификата", "количество"}.issubset(df.columns)
            else 0
        )

        no_cert_units = (
            pd.to_numeric(
                df.loc[df["статус_сертификата"] == "Нет сертификата", "количество"],
                errors="coerce",
            ).fillna(0).sum()
            if not df.empty and {"статус_сертификата", "количество"}.issubset(df.columns)
            else 0
        )

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        subtitle_cell = self.ws.cell(
            row=row,
            column=2,
            value=(
                f"Дата остатков: {report_date_formatted} | "
                f"Просрочено: {expired_count} | "
                f"Истекает в ближайшие 30 дней: {expiring_30_count} | "
                f"Без сертификата: {no_cert_count}"
            ),
        )
        subtitle_cell.font = Font(
            name="Roboto",
            size=11,
            color=COLORS["text_gray"],
        )
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 24
        row += 1

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        date_cell = self.ws.cell(
            row=row,
            column=2,
            value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
        )
        date_cell.font = Font(
            name="Roboto",
            size=9,
            italic=True,
            color=COLORS["text_gray"],
        )
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 2

        total_products = (
            len(df["nm_id"].unique())
            if not df.empty and "nm_id" in df.columns
            else 0
        )

        row1_cards = [
            {
                "title": "ТОВАРОВ С ПРОБЛЕМАМИ",
                "value": self._format_number(total_products),
                "subtitle": "артикулы требуют внимания",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "ПРОСРОЧЕННЫЕ",
                "value": self._format_number(expired_count),
                "subtitle": "артикулы",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "БЕЗ СЕРТИФИКАТА",
                "value": self._format_number(no_cert_count),
                "subtitle": "артикулы",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "ШТ. С ПРОСРОЧЕННЫМ",
                "value": self._format_number(expired_units),
                "subtitle": "единиц на остатке",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "ШТ. БЕЗ СЕРТИФИКАТА",
                "value": self._format_number(no_cert_units),
                "subtitle": "единиц на остатке",
                "color": COLORS["dark_green"],
                "width": 2,
            },
        ]

        row = self.kpi.draw_row(row, row1_cards)
        row += 2

        if df.empty:
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
            message_cell = self.ws.cell(
                row=row,
                column=2,
                value="✅ Нет товаров с просроченными, истекающими или отсутствующими сертификатами",
            )
            message_cell.font = Font(
                name="Roboto",
                size=14,
                bold=True,
                color=COLORS["dark_green"],
            )
            message_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            message_cell.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid",
            )
            message_cell.border = Border(
                left=Side(style="thin", color=COLORS["border_gray"]),
                right=Side(style="thin", color=COLORS["border_gray"]),
                top=Side(style="thin", color=COLORS["border_gray"]),
                bottom=Side(style="thin", color=COLORS["border_gray"]),
            )
            self.ws.row_dimensions[row].height = 60
            self.ws.sheet_view.showGridLines = False
            return

        headers = [
            "ID карточки WB",
            "Бренд",
            "Артикул продавца",
            "Категория",
            "Пол",
            "Наименование",
            "Размеры / остатки",
            "Кол-во, шт",
            "Дата окончания",
            "Статус",
        ]

        data_rows = []

        for _, row_data in df.iterrows():
            nm_id = self._safe_text(row_data.get("nm_id"))
            article = self._safe_text(row_data.get("артикул"))
            name = self._safe_text(row_data.get("наименование"), "не указано")
            size = self._safe_text(row_data.get("размер"), "не указан")
            status = self._safe_text(row_data.get("статус_сертификата"))

            if status == "Нет сертификата":
                status_text = "НЕТ СЕРТИФИКАТА"
            elif status == "Просрочен":
                status_text = "ПРОСРОЧЕН"
            else:
                status_text = "ИСТЕКАЕТ"

            cert_date = row_data.get("дата_окончания_сертификата")
            cert_date = self._safe_text(cert_date, "-")

            data_rows.append(
                [
                    f"'{nm_id}",
                    self._safe_text(row_data.get("бренд"), "не указан"),
                    f"'{article}",
                    self._safe_text(row_data.get("категория"), "не указана"),
                    self._safe_text(row_data.get("пол"), "не указан"),
                    name[:80],
                    size,
                    self._safe_number(row_data.get("количество"), 0),
                    cert_date,
                    status_text,
                ]
            )

        column_widths = {
            "A": 5,
            "B": 15,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 14,
            "G": 38,
            "H": 36,
            "I": 14,
            "J": 18,
            "K": 20,
        }

        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format="#,##0",
            highlight_cols=None,
            column_widths=column_widths,
        )

        first_data_row = row - len(data_rows)
        last_data_row = row - 1
        header_row = first_data_row - 1

        for r in range(first_data_row, row):
            for col in range(2, 9):
                self.ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )

            for col in range(9, 12):
                self.ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            qty_cell = self.ws.cell(row=r, column=9)
            qty_cell.number_format = "#,##0"
            qty_cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS["dark_green"],
            )
            qty_cell.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid",
            )

            date_cell = self.ws.cell(row=r, column=10)
            if str(date_cell.value).upper() == "НЕТ СЕРТИФИКАТА":
                date_cell.fill = PatternFill(
                    start_color=COLORS["delta_red"],
                    end_color=COLORS["delta_red"],
                    fill_type="solid",
                )
                date_cell.font = Font(
                    name="Roboto",
                    size=9,
                    bold=True,
                    color=COLORS["negative_brown"],
                )
            else:
                date_cell.number_format = "dd.mm.yyyy"

            status_cell = self.ws.cell(row=r, column=11)
            self._apply_status_style(status_cell)

        self.ws.auto_filter.ref = f"B{header_row}:K{last_data_row}"
        self.ws.freeze_panes = f"G{header_row + 1}"
        self.ws.sheet_view.showGridLines = False

        self.footnote.draw(
            row=row + 1,
            text=(
                "ПРОСРОЧЕН — дата окончания сертификата меньше даты отчета. "
                "ИСТЕКАЕТ — сертификат истекает в ближайшие 30 дней. "
                "НЕТ СЕРТИФИКАТА — у товара есть положительный остаток, но дата окончания сертификата не указана. "
                "Размеры сгруппированы по артикулу продавца и показаны списком с количеством остатков."
            ),
        )