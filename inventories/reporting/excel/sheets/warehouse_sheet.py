# inventories/reporting/excel/sheets/warehouse_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..components import create_table, Footnote, create_sheet_title, create_kpi_cards
from ..styles.theme import COLORS
from ..styles.helpers import draw_toc_button


class WarehouseSheet(BaseSheet):
    """Лист с остатками по складам"""
    
    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    def build(self, df, stats, report_date):
        row = 1
        
        # ============================================================
        # КНОПКА НАЗАД (стилизованная)
        # ============================================================
        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.hyperlink = "#'TOC'!A1"
        
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # ============================================================
        # ЗАГОЛОВОК
        # ============================================================
        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        
        # Основной заголовок
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        title_cell = self.ws.cell(row=row, column=2, value="ОСТАТКИ ПО СКЛАДАМ")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1
        
        # Подзаголовок с датой
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        subtitle_cell = self.ws.cell(row=row, column=2, value=f"Дата остатков: {report_date_formatted} (на 23:30 МСК)")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 24
        row += 1
        
        # Дата формирования
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        date_cell = self.ws.cell(row=row, column=2, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 2
        
        # ============================================================
        # KPI КАРТОЧКИ (один ряд)
        # ============================================================
        # Расчёт данных для карточек
        total_warehouses = stats.get('total_warehouses', 0) if stats else df['склад'].nunique() if 'склад' in df.columns else 0
        total_regions = df['регион'].nunique() if 'регион' in df.columns else 0
        on_hand = df['на_складе'].sum() if 'на_складе' in df.columns else 0
        in_transit_total = (df['в_пути_к_клиенту'].sum() + df['в_пути_от_клиента'].sum()) if 'в_пути_к_клиенту' in df.columns and 'в_пути_от_клиента' in df.columns else 0
        total_goods = df['итого'].sum() if 'итого' in df.columns else 0
        
        # Одна строка карточек (5 карточек)
        cards = [
            {'title': 'ВСЕГО СКЛАДОВ', 'value': self._format_number(total_warehouses), 
             'subtitle': 'точек хранения', 'color': COLORS["dark_green"], 'width': 1},
            {'title': 'ВСЕГО РЕГИОНОВ', 'value': self._format_number(total_regions), 
             'subtitle': 'географических зон', 'color': COLORS["dark_green"], 'width': 1},
            {'title': 'НА СКЛАДАХ', 'value': self._format_number(on_hand), 
             'subtitle': 'единиц в наличии', 'color': COLORS["dark_green"], 'width': 1},
            {'title': 'В ПУТИ', 'value': self._format_number(in_transit_total), 
             'subtitle': 'к клиенту + от клиента', 'color': COLORS["dark_green"], 'width': 2},
            {'title': 'ИТОГО', 'value': self._format_number(total_goods), 
             'subtitle': 'всего единиц', 'color': COLORS["dark_green"], 'width': 1},
        ]
        
        row = self.kpi.draw_row(row, cards)
        row += 2
        
        # ============================================================
        # ТАБЛИЦА (без колонки с номером)
        # ============================================================
        header_row = row
        
        headers = ['Склад', 'Регион', 'На складе', 'В пути к клиенту', 'В пути от клиента', 'ИТОГО']
        
        # Подготавливаем данные (без колонки с номером)
        data_rows = []
        for idx, row_data in df.iterrows():
            data_rows.append([
                row_data['склад'],
                row_data['регион'],
                row_data['на_складе'],
                row_data['в_пути_к_клиенту'],
                row_data['в_пути_от_клиента'],
                row_data['итого']
            ])
        
        # Добавляем итоговую строку
        total_row_data = [
            'ВСЕГО:',
            '',
            df['на_складе'].sum(),
            df['в_пути_к_клиенту'].sum(),
            df['в_пути_от_клиента'].sum(),
            df['итого'].sum()
        ]
        data_rows.append(total_row_data)
        
        column_widths = {
            'A': 5,   # отступ слева
            'B': 35,  # Склад
            'C': 25,  # Регион
            'D': 18,  # На складе
            'E': 20,  # В пути к клиенту
            'F': 20,  # В пути от клиента
            'G': 25,  # ИТОГО
        }
        
        # Рисуем таблицу
        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format='#,##0',
            highlight_cols=None,
            column_widths=column_widths
        )
        
        # ============================================================
        # ВЫРАВНИВАНИЕ КОЛОНОК
        # ============================================================
        for r in range(header_row + 1, row):
            # Левый край для текстовых колонок (Склад и Регион)
            for col in [2, 3]:
                cell = self.ws.cell(row=r, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Центр для числовых колонок (D, E, F, G)
            for col in [4, 5, 6, 7]:
                cell = self.ws.cell(row=r, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '#,##0'
        
        # Выделяем итоговую строку жирным шрифтом
        total_row_idx = header_row + len(df) + 1
        for col in range(2, 8):
            cell = self.ws.cell(row=total_row_idx, column=col)
            cell.font = Font(bold=True)
        
 
        
        # ============================================================
        # НАСТРОЙКИ
        # ============================================================
        # Фильтр для таблицы (колонки B:G)
        last_data_row = row - 2
        self.ws.auto_filter.ref = f'B{header_row}:G{last_data_row}'
        
        # Фиксируем шапку таблицы
        self.ws.freeze_panes = f'C{header_row + 1}'
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False
    
    def _format_number(self, num):
        """Форматирует число с разделителями тысяч"""
        try:
            return f"{int(num):,}".replace(",", " ")
        except (ValueError, TypeError):
            return str(num)