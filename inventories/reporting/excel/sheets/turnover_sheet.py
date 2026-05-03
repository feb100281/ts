# inventories/reporting/excel/sheets/turnover_sheet.py

from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .base_sheet import BaseSheet
from ..styles.theme import COLORS
from ..components import create_kpi_cards, create_table, Footnote, create_sheet_title


class TurnoverSheet(BaseSheet):
    """Лист с анализом оборачиваемости остатков"""

    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    @staticmethod
    def _safe_text(value, default=""):
        if pd.isna(value):
            return default
        return str(value)

    @staticmethod
    def _safe_number(value, default=0):
        if pd.isna(value):
            return default
        return value

    @staticmethod
    def _safe_optional_number(value):
        if pd.isna(value):
            return None
        return value

    @staticmethod
    def _normalize_size_token(value):
        return str(value).strip().upper().replace(" ", "")

    @classmethod
    def _is_one_size(cls, value):
        token = cls._normalize_size_token(value)
        return token in {"0", "ONESIZE", "ONE SIZE".replace(" ", ""), "OS", "OSFA"}

    @classmethod
    def _count_available_sizes(cls, value):
        """Считает количество размеров.

        Важно:
        - 0 / ONESIZE / ONE SIZE считаются валидным одним размером.
        - Пустое значение = 0 размеров.
        """
        if pd.isna(value) or value in ("", None):
            return 0

        text = str(value).strip()
        if not text:
            return 0

        for sep in [";", ",", "|", "/"]:
            if sep in text:
                tokens = [x.strip() for x in text.split(sep) if x.strip()]
                return len(tokens)

        return 1

    @classmethod
    def _has_size_problem(cls, value):
        """Определяет проблему с размерами.

        Не считаем проблемой:
        - размер 0;
        - ONESIZE;
        - ONE SIZE.

        Проблема — только если размеры пустые / не указаны.
        """
        if pd.isna(value) or str(value).strip() == "":
            return True

        text = str(value).strip()
        if cls._is_one_size(text):
            return False

        return cls._count_available_sizes(text) == 0

    @staticmethod
    def _normalize_status(status):
        if pd.isna(status):
            return "НЕ ОПРЕДЕЛЕН"

        status = str(status).strip().upper()

        mapping = {
            "NO SALES EVER": "НЕТ ПРОДАЖ",
            "NEW ITEM": "НОВЫЙ ТОВАР",
            "NO SALES PERIOD": "НЕТ ПРОДАЖ ЗА ПЕРИОД",
            "STALE": "ДАВНО НЕ ПРОДАВАЛСЯ",
            "SLOW STOCK": "ЗАТОВАРИВАНИЕ",
            "RISK OOS": "РИСК ЗАКОНЧИТЬСЯ",
            "ACTIVE": "АКТИВНЫЙ",
        }

        return mapping.get(status, status)

    @staticmethod
    def _derive_status(row_data):
        sold_qty = pd.to_numeric(row_data.get("продано_за_период"), errors="coerce")
        days_stock = pd.to_numeric(row_data.get("дней_остатка"), errors="coerce")
        last_sale_gap = pd.to_numeric(row_data.get("дней_с_последней_продажи"), errors="coerce")
        is_new = bool(row_data.get("новый_товар")) if not pd.isna(row_data.get("новый_товар")) else False

        if is_new:
            return "НОВЫЙ ТОВАР"

        if pd.isna(sold_qty) or sold_qty <= 0:
            return "НЕТ ПРОДАЖ ЗА ПЕРИОД"

        if not pd.isna(last_sale_gap) and last_sale_gap >= 60:
            return "ДАВНО НЕ ПРОДАВАЛСЯ"

        if not pd.isna(days_stock) and days_stock <= 14:
            return "РИСК ЗАКОНЧИТЬСЯ"

        if not pd.isna(days_stock) and days_stock > 90:
            return "ЗАТОВАРИВАНИЕ"

        return "АКТИВНЫЙ"

    def _build_conclusion_text(
        self,
        df,
        days,
        new_items,
        slow_items,
        no_sales,
        risk_oos,
        old_no_sales,
        size_problem,
    ):
        if df.empty:
            return "Нет данных для анализа оборачиваемости."

        parts = []

        if new_items > 0:
            parts.append(
                f"• {self._format_number(new_items)} товаров являются новыми: по ним уже были первые продажи, но истории недостаточно для вывода о затоваривании. Запас в днях по ним не рассчитывается."
            )

        if no_sales > 0:
            parts.append(
                f"• {self._format_number(no_sales)} товаров не имели продаж за последние {days} дней — их стоит проверить отдельно: актуальность карточки, цену, фото и продвижение."
            )

        if slow_items > 0:
            parts.append(
                f"• {self._format_number(slow_items)} товаров имеют расчетный запас более 90 дней — это зона риска затоваривания. Новые товары в этот показатель не включены."
            )

        if risk_oos > 0:
            parts.append(
                f"• {self._format_number(risk_oos)} товаров могут закончиться в ближайшее время — расчетного остатка хватает до 14 дней."
            )

        if old_no_sales > 0:
            parts.append(
                f"• {self._format_number(old_no_sales)} товаров давно не продавались: последняя продажа была 60+ дней назад."
            )

        if size_problem > 0:
            parts.append(
                f"• У {self._format_number(size_problem)} товаров размеры не указаны или не распознаны. Размеры 0, ONESIZE и ONE SIZE не считаются проблемой."
            )

        if not parts:
            parts.append("• Критичных отклонений по оборачиваемости не выявлено.")

        return "\n".join(parts)

    def _apply_status_style(self, cell, status):
        status = str(status).upper() if status else ""

        if status in ["НЕТ ПРОДАЖ", "NO SALES EVER"]:
            cell.fill = PatternFill(start_color="7A4E4E", end_color="7A4E4E", fill_type="solid")
            cell.font = Font(name="Roboto", size=9, bold=True, color="FFFFFF")

        elif status in ["НОВЫЙ ТОВАР", "NEW ITEM"]:
            cell.fill = PatternFill(start_color="EAF3FF", end_color="EAF3FF", fill_type="solid")
            cell.font = Font(name="Roboto", size=9, bold=True, color="1F4E79")

        elif status in ["НЕТ ПРОДАЖ ЗА ПЕРИОД", "NO SALES PERIOD", "ДАВНО НЕ ПРОДАВАЛСЯ", "STALE"]:
            cell.fill = PatternFill(start_color="FBEAEA", end_color="FBEAEA", fill_type="solid")
            cell.font = Font(name="Roboto", size=9, bold=True, color="7A4E4E")

        elif status in ["ЗАТОВАРИВАНИЕ", "SLOW STOCK"]:
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            cell.font = Font(name="Roboto", size=9, bold=True, color="7A4E00")

        elif status in ["РИСК ЗАКОНЧИТЬСЯ", "RISK OOS"]:
            cell.fill = PatternFill(start_color="EAF6EE", end_color="EAF6EE", fill_type="solid")
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])

        elif status in ["АКТИВНЫЙ", "ACTIVE"]:
            cell.fill = PatternFill(start_color="E7F1ED", end_color="E7F1ED", fill_type="solid")
            cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])

    def build(self, df, stats, report_date, days=90):
        row = 1

        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(
            start_color=COLORS["light_green"],
            end_color=COLORS["light_green"],
            fill_type="solid",
        )
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"]),
        )
        btn_cell.hyperlink = "#'TOC'!A1"

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2

        report_date_formatted = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d.%m.%Y")

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=21)
        title_cell = self.ws.cell(row=row, column=2, value="ОБОРАЧИВАЕМОСТЬ ОСТАТКОВ")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=21)
        subtitle_cell = self.ws.cell(
            row=row,
            column=2,
            value=(
                f"Дата остатков: {report_date_formatted}. "
                f"Продано за период = продажи за последние {days} дней до даты отчета включительно. "
                f"Дней без продаж = дней от последней продажи до даты отчета."
            ),
        )
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.ws.row_dimensions[row].height = 34
        row += 2

        if not df.empty:
            days_stock = pd.to_numeric(df.get("дней_остатка"), errors="coerce")
            sold_qty = pd.to_numeric(df.get("продано_за_период"), errors="coerce").fillna(0)
            days_no_sales = pd.to_numeric(df.get("дней_с_последней_продажи"), errors="coerce")

            if "новый_товар" in df.columns:
                new_mask = df["новый_товар"].fillna(False).astype(bool)
            else:
                new_mask = pd.Series(False, index=df.index)

            new_items = new_mask.sum()
            slow_items = days_stock.gt(90).sum()
            no_sales = sold_qty.le(0).sum()
            risk_oos = days_stock.gt(0).le(14).sum()
            old_no_sales = days_no_sales.fillna(0).ge(60).sum()

            if "доступные_размеры" in df.columns:
                size_problem = df["доступные_размеры"].apply(self._has_size_problem).sum()
            else:
                size_problem = 0
        else:
            new_items = 0
            slow_items = 0
            no_sales = 0
            risk_oos = 0
            old_no_sales = 0
            size_problem = 0

        row1_cards = [
            {
                "title": "ТОВАРОВ В АНАЛИЗЕ",
                "value": self._format_number(len(df)),
                "subtitle": "с положительным остатком",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "НОВЫЕ ТОВАРЫ",
                "value": self._format_number(new_items),
                "subtitle": "истории продаж пока мало",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": f"БЕЗ ПРОДАЖ {days} ДНЕЙ",
                "value": self._format_number(no_sales),
                "subtitle": "не продавались в периоде",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "ЗАПАС БОЛЕЕ 90 ДНЕЙ",
                "value": self._format_number(slow_items),
                "subtitle": "без учета новых товаров",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            {
                "title": "РИСК ЗАКОНЧИТЬСЯ",
                "value": self._format_number(risk_oos),
                "subtitle": "остатка до 14 дней",
                "color": COLORS["dark_green"],
                "width": 2,
            },
            
        ]

        row = self.kpi.draw_row(row, row1_cards)
        row += 1

        conclusion_text = self._build_conclusion_text(
            df=df,
            days=days,
            new_items=new_items,
            slow_items=slow_items,
            no_sales=no_sales,
            risk_oos=risk_oos,
            old_no_sales=old_no_sales,
            size_problem=size_problem,
        )

        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=21)
        conclusion_cell = self.ws.cell(row=row, column=2, value=conclusion_text)
        conclusion_cell.font = Font(name="Roboto", size=10, color=COLORS["black"])
        conclusion_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        conclusion_cell.fill = PatternFill(
            start_color=COLORS["summary_fill"],
            end_color=COLORS["summary_fill"],
            fill_type="solid",
        )
        conclusion_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"]),
        )
        self.ws.row_dimensions[row].height = 88
        row += 2

        headers = [
            "ID карточки WB",
            "Бренд",
            "Артикул",
            "Категория",
            "Пол",
            "Наименование",
            "Размеры",
            "Кол-во размеров",
            "Остаток, шт",
            f"Продано за последние {days} дней",
            "Продано всего",
            "Первая продажа",
            "Последняя продажа",
            "Дней без продаж",
            "Возраст продаж, дней",
            "Дней наблюдения",
            "Средн. продаж в день",
            "Дней остатка",
            "Месяцев остатка",
            "Статус",
        ]

        data_rows = []

        for _, row_data in df.iterrows():
            nm_id = self._safe_text(row_data.get("nm_id"))
            name = self._safe_text(row_data.get("наименование"))
            available_sizes = self._safe_text(row_data.get("доступные_размеры"), "не указаны")
            available_sizes_count = self._count_available_sizes(available_sizes)

            raw_status = row_data.get("статус_остатка")
            if pd.isna(raw_status):
                status = self._derive_status(row_data)
            else:
                status = self._normalize_status(raw_status)

            data_rows.append([
                f"'{nm_id}",
                self._safe_text(row_data.get("бренд"), "Бренд не указан"),
                self._safe_text(row_data.get("артикул")),
                self._safe_text(row_data.get("категория")),
                self._safe_text(row_data.get("пол"), "не указан"),
                name[:80],
                available_sizes,
                available_sizes_count,
                self._safe_number(row_data.get("остаток"), 0),
                self._safe_number(row_data.get("продано_за_период"), 0),
                self._safe_number(row_data.get("продано_за_все_время"), 0),
                row_data.get("первая_продажа"),
                row_data.get("последняя_продажа"),
                self._safe_optional_number(row_data.get("дней_с_последней_продажи")),
                self._safe_optional_number(row_data.get("возраст_продаж_дней")),
                self._safe_optional_number(row_data.get("дней_наблюдения")),
                self._safe_optional_number(row_data.get("средние_продажи_в_день")),
                self._safe_optional_number(row_data.get("дней_остатка")),
                self._safe_optional_number(row_data.get("месяцев_остатка")),
                status,
            ])

        column_widths = {
            "A": 5,
            "B": 15,
            "C": 18,
            "D": 16,
            "E": 18,
            "F": 14,
            "G": 38,
            "H": 24,
            "I": 16,
            "J": 16,
            "K": 22,
            "L": 16,
            "M": 14,
            "N": 16,
            "O": 16,
            "P": 18,
            "Q": 18,
            "R": 18,
            "S": 16,
            "T": 16,
            "U": 22,
        }

        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format="#,##0",
            highlight_cols=None,
            column_widths=column_widths,
        )

        first_data_row = row - len(data_rows)
        last_data_row = row - 1
        header_row = first_data_row - 1

        for r in range(first_data_row, row):
            for col in range(2, 9):
                self.ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )

            for col in range(9, 22):
                self.ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            # B:U
            self.ws.cell(row=r, column=9).number_format = "#,##0"      # Кол-во размеров
            self.ws.cell(row=r, column=10).number_format = "#,##0"     # Остаток
            stock_cell = self.ws.cell(row=r, column=10)
            stock_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
            stock_cell.fill = PatternFill(
                start_color=COLORS["light_green"],
                end_color=COLORS["light_green"],
                fill_type="solid",
            )
            self.ws.cell(row=r, column=11).number_format = "#,##0"     # Продано за период
            self.ws.cell(row=r, column=12).number_format = "#,##0"     # Продано всего

            self.ws.cell(row=r, column=13).number_format = "dd.mm.yyyy"  # Первая продажа
            self.ws.cell(row=r, column=14).number_format = "dd.mm.yyyy"  # Последняя продажа

            self.ws.cell(row=r, column=15).number_format = "#,##0"       # Дней без продаж
            self.ws.cell(row=r, column=16).number_format = "#,##0"       # Возраст продаж
            self.ws.cell(row=r, column=17).number_format = "#,##0"       # Дней наблюдения
            self.ws.cell(row=r, column=18).number_format = "#,##0.00"    # Средн. продаж
            self.ws.cell(row=r, column=19).number_format = "#,##0"       # Дней остатка
            self.ws.cell(row=r, column=20).number_format = "#,##0.00"    # Месяцев остатка

            status_cell = self.ws.cell(row=r, column=21)
            self._apply_status_style(status_cell, status_cell.value)

            sizes_count_cell = self.ws.cell(row=r, column=9)
            days_no_sales_cell = self.ws.cell(row=r, column=15)
            days_stock_cell = self.ws.cell(row=r, column=19)

            status_value = str(status_cell.value).upper() if status_cell.value else ""
            sizes_count_value = sizes_count_cell.value
            days_no_sales_value = days_no_sales_cell.value
            days_stock_value = days_stock_cell.value

            if days_stock_value is not None and days_stock_value > 90 and status_value != "НОВЫЙ ТОВАР":
                days_stock_cell.fill = PatternFill(
                    start_color=COLORS["stock_highlight"],
                    end_color=COLORS["stock_highlight"],
                    fill_type="solid",
                )
                days_stock_cell.font = Font(name="Roboto", size=9, bold=True, color="7A4E00")

            if days_stock_value is not None and 0 < days_stock_value <= 14 and status_value != "НОВЫЙ ТОВАР":
                days_stock_cell.fill = PatternFill(
                    start_color=COLORS["delta_green"],
                    end_color=COLORS["delta_green"],
                    fill_type="solid",
                )
                days_stock_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])

            if days_no_sales_value is not None and days_no_sales_value >= 60:
                days_no_sales_cell.fill = PatternFill(
                    start_color=COLORS["delta_red"],
                    end_color=COLORS["delta_red"],
                    fill_type="solid",
                )
                days_no_sales_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["negative_brown"])

            available_sizes_value = self.ws.cell(row=r, column=8).value
            if self._has_size_problem(available_sizes_value):
                sizes_count_cell.fill = PatternFill(
                    start_color=COLORS["delta_red"],
                    end_color=COLORS["delta_red"],
                    fill_type="solid",
                )
                sizes_count_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["negative_brown"])

        self.ws.auto_filter.ref = f"B{header_row}:U{last_data_row}"
        self.ws.freeze_panes = f"H{header_row + 1}"
        self.ws.sheet_view.showGridLines = False

        self.footnote.draw(
            row=row + 1,
            text=(
                f"Продано за последние {days} дней — количество продаж за период от даты отчета минус {days - 1} дней до даты отчета включительно. "
                "Дней без продаж — количество дней от последней продажи до даты отчета. "
                "НОВЫЙ ТОВАР — первая продажа была недавно, поэтому вывод о затоваривании по нему не делается и запас в днях не рассчитывается. "
                "ЗАТОВАРИВАНИЕ — расчетный запас более 90 дней, без учета новых товаров. "
               
            ),
        )

