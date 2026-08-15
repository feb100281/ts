from datetime import datetime

from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
)

from .base_sheet import BaseSheet
from ..styles.theme import COLORS
from ..components import (
    create_kpi_cards,
    create_table,
)


class FBSSheet(BaseSheet):
    """Физические остатки на нашем складе FBS."""

    def __init__(
        self,
        workbook,
        sheet_number,
    ):
        super().__init__(
            workbook,
            sheet_number,
        )

        self.kpi = create_kpi_cards(
            self.ws
        )

        self.table = create_table(
            self.ws
        )

    def build(
        self,
        df,
        stats,
        report_date,
    ):
        row = 1

        # ============================================================
        # НАЗАД
        # ============================================================

        btn_cell = self.ws.cell(
            row=row,
            column=2,
            value="←  ОГЛАВЛЕНИЕ",
        )

        btn_cell.font = Font(
            name="Roboto",
            size=9,
            bold=True,
            color=COLORS["dark_green"],
        )

        btn_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        btn_cell.fill = PatternFill(
            start_color=COLORS[
                "light_green"
            ],
            end_color=COLORS[
                "light_green"
            ],
            fill_type="solid",
        )

        btn_cell.border = Border(
            left=Side(
                style="thin",
                color=COLORS[
                    "border_gray"
                ],
            ),
            right=Side(
                style="thin",
                color=COLORS[
                    "border_gray"
                ],
            ),
            top=Side(
                style="thin",
                color=COLORS[
                    "border_gray"
                ],
            ),
            bottom=Side(
                style="thin",
                color=COLORS[
                    "border_gray"
                ],
            ),
        )

        btn_cell.hyperlink = (
            "#'TOC'!A1"
        )

        self.ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=4,
        )

        row += 2

        # ============================================================
        # ЗАГОЛОВОК
        # ============================================================

        report_date_formatted = (
            datetime.strptime(
                report_date,
                "%Y-%m-%d",
            ).strftime(
                "%d.%m.%Y"
            )
        )

        self.ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=9,
        )

        title_cell = self.ws.cell(
            row=row,
            column=2,
            value="ОСТАТКИ FBS",
        )

        title_cell.font = Font(
            name="Roboto",
            size=16,
            bold=True,
            color=COLORS[
                "dark_green"
            ],
        )

        row += 1

        self.ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=9,
        )

        self.ws.cell(
            row=row,
            column=2,
            value=(
                "Физические остатки "
                "на нашем складе "
                f"на {report_date_formatted}"
            ),
        )

        row += 2

        # ============================================================
        # KPI
        # ============================================================

        cards = [
            {
                "title": "ОСТАТОК FBS",
                "value": self._format_number(
                    stats.get(
                        "total_fbs",
                        0,
                    )
                ),
                "subtitle": "шт",
                "color": COLORS[
                    "dark_green"
                ],
                "width": 3,
            },
            {
                "title": "ТОВАРОВ",
                "value": self._format_number(
                    df["nm_id"].nunique()
                    if not df.empty
                    else 0
                ),
                "subtitle": (
                    "уникальных карточек"
                ),
                "color": COLORS[
                    "dark_green"
                ],
                "width": 3,
            },
            {
                "title": "СТРОК",
                "value": self._format_number(
                    len(df)
                ),
                "subtitle": (
                    "размерных позиций"
                ),
                "color": COLORS[
                    "dark_green"
                ],
                "width": 2,
            },
        ]

        row = self.kpi.draw_row(
            row,
            cards,
        )

        row += 2

        # ============================================================
        # ТАБЛИЦА
        # ============================================================

        headers = [
            "ID карточки WB",
            "Бренд",
            "Артикул продавца",
            "Категория",
            "Пол",
            "Наименование",
            "Размер",
            "Остаток FBS, шт",
        ]

        data_rows = []

        for _, item in df.iterrows():

            data_rows.append(
                [
                    str(
                        item["nm_id"]
                    ),
                    item["бренд"],
                    item["артикул"],
                    item["категория"],
                    item["пол"],
                    item["наименование"],
                    item["размер"],
                    item["количество"],
                ]
            )

        column_widths = {
            "A": 5,
            "B": 14,
            "C": 20,
            "D": 18,
            "E": 20,
            "F": 16,
            "G": 40,
            "H": 14,
            "I": 18,
        }

        header_row = row

        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format="#,##0",
            highlight_cols=None,
            column_widths=column_widths,
        )

        for r in range(
            header_row + 1,
            row,
        ):
            qty_cell = self.ws.cell(
                row=r,
                column=9,
            )

            qty_cell.number_format = (
                "#,##0"
            )

            qty_cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS[
                    "dark_green"
                ],
            )

            qty_cell.fill = PatternFill(
                start_color=COLORS[
                    "light_green"
                ],
                end_color=COLORS[
                    "light_green"
                ],
                fill_type="solid",
            )

        self.ws.auto_filter.ref = (
            f"B{header_row}:"
            f"I{row - 1}"
        )

        self.ws.freeze_panes = (
            f"C{header_row + 1}"
        )

        self.ws.sheet_view.showGridLines = (
            False
        )