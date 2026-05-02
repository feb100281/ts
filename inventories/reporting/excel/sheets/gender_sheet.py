# inventories/reporting/excel/sheets/gender_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..components import create_kpi_cards, create_table, Footnote, create_sheet_title
from ..styles.theme import COLORS
from ..styles.helpers import draw_toc_button


class GenderSheet(BaseSheet):
    """Лист с остатками по полу"""
    
    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    def build(self, df, stats, report_date):
        row = 1
        
        # ============================================================
        # КНОПКА НАЗАД (стилизованная)
        # ============================================================
        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.hyperlink = "#'TOC'!A1"
        
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # ============================================================
        # ЗАГОЛОВОК
        # ============================================================
        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        
        # Основной заголовок
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        title_cell = self.ws.cell(row=row, column=2, value="ОСТАТКИ ПО ПОЛУ")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1
        
        # Подзаголовок с датой
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        subtitle_cell = self.ws.cell(row=row, column=2, value=f"Дата остатков: {report_date_formatted} (на 23:30 МСК)")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 24
        row += 1
        
        # Дата формирования
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        date_cell = self.ws.cell(row=row, column=2, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 2
        
        # ============================================================
        # KPI КАРТОЧКИ
        # ============================================================
        # total_genders - уникальное количество значений в колонке 'пол'
        total_genders = df['пол'].nunique() if 'пол' in df.columns else 0
        total_quantity = df['количество'].sum() if 'количество' in df.columns else 0
        
        row1_cards = [
            {'title': 'ВСЕГО ГРУПП ПОЛА', 'value': self._format_number(total_genders), 
             'subtitle': 'категорий пола', 'color': COLORS["dark_green"], 'width': 1},
            {'title': 'ОБЩЕЕ КОЛИЧЕСТВО', 'value': self._format_number(total_quantity), 
             'subtitle': 'единиц на складах и в пути', 'color': COLORS["dark_green"], 'width': 2},
        ]
        
        row = self.kpi.draw_row(row, row1_cards)
        row += 2
        
        # ============================================================
        # ТАБЛИЦА
        # ============================================================
        header_row = row
        
        headers = ['Пол', 'Количество товаров', 'Всего, шт']
        
        # Подготавливаем данные (без колонки с номером)
        data_rows = []
        for idx, row_data in df.iterrows():
            data_rows.append([
                row_data['пол'],
                row_data['товаров'],
                row_data['количество']
            ])
        
        column_widths = {
            'A': 5,   # отступ слева
            'B': 25,  # Пол
            'C': 25,  # Количество товаров
            'D': 18,  # Всего, шт
        }
        
        # Рисуем таблицу
        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format='#,##0',
            highlight_cols=None,
            column_widths=column_widths
        )
        
        # ============================================================
        # ВЫРАВНИВАНИЕ КОЛОНОК
        # ============================================================
        for r in range(header_row + 1, row):
            # Левый край для колонки "Пол" (колонка B)
            cell_gender = self.ws.cell(row=r, column=2)
            cell_gender.alignment = Alignment(horizontal="left", vertical="center")
            
            # Центр для числовых колонок (C и D)
            for col in [3, 4]:
                cell = self.ws.cell(row=r, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '#,##0'
        
        # ============================================================
        # СНОСКА
        # ============================================================
        self.footnote.draw(row, "* Данные представлены по всем категориям товаров")
        row += 1
        
        # ============================================================
        # НАСТРОЙКИ
        # ============================================================
        # Фильтр для таблицы (колонки B:D)
        last_data_row = row - 2
        self.ws.auto_filter.ref = f'B{header_row}:D{last_data_row}'
        
        # Фиксируем шапку таблицы
        self.ws.freeze_panes = f'B{header_row + 1}'
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False
    
    def _format_number(self, num):
        """Форматирует число с разделителями тысяч"""
        try:
            return f"{int(num):,}".replace(",", " ")
        except (ValueError, TypeError):
            return str(num)




