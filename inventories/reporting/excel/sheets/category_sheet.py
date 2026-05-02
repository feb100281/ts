# inventories/reporting/excel/sheets/category_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, FONTS, BORDERS, ALIGNMENTS, FILLS
from ..components import create_table, Footnote, create_sheet_title, create_kpi_cards
from ..styles.helpers import draw_toc_button


class CategorySheet(BaseSheet):
    """Лист с остатками по категориям"""
    
    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    def build(self, df, stats, report_date):
        row = 1
        
        # ============================================================
        # КНОПКА НАЗАД (стилизованная)
        # ============================================================
        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.hyperlink = "#'TOC'!A1"
        
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # ============================================================
        # ЗАГОЛОВОК
        # ============================================================
        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        
        # Основной заголовок
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        title_cell = self.ws.cell(row=row, column=2, value="ОСТАТКИ ПО КАТЕГОРИЯМ")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1
        
        # Подзаголовок с датой
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        subtitle_cell = self.ws.cell(row=row, column=2, value=f"Дата остатков: {report_date_formatted} (на 23:30 МСК)")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 24
        row += 1
        
        # Дата формирования
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        date_cell = self.ws.cell(row=row, column=2, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 2
        
        # ============================================================
        # KPI КАРТОЧКИ (одна строка, 2 карточки)
        # ============================================================
        row1_cards = [
            {'title': 'ВСЕГО КАТЕГОРИЙ', 'value': self._format_number(stats['total_categories']), 
             'subtitle': 'товарных групп', 'color': COLORS["dark_green"], 'width': 1},
            {'title': 'ОБЩЕЕ КОЛИЧЕСТВО', 'value': self._format_number(stats['total_quantity']), 
             'subtitle': 'единиц на складах и в пути', 'color': COLORS["dark_green"], 'width': 2},
        ]
        
        row = self.kpi.draw_row(row, row1_cards)
        row += 2
        
        # ============================================================
        # ТАБЛИЦА
        # ============================================================
        # СОХРАНЯЕМ строку, где будет нарисован ЗАГОЛОВОК таблицы
        header_row = row
        
        headers = ['Категория', 'Разбивка по полу', 'Всего, шт']
        
        # Подготавливаем данные
        data_rows = []
        for idx, row_data in df.iterrows():
            data_rows.append([
                row_data['категория'],
                row_data['разбивка_по_полу'],
                row_data['всего']
            ])
        
        column_widths = {
            'A': 5,   # отступ слева
            'B': 35,  # Категория
            'C': 50,  # Разбивка по полу
            'D': 15,  # Всего, шт
        }
        
        # Рисуем таблицу
        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format='#,##0',
            highlight_cols=None,
            column_widths=column_widths
        )
        
        # ============================================================
        # ВЫРАВНИВАНИЕ КОЛОНОК И ПЕРЕНОС ТЕКСТА
        # ============================================================
        for r in range(header_row + 1, row):  # +1 чтобы пропустить заголовок
            # Левый край для текстовых колонок (B, C)
            for col in [2, 3]:
                cell = self.ws.cell(row=r, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Центр для количества (D)
            cell_qty = self.ws.cell(row=r, column=4)
            cell_qty.alignment = Alignment(horizontal="center", vertical="center")
            cell_qty.number_format = '#,##0'
            
            # Автовысота строки для колонки с разбивкой (если есть переносы)
            if self.ws.cell(row=r, column=3).value:
                lines = str(self.ws.cell(row=r, column=3).value).count('\n') + 1
                self.ws.row_dimensions[r].height = max(22, lines * 14)
        
        # ============================================================
        # СНОСКА
        # ============================================================
        self.footnote.draw(row, "* В разбивке указано: пол -> количество в штуках")
        row += 1
        
        # ============================================================
        # НАСТРОЙКИ
        # ============================================================
        # Фильтр для таблицы (от строки заголовка до последней строки данных)
        # header_row - строка с заголовками таблицы
        # row - текущая строка (после сноски, но данные закончились на row-2)
        last_data_row = row - 2  # минус строка со сноской
        self.ws.auto_filter.ref = f'B{header_row}:D{last_data_row}'
        
        # Фиксируем шапку таблицы (строка заголовка)
        self.ws.freeze_panes = f'C{header_row + 1}'  # +1 чтобы заморозить выше заголовка

        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False