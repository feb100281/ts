# inventories/reporting/excel/sheets/toc_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, BORDERS


class TOCSheet(BaseSheet):
    def __init__(self, workbook):
        super().__init__(workbook, "TOC")
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    def build(self, sheets_info, report_date):
        row = 1
        start_col = 2
        # Таблица занимает колонки B, C, D (2, 3, 4)
        table_start_col = 2
        table_end_col = 4
        
        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')

        
        # ============================================================
        # ЗАГОЛОВОЧНЫЙ БЛОК
        # ============================================================
        # Главный заголовок
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        title_cell = self.ws.cell(row=row, column=table_start_col, value="ОТЧЕТ ПО ОСТАТКАМ НА СКЛАДАХ")
        title_cell.font = Font(name="Roboto", size=18, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 35
        row += 1

        # Подзаголовок
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        subtitle_cell = self.ws.cell(row=row, column=table_start_col, value="Аналитика и детализация остатков")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 22
        row += 1

        # Дата формирования
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        date_cell = self.ws.cell(row=row, column=table_start_col, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 1
        
        # Дата остатков (выделенный блок)
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        stocks_date_cell = self.ws.cell(row=row, column=table_start_col, value=f"ДАННЫЕ ПО ОСТАТКАМ НА {report_date_formatted} (на 23:30 МСК)")
        stocks_date_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        stocks_date_cell.alignment = Alignment(horizontal="left", vertical="center")
        stocks_date_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        self.ws.row_dimensions[row].height = 28
        row += 2

        # ============================================================
        # ДЕКОРАТИВНЫЙ РАЗДЕЛИТЕЛЬ
        # ============================================================
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
            cell.border = Border(bottom=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 4
        row += 2

        # ============================================================
        # БЛОК НАВИГАЦИИ
        # ============================================================
        nav_cell = self.ws.cell(row=row, column=table_start_col, value="НАВИГАЦИЯ ПО ОТЧЕТУ")
        nav_cell.font = Font(name="Roboto", size=13, bold=True, color=COLORS["dark_green"])
        nav_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 28
        row += 1


        # ============================================================
        # ТАБЛИЦА ОГЛАВЛЕНИЯ
        # ============================================================
        headers = ["№", "РАЗДЕЛ", "ОПИСАНИЕ"]
        
        # Заголовки таблицы
        for col_idx, header in enumerate(headers, start=table_start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
            cell.border = Border(
                left=Side(style="thin", color=COLORS["border_gray"]),
                right=Side(style="thin", color=COLORS["border_gray"]),
                top=Side(style="thin", color=COLORS["border_gray"]),
                bottom=Side(style="medium", color=COLORS["dark_green"])
            )

        self.ws.row_dimensions[row].height = 32
        row += 1
        table_data_start_row = row

        # Данные таблицы
        for idx, sheet in enumerate(sheets_info):
            sheet_num = sheet['number']
            
            # Номер
            cell_num = self.ws.cell(row=row, column=table_start_col, value=f"{sheet_num:02d}")
            cell_num.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
            cell_num.alignment = Alignment(horizontal="center", vertical="center")
            cell_num.border = BORDERS["thin"]
            cell_num.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")

            # Название с гиперссылкой
            cell_name = self.ws.cell(row=row, column=table_start_col + 1, value=sheet['name'])
            cell_name.font = Font(name="Roboto", size=10, bold=True, color="0066CC")
            cell_name.alignment = Alignment(horizontal="left", vertical="center")
            cell_name.border = BORDERS["thin"]
            cell_name.hyperlink = f"#'{sheet['number']}'!A1"

            # Описание
            cell_desc = self.ws.cell(row=row, column=table_start_col + 2, value=sheet.get("description", ""))
            cell_desc.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            cell_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell_desc.border = BORDERS["thin"]

            # Чередование цвета строк
            if idx % 2 == 0:
                row_fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")
            else:
                row_fill = PatternFill(start_color=COLORS["white"], end_color=COLORS["white"], fill_type="solid")
            
            cell_name.fill = row_fill
            cell_desc.fill = row_fill

            self.ws.row_dimensions[row].height = 28
            row += 1

        # Нижняя граница таблицы
        for col in range(table_start_col, table_start_col + 3):
            cell = self.ws.cell(row=row, column=col)
            cell.border = Border(bottom=Side(style="medium", color=COLORS["dark_green"]))
        self.ws.row_dimensions[row].height = 8
        row += 1

        # ============================================================
        # ИТОГОВАЯ СТРОКА
        # ============================================================
        total_cell = self.ws.cell(row=row, column=table_start_col, value=f"Всего разделов в отчете: {len(sheets_info)}")
        total_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["white"])
        total_cell.fill = PatternFill(start_color=COLORS["dark_green"], end_color=COLORS["dark_green"], fill_type="solid")
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_start_col + 2)
        self.ws.row_dimensions[row].height = 32
        row += 2

        # ============================================================
        # БЛОК ИНФОРМАЦИИ
        # ============================================================
        # Декоративная линия
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.border = Border(top=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 6
        row += 1
        
        # Заголовок
        info_title = self.ws.cell(row=row, column=table_start_col, value="ИНФОРМАЦИЯ")
        info_title.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        info_title.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 22
        row += 1

        footer_texts = [
            f"• Данные по остаткам представлены на {report_date_formatted} (на 23:30 по МСК)",
            "• В отчете показаны только товары с ненулевым остатком на складах",
            "• Для перехода к разделу кликните на его название",

        ]
        
        for text in footer_texts:
            text_cell = self.ws.cell(row=row, column=table_start_col + 1, value=text)
            text_cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            text_cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.merge_cells(start_row=row, start_column=table_start_col + 1, end_row=row, end_column=table_end_col)
            self.ws.row_dimensions[row].height = 18
            row += 1

        row += 1

        # ============================================================
        # НИЖНЯЯ ДЕКОРАТИВНАЯ ПОЛОСА
        # ============================================================
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
            cell.border = Border(top=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 6

        # ============================================================
        # НАСТРОЙКА ШИРИНЫ КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 8
        self.ws.column_dimensions["C"].width = 40
        self.ws.column_dimensions["D"].width = 60
        self.ws.column_dimensions["E"].width = 5
        
        # Включаем перенос текста для описания
        for r in range(table_data_start_row, row - 8):
            self.ws.cell(row=r, column=table_start_col + 2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False