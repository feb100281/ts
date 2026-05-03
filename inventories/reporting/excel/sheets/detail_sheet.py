# inventories/reporting/excel/sheets/detail_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles.theme import COLORS, FONTS, BORDERS, ALIGNMENTS, FILLS
from ..components import create_kpi_cards, create_table, Footnote, create_sheet_title
from ..styles.helpers import draw_toc_button


class DetailSheet(BaseSheet):
    """Лист с детальными остатками по товарам"""
    
    def __init__(self, workbook, sheet_number):
        super().__init__(workbook, sheet_number)
        self.sheet_title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)
        self.footnote = Footnote(self.ws)

    def build(self, df, stats, report_date):
        row = 1
        
        # ============================================================
        # КНОПКА НАЗАД (стилизованная)
        # ============================================================
        btn_cell = self.ws.cell(row=row, column=2, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = PatternFill(start_color=COLORS["light_green"], end_color=COLORS["light_green"], fill_type="solid")
        btn_cell.border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.hyperlink = "#'TOC'!A1"
        
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # ============================================================
        # ЗАГОЛОВОК
        # ============================================================
        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        
        # Основной заголовок
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        title_cell = self.ws.cell(row=row, column=2, value="ДЕТАЛЬНЫЕ ОСТАТКИ ПО ТОВАРАМ")
        title_cell.font = Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1
        
        # Подзаголовок с датой
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        subtitle_cell = self.ws.cell(row=row, column=2, value=f"Дата остатков: {report_date_formatted} (на 23:30 МСК)")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 24
        row += 1
        
        # Дата формирования
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        date_cell = self.ws.cell(row=row, column=2, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 20
        row += 2
        
        # ============================================================
        # KPI КАРТОЧКИ (одна строка, 3 карточки)
        # ============================================================
        row1_cards = [
            {'title': 'ВСЕГО ТОВАРОВ', 'value': self._format_number(stats['total_products']), 
             'subtitle': 'уникальных позиций', 'color': COLORS["dark_green"], 'width': 3},
            {'title': 'ВСЕГО КАТЕГОРИЙ', 'value': self._format_number(stats['total_categories']), 
             'subtitle': 'товарных групп', 'color': COLORS["dark_green"], 'width': 3},
            {'title': 'ОБЩЕЕ КОЛИЧЕСТВО', 'value': self._format_number(stats['total_quantity']), 
             'subtitle': 'единиц на складах и в пути', 'color': COLORS["dark_green"], 'width': 2},
        ]
        
        row = self.kpi.draw_row(row, row1_cards)
        row += 2
        
        # ============================================================
        # ТАБЛИЦА
        # ============================================================
        headers = ['ID карточки WB','Бренд', 'Артикул продавца', 'Категория', 'Пол', 'Наименование', 'Размеры', 'Количество, шт']
        
        # Подготавливаем данные с правильным форматированием
        data_rows = []
        for idx, row_data in df.iterrows():
            # Форматируем nm_id как текст (чтобы Excel не превращал в число)
            nm_id_text = f"'{row_data['nm_id']}" if isinstance(row_data['nm_id'], (int, float)) else str(row_data['nm_id'])
            # Форматируем бренд/артикул продавца как текст
            brand_text = str(row_data['бренд']) if row_data['бренд'] and str(row_data['бренд']) != 'nan' else 'не указан'
            article_text = f"'{row_data['артикул']}" if isinstance(row_data['артикул'], (int, float)) else str(row_data['артикул'])
            
            data_rows.append([
                nm_id_text,
                brand_text,
                article_text,
                row_data['категория'],
                row_data['пол'] if row_data['пол'] and str(row_data['пол']) != 'nan' else 'не указан',
                row_data['наименование'][:80] if len(str(row_data['наименование'])) > 80 else row_data['наименование'],
                row_data['доступные_размеры'][:100] if len(str(row_data['доступные_размеры'])) > 100 else row_data['доступные_размеры'],
                row_data['количество']
            ])
        
        column_widths = {

            'A': 5,
            'B': 12,  # ID карточки WB
            'C': 18,  # Бренд
            'D': 16,  # Артикул продавца
            'E': 16,  # Категория
            'F': 16,  # Пол
            'G': 35,  # Наименование
            'H': 45,  # Размеры
            'I': 16,  # Количество, шт
        }
        
        # Рисуем таблицу
        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            number_format='#,##0',
            highlight_cols=None,
            column_widths=column_widths
        )
        
        # ============================================================
        # ВЫРАВНИВАНИЕ КОЛОНОК
        # ============================================================
        for r in range(row - len(data_rows), row):
            # Левый край для текстовых колонок (B-G)
            for col in [2, 3, 4, 5, 6, 7, 8]:
                cell = self.ws.cell(row=r, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # ПОДСВЕТКА КОЛОНКИ "Количество, шт" (колонка I, индекс 9)
                qty_cell = self.ws.cell(row=r, column=9)
                qty_cell.alignment = Alignment(horizontal="center", vertical="center")
                qty_cell.number_format = '#,##0'
                qty_cell.font = Font(
                    name="Roboto",
                    size=9,
                    bold=True,
                    color=COLORS["dark_green"],
                )
                qty_cell.fill = PatternFill(
                    start_color=COLORS["light_green"],
                    end_color=COLORS["light_green"],
                    fill_type="solid",
                )
            
            # Правый край для количества (H)
            cell_qty = self.ws.cell(row=r, column=9)
            cell_qty.alignment = Alignment(horizontal="center", vertical="center")
            cell_qty.number_format = '#,##0'
        
        
        # ============================================================
        # НАСТРОЙКИ
        # ============================================================
        # Фильтр только для таблицы (с 8 строки)
        self.ws.auto_filter.ref = f'B{row - len(data_rows) - 1}:I{row - 2}'
        

        header_row = row - len(data_rows) - 1
        # Замораживаем: колонки A-B и строки до header_row
        # Ячейка C{header_row + 2} означает:
        # - колонка C - первая незамороженная (A-B заморожены)
        # - строка header_row+2 - первая незамороженная (строки до header_row+1 заморожены)
        self.ws.freeze_panes = f'C{header_row + 1}'
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False