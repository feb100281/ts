# cards/exports/single_upd_exporter.py

import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from django.http import HttpResponse


# Цвета
COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED", 
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "back_text_green": "1F5E4E",
}

# Заливки
FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
}

# Шрифты
FONTS = {
    "title": Font(name="Helvetica Light", size=14, bold=False, color=COLORS["white"]),
    "subtitle": Font(name="Helvetica Light", size=10, color=COLORS["text_gray"]),
    "label": Font(name="Helvetica Light", size=10, color=COLORS["black"]),
    "header": Font(name="Helvetica Light", size=10, color=COLORS["white"]),
    "bold": Font(name="Helvetica Light", size=10, color=COLORS["black"]),
    "normal": Font(name="Helvetica", size=10, color=COLORS["black"]),
    "total": Font(name="Helvetica Light", size=11, bold=True, color=COLORS["black"]),
    "id": Font(name="Helvetica Light", size=10, bold=True, color=COLORS["dark_green"]),
    "mono": Font(name="Helvetica Light", size=10, color=COLORS["black"]),
    "link": Font(name="Helvetica Light", size=10, color=COLORS["dark_green"], underline="single"),
    "empty": Font(name="Helvetica Light", size=10, color="C62828"),
}

# Границы
thin = Side(style="thin", color=COLORS["border_gray"])
medium = Side(style="medium", color=COLORS["dark_green"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_medium": Border(bottom=medium),
    "total": Border(left=medium, right=medium, top=medium, bottom=medium),
}

# Выравнивание
ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
}


class SingleUpdExporter:
    """Класс для выгрузки отдельных УПД в Excel"""

    @classmethod
    def generate_response(cls, upd_document):
        """Генерация HTTP-ответа для одной УПД"""
        excel_buffer = cls._generate_excel_buffer(upd_document)
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"УПД_{upd_document.number}_от_{upd_document.date.strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    @classmethod
    def generate_zip_response(cls, upd_documents):
        """Генерация ZIP-архива для нескольких УПД"""
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for upd_document in upd_documents:
                excel_buffer = cls._generate_excel_buffer(upd_document)
                filename = f"УПД_{upd_document.number}_от_{upd_document.date.strftime('%Y-%m-%d')}.xlsx"
                zip_file.writestr(filename, excel_buffer.getvalue())
        
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename=УПД_пакет_{upd_documents.count()}_шт.zip'
        return response

    @classmethod
    def _generate_excel_buffer(cls, upd_document):
        """Генерация Excel-буфера для одной УПД"""
        wb = Workbook()
        ws = wb.active
        ws.title = "УПД"
        
        # ========== ОТКЛЮЧАЕМ ГРИДЛАЙНЫ ==========
        ws.sheet_view.showGridLines = False

        # ========== ЗАГОЛОВОК ==========
        # Строка 1 - УПД
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        cell = ws.cell(row=1, column=1)
        cell.value = f"УПД №{upd_document.number}"
        cell.font = FONTS["title"]
        cell.alignment = ALIGNMENTS["center"]
        cell.fill = FILLS["header"]

        # Строка 2 - дата и поставщик
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        cell = ws.cell(row=2, column=1)
        cell.value = f"от {upd_document.date.strftime('%d.%m.%Y')}"
        cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["white"])
        cell.alignment = ALIGNMENTS["center"]
        cell.fill = FILLS["header"]

        ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=14)
        cell = ws.cell(row=2, column=7)
        cell.value = f"Поставщик: {str(upd_document.counterparty) if upd_document.counterparty else 'Не указан'}"
        cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["white"])
        cell.alignment = ALIGNMENTS["center"]
        cell.fill = FILLS["header"]
        
        # Нижняя граница заголовка
        for col in range(1, 15):
            cell = ws.cell(row=2, column=col)
            cell.border = BORDERS["bottom_medium"]

        # ========== ИНФОРМАЦИЯ ==========
        row = 4
        

        # Сводка справа
        income_lines = upd_document.income_lines.all()
        total_qty = sum(float(l.upd_qty or 0) for l in income_lines)
        total_vatless = sum(float(l.upd_amount_vatless or 0) for l in income_lines)
        total_vatadd = sum(float(l.upd_amount_vatadd or 0) for l in income_lines)
        total_items = income_lines.count()
        
        summary_data = [
            ("Всего позиций:", f"{total_items} шт."),
            ("Общее количество:", f"{total_qty:,.2f}".replace(',', ' ')),
            ("Сумма без НДС:", f"{total_vatless:,.2f} ₽".replace(',', ' ')),
            ("Сумма с НДС:", f"{total_vatadd:,.2f} ₽".replace(',', ' ')),
        ]
        
        for label, value in summary_data:
            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=12)
            cell = ws.cell(row=row, column=12)
            cell.value = label
            cell.font = FONTS["label"]
            cell.alignment = ALIGNMENTS["right"]
            
            ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
            cell = ws.cell(row=row, column=13)
            cell.value = value
            cell.font = FONTS["bold"]
            cell.alignment = ALIGNMENTS["right"]
            row += 1

        # Пустая строка перед таблицей
        row += 1

        # ========== ТАБЛИЦА ==========
        headers = [
            (1, 'ID', 12, 'center'),
            (2, 'Позиция', 10, 'center'),
            (3, 'Артикул поставщика', 22, 'left'),
            (4, 'Наименование товара', 45, 'left'),
            (5, 'Размер', 12, 'center'),
            (6, 'Ставка НДС, %', 14, 'center'),
            (7, 'Ед. изм.', 10, 'center'),
            (8, 'Кол-во', 14, 'right'),
            (9, 'Цена без НДС', 18, 'right'),
            (10, 'Сумма без НДС', 18, 'right'),
            (11, 'Сумма с НДС', 18, 'right'),
            (12, 'Себестоимость', 18, 'right'),
            (13, 'Валюта', 10, 'center'),
            (14, 'Карточка WB', 20, 'center'),
        ]
        
        header_row = row
        for col_idx, header, width, align in headers:
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = FONTS["header"]
            cell.fill = FILLS["header"]
            cell.alignment = ALIGNMENTS["center_wrap"]
            cell.border = BORDERS["thin"]
            # Устанавливаем ширину колонок
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = width

        # ========== ДАННЫЕ ==========
        data_row = header_row + 1
        thin_border = BORDERS["thin"]
        
        for idx, line in enumerate(income_lines.order_by('upd_pos'), 1):
            # ID
            cell = ws.cell(row=data_row, column=1)
            cell.value = line.id
            cell.font = FONTS["id"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Позиция
            cell = ws.cell(row=data_row, column=2)
            cell.value = line.upd_pos or ''
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Артикул
            cell = ws.cell(row=data_row, column=3)
            cell.value = line.upd_sa_name or ''
            cell.font = FONTS["mono"]
            cell.alignment = ALIGNMENTS["left"]
            cell.border = thin_border
            
            # Наименование
            cell = ws.cell(row=data_row, column=4)
            cell.value = line.upd_title or ''
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["left_wrap"]
            cell.border = thin_border
            
            # Размер
            cell = ws.cell(row=data_row, column=5)
            cell.value = line.upd_size or ''
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Ставка НДС
            cell = ws.cell(row=data_row, column=6)
            cell.value = float(line.upd_vat_rate) if line.upd_vat_rate is not None else 0
            cell.number_format = numbers.FORMAT_NUMBER_00
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Ед. изм.
            cell = ws.cell(row=data_row, column=7)
            cell.value = line.upd_unit or ''
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Количество
            cell = ws.cell(row=data_row, column=8)
            cell.value = float(line.upd_qty) if line.upd_qty is not None else 0
            cell.number_format = '#,##0.00'
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = thin_border
            
            # Цена без НДС
            cell = ws.cell(row=data_row, column=9)
            cell.value = float(line.upd_price_vatless) if line.upd_price_vatless is not None else 0
            cell.number_format = '#,##0.00'
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = thin_border
            
            # Сумма без НДС
            cell = ws.cell(row=data_row, column=10)
            cell.value = float(line.upd_amount_vatless) if line.upd_amount_vatless is not None else 0
            cell.number_format = '#,##0.00'
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = thin_border
            
            # Сумма с НДС
            cell = ws.cell(row=data_row, column=11)
            cell.value = float(line.upd_amount_vatadd) if line.upd_amount_vatadd is not None else 0
            cell.number_format = '#,##0.00'
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = thin_border
            
            # Себестоимость
            cell = ws.cell(row=data_row, column=12)
            if line.man_cost_per_unit is not None:
                cell.value = float(line.man_cost_per_unit)
                cell.font = FONTS["normal"]
            else:
                cell.value = 0
                cell.font = FONTS["empty"]
                cell.fill = PatternFill("solid", fgColor="FDECEC")
            cell.number_format = '#,##0.00'
            cell.alignment = ALIGNMENTS["right"]
            cell.border = thin_border
            
            # Валюта
            cell = ws.cell(row=data_row, column=13)
            cell.value = line.currency_code or 'RUB'
            cell.font = FONTS["normal"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Карточка WB
            cell = ws.cell(row=data_row, column=14)
            cell.fill = FILLS["section"]  # Светло-зеленая заливка
            if line.nm:
                cell.value = str(line.nm_id)
                cell.hyperlink = f"https://www.wildberries.ru/catalog/{line.nm_id}/detail.aspx"
                cell.font = FONTS["link"]
                cell.alignment = ALIGNMENTS["center"]
            else:
                cell.value = '-'
                cell.font = FONTS["normal"]
                cell.alignment = ALIGNMENTS["center"]
            cell.border = thin_border
            
            # Чередование цветов
            if idx % 2 == 0:
                for col in range(1, 15):
                    current_cell = ws.cell(row=data_row, column=col)
                    if not current_cell.fill or current_cell.fill.patternType is None:
                        current_cell.fill = FILLS["alt"]
            
            data_row += 1

        # ========== ИТОГО ==========
        if income_lines.exists():
            total_row = data_row
            total_border = BORDERS["total"]
            
            # Сначала объединяем ячейки для "ИТОГО"
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
            
            # Затем устанавливаем значения
            cell = ws.cell(row=total_row, column=1)
            cell.value = "ИТОГО:"
            cell.alignment = ALIGNMENTS["right"]
            cell.font = FONTS["total"]
            cell.fill = FILLS["total"]
            cell.border = total_border
            
            # Для остальных колонок устанавливаем значения по отдельности
            total_cost_sum = sum((float(l.man_cost_per_unit or 0) * float(l.upd_qty or 0)) for l in income_lines)
            
            # Колонка 8 - Количество
            cell = ws.cell(row=total_row, column=8)
            cell.value = total_qty
            cell.number_format = '#,##0.00'
            cell.font = FONTS["total"]
            cell.fill = FILLS["total"]
            cell.border = total_border
            cell.alignment = ALIGNMENTS["right"]
            
            # Колонка 10 - Сумма без НДС
            cell = ws.cell(row=total_row, column=10)
            cell.value = total_vatless
            cell.number_format = '#,##0.00'
            cell.font = FONTS["total"]
            cell.fill = FILLS["total"]
            cell.border = total_border
            cell.alignment = ALIGNMENTS["right"]
            
            # Колонка 11 - Сумма с НДС
            cell = ws.cell(row=total_row, column=11)
            cell.value = total_vatadd
            cell.number_format = '#,##0.00'
            cell.font = FONTS["total"]
            cell.fill = FILLS["total"]
            cell.border = total_border
            cell.alignment = ALIGNMENTS["right"]
            
            # Колонка 12 - Себестоимость
            cell = ws.cell(row=total_row, column=12)
            cell.value = total_cost_sum
            cell.number_format = '#,##0.00'
            cell.font = FONTS["total"]
            cell.fill = FILLS["total"]
            cell.border = total_border
            cell.alignment = ALIGNMENTS["right"]

            # Для пустых колонок (2-7, 9, 13-14) просто устанавливаем фон и границы
            for col in [2, 3, 4, 5, 6, 7, 9, 13, 14]:
                cell = ws.cell(row=total_row, column=col)
                cell.fill = FILLS["total"]
                cell.border = total_border

        # ========== ФИНАЛЬНЫЕ НАСТРОЙКИ ==========
        # Замораживаем панель
        ws.freeze_panes = ws.cell(row=header_row + 1, column=5)
        
        # Высота строк
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[header_row].height = 30

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer