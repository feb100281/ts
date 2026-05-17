# cards/management/commands/reconcile_upd_1c.py
import re
from decimal import Decimal

import pandas as pd

from django.core.management.base import BaseCommand
from django.db.models import Sum

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from cards.models import UpdDocument


IGNORE_COUNTERPARTIES = {
    "ВАЙЛДБЕРРИЗ ООО",
    "ЗДОРОВАЯ ВОДА ООО",
    "РВБ ООО",
    "ТОРГОВЫЙ ДОМ ГЛАВРУС-РЕКЛАМА ООО",
    "ВАНПЛАСТ ООО",
    "ООО ЛУКОЙЛ-ИНТЕР-КАРД",
    "СИТИЛИНК ООО",
    "ТКС ООО",
    "Чащина Татьяна Владимировна",
    "СПЕЦМОНТАЖ ТЕХНОЛОДЖИ ООО",
    "ДОМАШНИЙ ИНТЕРЬЕР ООО",
    "ГК ГАЛА-ПРОДЖЕКТ ООО",
    "ИП Васильев Данил Андреевич",
    'ГАЛА ООО',
    'ТВИНЛАЙТ ООО'
}


BAD_AMOUNTS = []


def normalize_number(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()
    value = value.replace("№", "")
    value = value.replace(" ", "")
    value = value.replace("\u00a0", "")

    return value


def normalize_counterparty_name(value):
    if pd.isna(value):
        return ""

    value = str(value).upper().strip()

    value = value.replace('"', "")
    value = value.replace("«", "")
    value = value.replace("»", "")
    value = value.replace(".", "")
    value = value.replace(",", "")
    value = value.replace("(", " ")
    value = value.replace(")", " ")

    for word in ["ООО", "ЗАО", "АО", "ПАО", "ИП"]:
        value = re.sub(rf"\b{word}\b", " ", value)

    value = re.sub(r"\s+", " ", value).strip()

    return value


def clean_amount(value):
    original_value = value

    if pd.isna(value):
        return Decimal("0")

    value = str(value).strip()

    if not value or value.lower() in ["nan", "none", "-", "—"]:
        return Decimal("0")

    value = value.replace("\u00a0", "")
    value = value.replace("\u202f", "")
    value = value.replace("₽", "")
    value = value.replace("руб.", "")
    value = value.replace("руб", "")
    value = value.strip()

    if "," in value and "." in value:
        value = value.replace(",", "")
    elif "," in value and "." not in value:
        value = value.replace(" ", "")
        value = value.replace(",", ".")
    else:
        value = value.replace(" ", "")

    value = re.sub(r"[^0-9.\-]", "", value)

    try:
        return Decimal(value)
    except Exception:
        BAD_AMOUNTS.append(original_value)
        return Decimal("0")


def make_numeric_columns(df, columns):
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def beautify_excel(output_path):
    wb = None

    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    status_fills = {
        "OK": PatternFill("solid", fgColor="E8F5E9"),
        "SUM_DIFF": PatternFill("solid", fgColor="FFF4E5"),
        "ONLY_IN_US": PatternFill("solid", fgColor="E8F0FE"),
        "ONLY_IN_1C": PatternFill("solid", fgColor="FDECEC"),
    }

    money_cols = {
        "amount_our",
        "amount_1c",
        "diff",
    }

    integer_cols = {
        "rows_1c",
        "count",
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        headers = {
            cell.value: cell.column
            for cell in ws[1]
            if cell.value
        }

        status_col = headers.get("status")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_status = None

            if status_col:
                row_status = ws.cell(row=row[0].row, column=status_col).value

            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                header = ws.cell(row=1, column=cell.column).value

                if header in money_cols:
                    cell.number_format = '#,##0.00 ₽'

                if header in integer_cols:
                    cell.number_format = '#,##0'

                if row_status in status_fills:
                    cell.fill = status_fills[row_status]

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                value = str(cell.value)
                max_length = max(max_length, len(value))

            width = min(max(max_length + 2, 10), 42)

            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 28

    wb.save(output_path)


class Command(BaseCommand):
    help = "Сверка УПД из Excel 1С с УПД в базе"

    def handle(self, *args, **options):
        excel_path = "/Users/daria/Desktop/New.xlsx"
        output_path = "/Users/daria/Desktop/upd_reconciliation_result.xlsx"

        self.stdout.write(f"Читаю файл 1С: {excel_path}")

        df_1c = pd.read_excel(
            excel_path,
            header=3,
        )

        self.stdout.write(f"Строк в исходном файле: {len(df_1c)}")

        required_columns = [
            "Дата вх.",
            "Номер вх.",
            "Сумма",
            "Контрагент",
            "Счет фактура",
            "Вид операции",
        ]

        missing_columns = [
            col for col in required_columns
            if col not in df_1c.columns
        ]

        if missing_columns:
            self.stderr.write(
                "Не найдены колонки: "
                + ", ".join(missing_columns)
            )

            self.stderr.write(
                "Фактические колонки:\n"
                + "\n".join(map(str, df_1c.columns))
            )
            return

        df_1c = df_1c[
            (
                df_1c["Вид операции"]
                .astype(str)
                .str.strip()
                == "Товары"
            )
            &
            (
                df_1c["Счет фактура"]
                .astype(str)
                .str.strip()
                .isin([
                    "Проведен",
                    "Не требуется",
                ])
            )
        ].copy()

        self.stdout.write(f"После фильтра: {len(df_1c)}")

        df_1c["date"] = pd.to_datetime(
            df_1c["Дата вх."],
            dayfirst=True,
            errors="coerce",
        ).dt.date

        df_1c["number_norm"] = (
            df_1c["Номер вх."]
            .apply(normalize_number)
        )

        df_1c["counterparty_norm"] = (
            df_1c["Контрагент"]
            .apply(normalize_counterparty_name)
        )

        ignore_norm = {
            normalize_counterparty_name(name)
            for name in IGNORE_COUNTERPARTIES
        }

        before_ignore = len(df_1c)

        df_1c = df_1c[
            ~df_1c["counterparty_norm"].isin(ignore_norm)
        ].copy()

        self.stdout.write(
            f"После исключения контрагентов: {len(df_1c)} "
            f"(исключено {before_ignore - len(df_1c)})"
        )

        df_1c["amount_1c"] = (
            df_1c["Сумма"]
            .apply(clean_amount)
        )

        if BAD_AMOUNTS:
            self.stdout.write("Проблемные значения суммы:")
            for item in BAD_AMOUNTS[:20]:
                self.stdout.write(f"  {repr(item)}")

        df_1c_grouped = (
            df_1c
            .groupby(
                [
                    "counterparty_norm",
                    "date",
                    "number_norm",
                ],
                dropna=False,
            )
            .agg(
                amount_1c=("amount_1c", "sum"),
                rows_1c=("amount_1c", "count"),
                counterparty_1c=("Контрагент", "first"),
            )
            .reset_index()
        )

        self.stdout.write(
            f"УПД после группировки 1С: {len(df_1c_grouped)}"
        )

        qs = (
            UpdDocument.objects
            .select_related("counterparty")
            .annotate(
                amount_our=Sum("income_lines__upd_amount_vatadd")
            )
        )

        our_rows = []

        for upd in qs:
            counterparty_name = (
                str(upd.counterparty)
                if upd.counterparty
                else ""
            )

            if " (ИНН:" in counterparty_name:
                counterparty_name = counterparty_name.split(" (ИНН:")[0]

            our_rows.append({
                "upd_id": upd.id,
                "counterparty_our": counterparty_name,
                "counterparty_norm": normalize_counterparty_name(counterparty_name),
                "date": upd.date,
                "number_norm": normalize_number(upd.number),
                "number_our": upd.number,
                "amount_our": upd.amount_our or Decimal("0"),
            })

        df_our = pd.DataFrame(our_rows)

        self.stdout.write(f"УПД в нашей базе: {len(df_our)}")

        duplicate_keys = [
            "counterparty_norm",
            "date",
            "number_norm",
        ]

        df_our_duplicates = (
            df_our[
                df_our.duplicated(
                    subset=duplicate_keys,
                    keep=False,
                )
            ]
            .sort_values(duplicate_keys)
            .copy()
        )

        self.stdout.write(
            f"Дубликатов УПД в нашей базе: {len(df_our_duplicates)}"
        )

        df_result = df_our.merge(
            df_1c_grouped,
            on=[
                "counterparty_norm",
                "date",
                "number_norm",
            ],
            how="outer",
            indicator=True,
        )

        def define_status(row):
            if row["_merge"] == "left_only":
                return "ONLY_IN_US"

            if row["_merge"] == "right_only":
                return "ONLY_IN_1C"

            amount_our = row.get("amount_our") or Decimal("0")
            amount_1c = row.get("amount_1c") or Decimal("0")

            diff = abs(
                Decimal(str(amount_our))
                -
                Decimal(str(amount_1c))
            )

            if diff <= Decimal("1"):
                return "OK"

            return "SUM_DIFF"

        df_result["status"] = df_result.apply(
            define_status,
            axis=1,
        )

        df_result["amount_our"] = df_result["amount_our"].fillna(0)
        df_result["amount_1c"] = df_result["amount_1c"].fillna(0)

        df_result["diff"] = (
            df_result["amount_our"].apply(lambda x: Decimal(str(x)))
            -
            df_result["amount_1c"].apply(lambda x: Decimal(str(x)))
        )

        df_result = make_numeric_columns(
            df_result,
            [
                "amount_our",
                "amount_1c",
                "diff",
            ],
        )

        df_our_duplicates = make_numeric_columns(
            df_our_duplicates,
            [
                "amount_our",
            ],
        )

        summary = (
            df_result
            .groupby("status")
            .size()
            .reset_index(name="count")
        )

        summary_order = [
            "OK",
            "SUM_DIFF",
            "ONLY_IN_US",
            "ONLY_IN_1C",
        ]

        summary["status"] = pd.Categorical(
            summary["status"],
            categories=summary_order,
            ordered=True,
        )

        summary = summary.sort_values("status")

        self.stdout.write("Результат сверки:")

        for _, row in summary.iterrows():
            self.stdout.write(f"{row['status']}: {row['count']}")

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
        ) as writer:
            summary.to_excel(
                writer,
                sheet_name="summary",
                index=False,
            )

            df_result.to_excel(
                writer,
                sheet_name="all",
                index=False,
            )

            df_our_duplicates.to_excel(
                writer,
                sheet_name="DUPLICATES_IN_US",
                index=False,
            )

            for status in [
                "OK",
                "SUM_DIFF",
                "ONLY_IN_US",
                "ONLY_IN_1C",
            ]:
                df_result[
                    df_result["status"] == status
                ].to_excel(
                    writer,
                    sheet_name=status,
                    index=False,
                )

        beautify_excel(output_path)

        self.stdout.write(
            self.style.SUCCESS(
                f"Готово.\n"
                f"Отчет: {output_path}"
            )
        )