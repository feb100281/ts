# cards/reporting/reconcile/styles/theme.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Цвета для отчета по сверке УПД
COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED", 
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "back_text_green": "1F5E4E",
    
    # Статусы
    "ok_green": "E8F5E9",
    "ok_text": "2E7D32",
    "warning_orange": "FFF4E5",
    "warning_text": "ED6C02",
    "info_blue": "E8F0FE",
    "info_text": "1976D2",
    "error_red": "FDECEC",
    "error_text": "D32F2F",
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "none": PatternFill(fill_type=None),
    
    # Статусы
    "ok": PatternFill("solid", fgColor=COLORS["ok_green"]),
    "warning": PatternFill("solid", fgColor=COLORS["warning_orange"]),
    "info": PatternFill("solid", fgColor=COLORS["info_blue"]),
    "error": PatternFill("solid", fgColor=COLORS["error_red"]),
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=12, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=11, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    
    # Статусы
    "ok": Font(name="Roboto", size=10, bold=True, color=COLORS["ok_text"]),
    "warning": Font(name="Roboto", size=10, bold=True, color=COLORS["warning_text"]),
    "info": Font(name="Roboto", size=10, bold=True, color=COLORS["info_text"]),
    "error": Font(name="Roboto", size=10, bold=True, color=COLORS["error_text"]),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium = Side(style="medium", color=COLORS["dark_green"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium),
    "none": Border(),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
}

FORMATS = {
    "number": '#,##0',
    "decimal": '#,##0.00',
    "date": 'dd.mm.yyyy',
    "money": '#,##0.00 ₽',
}