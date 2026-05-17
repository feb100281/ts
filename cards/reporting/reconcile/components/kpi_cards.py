# cards/reporting/reconcile/components/kpi_cards.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles.theme import COLORS, BORDERS, FILLS


class KPICards:
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw_row(self, start_row, cards):
        """Рисует строку KPI карточек"""
        current_col = 2
        
        for card in cards:
            width = card.get('width', 2)  # Теперь все карточки width=2, одинаковые
            
            # Объединяем ячейки для карточки
            self.ws.merge_cells(
                start_row=start_row, 
                start_column=current_col,
                end_row=start_row, 
                end_column=current_col + width - 1
            )
            
            # Контейнер карточки
            container_cell = self.ws.cell(row=start_row, column=current_col)
            container_cell.fill = FILLS["section"]
            container_cell.border = BORDERS["thin"]
            
            # Заголовок
            title_cell = self.ws.cell(row=start_row, column=current_col, value=card['title'])
            title_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["text_gray"])
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Значение
            value_cell = self.ws.cell(row=start_row + 1, column=current_col, value=card['value'])
            value_cell.font = Font(name="Roboto", size=20, bold=True, color=card.get('color', COLORS["dark_green"]))
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Подзаголовок
            subtitle_cell = self.ws.cell(row=start_row + 2, column=current_col, value=card.get('subtitle', ''))
            subtitle_cell.font = Font(name="Roboto", size=8, color=COLORS["text_gray"])
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Устанавливаем ширину колонок под карточки
            for col in range(current_col, current_col + width):
                col_letter = self._get_column_letter(col)
                self.ws.column_dimensions[col_letter].width = 18
            
            # Высота строк для карточки
            self.ws.row_dimensions[start_row].height = 24
            self.ws.row_dimensions[start_row + 1].height = 44
            self.ws.row_dimensions[start_row + 2].height = 20
            
            current_col += width
        
        return start_row + 3
    
    def _get_column_letter(self, col_idx):
        """Конвертирует номер колонки в букву"""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + col_idx % 26) + result
            col_idx //= 26
        return result


def create_kpi_cards(worksheet):
    """Создает экземпляр KPICards"""
    return KPICards(worksheet)