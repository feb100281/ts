# # cards/reporting/reconcile/report_builder.py
# from io import BytesIO
# from openpyxl import Workbook

# from .sheets.toc_sheet import create_toc_sheet
# from .sheets.summary_sheet import create_summary_sheet
# from .sheets.only_in_us_sheet import create_only_in_us_sheet
# from .sheets.only_in_1c_sheet import create_only_in_1c_sheet
# from .sheets.sum_diff_sheet import create_sum_diff_sheet
# from .sheets.duplicates_sheet import create_duplicates_sheet


# class ReportBuilder:
#     def __init__(self):
#         self.wb = Workbook()
#         # Удаляем дефолтный лист
#         if "Sheet" in self.wb.sheetnames:
#             self.wb.remove(self.wb["Sheet"])
    
#     def build(self, df_result, df_only_in_us, df_only_in_1c, df_sum_diff, df_duplicates, stats, total_amount_our, total_amount_1c, total_diff):
#         """Строит отчет"""
        
#         # Сводка
#         create_summary_sheet(self.wb, 1, stats, total_amount_our, total_amount_1c, total_diff)
        
#         # Только у нас
#         create_only_in_us_sheet(self.wb, 2, df_only_in_us)
        
#         # Только в 1С
#         create_only_in_1c_sheet(self.wb, 3, df_only_in_1c)
        
#         # Расхождения сумм
#         create_sum_diff_sheet(self.wb, 4, df_sum_diff)
        
#         # Дубликаты
#         create_duplicates_sheet(self.wb, 5, df_duplicates)
        
#         # Оглавление (добавляем последним, но оно будет первым)
#         sheets_info = [
#             {'number': '01', 'name': 'Сводка', 'description': 'Общая статистика по сверке и суммы'},
#             {'number': '02', 'name': 'Только у нас', 'description': 'УПД, которые есть в системе, но отсутствуют в 1С'},
#             {'number': '03', 'name': 'Только в 1С', 'description': 'УПД, которые есть в 1С, но отсутствуют в системе'},
#             {'number': '04', 'name': 'Расхождения сумм', 'description': 'УПД с несовпадающими суммами'},
#             {'number': '05', 'name': 'Дубликаты', 'description': 'Дубликаты УПД в системе'},
#         ]
#         create_toc_sheet(self.wb, sheets_info)
        
#         # Перемещаем оглавление в начало
#         self.wb.move_sheet(self.wb["TOC"], offset=-len(self.wb.sheetnames) + 1)
        
#         return self.wb
    
#     def save(self):
#         """Сохраняет в BytesIO"""
#         output = BytesIO()
#         self.wb.save(output)
#         output.seek(0)
#         return output


# def build_reconciliation_report(df_result, df_only_in_us, df_only_in_1c, df_sum_diff, df_duplicates, stats, total_amount_our, total_amount_1c, total_diff):
#     """Строит отчет по сверке"""
#     builder = ReportBuilder()
#     builder.build(df_result, df_only_in_us, df_only_in_1c, df_sum_diff, df_duplicates, stats, total_amount_our, total_amount_1c, total_diff)
#     return builder.save()



from io import BytesIO
from openpyxl import Workbook

from .sheets.toc_sheet import create_toc_sheet
from .sheets.summary_sheet import create_summary_sheet
from .sheets.only_in_us_sheet import create_only_in_us_sheet
from .sheets.only_in_1c_sheet import create_only_in_1c_sheet
from .sheets.sum_diff_sheet import create_sum_diff_sheet
from .sheets.duplicates_sheet import create_duplicates_sheet


class ReportBuilder:
    def __init__(self):
        self.wb = Workbook()
        # Удаляем дефолтный лист
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
    
    def build(self, df_result, df_only_in_us, df_only_in_1c, df_sum_diff, df_duplicates, stats, total_amount_our, total_amount_1c, total_diff):
        """Строит отчет"""
        
        # Создаем листы и запоминаем их точные имена
        sheet_names = {}
        
        # Сводка
        sheet_names['summary'] = "01_Сводка"
        create_summary_sheet(self.wb, 1, stats, total_amount_our, total_amount_1c, total_diff)
        
        # Только у нас
        sheet_names['only_in_us'] = "02_Только_у_нас"
        create_only_in_us_sheet(self.wb, 2, df_only_in_us)
        
        # Только в 1С
        sheet_names['only_in_1c'] = "03_Только_в_1С"
        create_only_in_1c_sheet(self.wb, 3, df_only_in_1c)
        
        # Расхождения сумм
        sheet_names['sum_diff'] = "04_Расхождения_сумм"
        create_sum_diff_sheet(self.wb, 4, df_sum_diff)
        
        # Дубликаты
        sheet_names['duplicates'] = "05_Дубликаты"
        create_duplicates_sheet(self.wb, 5, df_duplicates)
        
        # Оглавление (добавляем последним, но оно будет первым)
        sheets_info = [
            {
                'number': '01', 
                'name': 'Сводка', 
                'sheet_name': sheet_names['summary'],  # точное имя листа
                'description': 'Общая статистика по сверке и суммы'
            },
            {
                'number': '02', 
                'name': 'Только у нас', 
                'sheet_name': sheet_names['only_in_us'],
                'description': 'УПД, которые есть в системе, но отсутствуют в 1С'
            },
            {
                'number': '03', 
                'name': 'Только в 1С', 
                'sheet_name': sheet_names['only_in_1c'],
                'description': 'УПД, которые есть в 1С, но отсутствуют в системе'
            },
            {
                'number': '04', 
                'name': 'Расхождения сумм', 
                'sheet_name': sheet_names['sum_diff'],
                'description': 'УПД с несовпадающими суммами'
            },
            {
                'number': '05', 
                'name': 'Дубликаты', 
                'sheet_name': sheet_names['duplicates'],
                'description': 'Дубликаты УПД в системе'
            },
        ]
        create_toc_sheet(self.wb, sheets_info)
        
        # Перемещаем оглавление в начало
        self.wb.move_sheet(self.wb["TOC"], offset=-len(self.wb.sheetnames) + 1)
        
        return self.wb
    
    def save(self):
        """Сохраняет в BytesIO"""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output


def build_reconciliation_report(df_result, df_only_in_us, df_only_in_1c, df_sum_diff, df_duplicates, stats, total_amount_our, total_amount_1c, total_diff):
    """Строит отчет по сверке"""
    builder = ReportBuilder()
    builder.build(df_result, df_only_in_us, df_only_in_1c, df_sum_diff, df_duplicates, stats, total_amount_our, total_amount_1c, total_diff)
    return builder.save()