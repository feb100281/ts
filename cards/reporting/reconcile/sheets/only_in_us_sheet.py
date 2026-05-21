# cards/reporting/reconcile/sheets/only_in_us_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment
from ..components.sheet_title import create_sheet_title
from ..components.tables import create_table
from ..styles.theme import COLORS, FILLS, BORDERS
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class OnlyInUsSheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        sheet_name = f"{sheet_number:02d}_Только_у_нас"
        
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)
        
        self.title = create_sheet_title(self.ws)
        self.table = create_table(self.ws)

    def build(self, df):
        row = 1
        
        # Устанавливаем ширину колонок
        self.ws.column_dimensions['A'].width = 3   # Колонка A - узкая для отступа
        self.ws.column_dimensions['B'].width = 10  # Колонка B - ID
        self.ws.column_dimensions['C'].width = 25  # Колонка C - Номер
        self.ws.column_dimensions['D'].width = 15  # Колонка D - Дата
        self.ws.column_dimensions['E'].width = 40  # Колонка E - Контрагент
        self.ws.column_dimensions['F'].width = 22  # Колонка F - Сумма
        
        # Кнопка назад - объединяем A1:B1 (две ячейки)
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        btn_cell = self.ws.cell(row=row, column=1, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["back_text_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = FILLS.get("section", PatternFill(fill_type=None))
        thin_border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.border = thin_border
        btn_cell.hyperlink = "#'TOC'!A1"
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # Заголовок
        row = self.title.draw(
            row=row,
            title="ДОКУМЕНТЫ ТОЛЬКО В СИСТЕМЕ",
            subtitle="УПД, которые есть в системе, но отсутствуют в выгрузке из 1С",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
            start_col=2,
            end_col=8
        )
        
        if len(df) == 0:
            # Нет данных
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            cell = self.ws.cell(row=row, column=2, value="✅ Нет документов, которые есть только в системе")
            cell.font = Font(name="Roboto", size=12, color=COLORS["ok_text"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return
        
        # Таблица
        headers = ['ID', 'Номер', 'Дата', 'Контрагент', 'Сумма']
        data_rows = []
        
        for _, row_data in df.iterrows():
            # Форматируем дату
            date_val = row_data.get('date', '')
            if hasattr(date_val, 'strftime'):
                date_val = date_val.strftime('%d.%m.%Y')
            
            data_rows.append([
                row_data.get('upd_id', ''),
                row_data.get('number_our', ''),
                date_val,
                row_data.get('counterparty_our', ''),
                row_data.get('amount_our', 0)
            ])
        
        self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            money_cols=[4],  # Колонка Сумма (индекс 4, т.к. start_col=2 -> B=0, C=1, D=2, E=3, F=4)
            column_widths={'B': 10, 'C': 25, 'D': 15, 'E': 40, 'F': 22}
        )
        
        # Дополнительное форматирование для колонки с суммами
        for r in range(row, row + len(data_rows)):
            cell = self.ws.cell(row=r, column=6)  # Колонка F (индекс 6)
            cell.number_format = '#,##0.00 ₽'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Настройки
        self.ws.sheet_view.showGridLines = False
        self.ws.freeze_panes = 'A2'


def create_only_in_us_sheet(workbook, sheet_number, df):
    """Создает лист с документами только в системе"""
    sheet = OnlyInUsSheet(workbook, sheet_number)
    sheet.build(df)
    return sheet.ws