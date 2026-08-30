# cards/reporting/reconcile/sheets/sum_diff_sheet.py

from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from ..components.sheet_title import create_sheet_title
from ..components.tables import create_table
from ..styles.theme import COLORS, FILLS, BORDERS
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class SumDiffSheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        sheet_name = f"{sheet_number:02d}_Расхождения_сумм"
        
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)
        
        self.title = create_sheet_title(self.ws)
        self.table = create_table(self.ws)

    def build(self, df):
        row = 1
        
        # Устанавливаем ширину колонок
        self.ws.column_dimensions['A'].width = 3   # Колонка A - узкая для отступа
        self.ws.column_dimensions['B'].width = 25  # Колонка B - Номер
        self.ws.column_dimensions['C'].width = 15  # Колонка C - Дата
        self.ws.column_dimensions['D'].width = 35  # Колонка D - Контрагент (система)
        self.ws.column_dimensions['E'].width = 35  # Колонка E - Контрагент (1С)
        self.ws.column_dimensions['F'].width = 20  # Колонка F - Сумма система
        self.ws.column_dimensions['G'].width = 20  # Колонка G - Сумма 1С
        self.ws.column_dimensions['H'].width = 20  # Колонка H - Разница
        
        # Кнопка назад - объединяем A1:B1 (две ячейки)
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        btn_cell = self.ws.cell(row=row, column=1, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["back_text_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = FILLS.get("section", PatternFill(fill_type=None))
        thin_border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.border = thin_border
        btn_cell.hyperlink = "#'TOC'!A1"
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # Заголовок
        row = self.title.draw(
            row=row,
            title="ДОКУМЕНТЫ С РАСХОЖДЕНИЕМ СУММ",
            subtitle="УПД, которые есть и в системе, и в 1С, но суммы не совпадают",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
            start_col=2,
            end_col=10
        )
        
        if len(df) == 0:
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            cell = self.ws.cell(row=row, column=2, value="✅ Нет документов с расхождением сумм")
            cell.font = Font(name="Roboto", size=12, color=COLORS["ok_text"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return
        
        # Таблица
        headers = ['Номер', 'Дата', 'Контрагент (система)', 'Контрагент (1С)', 'Сумма система', 'Сумма 1С', 'Разница']
        data_rows = []
        
        for _, row_data in df.iterrows():
            # Форматируем дату
            date_val = row_data.get('date', '')
            if hasattr(date_val, 'strftime'):
                date_val = date_val.strftime('%d.%m.%Y')
            
            data_rows.append([
                row_data.get('number_our', ''),
                date_val,
                row_data.get('counterparty_our', ''),
                row_data.get('counterparty_1c', ''),
                row_data.get('amount_our', 0),
                row_data.get('amount_1c', 0),
                row_data.get('diff', 0)
            ])
        
        self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            money_cols=[4, 5, 6],  # Колонки Сумма система, Сумма 1С, Разница
            column_widths={'B': 25, 'C': 15, 'D': 35, 'E': 35, 'F': 20, 'G': 20, 'H': 20}
        )
        
        # Дополнительное форматирование для колонок с суммами
        for r in range(row, row + len(data_rows)):
            # Сумма система (колонка F, индекс 6)
            cell_system = self.ws.cell(row=r, column=6)
            cell_system.number_format = '#,##0.00 ₽'
            cell_system.alignment = Alignment(horizontal="right", vertical="center")
            
            # Сумма 1С (колонка G, индекс 7)
            cell_1c = self.ws.cell(row=r, column=7)
            cell_1c.number_format = '#,##0.00 ₽'
            cell_1c.alignment = Alignment(horizontal="right", vertical="center")
            
            # Разница (колонка H, индекс 8)
            cell_diff = self.ws.cell(row=r, column=8)
            cell_diff.number_format = '#,##0.00 ₽'
            cell_diff.alignment = Alignment(horizontal="right", vertical="center")
            
            # Подсветка разницы: красным если не ноль
            diff_value = row_data.get('diff', 0)
            if abs(diff_value) > 1:
                cell_diff.font = Font(name="Roboto", size=10, bold=True, color=COLORS["error_text"])
                cell_diff.fill = FILLS.get("error", PatternFill(fill_type=None))
        
        # Настройки
        self.ws.sheet_view.showGridLines = False
        self.ws.freeze_panes = 'A2'


def create_sum_diff_sheet(workbook, sheet_number, df):
    """Создает лист с расхождениями сумм"""
    sheet = SumDiffSheet(workbook, sheet_number)
    sheet.build(df)
    return sheet.ws