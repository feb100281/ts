# cards/reporting/reconcile/sheets/summary_sheet.py

from datetime import datetime
from openpyxl.styles import Font, Alignment
from ..components.sheet_title import create_sheet_title
from ..components.tables import create_table
from ..components.kpi_cards import create_kpi_cards
from ..styles.theme import COLORS, FILLS, BORDERS
from ..styles.helpers import draw_toc_button, apply_money
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class SummarySheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        sheet_name = f"{sheet_number:02d}_Сводка"
        
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)
        
        self.title = create_sheet_title(self.ws)
        self.kpi = create_kpi_cards(self.ws)
        self.table = create_table(self.ws)

    def build(self, stats, total_amount_our, total_amount_1c, total_diff):
        row = 1
        
        # Устанавливаем ширину колонок
        self.ws.column_dimensions['A'].width = 3   # Колонка A - узкая для отступа
        self.ws.column_dimensions['B'].width = 18  # Колонка B для карточек
        self.ws.column_dimensions['C'].width = 18  # Колонка C для карточек
        self.ws.column_dimensions['D'].width = 18  # Колонка D для карточек
        self.ws.column_dimensions['E'].width = 18  # Колонка E для карточек
        self.ws.column_dimensions['F'].width = 18  # Колонка F для карточек
        self.ws.column_dimensions['G'].width = 18  # Колонка G для карточек
        self.ws.column_dimensions['H'].width = 35  # Колонка H для заголовка таблицы
        self.ws.column_dimensions['I'].width = 30  # Колонка I для сумм
        
        # Кнопка назад - объединяем A1:B1 (две ячейки)
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        btn_cell = self.ws.cell(row=row, column=1, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["back_text_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = FILLS.get("section", PatternFill(fill_type=None))
        btn_cell.border = BORDERS["thin"]
        btn_cell.hyperlink = "#'TOC'!A1"
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # Заголовок
        row = self.title.draw(
            row=row,
            title="СВОДКА ПО СВЕРКЕ",
            subtitle="Результаты сверки документов УПД с выгрузкой из 1С",
            date_text=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
            start_col=2,
            end_col=8
        )
        
        # KPI карточки - все одинаковой ширины
        cards = [
            {'title': 'ВСЕГО УПД В СИСТЕМЕ', 'value': f"{stats.get('ONLY_IN_US', 0) + stats.get('OK', 0):,}", 
             'subtitle': 'документов', 'color': COLORS["dark_green"], 'width': 2},
            {'title': 'ВСЕГО УПД В 1С', 'value': f"{stats.get('ONLY_IN_1C', 0) + stats.get('OK', 0):,}", 
             'subtitle': 'документов', 'color': COLORS["dark_green"], 'width': 2},
            {'title': 'СОВПАДАЕТ', 'value': f"{stats.get('OK', 0):,}", 
             'subtitle': 'документов', 'color': COLORS["ok_text"], 'width': 2},
            {'title': 'РАСХОДИТСЯ СУММА', 'value': f"{stats.get('SUM_DIFF', 0):,}", 
             'subtitle': 'документов', 'color': COLORS["warning_text"], 'width': 2},
            {'title': 'ТОЛЬКО У НАС', 'value': f"{stats.get('ONLY_IN_US', 0):,}", 
             'subtitle': 'документов', 'color': COLORS["info_text"], 'width': 2},
            {'title': 'ТОЛЬКО В 1С', 'value': f"{stats.get('ONLY_IN_1C', 0):,}", 
             'subtitle': 'документов', 'color': COLORS["error_text"], 'width': 2},
        ]
        
        row = self.kpi.draw_row(row, cards)
        row += 2
        
        # Таблица с суммами
        headers = ['Показатель', 'Сумма']
        data_rows = [
            ['Сумма в системе', total_amount_our],
            ['Сумма в 1С', total_amount_1c],
            ['Разница', total_diff],
        ]
        
        # Рисуем таблицу, но без автоматической установки ширины
        # Используем готовые настройки column_widths
        row = self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=8,  # Начинаем с колонки H (индекс 8)
            money_cols=[1],
            column_widths=None  # Не устанавливаем через таблицу, у нас уже есть ширина
        )
        
        # Дополнительное форматирование для колонки с суммами
        for r in range(row - len(data_rows), row):
            cell = self.ws.cell(row=r, column=9)  # Колонка I (индекс 9)
            cell.number_format = '#,##0.00 ₽'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Настройки
        self.ws.sheet_view.showGridLines = False
        self.ws.freeze_panes = 'A2'


def create_summary_sheet(workbook, sheet_number, stats, total_amount_our, total_amount_1c, total_diff):
    """Создает лист со сводкой"""
    sheet = SummarySheet(workbook, sheet_number)
    sheet.build(stats, total_amount_our, total_amount_1c, total_diff)
    return sheet.ws