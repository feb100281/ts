# cards/reporting/registry_exporter.py

# cards/reporting/registry_exporter.py

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Sum, Count
from ..models import UpdDocument


def generate_registry_excel(upd_ids):
    """
    Генерирует реестр документов в формате Excel
    """
    
    # Получаем данные с агрегацией
    queryset = UpdDocument.objects.filter(
        id__in=upd_ids
    ).select_related(
        'counterparty', 'lot'
    ).annotate(
        total_amount_vatadd=Sum('income_lines__upd_amount_vatadd'),
        total_qty=Sum('income_lines__upd_qty'),
        lines_count=Count('income_lines')
    ).order_by('date', 'number')
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр документов"
    
    # Определяем стили
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Заголовки
    headers = [
        ('№ п/п', 8),
        ('Контрагент', 35),
        ('Номер УПД', 20),
        ('Дата УПД', 15),
        ('ID УПД', 10),
        ('Сумма с НДС', 18),
        ('Кол-во товаров', 15),
        ('Кол-во строк в УПД', 18),
    ]
    
    # Записываем заголовки
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Записываем данные
    total_sum = 0
    total_qty = 0
    total_lines = 0
    
    for idx, doc in enumerate(queryset, 1):
        row = idx + 1
        
        # Контрагент
        counterparty_name = str(doc.counterparty) if doc.counterparty else '-'
        if ' (ИНН:' in counterparty_name:
            counterparty_name = counterparty_name.split(' (ИНН:')[0]
        
        # Сумма
        amount = doc.total_amount_vatadd or 0
        total_sum += amount
        
        # Количество
        qty = doc.total_qty or 0
        total_qty += qty
        
        # Количество строк
        lines = doc.lines_count or 0
        total_lines += lines
        
        # Записываем ячейки
        ws.cell(row=row, column=1, value=idx).alignment = center_alignment
        ws.cell(row=row, column=1).font = cell_font
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=counterparty_name).alignment = cell_alignment
        ws.cell(row=row, column=2).font = cell_font
        ws.cell(row=row, column=2).border = border
        
        ws.cell(row=row, column=3, value=doc.number).alignment = center_alignment
        ws.cell(row=row, column=3).font = cell_font
        ws.cell(row=row, column=3).border = border
        
        ws.cell(row=row, column=4, value=doc.date.strftime('%d.%m.%Y')).alignment = center_alignment
        ws.cell(row=row, column=4).font = cell_font
        ws.cell(row=row, column=4).border = border
        
        ws.cell(row=row, column=5, value=doc.id).alignment = center_alignment
        ws.cell(row=row, column=5).font = cell_font
        ws.cell(row=row, column=5).border = border
        
        ws.cell(row=row, column=6, value=round(amount, 2)).alignment = number_alignment
        ws.cell(row=row, column=6).font = cell_font
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).number_format = '#,##0.00 "₽"'
        
        ws.cell(row=row, column=7, value=round(qty, 2)).alignment = number_alignment
        ws.cell(row=row, column=7).font = cell_font
        ws.cell(row=row, column=7).border = border
        
        ws.cell(row=row, column=8, value=lines).alignment = number_alignment
        ws.cell(row=row, column=8).font = cell_font
        ws.cell(row=row, column=8).border = border
    
    # Добавляем строку итогов
    footer_row = len(queryset) + 2
    
    # Объединяем ячейки для текста "ИТОГО"
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
    ws.cell(row=footer_row, column=1, value='ИТОГО:').font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=footer_row, column=1).border = border
    
    # Итоговые суммы
    ws.cell(row=footer_row, column=6, value=round(total_sum, 2)).font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=footer_row, column=6).alignment = number_alignment
    ws.cell(row=footer_row, column=6).border = border
    ws.cell(row=footer_row, column=6).number_format = '#,##0.00 "₽"'
    
    ws.cell(row=footer_row, column=7, value=round(total_qty, 2)).font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=footer_row, column=7).alignment = number_alignment
    ws.cell(row=footer_row, column=7).border = border
    
    ws.cell(row=footer_row, column=8, value=total_lines).font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=footer_row, column=8).alignment = number_alignment
    ws.cell(row=footer_row, column=8).border = border
    
    # Добавляем информацию о дате формирования
    info_row = footer_row + 2
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=8)
    ws.cell(
        row=info_row, 
        column=1, 
        value=f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")} | Всего документов: {len(queryset)}'
    ).font = Font(name='Arial', size=9, italic=True)
    ws.cell(row=info_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_registry_response(upd_ids, format_type='excel'):
    """
    Генерирует HTTP ответ с реестром документов
    """
    
    if format_type == 'pdf':
        from .registry_pdf import generate_registry_pdf
        return generate_registry_pdf(upd_ids)
    
    # По умолчанию Excel
    output = generate_registry_excel(upd_ids)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Реестр_УПД_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    return response