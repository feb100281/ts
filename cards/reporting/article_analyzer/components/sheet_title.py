from openpyxl.styles import Font, Alignment
from ..styles.theme import COLORS, ALIGNMENTS, FILLS, BORDERS


class SheetTitle:
    def __init__(self, ws):
        self.ws = ws
    
    def draw(self, row, title, subtitle, start_col=2, end_col=8):
        """Рисует заголовок листа"""
        # Заголовок
        self.ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col
        )
        title_cell = self.ws.cell(row=row, column=start_col, value=title)
        title_cell.font = Font(name="Roboto", size=14, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 30
        row += 1
        
        # Подзаголовок
        self.ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col
        )
        subtitle_cell = self.ws.cell(row=row, column=start_col, value=subtitle)
        subtitle_cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
        subtitle_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 20
        row += 1
        
        return row


def create_sheet_title(ws):
    """Создает объект заголовка листа"""
    return SheetTitle(ws)