# cards/reporting/article_analyzer/sheets/summary_sheet.py
from datetime import datetime
from ..components.sheet_title import create_sheet_title
from ..components.tables import create_table
from ..styles.theme import COLORS, FILLS, ALIGNMENTS, FONTS, BORDERS
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class SummarySheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        sheet_name = f"{sheet_number:02d}_Сводка_по_артиклям"
        
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)
        
        self.title = create_sheet_title(self.ws)
        self.table = create_table(self.ws)
    
    def build(self, df, articles_not_found):
        row = 1
        
        # Устанавливаем ширину колонок
        col_widths = {
            'A': 3,   # Отступ
            'B': 20,  # Артикль
            'C': 12,  # Статус
            'D': 20,  # Кол-во позиций в УПД
            'E': 15,  # Общее кол-во товара
            'F': 20,  # Общая сумма
            'G': 18,  # Мин. цена (без НДС)
            'H': 18,  # Макс. цена (без НДС)
            'I': 18,  # Средняя цена (без НДС)
            'J': 18,  # Медиана (без НДС)
        }
        
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Кнопка назад
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        btn_cell = self.ws.cell(row=row, column=1, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["back_text_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = FILLS.get("section", PatternFill(fill_type=None))
        thin_border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.border = thin_border
        btn_cell.hyperlink = "#'TOC'!A1"
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # Заголовок
        row = self.title.draw(
            row=row,
            title="СВОДНЫЙ АНАЛИЗ ПО АРТИКЛЯМ",
            subtitle="Агрегированная статистика по каждому артиклю (цены указаны без НДС)",
            start_col=2,
            end_col=10
        )
        row += 1
        
        # Информационная панель
        info_row = row
        total_found = len(df) if not df.empty else 0
        
        self.ws.merge_cells(start_row=info_row, start_column=2, end_row=info_row, end_column=3)
        found_cell = self.ws.cell(row=info_row, column=2, value=f"Найдено артиклей: {total_found}")
        found_cell.font = FONTS["found"]
        found_cell.fill = FILLS["found"]
        found_cell.alignment = ALIGNMENTS["left"]
   
        
        self.ws.merge_cells(start_row=info_row, start_column=4, end_row=info_row, end_column=5)
        not_found_cell = self.ws.cell(row=info_row, column=4, value=f"Не найдено артиклей: {len(articles_not_found)}")
        not_found_cell.font = FONTS["not_found"]
        not_found_cell.fill = FILLS["not_found"]
        not_found_cell.alignment = ALIGNMENTS["left"]
  
        
        self.ws.row_dimensions[info_row].height = 30
        row += 2
        
        if df.empty:
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
            cell = self.ws.cell(row=row, column=2, value="Нет данных по указанным артиклям")
            cell.font = Font(name="Roboto", size=12, color=COLORS["not_found_text"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return
        
        # Таблица сводки с пояснениями в заголовках
        headers = [
            'Артикль', 
            'Статус', 
            'Позиций\n(в УПД)', 
            'Кол-во\n(ед.)', 
            'Сумма\n(с НДС)', 
            'Мин. цена', 
            'Макс. цена', 
            'Ср. цена', 
            'Медиана'
        ]
        data_rows = []
        
        for _, row_data in df.iterrows():
            status = "Найден" if row_data.get('Найден в системе') == 'Да' else "Не найден"
            data_rows.append([
                row_data.get('Артикль', ''),
                status,
                row_data.get('Кол-во позиций', 0),
                row_data.get('Общее кол-во товара', 0),
                row_data.get('Общая стоимость (с НДС)', 0),
                row_data.get('Мин. цена', '-'),
                row_data.get('Макс. цена', '-'),
                row_data.get('Средняя цена', '-'),
                row_data.get('Медианная цена', '-'),
            ])
        
        self.table.draw(
            start_row=row,
            headers=headers,
            data_rows=data_rows,
            start_col=2,
            money_cols=[4, 5, 6, 7, 8],  # Сумма, мин/макс/ср/медиана
            center_cols=[2, 3],  # Статус, Позиций
        )
        
        # Добавляем примечание внизу
        last_row = row + len(data_rows) + 1
        note_cell = self.ws.cell(row=last_row, column=2, value="Примечание: цены указаны без НДС")
        note_cell.font = Font(name="Roboto", size=8, italic=True, color=COLORS["text_gray"])
        note_cell.alignment = ALIGNMENTS["left"]
        
        # Настройки
        self.ws.sheet_view.showGridLines = False
        # Заморозка: строка 4 (заголовки таблицы) и колонка B (Артикль)
        self.ws.freeze_panes = 'C9'  

def create_summary_sheet(workbook, sheet_number, df, articles_not_found):
    """Создает лист со сводкой по артиклям"""
    sheet = SummarySheet(workbook, sheet_number)
    sheet.build(df, articles_not_found)
    return sheet.ws