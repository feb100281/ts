# cards/reporting/article_analyzer/sheets/not_found_sheet.py
from ..components.sheet_title import create_sheet_title
from ..styles.theme import COLORS, FILLS, ALIGNMENTS, FONTS, BORDERS
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class NotFoundSheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        sheet_name = f"{sheet_number:02d}_Артикли_не_найдены"

        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)

        self.title = create_sheet_title(self.ws)

    def _thin_border(self):
        return Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"]),
        )

    def _draw_back_button(self, row):
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        btn_cell = self.ws.cell(row=row, column=1, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(
            name="Roboto",
            size=9,
            bold=True,
            color=COLORS["back_text_green"]
        )
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = FILLS.get("section", PatternFill(fill_type=None))
        btn_cell.border = self._thin_border()
        btn_cell.hyperlink = "#'TOC'!A1"

        self.ws.row_dimensions[row].height = 24

    def _draw_success_block(self, row):
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=4)

        cell = self.ws.cell(row=row, column=2, value="✅ Все артикли найдены в системе")
        cell.font = Font(
            name="Roboto",
            size=13,
            bold=True,
            color=COLORS["found_text"]
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.fill = FILLS.get("alt", PatternFill(fill_type=None))
        cell.border = self._thin_border()

        for r in range(row, row + 2):
            self.ws.row_dimensions[r].height = 28
            for col in range(2, 5):
                self.ws.cell(row=r, column=col).border = self._thin_border()

    def _draw_table_headers(self, row):
        headers = {
            2: "№",
            3: "Артикль",
            4: "Статус",
        }

        for col_idx, header in headers.items():
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.fill = FILLS["header"]
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 34

    def build(self, articles_not_found):
        row = 1

        # Ширина колонок
        col_widths = {
            "A": 3,
            "B": 8,    # №
            "C": 34,   # Артикль
            "D": 28,   # Статус
        }

        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

        # Кнопка назад
        self._draw_back_button(row)
        row += 2

        # Заголовок
        row = self.title.draw(
            row=row,
            title="АРТИКЛИ, НЕ НАЙДЕННЫЕ В СИСТЕМЕ",
            subtitle="Артикли из загруженного файла, которые отсутствуют в УПД",
            start_col=2,
            end_col=4
        )
        row += 1

        if not articles_not_found:
            self._draw_success_block(row)
            self.ws.sheet_view.showGridLines = False
            return

        # Заголовки таблицы
        self._draw_table_headers(row)
        row += 1

        # Данные
        for idx, article in enumerate(articles_not_found, start=1):
            current_row = row + idx - 1
            is_alt = idx % 2 == 1
            fill = FILLS["alt"] if is_alt else FILLS["none"]

            # №
            num_cell = self.ws.cell(row=current_row, column=2, value=idx)
            num_cell.font = Font(
                name="Roboto",
                size=9,
                bold=True,
                color=COLORS.get("text_gray", "666666")
            )
            num_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            num_cell.border = BORDERS["thin"]
            num_cell.fill = fill

            # Артикль
            art_cell = self.ws.cell(row=current_row, column=3, value=article)
            art_cell.font = Font(
                name="Roboto",
                size=10,
                bold=True,
                color=COLORS.get("not_found_text", "C00000")
            )
            art_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )
            art_cell.border = BORDERS["thin"]
            art_cell.fill = fill

            # Статус
            status_cell = self.ws.cell(row=current_row, column=4, value="Не найден в УПД")
            status_cell.font = Font(
                name="Roboto",
                size=9,
                italic=True,
                color=COLORS.get("text_gray", "808080")
            )
            status_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            status_cell.border = BORDERS["thin"]
            status_cell.fill = fill

            self.ws.row_dimensions[current_row].height = 26

        # Примечание
        note_row = row + len(articles_not_found) + 1
        self.ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=4)

        note_cell = self.ws.cell(
            row=note_row,
            column=2,
            value="Примечание: данные артикли присутствуют в загруженном файле, но не найдены среди позиций УПД."
        )
        note_cell.font = Font(
            name="Roboto",
            size=8,
            italic=True,
            color=COLORS.get("text_gray", "808080")
        )
        note_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        self.ws.row_dimensions[note_row].height = 24

        # Настройки листа
        self.ws.sheet_view.showGridLines = False
        self.ws.freeze_panes = "B7"


def create_not_found_sheet(workbook, sheet_number, articles_not_found):
    """Создает лист с ненайденными артиклями"""
    sheet = NotFoundSheet(workbook, sheet_number)
    sheet.build(articles_not_found)
    return sheet.ws