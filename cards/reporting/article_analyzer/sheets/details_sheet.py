# cards/reporting/article_analyzer/sheets/details_sheet.py
import pandas as pd
from datetime import datetime
from ..components.sheet_title import create_sheet_title
from ..components.tables import create_table
from ..styles.theme import COLORS, FILLS, ALIGNMENTS, FONTS, BORDERS
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class DetailsSheet:
    def __init__(self, workbook, sheet_number):
        self.wb = workbook
        self.sheet_number = sheet_number
        sheet_name = f"{sheet_number:02d}_Детали_по_позициям"
        
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.create_sheet(sheet_name)
        
        self.title = create_sheet_title(self.ws)
        self.table = create_table(self.ws)
    
    def _extract_date(self, date_val):
        """Извлекает дату для сортировки"""
        if date_val is None or pd.isna(date_val):
            return datetime.min
        if hasattr(date_val, 'strftime'):
            return date_val
        try:
            return datetime.strptime(str(date_val), '%d.%m.%Y')
        except:
            return datetime.min
    
    def _group_upd_documents(self, df):
        """Группирует УПД по артиклю с сортировкой по дате и объединением дубликатов"""
        grouped_data = []
        
        for article in df['Артикул (из УПД)'].unique():
            article_data = df[df['Артикул (из УПД)'] == article]
            
            # Собираем уникальные комбинации (УПД + дата + цена + контрагент)
            unique_positions = {}
            
            for _, row in article_data.iterrows():
                upd_val = row.get('УПД', '')
                date_val = row.get('Дата УПД', '')
                price = row.get('Цена без НДС', 0)
                counterparty = row.get('Контрагент', '')  # Берём контрагента из DataFrame
                
                if upd_val and pd.notna(upd_val):
                    # Форматируем дату
                    if date_val and pd.notna(date_val):
                        if hasattr(date_val, 'strftime'):
                            date_obj = date_val
                            date_str = date_val.strftime('%d.%m.%Y')
                        else:
                            date_obj = self._extract_date(date_val)
                            date_str = str(date_val)
                    else:
                        date_obj = datetime.min
                        date_str = ''
                    
                    # Ключ для уникальности (УПД + дата + цена + контрагент)
                    key = (upd_val, date_str, price, counterparty)
                    
                    if key not in unique_positions:
                        unique_positions[key] = {
                            'upd': upd_val,
                            'date_obj': date_obj,
                            'date_str': date_str,
                            'price': price,
                            'counterparty': counterparty,
                        }
            
            # Сортируем по дате
            sorted_positions = sorted(unique_positions.values(), key=lambda x: x['date_obj'])
            
            # Формируем списки с красивым форматированием
            upd_list = []
            price_list = []
            for pos in sorted_positions:
                # Форматируем строку УПД с контрагентом
                if pos['counterparty']:
                    upd_str = f"УПД №{pos['upd']} • {pos['date_str']} • {pos['counterparty']}"
                else:
                    upd_str = f"УПД № {pos['upd']} от {pos['date_str']}"
                upd_list.append(upd_str)
                price_list.append(pos['price'])
            
            # Общие суммы
            total_qty = article_data['Количество'].sum()
            total_amount = article_data['Стоимость с НДС'].sum()
            
            grouped_data.append({
                'Артикль': article,
                'Название': article_data['Название из УПД'].iloc[0] if not article_data.empty else '',
                'УПД': upd_list,
                'Цены': price_list,
                'Количество': total_qty,
                'Сумма (с НДС)': total_amount,
                'Кол-во УПД': len(upd_list),
            })
        
        return pd.DataFrame(grouped_data)
    
    def _format_upd_with_prices(self, upd_list, price_list):
        """Форматирует УПД с соответствующими ценами"""
        if not upd_list:
            return "-", "-"
        
        upd_lines = []
        price_lines = []
        
        for i, upd_str in enumerate(upd_list):
            upd_lines.append(f"• {upd_str}")
            price = price_list[i] if i < len(price_list) else 0
            price_lines.append(f"• {price:,.2f} ₽")
        
        return "\n".join(upd_lines), "\n".join(price_lines)
    
    def build(self, df):
        row = 1
        
        # Устанавливаем ширину колонок
        col_widths = {
            'A': 3,   # Отступ
            'B': 20,  # Артикль
            'C': 45,  # Название
            'D': 55,  # УПД (увеличили для длинных названий)
            'E': 25,  # Цены
            'F': 15,  # Кол-во
            'G': 20,  # Сумма
            'H': 12,  # Кол-во УПД
        }
        
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Кнопка назад
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        btn_cell = self.ws.cell(row=row, column=1, value="←  ОГЛАВЛЕНИЕ")
        btn_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["back_text_green"])
        btn_cell.alignment = Alignment(horizontal="left", vertical="center")
        btn_cell.fill = FILLS.get("section", PatternFill(fill_type=None))
        thin_border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        btn_cell.border = thin_border
        btn_cell.hyperlink = "#'TOC'!A1"
        self.ws.row_dimensions[row].height = 24
        row += 2
        
        # Заголовок
        row = self.title.draw(
            row=row,
            title="ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО АРТИКЛЯМ",
            subtitle="Сгруппировано по артиклям (цены указаны без НДС, УПД отсортированы по дате)",
            start_col=2,
            end_col=8
        )
        row += 1
        
        if df.empty:
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            cell = self.ws.cell(row=row, column=2, value="Нет данных по указанным артиклям")
            cell.font = Font(name="Roboto", size=12, color=COLORS["not_found_text"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return
        
        # Группируем данные
        grouped_df = self._group_upd_documents(df)
        
        # Заголовки таблицы
        headers = ['Артикль', 'Название', 'УПД', 'Цены\n(без НДС)', 'Кол-во\n(ед.)', 'Сумма\n(с НДС)', 'Кол-во\nУПД']
        
        # Рисуем заголовки
        for col_idx, header in enumerate(headers, start=2):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.alignment = ALIGNMENTS["center_wrap"]
            cell.fill = FILLS["header"]
            cell.border = BORDERS["thin"]
        
        self.ws.row_dimensions[row].height = 40
        row += 1
        
        # Данные
        for idx, (_, row_data) in enumerate(grouped_df.iterrows()):
            upd_list = row_data.get('УПД', [])
            price_list = row_data.get('Цены', [])
            
            # Форматируем УПД и цены
            upd_text, price_text = self._format_upd_with_prices(upd_list, price_list)
            
            # Артикль
            art_cell = self.ws.cell(row=row, column=2, value=row_data.get('Артикль', ''))
            art_cell.font = FONTS["bold"]
            art_cell.alignment = Alignment(horizontal="left", vertical="center")
            art_cell.border = BORDERS["thin"]
            art_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # Название
            name_cell = self.ws.cell(row=row, column=3, value=row_data.get('Название', ''))
            name_cell.font = FONTS["normal"]
            name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            name_cell.border = BORDERS["thin"]
            name_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # УПД
            upd_cell = self.ws.cell(row=row, column=4, value=upd_text)
            upd_cell.font = Font(name="Roboto", size=9)
            upd_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            upd_cell.border = BORDERS["thin"]
            upd_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # Цены
            price_cell = self.ws.cell(row=row, column=5, value=price_text)
            price_cell.font = Font(name="Roboto", size=9)
            price_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            price_cell.border = BORDERS["thin"]
            price_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # Количество
            qty_cell = self.ws.cell(row=row, column=6, value=row_data.get('Количество', 0))
            qty_cell.font = FONTS["bold"]
            qty_cell.alignment = Alignment(horizontal="right", vertical="center")
            qty_cell.number_format = '#,##0'
            qty_cell.border = BORDERS["thin"]
            qty_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # Сумма
            amount_cell = self.ws.cell(row=row, column=7, value=row_data.get('Сумма (с НДС)', 0))
            amount_cell.font = FONTS["bold"]
            amount_cell.alignment = Alignment(horizontal="right", vertical="center")
            amount_cell.number_format = '#,##0.00 ₽'
            amount_cell.border = BORDERS["thin"]
            amount_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # Кол-во УПД
            count_cell = self.ws.cell(row=row, column=8, value=row_data.get('Кол-во УПД', 0))
            count_cell.font = FONTS["normal"]
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.border = BORDERS["thin"]
            count_cell.fill = FILLS["alt"] if idx % 2 == 0 else FILLS["none"]
            
            # Высота строки
            max_lines = len(upd_list)
            row_height = max(25, min(120, 16 * max_lines + 10))
            self.ws.row_dimensions[row].height = row_height
            
            row += 1
        
        # Добавляем примечание
        note_row = row + 1
        note_cell = self.ws.cell(row=note_row, column=2, value="Примечание: цены указаны без НДС, УПД отсортированы по дате. Каждая цена соответствует УПД на той же строке.")
        note_cell.font = Font(name="Roboto", size=8, italic=True, color=COLORS["text_gray"])
        note_cell.alignment = ALIGNMENTS["left"]
        
        # Настройки
        self.ws.sheet_view.showGridLines = False
        self.ws.freeze_panes = 'C7'


def create_details_sheet(workbook, sheet_number, df):
    """Создает лист с детальной информацией"""
    sheet = DetailsSheet(workbook, sheet_number)
    sheet.build(df)
    return sheet.ws