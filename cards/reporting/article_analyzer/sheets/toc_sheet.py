# cards/reporting/article_analyzer/sheets/toc_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles.theme import COLORS, BORDERS, ALIGNMENTS, FILLS, FONTS


class TOCSheet:
    def __init__(self, workbook):
        self.wb = workbook
        if "TOC" in self.wb.sheetnames:
            self.ws = self.wb["TOC"]
        else:
            self.ws = self.wb.create_sheet("TOC")
        # Перемещаем TOC на первый лист
        self.wb.move_sheet(self.ws, offset=-self.wb.sheetnames.index(self.ws.title))

    def build(self, sheets_info, total_articles=0, found_count=0, not_found_count=0):
        row = 1
        start_col = 2
        table_start_col = 2
        table_end_col = 4
        
        # ============================================================
        # ЗАГОЛОВОЧНЫЙ БЛОК
        # ============================================================
        # Главный заголовок
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        title_cell = self.ws.cell(row=row, column=table_start_col, value="ОТЧЕТ ПО АНАЛИЗУ АРТИКЛЕЙ")
        title_cell.font = Font(name="Roboto", size=18, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 35
        row += 1

        # Подзаголовок
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        subtitle_cell = self.ws.cell(row=row, column=table_start_col, value="Анализ по загруженным артиклям из Excel")
        subtitle_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        subtitle_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 22
        row += 1

        # Дата формирования
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        date_cell = self.ws.cell(row=row, column=table_start_col, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}")
        date_cell.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
        date_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 20
        row += 2

        # ============================================================
        # БЛОК НАВИГАЦИИ
        # ============================================================
        nav_cell = self.ws.cell(row=row, column=table_start_col, value="СОДЕРЖАНИЕ ОТЧЕТА")
        nav_cell.font = Font(name="Roboto", size=13, bold=True, color=COLORS["dark_green"])
        nav_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 28
        row += 1

        # ============================================================
        # ТАБЛИЦА ОГЛАВЛЕНИЯ
        # ============================================================
        headers = ["№", "РАЗДЕЛ", "ОПИСАНИЕ"]
        
        # Заголовки таблицы
        for col_idx, header in enumerate(headers, start=table_start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header_white"]
            cell.alignment = ALIGNMENTS["center"]
            cell.fill = FILLS["header"]
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 32
        row += 1
        table_data_start_row = row

        # Данные таблицы
        for idx, sheet in enumerate(sheets_info):
            # Номер
            cell_num = self.ws.cell(row=row, column=table_start_col, value=sheet['number'])
            cell_num.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
            cell_num.alignment = ALIGNMENTS["center"]
            cell_num.border = BORDERS["thin"]
            cell_num.fill = FILLS["section"]

            # Название с гиперссылкой
            cell_name = self.ws.cell(row=row, column=table_start_col + 1, value=sheet['name'])
            cell_name.font = Font(name="Roboto", size=10, bold=True, color="0066CC")
            cell_name.alignment = ALIGNMENTS["left"]
            cell_name.border = BORDERS["thin"]
            cell_name.hyperlink = f"#'{sheet['sheet_name']}'!A1"

            # Описание
            cell_desc = self.ws.cell(row=row, column=table_start_col + 2, value=sheet.get("description", ""))
            cell_desc.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            cell_desc.alignment = ALIGNMENTS["left_wrap"]
            cell_desc.border = BORDERS["thin"]

            # Чередование цвета строк
            if idx % 2 == 0:
                row_fill = FILLS["alt"]
            else:
                row_fill = FILLS["none"]
            
            cell_name.fill = row_fill
            cell_desc.fill = row_fill

            self.ws.row_dimensions[row].height = 28
            row += 1

        row += 1

        # ============================================================
        # СТАТИСТИКА ЗАПРОСА (аккуратная, как в сводке)
        # ============================================================
        # Заголовок блока статистики
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_end_col)
        stats_title = self.ws.cell(row=row, column=table_start_col, value="СТАТИСТИКА ЗАПРОСА")
        stats_title.font = Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"])
        stats_title.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 25
        row += 1
        
        # Блок статистики с светло-зеленой заливкой (как в сводке)
        stats_start_row = row
        
        # Строка 1: Всего артиклей
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_start_col + 1)
        total_label = self.ws.cell(row=row, column=table_start_col, value="Всего артиклей в запросе")
        total_label.font = Font(name="Roboto", size=10, bold=True)
        total_label.alignment = ALIGNMENTS["left"]
        total_label.fill = FILLS["section"]

        
        self.ws.merge_cells(start_row=row, start_column=table_start_col + 2, end_row=row, end_column=table_end_col)
        total_value = self.ws.cell(row=row, column=table_start_col + 2, value=total_articles)
        total_value.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        total_value.alignment = ALIGNMENTS["left"]
        total_value.fill = FILLS["section"]
  
        row += 1
        
        # Строка 2: Найдено в системе
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_start_col + 1)
        found_label = self.ws.cell(row=row, column=table_start_col, value="Найдено в системе")
        found_label.font = Font(name="Roboto", size=10, bold=True)
        found_label.alignment = ALIGNMENTS["left"]
        found_label.fill = FILLS["found"]

        
        self.ws.merge_cells(start_row=row, start_column=table_start_col + 2, end_row=row, end_column=table_end_col)
        found_value = self.ws.cell(row=row, column=table_start_col + 2, value=found_count)
        found_value.font = Font(name="Roboto", size=10, bold=True, color=COLORS["found_text"])
        found_value.alignment = ALIGNMENTS["left"]
        found_value.fill = FILLS["found"]

        row += 1
        
        # Строка 3: Не найдено в системе
        self.ws.merge_cells(start_row=row, start_column=table_start_col, end_row=row, end_column=table_start_col + 1)
        not_found_label = self.ws.cell(row=row, column=table_start_col, value="Не найдено в системе")
        not_found_label.font = Font(name="Roboto", size=10, bold=True)
        not_found_label.alignment = ALIGNMENTS["left"]
        not_found_label.fill = FILLS["not_found"]

        
        self.ws.merge_cells(start_row=row, start_column=table_start_col + 2, end_row=row, end_column=table_end_col)
        not_found_value = self.ws.cell(row=row, column=table_start_col + 2, value=not_found_count)
        not_found_value.font = Font(name="Roboto", size=10, bold=True, color=COLORS["not_found_text"])
        not_found_value.alignment = ALIGNMENTS["left"]
        not_found_value.fill = FILLS["not_found"]
    
        row += 1
        
        # Высота строк статистики
        for r in range(stats_start_row, row):
            self.ws.row_dimensions[r].height = 28
        
        row += 1

        # ============================================================
        # НАСТРОЙКА ШИРИНЫ КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 8
        self.ws.column_dimensions["C"].width = 30
        self.ws.column_dimensions["D"].width = 50
        
        # Скрываем сетку
        self.ws.sheet_view.showGridLines = False


def create_toc_sheet(workbook, sheets_info, total_articles=0, found_count=0, not_found_count=0):
    """Создает лист оглавления"""
    toc = TOCSheet(workbook)
    toc.build(sheets_info, total_articles, found_count, not_found_count)
    return toc.ws