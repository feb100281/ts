# # cards/reporting/article_analyzer/report_builder.py
# from openpyxl import Workbook
# from .sheets.summary_sheet import create_summary_sheet
# from .sheets.details_sheet import create_details_sheet
# from .sheets.not_found_sheet import create_not_found_sheet
# from .sheets.toc_sheet import create_toc_sheet


# class ArticleAnalysisReportBuilder:
#     """Построитель отчета по анализу артиклей"""
    
#     def __init__(self):
#         self.wb = Workbook()
#         # Удаляем стандартный лист
#         if "Sheet" in self.wb.sheetnames:
#             self.wb.remove(self.wb["Sheet"])
    
#     def build(self, summary_df, details_df, articles_not_found, total_articles=0, articles_found=0):
#         """Строит отчет"""
#         sheets_info = []
#         sheet_counter = 1
        
#         found_count = len(summary_df) if not summary_df.empty else 0
#         not_found_count = len(articles_not_found)
        
#         # Лист 1: Сводка по артиклям
#         create_summary_sheet(self.wb, sheet_counter, summary_df, articles_not_found)
#         sheets_info.append({
#             'number': sheet_counter,
#             'name': 'Сводка по артиклям',
#             'description': f'Агрегированная статистика: количество, суммы, мин/макс/ср/медиана цен (найдено {found_count} артиклей)',
#             'sheet_name': f"{sheet_counter:02d}_Сводка_по_артиклям"
#         })
#         sheet_counter += 1
        
#         # Лист 2: Детали с группировкой по артиклям
#         create_details_sheet(self.wb, sheet_counter, details_df)
#         sheets_info.append({
#             'number': sheet_counter,
#             'name': 'Детали по артиклям',
#             'description': f'Артикль, УПД (через запятую), цены (все уникальные значения), сумма, кол-во ({found_count} артиклей с данными)',
#             'sheet_name': f"{sheet_counter:02d}_Детали_по_позициям"
#         })
#         sheet_counter += 1
        
#         # Лист 3: Не найденные артикли (если есть)
#         if articles_not_found:
#             create_not_found_sheet(self.wb, sheet_counter, articles_not_found)
#             sheets_info.append({
#                 'number': sheet_counter,
#                 'name': 'Не найденные артикли',
#                 'description': f'Артикли, которые не найдены в системе ({not_found_count} шт.)',
#                 'sheet_name': f"{sheet_counter:02d}_Артикли_не_найдены"
#             })
#             sheet_counter += 1
        
#         # Лист TOC 
#         create_toc_sheet(self.wb, sheets_info, total_articles, found_count, not_found_count)
        
#         return self.wb
    
#     def save(self, filename):
#         """Сохраняет отчет в файл"""
#         self.wb.save(filename)
#         return filename
    
#     def get_workbook(self):
#         """Возвращает workbook"""
#         return self.wb



# cards/reporting/article_analyzer/report_builder.py
from openpyxl import Workbook
from .sheets.summary_sheet import create_summary_sheet
from .sheets.details_sheet import create_details_sheet
from .sheets.not_found_sheet import create_not_found_sheet
from .sheets.toc_sheet import create_toc_sheet


class ArticleAnalysisReportBuilder:
    """Построитель отчета по анализу артиклей"""
    
    def __init__(self):
        self.wb = Workbook()
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
    
    def build(self, summary_df, details_df, articles_not_found, total_articles=0, articles_found=0, not_mapped_articles=None):
        """Строит отчет"""
        sheets_info = []
        sheet_counter = 1
        
        found_count = len(summary_df) if not summary_df.empty else 0
        not_found_count = len(articles_not_found)
        
        # Лист 1: Сводка по артиклям
        create_summary_sheet(self.wb, sheet_counter, summary_df, articles_not_found)
        sheets_info.append({
            'number': sheet_counter,
            'name': 'Сводка по артиклям',
            'description': f'Агрегированная статистика: количество, суммы, мин/макс/ср/медиана цен (найдено {found_count} артиклей)',
            'sheet_name': f"{sheet_counter:02d}_Сводка_по_артиклям"
        })
        sheet_counter += 1
        
        # Лист 2: Детали с группировкой по артиклям
        create_details_sheet(self.wb, sheet_counter, details_df)
        sheets_info.append({
            'number': sheet_counter,
            'name': 'Детали по артиклям',
            'description': f'Артикль, УПД (через запятую), цены (все уникальные значения), сумма, кол-во ({found_count} артиклей с данными)',
            'sheet_name': f"{sheet_counter:02d}_Детали_по_позициям"
        })
        sheet_counter += 1
        
        # Лист 3: Не найденные артикли (если есть)
        if articles_not_found:
            create_not_found_sheet(self.wb, sheet_counter, articles_not_found)
            sheets_info.append({
                'number': sheet_counter,
                'name': 'Не найденные артикли',
                'description': f'Артикли, которые не найдены в системе ({not_found_count} шт.)',
                'sheet_name': f"{sheet_counter:02d}_Артикли_не_найдены"
            })
            sheet_counter += 1
        
        # Лист 4: Артикли без маппинга в USK (если есть)
        if not_mapped_articles:
            ws = self.wb.create_sheet(f"{sheet_counter:02d}_Артикли_без_USK")
            ws.append(["Артикль", "Статус"])
            for article in not_mapped_articles:
                ws.append([article, "Не найден в таблице USK"])
            
            # Форматирование
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            
            sheets_info.append({
                'number': sheet_counter,
                'name': 'Артикли без USK',
                'description': f'Артикли, которые отсутствуют в таблице USK ({len(not_mapped_articles)} шт.)',
                'sheet_name': f"{sheet_counter:02d}_Артикли_без_USK"
            })
            sheet_counter += 1
        
        # Лист TOC 
        create_toc_sheet(self.wb, sheets_info, total_articles, found_count, not_found_count)
        
        return self.wb