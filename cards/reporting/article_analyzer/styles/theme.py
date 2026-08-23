from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Цвета для отчета по анализу артиклей (в стиле основного отчета)
COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED", 
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "back_text_green": "1F5E4E",
    
    # Статусы
    "found_green": "E8F5E9",
    "found_text": "2E7D32",
    "not_found_red": "FDECEC",
    "not_found_text": "D32F2F",
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "none": PatternFill(fill_type=None),
    
    # Статусы
    "found": PatternFill("solid", fgColor=COLORS["found_green"]),
    "not_found": PatternFill("solid", fgColor=COLORS["not_found_red"]),
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=12, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=11, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    
    # Статусы
    "found": Font(name="Roboto", size=10, bold=True, color=COLORS["found_text"]),
    "not_found": Font(name="Roboto", size=10, bold=True, color=COLORS["not_found_text"]),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium = Side(style="medium", color=COLORS["dark_green"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium),
    "none": Border(),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
}

FORMATS = {
    "number": '#,##0',
    "decimal": '#,##0.00',
    "date": 'dd.mm.yyyy',
    "money": '#,##0.00 ₽',
}