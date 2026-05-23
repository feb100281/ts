# cards/reporting/components/kpi_cards.py

from openpyxl.styles import Font, PatternFill
from ..styles import BORDERS, ALIGNMENTS


DEFAULT_COLORS = {
    "bg": "F8F9FA",
    "value": "1F4E3D",
    "title": "666666",
}


class KPICards:
    """Компонент для отрисовки KPI-карточек в Excel."""

    def __init__(self, worksheet):
        self.ws = worksheet

    def draw_compact_card(
        self,
        row,
        col,
        title,
        value,
        subtitle=None,
        width=2,
        bg_color=None,
        value_color=None,
        title_color=None,
    ):
        bg_color = bg_color or DEFAULT_COLORS["bg"]
        value_color = value_color or DEFAULT_COLORS["value"]
        title_color = title_color or DEFAULT_COLORS["title"]

        height = 3 if subtitle else 2

        for r in range(row, row + height):
            for c in range(col, col + width):
                cell = self.ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.border = BORDERS["thin"]

        if width > 1:
            self.ws.merge_cells(
                start_row=row,
                start_column=col,
                end_row=row,
                end_column=col + width - 1,
            )

        value_cell = self.ws.cell(row=row, column=col, value=value)
        value_cell.font = Font(
            name="Roboto",
            size=14,
            bold=True,
            color=value_color,
        )
        value_cell.alignment = ALIGNMENTS["center"]

        if width > 1:
            self.ws.merge_cells(
                start_row=row + 1,
                start_column=col,
                end_row=row + 1,
                end_column=col + width - 1,
            )

        title_cell = self.ws.cell(row=row + 1, column=col, value=title)
        title_cell.font = Font(
            name="Roboto",
            size=8,
            color=title_color,
        )
        title_cell.alignment = ALIGNMENTS["center"]

        if subtitle:
            if width > 1:
                self.ws.merge_cells(
                    start_row=row + 2,
                    start_column=col,
                    end_row=row + 2,
                    end_column=col + width - 1,
                )

            subtitle_cell = self.ws.cell(row=row + 2, column=col, value=subtitle)
            subtitle_cell.font = Font(
                name="Roboto",
                size=8,
                bold=True,
                color=value_color,
            )
            subtitle_cell.alignment = ALIGNMENTS["center"]

        self.ws.row_dimensions[row].height = 24
        self.ws.row_dimensions[row + 1].height = 18

        if subtitle:
            self.ws.row_dimensions[row + 2].height = 18

        return height

    def draw_row(self, start_row, cards, start_col=1, gap=0):
        current_col = start_col
        max_height = 2

        for card in cards:
            width = card.get("width", 2)

            height = self.draw_compact_card(
                row=start_row,
                col=current_col,
                title=card.get("title", ""),
                value=card.get("value", ""),
                subtitle=card.get("subtitle"),
                width=width,
                bg_color=card.get("bg_color"),
                value_color=card.get("value_color"),
                title_color=card.get("title_color"),
            )

            max_height = max(max_height, height)
            current_col += width + gap

        return start_row + max_height


def create_kpi_cards(worksheet):
    return KPICards(worksheet)