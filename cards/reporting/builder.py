# # cards/reporting/builder.py

# import zipfile
# from io import BytesIO
# from openpyxl import Workbook
# from django.http import HttpResponse
# from typing import List

# from .queries import MissingFieldsQueries
# from .sheets import MissingNmSheet, MissingChrtSheet
# from .pdf_exporter import build_missing_fields_pdf_response


# class MissingFieldsReportGenerator:
#     """Генератор отчетов по недостающим полям"""

#     def __init__(self):
#         self.queries = MissingFieldsQueries()

#     def _create_workbook(self) -> Workbook:
#         wb = Workbook()

#         if "Sheet" in wb.sheetnames:
#             wb.remove(wb["Sheet"])

#         return wb

#     def _save_workbook_to_buffer(self, wb: Workbook) -> BytesIO:
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         return buffer

#     def generate_missing_nm_report(self, upd_ids: List[int] = None) -> BytesIO:
#         """Сгенерировать Excel-отчет по товарам без NM_ID"""

#         df = self.queries.get_grouped_missing_nm(upd_ids)
#         stats = self.queries.get_summary_stats(upd_ids)

#         wb = self._create_workbook()

#         sheet = MissingNmSheet(wb, df, stats)
#         sheet.build()

#         return self._save_workbook_to_buffer(wb)

#     def generate_missing_chrt_report(self, upd_ids: List[int] = None) -> BytesIO:
#         """Сгенерировать Excel-отчет по товарам без CHRT_ID / размера"""

#         df = self.queries.get_grouped_missing_chrt(upd_ids)
#         stats = self.queries.get_summary_stats(upd_ids)

#         wb = self._create_workbook()

#         sheet = MissingChrtSheet(wb, df, stats)
#         sheet.build()

#         return self._save_workbook_to_buffer(wb)

#     def generate_complete_package(self, upd_ids: List[int] = None) -> BytesIO:
#         """Сгенерировать ZIP-архив с отчетами и PDF-сопроводилкой"""

#         stats = self.queries.get_summary_stats(upd_ids)

#         zip_buffer = BytesIO()

#         with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
#             pdf_response = build_missing_fields_pdf_response(stats)
#             zip_file.writestr("Сопроводительное_письмо.pdf", pdf_response.content)

#             if stats.get("missing_nm_count", 0) > 0:
#                 nm_report = self.generate_missing_nm_report(upd_ids)
#                 zip_file.writestr("Отчет_товары_без_NM_ID.xlsx", nm_report.getvalue())

#             if stats.get("missing_chrt_count", 0) > 0:
#                 chrt_report = self.generate_missing_chrt_report(upd_ids)
#                 zip_file.writestr("Отчет_товары_без_CHRT_ID.xlsx", chrt_report.getvalue())

#         zip_buffer.seek(0)
#         return zip_buffer

#     def get_report_response(self, report_type: str = "both", upd_ids: List[int] = None) -> HttpResponse:
#         """
#         Получить HTTP-ответ:
#         - nm — только отчет без NM_ID
#         - chrt — только отчет без CHRT_ID
#         - both — ZIP с обоими отчетами и PDF
#         """

#         from datetime import datetime

#         timestamp = datetime.now().strftime("%Y%m%d_%H%M")

#         if report_type == "nm":
#             buffer = self.generate_missing_nm_report(upd_ids)
#             filename = f"report_missing_nm_id_{timestamp}.xlsx"
#             content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#         elif report_type == "chrt":
#             buffer = self.generate_missing_chrt_report(upd_ids)
#             filename = f"report_missing_chrt_id_{timestamp}.xlsx"
#             content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#         else:
#             buffer = self.generate_complete_package(upd_ids)
#             filename = f"reports_missing_fields_{timestamp}.zip"
#             content_type = "application/zip"

#         response = HttpResponse(
#             buffer.getvalue(),
#             content_type=content_type,
#         )

#         response["Content-Disposition"] = f'attachment; filename="{filename}"'
#         return response



# cards/reporting/builder.py

import zipfile
from io import BytesIO
from openpyxl import Workbook
from django.http import HttpResponse
from typing import List

from .queries import MissingFieldsQueries
from .sheets import MissingNmSheet, MissingChrtSheet
from .pdf_exporter import build_missing_fields_pdf_response


class MissingFieldsReportGenerator:
    """Генератор отчетов по недостающим полям"""

    def __init__(self):
        self.queries = MissingFieldsQueries()

    def _create_workbook(self) -> Workbook:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        return wb

    def _save_workbook_to_buffer(self, wb: Workbook) -> BytesIO:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_missing_nm_report(self, upd_ids: List[int] = None) -> BytesIO:
        """Сгенерировать Excel-отчет по товарам без NM_ID"""

        df = self.queries.get_grouped_missing_nm(upd_ids)
        stats = self.queries.get_summary_stats(upd_ids)

        wb = self._create_workbook()

        sheet = MissingNmSheet(wb, df, stats)
        sheet.build()

        return self._save_workbook_to_buffer(wb)

    def generate_missing_chrt_report(self, upd_ids: List[int] = None) -> BytesIO:
        """Сгенерировать Excel-отчет по товарам без CHRT_ID / размера"""

        df = self.queries.get_grouped_missing_chrt(upd_ids)
        stats = self.queries.get_summary_stats(upd_ids)

        wb = self._create_workbook()

        sheet = MissingChrtSheet(wb, df, stats)
        sheet.build()

        return self._save_workbook_to_buffer(wb)

    def generate_complete_package(self, upd_ids: List[int] = None) -> BytesIO:
        """Сгенерировать ZIP-архив с отчетами и PDF-сопроводилкой"""

        stats = self.queries.get_summary_stats(upd_ids)
        
        # Получаем дополнительную информацию для PDF
        upd_list = self.queries.get_upd_list(upd_ids) if upd_ids else []
        missing_nm_by_upd = self.queries.get_missing_nm_by_upd(upd_ids) if upd_ids else []
        missing_chrt_by_upd = self.queries.get_missing_chrt_by_upd(upd_ids) if upd_ids else []

        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            pdf_response = build_missing_fields_pdf_response(
                stats,
                upd_list=upd_list,
                missing_nm_by_upd=missing_nm_by_upd,
                missing_chrt_by_upd=missing_chrt_by_upd
            )
            zip_file.writestr("Сопроводительное_письмо.pdf", pdf_response.content)

            if stats.get("missing_nm_count", 0) > 0:
                nm_report = self.generate_missing_nm_report(upd_ids)
                zip_file.writestr("Отчет_товары_без_NM_ID.xlsx", nm_report.getvalue())

            if stats.get("missing_chrt_count", 0) > 0:
                chrt_report = self.generate_missing_chrt_report(upd_ids)
                zip_file.writestr("Отчет_товары_без_CHRT_ID.xlsx", chrt_report.getvalue())

        zip_buffer.seek(0)
        return zip_buffer

    def get_report_response(self, report_type: str = "both", upd_ids: List[int] = None) -> HttpResponse:
        """
        Получить HTTP-ответ:
        - nm — только отчет без NM_ID
        - chrt — только отчет без CHRT_ID
        - both — ZIP с обоими отчетами и PDF
        """

        from datetime import datetime

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")

        if report_type == "nm":
            buffer = self.generate_missing_nm_report(upd_ids)
            filename = f"report_missing_nm_id_{timestamp}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif report_type == "chrt":
            buffer = self.generate_missing_chrt_report(upd_ids)
            filename = f"report_missing_chrt_id_{timestamp}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:
            buffer = self.generate_complete_package(upd_ids)
            filename = f"reports_missing_fields_{timestamp}.zip"
            content_type = "application/zip"

        response = HttpResponse(
            buffer.getvalue(),
            content_type=content_type,
        )

        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response