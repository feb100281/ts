# cards/reporting/sheets.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from .components.kpi_cards import create_kpi_cards

from .styles import FONTS, FILLS, BORDERS, ALIGNMENTS


def safe_text(value, default="—"):
    if pd.isna(value):
        return default
    return str(value)


def safe_number(value, default=0):
    if pd.isna(value):
        return default
    try:
        return float(value)
    except Exception:
        return default


def short_text(value, max_len=100):
    text = safe_text(value)
    return text[:max_len] + "..." if len(text) > max_len else text


class BaseReportSheet:
    """Базовый класс для аккуратных Excel-листов отчета."""

    sheet_name = "Отчет"
    title = "ОТЧЕТ"

    end_col = 8
    table_headers = []

    numeric_cols = []
    money_cols = []
    manual_fill_cols = []

    total_label_col = None
    total_cols = []

    column_widths = {}

    def __init__(self, wb: Workbook, df, stats: dict):
        self.wb = wb
        self.df = df
        self.stats = stats
        self.ws = wb.create_sheet(self.sheet_name)

    def build(self):
        row = 1

        row = self._build_title(row)
        row = self._build_summary(row)
        row += 1

        table_header_row = row
        self._build_table_header(row)
        row += 1

        row = self._build_table_body(row)

        if not self.df.empty:
            self._build_total_row(row)
        else:
            self._build_empty_row(row)
            row += 1

        self._finalize(table_header_row, row)

    def _build_title(self, row):
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.end_col)
        cell = self.ws.cell(row=row, column=1, value=self.title)
        cell.font = FONTS["title"]
        cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 32
        row += 1

        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.end_col)
        cell = self.ws.cell(
            row=row,
            column=1,
            value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}",
        )
        cell.font = FONTS["subtitle"]
        cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 22

        return row + 2

    def _build_summary(self, row):
        cards = self._get_stats_rows()

        if not cards:
            return row

        kpi = create_kpi_cards(self.ws)

        row = kpi.draw_row(
            start_row=row,
            cards=cards,
            start_col=1,
            gap=0,
        )

        return row
    
    def _build_table_header(self, row):
        for col_idx, header in enumerate(self.table_headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx, value=header)

            if col_idx in self.manual_fill_cols:
                cell.font = FONTS["total"]
                cell.fill = FILLS["section"]
            else:
                cell.font = FONTS["header"]
                cell.fill = FILLS["header"]

            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]

        self.ws.row_dimensions[row].height = 32

    def _build_table_body(self, row):
        for i, (_, item) in enumerate(self.df.iterrows()):
            values = self._get_row_values(item)
            base_fill = FILLS["alt"] if i % 2 else FILLS["none"]

            for col_idx, value in enumerate(values, start=1):
                cell = self.ws.cell(row=row, column=col_idx, value=value)
                cell.font = FONTS["normal"]
                cell.fill = FILLS["section"] if col_idx in self.manual_fill_cols else base_fill
                cell.border = BORDERS["thin"]

                if col_idx in self.numeric_cols or col_idx in self.money_cols:
                    cell.alignment = ALIGNMENTS["right"]
                elif col_idx in self.manual_fill_cols:
                    cell.alignment = ALIGNMENTS["center"]
                else:
                    cell.alignment = ALIGNMENTS["left"]

                if col_idx in self.numeric_cols:
                    cell.number_format = "#,##0.00"

                if col_idx in self.money_cols:
                    cell.number_format = '#,##0.00 ₽'

            self.ws.row_dimensions[row].height = 24
            row += 1

        return row

    def _build_total_row(self, row):
        for col_idx in range(1, self.end_col + 1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.fill = FILLS["total"]
            cell.border = BORDERS["thin"]
            cell.font = FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]

        if self.total_label_col:
            self.ws.cell(row=row, column=self.total_label_col, value="ИТОГО:")

        for col_idx, field in self.total_cols:
            total = pd.to_numeric(self.df[field], errors="coerce").fillna(0).sum()
            cell = self.ws.cell(row=row, column=col_idx, value=float(total))
            cell.font = FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.number_format = '#,##0.00 ₽' if col_idx in self.money_cols else "#,##0.00"

        self.ws.row_dimensions[row].height = 28

    def _build_empty_row(self, row):
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.end_col)
        cell = self.ws.cell(row=row, column=1, value="Данные отсутствуют")
        cell.font = FONTS["muted"]
        cell.fill = FILLS["summary"]
        cell.alignment = ALIGNMENTS["center"]
        cell.border = BORDERS["thin"]
        self.ws.row_dimensions[row].height = 26

    def _finalize(self, table_header_row, last_row):
        for col_idx, width in self.column_widths.items():
            letter = get_column_letter(col_idx) if isinstance(col_idx, int) else col_idx
            self.ws.column_dimensions[letter].width = width

        last_col = get_column_letter(self.end_col)

        self.ws.auto_filter.ref = f"A{table_header_row}:{last_col}{last_row}"
     
        self.ws.freeze_panes = f"D{table_header_row + 1}"
        
        self.ws.sheet_view.showGridLines = False
        

    def _format_currency(self, value):
        try:
            return f"{float(value):,.2f} ₽".replace(",", " ").replace(".", ",")
        except Exception:
            return "0,00 ₽"

    def _format_number(self, value):
        try:
            return f"{float(value):,.2f}".replace(",", " ").replace(".", ",")
        except Exception:
            return "0"

    def _get_stats_rows(self):
        return []

    def _get_row_values(self, item):
        return []


class MissingNmSheet(BaseReportSheet):
    """Лист с товарами без nm_id."""

    sheet_name = "Нет_NM_ID"
    title = "ОТЧЕТ: ТОВАРЫ БЕЗ NM_ID"

    end_col = 10

    table_headers = [
        "Артикул в УПД",
        "NM_ID карточки WB",
        "Бренд",
        "Название товара",
        "Размер",
        "УПД",
        "Контрагенты",
        "Кол-во",
        "Цена за ед. с НДС",
        "Сумма с НДС",
    ]

    manual_fill_cols = [2]

    numeric_cols = [8]
    money_cols = [9, 10]

    total_label_col = 7
    total_cols = [
        (8, "upd_qty"),
        (10, "upd_amount_vatadd"),
    ]

    column_widths = {
        "A": 18,
        "B": 20,
        "C": 15,
        "D": 50,
        "E": 12,
        "F": 35,
        "G": 30,
        "H": 15,
        "I": 18,
        "J": 18,
    }
    
    def _get_stats_rows(self):
        return [
            {
                "title": "УПД",
                "value": self.stats.get("total_upd_count", 0),
                "width": 2,
            },
            {
                "title": "Строк всего в УПД",
                "value": self.stats.get("total_lines", 0),
                "width": 2,
            },
            {
                "title": "Строк в УПД без NM_ID",
                "value": self.stats.get("missing_nm_count", 0),
                "width": 2,
            },
            {
                "title": "Кол-во",
                "value": self._format_number(self.stats.get("missing_nm_qty", 0)),
                "width": 2,
            },
            {
                "title": "Сумма с НДС",
                "value": self._format_currency(self.stats.get("missing_nm_amount", 0)),
                "width": 2,
            },
        ]

    

    def _get_row_values(self, item):
        return [
            safe_text(item.get("upd_sa_name")),
            "",
            safe_text(item.get("brand")),
            short_text(item.get("upd_title")),
            safe_text(item.get("upd_size")),
            safe_text(item.get("upd_info")),
            safe_text(item.get("counterparty_name")),
            safe_number(item.get("upd_qty")),
            safe_number(item.get("upd_price_vatadd")),
            safe_number(item.get("upd_amount_vatadd")),
        ]


class MissingChrtSheet(BaseReportSheet):
    """Лист с товарами без chrt_id."""

    sheet_name = "Нет_CHRT_ID"
    title = "ОТЧЕТ: ТОВАРЫ БЕЗ CHRT_ID / РАЗМЕРА"

    end_col = 13

    table_headers = [
        "Артикул в УПД",
        "Бренд",
        "Название товара",
        "NM_ID",
        "Артикул WB",
        "Размер в УПД",
        "Доступные размеры WB",
        "Правильный размер",
        "УПД",
        "Контрагенты",
        "Кол-во",
        "Цена за ед. с НДС",
        "Сумма с НДС",
    ]

    manual_fill_cols = [8]

    numeric_cols = [11]
    money_cols = [12, 13]

    total_label_col = 9
    total_cols = [
        (11, "upd_qty"),
        (13, "upd_amount_vatadd"),
    ]

    column_widths = {
        "A": 18,
        "B": 15,
        "C": 45,
        "D": 14,
        "E": 18,
        "F": 15,
        "G": 35,
        "H": 22,
        "I": 35,
        "J": 30,
        "K": 15,
        "L": 18,
        "M": 18,
    }
    
    def _get_stats_rows(self):
        return [
            {
                "title": "УПД",
                "value": self.stats.get("total_upd_count", 0),
                "width": 2,
            },
            {
                "title": "Строк всего в УПД",
                "value": self.stats.get("total_lines", 0),
                "width": 2,
            },
            {
                "title": "Строк в УПД без CHRT_ID",
                "value": self.stats.get("missing_chrt_count", 0),
                "width": 3,
            },
            {
                "title": "Кол-во",
                "value": self._format_number(self.stats.get("missing_chrt_qty", 0)),
                "width": 2,
            },
            {
                "title": "Сумма с НДС",
                "value": self._format_currency(self.stats.get("missing_chrt_amount", 0)),
                "width": 4,
            },
        ]
        

    def _get_row_values(self, item):
        return [
            safe_text(item.get("upd_sa_name")),
            safe_text(item.get("brand")),
            short_text(item.get("upd_title")),
            safe_text(item.get("nm_id")),
            safe_text(item.get("wb_sa_name")),
            safe_text(item.get("upd_size")),
            safe_text(item.get("available_sizes")),
            "",
            safe_text(item.get("upd_info")),
            safe_text(item.get("counterparty_name")),
            safe_number(item.get("upd_qty")),
            safe_number(item.get("upd_price_vatadd")),
            safe_number(item.get("upd_amount_vatadd")),
        ]
        

class VatMismatchSheet(BaseReportSheet):
    """Лист с товарами, где не совпадает НДС в УПД и карточке WB"""

    sheet_name = "Несовпадение_НДС"
    title = "ОТЧЕТ: НЕСОВПАДЕНИЕ СТАВОК НДС"

    end_col = 13

    table_headers = [
        "Артикул в УПД",
        "Бренд",
        "Название товара",
        "Размер",
        "NM_ID",
        "Артикул WB",
        "Ставка НДС в УПД",
        "Ставка НДС в карточке",
        "УПД",
        "Контрагенты",
        "Кол-во",
        "Цена за ед. с НДС",
        "Сумма с НДС",
    ]

    numeric_cols = [11]
    money_cols = [12, 13]
    manual_fill_cols = [7, 8]

    total_label_col = 9
    total_cols = [
        (11, "upd_qty"),
        (13, "upd_amount_vatadd"),
    ]

    column_widths = {
        "A": 18,
        "B": 15,
        "C": 45,
        "D": 12,
        "E": 14,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 35,
        "J": 30,
        "K": 15,
        "L": 18,
        "M": 18,
    }
    
    def _get_stats_rows(self):
        return [
            {
                "title": "УПД",
                "value": self.stats.get("total_upd_count", 0),
                "width": 2,
            },
            {
                "title": "Строк всего в УПД",
                "value": self.stats.get("total_lines", 0),
                "width": 2,
            },
            {
                "title": "Строк в УПД с несовпадением НДС",
                "value": self.stats.get("vat_mismatch_count", 0),
                "width": 3,
            },
            {
                "title": "Кол-во с несовпадением НДС",
                "value": self._format_number(self.stats.get("vat_mismatch_qty", 0)),
                "width": 2,
            },
            {
                "title": "Сумма с НДС (с несовпадением НДС)",
                "value": self._format_currency(self.stats.get("vat_mismatch_amount", 0)),
                "width": 4,
            },
        ]
        
    def _get_row_values(self, item):
        upd_vat = item.get("upd_vat_rate", 0)
        card_vat = item.get("card_vat_rate", 0)
        
        upd_vat_str = f"{int(upd_vat)}%" if upd_vat and upd_vat > 0 else "—"
        card_vat_str = f"{int(card_vat)}%" if card_vat and card_vat > 0 else "—"
        
        return [
            safe_text(item.get("upd_sa_name")),
            safe_text(item.get("brand")),
            short_text(item.get("upd_title")),
            safe_text(item.get("upd_size")),
            safe_text(item.get("nm_id")),
            safe_text(item.get("wb_sa_name")),
            upd_vat_str,
            card_vat_str,
            safe_text(item.get("upd_info")),
            safe_text(item.get("counterparty_name")),
            safe_number(item.get("upd_qty")),
            safe_number(item.get("upd_price_vatadd")),
            safe_number(item.get("upd_amount_vatadd")),
        ]