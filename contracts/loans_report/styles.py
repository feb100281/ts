# contracts/loans_report/styles.py 
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED",
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "back_text_green": "1F5E4E",
    "accent_orange": "FF6F00",
    "accent_orange_light": "FFF3E0",
    "error_red": "C62828",
    "error_bg": "FFEBEE",
    "primary_bg": "F1F8E9",  
}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "accent": PatternFill("solid", fgColor=COLORS["accent_orange_light"]),
    "error": PatternFill("solid", fgColor=COLORS["error_bg"]),
    "none": PatternFill(fill_type=None),
    "month_header": PatternFill("solid", fgColor=COLORS["dark_green"]), 
}

FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"]),
    "subtitle": Font(name="Roboto", size=11, color=COLORS["text_gray"]),
    "header": Font(name="Roboto", size=10, bold=True, color=COLORS["white"]),
    "error": Font(name="Roboto", size=9, bold=True, color=COLORS["error_red"]),
    "normal": Font(name="Roboto", size=9, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["dark_green"]),
    "accent": Font(name="Roboto", size=10, bold=True, color=COLORS["accent_orange"]),
    "link": Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"], underline="single"),
    "small": Font(name="Roboto", size=8, color=COLORS["text_gray"]),
    "kpi_value": Font(name="Roboto", size=16, bold=True, color=COLORS["dark_green"]),  # Добавлено
    "kpi_label": Font(name="Roboto", size=9, color=COLORS["text_gray"]),  # Добавлено
    "danger": Font(name="Roboto", size=9, bold=True, color=COLORS["error_red"]),  # Добавлено
    "header_white": Font(name="Roboto", size=10, bold=True, color=COLORS["white"]),  # Добавлено
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium = Side(style="medium", color=COLORS["dark_green"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_medium": Border(bottom=medium),
    "top_bottom_thin": Border(top=thin, bottom=thin),
}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
}

FORMATS = {
    "currency": '#,##0.00',
    "percentage": '0.00%',
    "integer": '#,##0',
    "decimal": '#,##0.00', 

    
}