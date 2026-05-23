# contracts/loans_report/components/kpi_cards.py
from openpyxl.styles import Font, PatternFill
from ..styles import COLORS, BORDERS, ALIGNMENTS


class KPICards:
    """Класс для отрисовки KPI-карточек"""

    def __init__(self, worksheet):
        self.ws = worksheet

    def format_number(self, value, title):
        """Форматирует число с разделителями и процентами"""
        if value is None:
            return "—"
        
        # Преобразуем в float (если Decimal или str)
        try:
            num_value = float(value)
        except (ValueError, TypeError):
            return str(value)
        
        if title == "СТАВКА":
            # Если ставка пришла как 0.155 -> 15.5%, если 15.5 -> 15.5%
            if num_value <= 1:
                num_value = num_value * 100
            return f"{num_value:.2f}%"
        
        # Для валют - разделители тысяч и 2 знака после запятой
        return f"{num_value:,.2f}"

    def draw_compact_card(self, row, col, title, value, subtitle=None, color=None, width=2):
        # Используем светло-серый фон
        bg_color = COLORS.get("light_gray", "F7F7F7")

        height = 2

        # Заливаем фон и добавляем границы
        for r in range(row, row + height):
            for c in range(col, col + width):
                cell = self.ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.border = BORDERS["thin"]

        # Ячейка со значением (верхняя строка)
        self.ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + width - 1,
        )

        # Форматируем значение
        formatted_value = self.format_number(value, title)
        value_cell = self.ws.cell(row=row, column=col, value=formatted_value)
        value_cell.font = Font(
            name="Roboto",
            size=13,
            bold=True,
            color=color or COLORS["dark_green"],
        )
        value_cell.alignment = ALIGNMENTS["center"]

        # Ячейка с заголовком
        self.ws.merge_cells(
            start_row=row + 1,
            start_column=col,
            end_row=row + 1,
            end_column=col + width - 1,
        )

        title_cell = self.ws.cell(row=row + 1, column=col, value=title.upper())
        title_cell.font = Font(name="Roboto", size=8, bold=True, color=COLORS["text_gray"])
        title_cell.alignment = ALIGNMENTS["center"]

        return height

    def draw_row(self, start_row, cards, start_col=2):
        current_col = start_col
        max_height = 2

        for card in cards:
            width = card.get("width", 2)

            height = self.draw_compact_card(
                row=start_row,
                col=current_col,
                title=card.get("title", ""),
                value=card.get("value", 0),
                subtitle=card.get("subtitle"),
                color=card.get("color"),
                width=width,
            )

            max_height = max(max_height, height)
            current_col += width

        return start_row + max_height


def create_kpi_cards(worksheet):
    return KPICards(worksheet)