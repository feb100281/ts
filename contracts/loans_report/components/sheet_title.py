# contracts/loans_report/components/sheet_title.py

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..styles import COLORS, FILLS, BORDERS


class SheetTitle:
    """Класс для отрисовки профессионального заголовка листа"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def draw(self, row, title, counterparty="", inn="", date_text="", start_col=2, end_col=16):
        """Рисует профессиональный заголовок листа с блоком информации"""
        
        thin_border = Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"])
        )
        
        # ============================================================
        # СТРОКА 1: КНОПКА НАЗАД (ТОЛЬКО В B-C)
        # ============================================================
        back_cell = self.ws.cell(row=row, column=start_col, value="← НАЗАД")
        back_cell.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
        back_cell.alignment = Alignment(horizontal="center", vertical="center")
        back_cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])
        back_cell.border = thin_border
        back_cell.hyperlink = "#'Справка'!A1"
        
        self.ws.merge_cells(
            start_row=row, start_column=start_col, 
            end_row=row, end_column=start_col + 1
        )
        self.ws.row_dimensions[row].height = 24
        row += 1
        
        # ============================================================
        # ИНФОРМАЦИОННЫЙ БЛОК (ОБЪЕДИНЯЕМ ВСЕ КОЛОНКИ)
        # ============================================================
        info_start_row = row
        
        # Блок с заливкой (3 строки)
        for i in range(3):
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row=row + i, column=col)
                cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])
                cell.border = thin_border
        
        # Договор (жирный, крупный)
        self.ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col
        )
        title_cell = self.ws.cell(row=row, column=start_col, value=title)
        title_cell.font = Font(name="Roboto", size=14, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[row].height = 32
        row += 1
        
        # Контрагент
        if counterparty:
            self.ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col
            )
            cp_cell = self.ws.cell(row=row, column=start_col, value=f"Контрагент: {counterparty}")
            cp_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
            cp_cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.row_dimensions[row].height = 26
            row += 1
        
        # ИНН
        if inn:
            self.ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col
            )
            inn_cell = self.ws.cell(row=row, column=start_col, value=f"ИНН: {inn}")
            inn_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
            inn_cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.row_dimensions[row].height = 26
            row += 1
        
        # Дата отчета (строка-разделитель)
        if date_text:
            self.ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col
            )
            date_cell = self.ws.cell(row=row, column=start_col, value=date_text)
            date_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["white"])
            date_cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
            date_cell.alignment = Alignment(horizontal="left", vertical="center")
            self.ws.row_dimensions[row].height = 28
            row += 1
            
            # Тонкая линия-разделитель после даты
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = Border(bottom=Side(style="thin", color=COLORS["border_gray"]))
            self.ws.row_dimensions[row].height = 4
            row += 1
        
        return row


def create_sheet_title(worksheet):
    """Создает экземпляр SheetTitle"""
    return SheetTitle(worksheet)