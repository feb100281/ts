# contracts/loans_report/sheets/loan_sheet.py

from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from openpyxl.utils import column_index_from_string
from ..styles import COLORS, FONTS, BORDERS, ALIGNMENTS, FILLS, FORMATS, thin, medium
from ..components.kpi_cards import create_kpi_cards


class LoanSheet(BaseSheet):
    """Детальный лист по конкретному договору займа"""
    
    def __init__(self, workbook, sheet_name: str):
        super().__init__(workbook, sheet_name)
    
    def build(self, loan_data: dict, report_date: str, transactions: list, summary: dict):
        """Строит детальный лист по договору"""
        row = 1
        start_col = 2
        end_col = start_col + 10  # Всего 11 колонок (B-L)
        
        # ============================================================
        # ВЕРХНЯЯ СТРОКА С КНОПКОЙ НАЗАД
        # ============================================================
        back_cell = self.ws.cell(row=row, column=start_col, value="← НАЗАД")
        back_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        back_cell.alignment = Alignment(horizontal="center", vertical="center")
        back_cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])
        back_cell.border = BORDERS["thin"]
        back_cell.hyperlink = "#'Справка'!A1"
        
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 1)
        self.ws.row_dimensions[row].height = 25
        row += 1
        row += 1
        
        # ============================================================
        # ЗАГОЛОВОЧНЫЙ БЛОК
        # ============================================================
        contract_number = loan_data.get('contract_number', 'Без номера')
        contract_date = loan_data.get('contract_date', '')
        if contract_date:
            contract_date_str = contract_date.strftime('%d.%m.%Y') if hasattr(contract_date, 'strftime') else str(contract_date)
        else:
            contract_date_str = ''

        counterparty = loan_data.get('counterparty_name', '—')
        currency = loan_data.get('currency', '₽')

        contract_info = f"{counterparty} | ДОГОВОР {contract_number}"
        if contract_date_str:
            contract_info += f" от {contract_date_str} | {currency}"

        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        title_cell = self.ws.cell(row=row, column=start_col, value=contract_info)
        title_cell.font = Font(name="Roboto", size=12, bold=True, color=COLORS["dark_green"])
        title_cell.alignment = ALIGNMENTS["left"]
        title_cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])
        self.ws.row_dimensions[row].height = 30
        row += 1

        inn = loan_data.get('inn', '—')
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        inn_cell = self.ws.cell(row=row, column=start_col, value=f"ИНН: {inn}")
        inn_cell.font = Font(name="Roboto", size=11, color=COLORS["text_gray"])
        inn_cell.alignment = ALIGNMENTS["left"]
        inn_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        self.ws.row_dimensions[row].height = 26
        row += 1

        report_date_formatted = datetime.strptime(report_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        date_cell = self.ws.cell(row=row, column=start_col, value=f"Данные на {report_date_formatted}")
        date_cell.font = Font(name="Roboto", size=10, bold=True, color="FFFFFF")
        date_cell.fill = PatternFill("solid", fgColor=COLORS["dark_green"])
        date_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 28
        row += 1

        for col in range(start_col, end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = FILLS["section"]
            cell.border = Border(bottom=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 4
        row += 2

        # ============================================================
        # KPI БЛОК С КАРТОЧКАМИ 
        # ============================================================
        total_drawdown = loan_data.get('total_drawdown', 0) or 0
        total_repaid = loan_data.get('total_repaid', 0) or 0
        rate = loan_data.get('rate', 0) or 0
        ending_balance = loan_data.get('ending_balance', 0) or 0
        interest_balance = loan_data.get('interest_balance', 0) or 0
        total_debt = ending_balance + interest_balance
        repayment_date = loan_data.get('repayment_date', '')
        penalty_rate = loan_data.get('penalty_rate', 0)
        
        if repayment_date:
            try:
                if isinstance(repayment_date, str):
                    repayment_date_obj = datetime.strptime(repayment_date, '%Y-%m-%d')
                    repayment_date_display = repayment_date_obj.strftime('%d.%m.%Y')
                else:
                    repayment_date_display = str(repayment_date)
            except:
                repayment_date_display = str(repayment_date)
        else:
            repayment_date_display = "—"
        
        penalty_display = f"{penalty_rate:.2f}%" if penalty_rate > 0 else "—"

        kpi = create_kpi_cards(self.ws)

        cards_row1 = [
            {"title": "ПОЛУЧЕНО", "value": total_drawdown, "subtitle": None, "width": 2, "color": COLORS["dark_green"]},
            {"title": "ОПЛАЧЕНО", "value": total_repaid, "subtitle": "тело + проценты", "width": 2, "color": COLORS["dark_green"]},
            {"title": "СТАВКА", "value": rate / 100 if rate > 1 else rate, "subtitle": None, "width": 3, "color": COLORS["dark_green"]},
            {"title": "СРОК ПОГАШЕНИЯ", "value": repayment_date_display, "subtitle": None, "width": 3, "color": COLORS["dark_green"]}
        ]

        cards_row2 = [
            {"title": "ДОЛГ ПО ТЕЛУ", "value": ending_balance, "subtitle": None, "width": 2, "color": COLORS["dark_green"]},
            {"title": "ДОЛГ ПО ПРОЦЕНТАМ", "value": interest_balance, "subtitle": None, "width": 2, "color": COLORS["dark_green"]},
            {"title": "ИТОГО ДОЛГ", "value": total_debt, "subtitle": None, "width": 3, "color": COLORS["dark_green"]},
            {"title": "ШТРАФНОЙ % в день", "value": penalty_display, "subtitle": None, "width": 3, "color": COLORS["error_red"] if penalty_rate > 0 else COLORS["dark_green"]}
        ]

        row = kpi.draw_row(start_row=row, cards=cards_row1, start_col=start_col)
        row += 1
        row = kpi.draw_row(start_row=row, cards=cards_row2, start_col=start_col)
        row += 2
        
        # ============================================================
        # ДВИЖЕНИЯ ПО ДОГОВОРУ
        # ============================================================
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        movements_title = self.ws.cell(row=row, column=start_col, value="ДВИЖЕНИЯ ПО ДОГОВОРУ")
        movements_title.font = FONTS["header"]
        movements_title.fill = FILLS["header"]
        movements_title.alignment = ALIGNMENTS["center"]
        self.ws.row_dimensions[row].height = 28
        row += 1
        
        # ЗАГОЛОВКИ - добавлена колонка "ИТОГО ОПЛАЧЕНО"
        headers = ["МЕСЯЦ", "ДАТА", "ОПЕРАЦИЯ", "ВЫДАНО", "ВОЗВРАТ ТЕЛА", 
                   "НАЧИСЛЕНО %", "ОПЛАЧЕНО %", "ИТОГО ОПЛАЧЕНО", 
                   "ОСТАТОК ТЕЛА", "ОСТАТОК %", "ИТОГО ДОЛГ"]
        
        for col_idx, header in enumerate(headers, start=start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header"]
            cell.alignment = ALIGNMENTS["center"]
            cell.fill = FILLS["header"]
            cell.border = Border(left=thin, right=thin, top=thin, bottom=medium)
        
        self.ws.row_dimensions[row].height = 30
        row += 1
        
        monthly_data = self._group_by_month(transactions)
        
        if monthly_data:
            for month in monthly_data:
                self._add_month_row(row, start_col, month, currency)
                row += 1
                
                if month['transactions']:
                    detail_start = row
                    for op in month['transactions']:
                        self._add_transaction_row(row, start_col, op, currency)
                        row += 1
                    detail_end = row - 1
                    
                    if detail_start <= detail_end:
                        self.ws.row_dimensions.group(detail_start, detail_end, outline_level=1, hidden=True)
                
                row += 1
        else:
            no_data_cell = self.ws.cell(row=row, column=start_col, value="Нет операций за период")
            no_data_cell.font = Font(name="Roboto", size=10, color=COLORS["text_gray"])
            no_data_cell.alignment = ALIGNMENTS["center"]
            self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            row += 1
        
        # Ширина колонок
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 10   # Месяц
        self.ws.column_dimensions["C"].width = 12   # Дата
        self.ws.column_dimensions["D"].width = 65   # Операция
        self.ws.column_dimensions["E"].width = 14   # Выдано
        self.ws.column_dimensions["F"].width = 14   # Возврат тела
        self.ws.column_dimensions["G"].width = 14   # Начислено %
        self.ws.column_dimensions["H"].width = 14   # Оплачено %
        self.ws.column_dimensions["I"].width = 18   # Итого оплачено
        self.ws.column_dimensions["J"].width = 14   # Остаток тела
        self.ws.column_dimensions["K"].width = 14   # Остаток %
        self.ws.column_dimensions["L"].width = 14   # Итого долг
        
        self.ws.sheet_view.showGridLines = False
    
    def _add_month_row(self, row: int, start_col: int, month: dict, currency: str):
        """Строка с месяцем"""
        self._add_cell(row, start_col, month['month'], "text", align="center", 
                      bold=True, font_color=COLORS["dark_green"])
        
        self._add_cell(row, start_col + 1, "", "text", align="center", font_color=COLORS["text_gray"])
        
        self._add_cell(row, start_col + 2, "ИТОГО ЗА МЕСЯЦ", "text", align="right", 
                      bold=True, font_color=COLORS["text_gray"], bg="FFF3E0")
        
        total_drawdown = month.get('total_drawdown', 0) or 0
        self._add_cell(row, start_col + 3, total_drawdown if total_drawdown != 0 else None, 
                      "currency", bold=True, font_color=COLORS["dark_green"], bg="FFF3E0")
        
        total_principal = month.get('total_principal', 0) or 0
        self._add_cell(row, start_col + 4, total_principal if total_principal != 0 else None,
                      "currency", bold=True, font_color=COLORS["error_red"], bg="FFF3E0")
        
        total_interest_accrued = month.get('total_interest_accrued', 0) or 0
        self._add_cell(row, start_col + 5, total_interest_accrued if total_interest_accrued != 0 else None,
                      "currency", bold=True, font_color=COLORS["accent_orange"], bg="FFF3E0")
        
        total_interest_paid = month.get('total_interest_paid', 0) or 0
        self._add_cell(row, start_col + 6, total_interest_paid if total_interest_paid != 0 else None,
                      "currency", bold=True, font_color=COLORS["error_red"], bg="FFF3E0")
        
        # Итого оплачено (тело + проценты)
        total_repaid = month.get('total_repaid', 0) or 0
        self._add_cell(row, start_col + 7, total_repaid if total_repaid != 0 else None,
                      "currency", bold=True, font_color=COLORS["dark_green"], bg="FFF3E0")
        
        end_balance = month.get('end_balance', 0) or 0
        self._add_cell(row, start_col + 8, end_balance, "currency", bold=True, 
                      font_color=COLORS["dark_green"], bg="FFF3E0")
        
        end_interest = month.get('end_interest', 0) or 0
        self._add_cell(row, start_col + 9, end_interest, "currency", bold=True, 
                      font_color=COLORS["accent_orange"], bg="FFF3E0")
        
        end_total = month.get('end_total_debt', 0) or 0
        self._add_cell(row, start_col + 10, end_total, "currency", bold=True, 
                      font_color=COLORS["dark_green"], bg="FFF3E0")
        
        self.ws.row_dimensions[row].height = 24
    
    def _add_transaction_row(self, row: int, start_col: int, op: dict, currency: str):
        """Строка детализации операции"""
        self._add_cell(row, start_col, "", "text", align="center", font_color=COLORS["text_gray"])
        
        date_val = op.get('date')
        if date_val:
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%d.%m.%Y')
            else:
                date_str = str(date_val)
        else:
            date_str = ''
        self._add_cell(row, start_col + 1, date_str, "text", align="center", font_color=COLORS["text_gray"])
        
        description = op.get('description', '—')
        if not description or description == '':
            description = '—'
        
        cell = self.ws.cell(row=row, column=start_col + 2, value=description)
        cell.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = BORDERS["thin"]
        
        drawdown = op.get('drawdown', 0) or 0
        if drawdown != 0:
            self._add_cell(row, start_col + 3, drawdown, "currency", align="right", 
                        font_color=COLORS["dark_green"])
        else:
            self._add_cell(row, start_col + 3, None, "currency", align="right", font_color=COLORS["text_gray"])
        
        principal = op.get('principal', 0) or 0
        if principal != 0:
            self._add_cell(row, start_col + 4, principal, "currency", align="right", 
                        font_color=COLORS["error_red"])
        else:
            self._add_cell(row, start_col + 4, None, "currency", align="right", font_color=COLORS["text_gray"])
        
        interest_accrued = op.get('interest_accrued', 0) or 0
        if interest_accrued != 0:
            self._add_cell(row, start_col + 5, interest_accrued, "currency", align="right", 
                        font_color=COLORS["accent_orange"])
        else:
            self._add_cell(row, start_col + 5, None, "currency", align="right", font_color=COLORS["text_gray"])
        
        interest_paid = op.get('interest_paid', 0) or 0
        if interest_paid != 0:
            self._add_cell(row, start_col + 6, interest_paid, "currency", align="right", 
                        font_color=COLORS["error_red"])
        else:
            self._add_cell(row, start_col + 6, None, "currency", align="right", font_color=COLORS["text_gray"])
        
        # Итого оплачено (тело + проценты за эту операцию)
        total_repaid_op = op.get('total_repaid', 0) or 0
        if total_repaid_op != 0:
            self._add_cell(row, start_col + 7, total_repaid_op, "currency", align="right", 
                        font_color=COLORS["dark_green"])
        else:
            self._add_cell(row, start_col + 7, None, "currency", align="right", font_color=COLORS["text_gray"])
        
        balance_principal = op.get('balance_principal', 0) or 0
        self._add_cell(row, start_col + 8, balance_principal, "currency", align="right", 
                    font_color=COLORS["text_gray"])
        
        balance_interest = op.get('balance_interest', 0) or 0
        self._add_cell(row, start_col + 9, balance_interest, "currency", align="right", 
                    font_color=COLORS["text_gray"])
        
        balance_total = op.get('balance_total', 0) or 0
        self._add_cell(row, start_col + 10, balance_total, "currency", align="right", 
                    font_color=COLORS["text_gray"])
        
        if row % 2 == 0:
            for col in range(start_col, start_col + 11):
                cell = self.ws.cell(row=row, column=col)
                if not cell.fill or cell.fill.fill_type == 'none':
                    cell.fill = FILLS["alt"]
        
        self.ws.row_dimensions[row].height = 20
    
    def _add_cell(self, row: int, col: int, value, value_type: str, 
                  align: str = "right", bold: bool = False, font_color: str = None, bg: str = None):
        """Универсальная функция добавления ячейки"""
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.border = BORDERS["thin"]
        
        if align == "center":
            cell.alignment = ALIGNMENTS["center"]
        elif align == "left":
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.alignment = ALIGNMENTS["right"]
        
        font_size = 10 if bold else 9
        font_bold = bold
        font_color = font_color or COLORS["text_gray"]
        
        cell.font = Font(name="Roboto", size=font_size, bold=font_bold, color=font_color)
        
        if value_type == "currency" and value is not None:
            cell.number_format = FORMATS["currency"]
        elif value_type == "percentage" and value is not None:
            cell.number_format = FORMATS["percentage"]
        
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
    
    def _group_by_month(self, transactions: list) -> list:
        """Группирует операции по месяцам."""
        if not transactions:
            return []
        
        sorted_trans = sorted(transactions, key=lambda x: x.get('date_from', ''))
        
        for t in sorted_trans:
            date_from = t.get('date_from')
            if isinstance(date_from, str):
                try:
                    t['date_from'] = datetime.strptime(date_from, '%Y-%m-%d').date()
                except:
                    t['date_from'] = datetime.now().date()
            elif isinstance(date_from, datetime):
                t['date_from'] = date_from.date()
            elif not date_from:
                t['date_from'] = datetime.now().date()
        
        months = defaultdict(lambda: {
            'transactions': [],
            'total_drawdown': 0,
            'total_principal': 0,
            'total_interest_accrued': 0,
            'total_interest_paid': 0,
            'total_repaid': 0,
            'end_balance': None,
            'end_interest': None,
            'end_total_debt': None
        })
        
        for trans in sorted_trans:
            current_date = trans['date_from']
            month_key = f"{current_date.year}-{current_date.month:02d}"
            
            drawdown = float(trans.get('drawdown_amount', 0) or 0)
            principal = float(trans.get('principal_repayment', 0) or 0)
            interest_accrued = float(trans.get('interest_accrued', 0) or 0)
            interest_paid = float(trans.get('interest_repayment', 0) or 0)
            total_repaid_op = principal + interest_paid
            
            balance_principal = float(trans.get('ending_balance', 0) or 0)
            balance_interest = float(trans.get('interest_balance', 0) or 0)
            balance_total = float(trans.get('total_debt', 0) or 0)
            
            op_desc = trans.get('operation_description')
            int_desc = trans.get('interest_description')
            
            description = ''
            if op_desc and str(op_desc) != '0':
                description = str(op_desc)
            elif int_desc and str(int_desc) != '0':
                description = str(int_desc)
            
            if not description or description == '' or description == '0':
                if drawdown != 0:
                    description = f"Выдача займа: {drawdown:,.2f} ₽".replace(',', ' ')
                elif principal != 0:
                    description = f"Возврат основного долга: {principal:,.2f} ₽".replace(',', ' ')
                elif interest_accrued != 0:
                    rate = float(trans.get('rate', 0) or 0)
                    rate_rounded = round(rate * 100, 2)
                    interest_fmt = f"{interest_accrued:,.2f}".replace(',', ' ')
                    description = f"Начисление процентов за {current_date.strftime('%d.%m.%Y')} (ставка {rate_rounded}%): {interest_fmt}"
                elif interest_paid != 0:
                    description = f"Оплата процентов: {interest_paid:,.2f} ₽".replace(',', ' ')
                else:
                    description = '—'
            
            months[month_key]['transactions'].append({
                'date': current_date,
                'description': description,
                'drawdown': drawdown,
                'principal': principal,
                'interest_accrued': interest_accrued,
                'interest_paid': interest_paid,
                'total_repaid': total_repaid_op,
                'balance_principal': balance_principal,
                'balance_interest': balance_interest,
                'balance_total': balance_total
            })
            
            months[month_key]['total_drawdown'] += drawdown
            months[month_key]['total_principal'] += principal
            months[month_key]['total_interest_accrued'] += interest_accrued
            months[month_key]['total_interest_paid'] += interest_paid
            months[month_key]['total_repaid'] += total_repaid_op
        
        result = []
        for month_key in sorted(months.keys()):
            year, month_num = month_key.split('-')
            year = int(year)
            month_num = int(month_num)
            month_name = self._get_month_name(month_num, year)
            
            month_data = months[month_key]
            ops = month_data['transactions']
            ops.sort(key=lambda x: x['date'])
            
            if ops:
                month_data['end_balance'] = ops[-1]['balance_principal']
                month_data['end_interest'] = ops[-1]['balance_interest']
                month_data['end_total_debt'] = ops[-1]['balance_total']
            else:
                month_data['end_balance'] = 0
                month_data['end_interest'] = 0
                month_data['end_total_debt'] = 0
            
            result.append({
                'month': month_name,
                'transactions': month_data['transactions'],
                'total_drawdown': month_data['total_drawdown'],
                'total_principal': month_data['total_principal'],
                'total_interest_accrued': month_data['total_interest_accrued'],
                'total_interest_paid': month_data['total_interest_paid'],
                'total_repaid': month_data['total_repaid'],
                'end_balance': month_data['end_balance'],
                'end_interest': month_data['end_interest'],
                'end_total_debt': month_data['end_total_debt']
            })
        
        return result
    
    def _get_month_name(self, month: int, year: int) -> str:
        """Возвращает название месяца"""
        months = {
            1: "ЯНВ", 2: "ФЕВ", 3: "МАР", 4: "АПР",
            5: "МАЙ", 6: "ИЮН", 7: "ИЮЛ", 8: "АВГ",
            9: "СЕН", 10: "ОКТ", 11: "НОЯ", 12: "ДЕК"
        }
        return f"{months[month]} {year}"