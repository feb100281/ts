# contracts/loans_report/sheets/toc_sheet.py
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .base_sheet import BaseSheet
from ..styles import COLORS, FONTS, BORDERS, ALIGNMENTS, FILLS, FORMATS, thin


class TOCSheet(BaseSheet):
    """Лист с оглавлением отчета по займам"""
    
    def __init__(self, workbook):
        super().__init__(workbook, "Справка")
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
    
    def build(self, loans_list: list, report_date: str):
        """Строит оглавление"""
        row = 1
        table_start_col = 2
        table_end_col = 13
        
        report_date_formatted = datetime.strptime(
            report_date, '%Y-%m-%d'
        ).strftime('%d.%m.%Y')
        
        # ============================================================
        # ЗАГОЛОВОЧНЫЙ БЛОК
        # ============================================================
        self.ws.merge_cells(
            start_row=row, start_column=table_start_col, 
            end_row=row, end_column=table_end_col
        )
        title_cell = self.ws.cell(
            row=row, column=table_start_col,
            value="ОТЧЕТ ПО ДОГОВОРАМ ЗАЙМА И КРЕДИТНЫМ ДОГОВОРАМ"
        )
        title_cell.font = FONTS["title"]
        title_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 35
        row += 1
        
        self.ws.merge_cells(
            start_row=row, start_column=table_start_col,
            end_row=row, end_column=table_end_col
        )
        subtitle_cell = self.ws.cell(
            row=row, column=table_start_col,
            value="Детализация по каждому договору: остатки, проценты, график платежей"
        )
        subtitle_cell.font = FONTS["subtitle"]
        subtitle_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 22
        row += 1
        
        self.ws.merge_cells(
            start_row=row, start_column=table_start_col,
            end_row=row, end_column=table_end_col
        )
        date_cell = self.ws.cell(
            row=row, column=table_start_col,
            value=f"Данные на {report_date_formatted}"
        )
        date_cell.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
        date_cell.fill = FILLS["section"]
        date_cell.alignment = ALIGNMENTS["left"]
        self.ws.row_dimensions[row].height = 28
        row += 1
        
        # ============================================================
        # ДЕКОРАТИВНЫЙ РАЗДЕЛИТЕЛЬ
        # ============================================================
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = FILLS["section"]
            cell.border = Border(bottom=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 4
        row += 2
        
        # ============================================================
        # ЗАГОЛОВОК ТАБЛИЦЫ
        # ============================================================
        headers = [
            "№", "КОНТРАГЕНТ", "ДОГОВОР", "СТАВКА",
            "ПОЛУЧЕНО\nИТОГО", "ОПЛАЧЕНО\nТЕЛО + %",
            "ДОЛГ\n(ТЕЛО)", "ДОЛГ\n(ПРОЦЕНТЫ)",
            "ИТОГО ДОЛГ", "СРОК ПОГАШЕНИЯ",
            "ШТРАФНОЙ %\nв день", "ПЕРЕЙТИ"
        ]
        
        for col_idx, header in enumerate(headers, start=table_start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.font = FONTS["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = FILLS["header"]
            cell.border = Border(
                left=thin, right=thin, top=thin,
                bottom=Side(style="medium", color=COLORS["dark_green"])
            )
        
        self.ws.row_dimensions[row].height = 40
        row += 1
        
        # ============================================================
        # ДАННЫЕ ТАБЛИЦЫ
        # ============================================================
        for idx, loan in enumerate(loans_list, start=1):
  
            principal_debt = loan.get('ending_balance', 0)
            interest_debt = loan.get('interest_balance', 0)
            total_debt = principal_debt + interest_debt
            
            # Номер
            cell_num = self.ws.cell(row=row, column=table_start_col, value=idx)
            cell_num.font = Font(name="Roboto", size=10, bold=True, color=COLORS["dark_green"])
            cell_num.alignment = ALIGNMENTS["center"]
            cell_num.border = BORDERS["thin"]
            
            # Контрагент
            cell_cp = self.ws.cell(row=row, column=table_start_col + 1, value=loan.get('counterparty_name', '—'))
            cell_cp.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            cell_cp.alignment = ALIGNMENTS["left"]
            cell_cp.border = BORDERS["thin"]
            
            # Договор
            contract_number = loan.get('contract_number', 'Без номера')
            contract_date = loan.get('contract_date', '')
            if contract_date:
                contract_date_str = contract_date.strftime('%d.%m.%Y') if hasattr(contract_date, 'strftime') else str(contract_date)
            else:
                contract_date_str = ''
            
            contract_name = f"{contract_number} от {contract_date_str}" if contract_date_str else contract_number
            cell_name = self.ws.cell(row=row, column=table_start_col + 2, value=contract_name)
            
            if contract_number in ['None', 'Без номера']:
                cell_name.font = Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"])
            else:
                cell_name.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            
            cell_name.alignment = ALIGNMENTS["left"]
            cell_name.border = BORDERS["thin"]
            
            # Ставка
            rate = loan.get('rate', 0) or 0
            cell_rate = self.ws.cell(row=row, column=table_start_col + 3, value=rate)
            cell_rate.font = Font(name="Roboto", size=9, bold=True, color=COLORS["dark_green"])
            cell_rate.alignment = ALIGNMENTS["center"]
            cell_rate.number_format = FORMATS["percentage"]
            cell_rate.border = BORDERS["thin"]
            
            # Получено (total_drawdown)
            total_drawdown = loan.get('total_drawdown', 0)
            cell_drawdown = self.ws.cell(row=row, column=table_start_col + 4, value=total_drawdown)
            cell_drawdown.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            cell_drawdown.alignment = ALIGNMENTS["right"]
            cell_drawdown.number_format = FORMATS["currency"]
            cell_drawdown.border = BORDERS["thin"]
            
            # Оплачено (total_repaid)
            total_repaid = loan.get('total_repaid', 0)
            cell_repaid = self.ws.cell(row=row, column=table_start_col + 5, value=total_repaid)
            cell_repaid.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            cell_repaid.alignment = ALIGNMENTS["right"]
            cell_repaid.number_format = FORMATS["currency"]
            cell_repaid.border = BORDERS["thin"]
            
            # Задолженность по телу (светло-красная заливка + цвет в зависимости от знака)
            cell_principal = self.ws.cell(row=row, column=table_start_col + 6, value=principal_debt)
            cell_principal.fill = FILLS["error"]  # Светло-красная заливка
            cell_principal.alignment = ALIGNMENTS["right"]
            cell_principal.number_format = FORMATS["integer"]
            cell_principal.border = BORDERS["thin"]
            
            # Цвет для тела долга
            if principal_debt < 0:
                cell_principal.font = FONTS["error"]  # Красный жирный для отрицательного
            else:
                cell_principal.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            
            # Задолженность по процентам (светло-красная заливка + цвет в зависимости от знака)
            cell_interest = self.ws.cell(row=row, column=table_start_col + 7, value=interest_debt)
            cell_interest.fill = FILLS["error"]  # Светло-красная заливка
            cell_interest.alignment = ALIGNMENTS["right"]
            cell_interest.number_format = FORMATS["integer"]
            cell_interest.border = BORDERS["thin"]
            
            # Цвет для процентов
            if interest_debt < 0:
                cell_interest.font = FONTS["error"]  # Красный жирный для отрицательного
            else:
                cell_interest.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            
            # Итого задолженность (светло-красная заливка + цвет в зависимости от знака)
            cell_total = self.ws.cell(row=row, column=table_start_col + 8, value=total_debt)
            cell_total.fill = FILLS["error"]  # Светло-красная заливка
            cell_total.alignment = ALIGNMENTS["right"]
            cell_total.number_format = FORMATS["integer"]
            cell_total.border = BORDERS["thin"]
            
            # Цвет для итогового долга
            if total_debt < 0:
                cell_total.font = FONTS["error"]  # Красный жирный для отрицательного
            else:
                cell_total.font = Font(name="Roboto", size=9, bold=True, color=COLORS["text_gray"])
            
            # Дата погашения
            repayment_date = loan.get('repayment_date', '')
            if repayment_date:
                try:
                    if isinstance(repayment_date, str):
                        repayment_date_obj = datetime.strptime(repayment_date, '%Y-%m-%d')
                        repayment_date_str = repayment_date_obj.strftime('%d.%m.%Y')
                    else:
                        repayment_date_str = str(repayment_date)
                except:
                    repayment_date_str = repayment_date
            else:
                repayment_date_str = '—'
            
            cell_repayment = self.ws.cell(row=row, column=table_start_col + 9, value=repayment_date_str)
            
            # Подсветка просрочки
            if repayment_date and repayment_date != '—':
                try:
                    if isinstance(repayment_date, str):
                        due_date = datetime.strptime(repayment_date, '%Y-%m-%d')
                        report_dt = datetime.strptime(report_date, '%Y-%m-%d')
                        if due_date < report_dt and total_debt > 0:
                            cell_repayment.font = FONTS["error"]
                            cell_repayment.fill = FILLS["error"]  # Подсветка просрочки
                        else:
                            cell_repayment.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
                except:
                    cell_repayment.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            else:
                cell_repayment.font = Font(name="Roboto", size=9, color=COLORS["text_gray"])
            
            cell_repayment.alignment = ALIGNMENTS["center"]
            cell_repayment.border = BORDERS["thin"]
            
            # Штрафной процент
            penalty_rate = loan.get('penalty_rate', 0)
            penalty_display = f"{penalty_rate:.2f}%" if penalty_rate > 0 else "0.00%"
            cell_penalty = self.ws.cell(row=row, column=table_start_col + 10, value=penalty_display)
            cell_penalty.font = Font(name="Roboto", size=9, color=COLORS["error_red"] if penalty_rate > 0 else COLORS["text_gray"])
            cell_penalty.alignment = ALIGNMENTS["center"]
            cell_penalty.number_format = "0.00"
            cell_penalty.border = BORDERS["thin"]
            
            # Ссылка
            cell_link = self.ws.cell(row=row, column=table_start_col + 11, value="Открыть")
            cell_link.font = FONTS["link"]
            cell_link.alignment = ALIGNMENTS["center"]
            cell_link.border = BORDERS["thin"]
            cell_link.hyperlink = f"#'{idx}'!A1"
            
            # Чередование цвета строк (поверх заливки долговых колонок)
            if idx % 2 == 0:
                row_fill = FILLS["alt"]
                for col in range(table_start_col, table_end_col + 1):
                    # Для долговых колонок (6,7,8) сохраняем красную заливку, но с наложением
                    if col not in [table_start_col + 6, table_start_col + 7, table_start_col + 8]:
                        self.ws.cell(row=row, column=col).fill = row_fill
            else:
                # Для нечетных строк - белый фон, но долговые колонки уже красные
                pass
            
            self.ws.row_dimensions[row].height = 26
            row += 1
        
        # ============================================================
        # ИТОГОВАЯ СТРОКА
        # ============================================================
        total_cell = self.ws.cell(
            row=row, column=table_start_col,
            value=f"ВСЕГО ДОГОВОРОВ: {len(loans_list)}"
        )
        total_cell.font = Font(name="Roboto", size=11, bold=True, color=COLORS["white"])
        total_cell.fill = FILLS["header"]
        total_cell.alignment = ALIGNMENTS["center"]
        self.ws.merge_cells(
            start_row=row, start_column=table_start_col,
            end_row=row, end_column=table_end_col
        )
        self.ws.row_dimensions[row].height = 32
        row += 2
        
        # ============================================================
        # НИЖНЯЯ ДЕКОРАТИВНАЯ ПОЛОСА
        # ============================================================
        for col in range(table_start_col, table_end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = FILLS["section"]
            cell.border = Border(top=Side(style="thin", color=COLORS["border_gray"]))
        self.ws.row_dimensions[row].height = 6
        
        # ============================================================
        # НАСТРОЙКА ШИРИНЫ КОЛОНОК
        # ============================================================
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 6       # №
        self.ws.column_dimensions["C"].width = 35      # Контрагент
        self.ws.column_dimensions["D"].width = 28      # Договор
        self.ws.column_dimensions["E"].width = 10      # Ставка
        self.ws.column_dimensions["F"].width = 16      # Получено
        self.ws.column_dimensions["G"].width = 16      # Оплачено
        self.ws.column_dimensions["H"].width = 16      # Долг тело
        self.ws.column_dimensions["I"].width = 16      # Долг проценты
        self.ws.column_dimensions["J"].width = 16      # Итого долг
        self.ws.column_dimensions["K"].width = 20      # Дата погашения
        self.ws.column_dimensions["L"].width = 16      # Штрафной %
        self.ws.column_dimensions["M"].width = 10      # Перейти
        
        self.ws.sheet_view.showGridLines = False