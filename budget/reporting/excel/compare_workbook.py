# budget/reporting/excel/compare_workbook.py

from openpyxl import Workbook

from .data_loader import load_budget_export_data
from .sheets.compare_summary_sheet import build_compare_summary_sheet
from .sheets.compare_articles_sheet import build_compare_articles_sheet
from .sheets.compare_gl_detail_sheet import build_compare_gl_detail_sheets
from .styles.theme import COLORS


def _set_tab_color_if_exists(wb, sheet_name, color):
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = color

def _get_scenario_code(item):
    return str((item["data"].get("revenue_param") or {}).get("scenario", "base")).lower()


def _reorder_versions_for_compare(versions_data):
    """
    Если среди выбранных версий есть base-сценарий,
    он становится базой.
    Иначе база = первая выбранная версия.
    """
    if not versions_data:
        return versions_data

    base_idx = None
    for idx, item in enumerate(versions_data):
        if _get_scenario_code(item) == "base":
            base_idx = idx
            break

    if base_idx is None or base_idx == 0:
        return versions_data

    base_item = versions_data[base_idx]
    rest = [item for i, item in enumerate(versions_data) if i != base_idx]
    return [base_item] + rest


def _build_detail_sheet_map(versions_data):
    """
    Для каждой статьи создаем 3 вида detail-листов:
    total   -> T_1, T_2, ...
    months  -> M_1, M_2, ...
    quarters-> Q_1, Q_2, ...
    """
    detail_items = {}

    for item in versions_data:
        pivot = item["data"].get("compare_pivot") or {}
        for row in pivot.get("rows", []):
            if row.get("row_type") != "item":
                continue

            path_key = row.get("path_key")
            if not path_key:
                continue

            if path_key not in detail_items:
                detail_items[path_key] = {
                    "activity": row.get("activity") or "",
                    "operation": row.get("operation") or "",
                    "item": row.get("item") or "",
                }

    sorted_keys = sorted(
        detail_items.keys(),
        key=lambda k: (
            detail_items[k]["activity"],
            detail_items[k]["operation"],
            detail_items[k]["item"],
        )
    )

    result = {}
    for idx, path_key in enumerate(sorted_keys, start=1):
        result[path_key] = {
            "total": f"T_{idx}",
            "months": f"M_{idx}",
            "quarters": f"Q_{idx}",
        }

    return result



def build_budget_compare_workbook(versions):
    wb = Workbook()
    wb.remove(wb.active)

    versions_data = []
    for version in versions:
        data = load_budget_export_data(version)
        versions_data.append({
            "version_obj": version,
            "data": data,
        })

    versions_data = _reorder_versions_for_compare(versions_data)
    detail_sheet_map = _build_detail_sheet_map(versions_data)

    build_compare_summary_sheet(wb, versions_data)
    build_compare_articles_sheet(wb, versions_data, detail_sheet_map)
    build_compare_gl_detail_sheets(wb, versions_data, detail_sheet_map)

    main_sheets = [
        "SUMMARY_COMPARE",
        "ARTICLES_COMPARE",
        "MONTHS_COMPARE",
        "QUARTERS_COMPARE",
    ]

    for sheet_name in main_sheets:
        _set_tab_color_if_exists(wb, sheet_name, COLORS["dark_green"])

    for sheet_name in wb.sheetnames:
        if sheet_name not in main_sheets:
            _set_tab_color_if_exists(wb, sheet_name, COLORS["light_green"])

    wb.active = wb.sheetnames.index("SUMMARY_COMPARE")
    return wb