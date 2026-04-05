# budget/reporting/excel/workbook.py

from openpyxl import Workbook

from .data_loader import load_budget_export_data
from .sheets.summary_sheet import build_summary_sheet
from .sheets.gl_sheet import build_gl_sheet
from .sheets.json_sheet import build_json_sheet
from .sheets.pivot_sheet import build_pivot_sheet
from .sheets.gl_detail_sheet import build_gl_detail_sheet

from .styles.theme import COLORS


def _hide_sheet_if_exists(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_state = "hidden"


def _set_tab_color_if_exists(wb, sheet_name, color):
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = color


def build_budget_workbook(version):
    wb = Workbook()
    wb.remove(wb.active)

    data = load_budget_export_data(version)

    build_summary_sheet(wb, data)
    build_pivot_sheet(wb, data)

    for detail in data["gl_pivot"].get("detail_sheets", []):
        build_gl_detail_sheet(wb, detail)

    build_gl_sheet(wb, data)

    build_json_sheet(
        wb=wb,
        title="REVENUE_PARAMS",
        json_data=data["revenue_param"],
        description="Параметры доходной части бюджета",
    )
    build_json_sheet(
        wb=wb,
        title="WB_COSTS_PARAMS",
        json_data=data["wb_costs_params"],
        description="Параметры расходной части WB",
    )
    build_json_sheet(
        wb=wb,
        title="CF_PARAMS",
        json_data=data["cf_params"],
        description="Параметры Cash Flow",
    )

    if data["report"]:
        build_json_sheet(
            wb=wb,
            title="REPORT",
            json_data=data["report"],
            description="Отчет по бюджету",
        )

    # ---------------------------
    # Цвета табов
    # ---------------------------

    # Главные листы — темно-зеленые
    main_sheets = ["SUMMARY", "БЮДЖЕТ"]

    for sheet_name in main_sheets:
        _set_tab_color_if_exists(wb, sheet_name, COLORS["dark_green"])

    # Расшифровки — светло-зеленые
    detail_sheets = [detail["sheet_name"] for detail in data["gl_pivot"].get("detail_sheets", [])]

    for sheet_name in detail_sheets:
        _set_tab_color_if_exists(wb, sheet_name, COLORS["light_green"])

    # Технические листы — можно тоже покрасить в светло-зеленый
    technical_sheets = ["GL", "REVENUE_PARAMS", "WB_COSTS_PARAMS", "CF_PARAMS", "REPORT"]

    for sheet_name in technical_sheets:
        _set_tab_color_if_exists(wb, sheet_name, COLORS["light_green"])

    # ---------------------------
    # Скрытие технических листов
    # ---------------------------
    _hide_sheet_if_exists(wb, "GL")
    _hide_sheet_if_exists(wb, "REVENUE_PARAMS")
    _hide_sheet_if_exists(wb, "WB_COSTS_PARAMS")
    _hide_sheet_if_exists(wb, "CF_PARAMS")
    _hide_sheet_if_exists(wb, "REPORT")

    wb.active = wb.sheetnames.index("SUMMARY")

    return wb