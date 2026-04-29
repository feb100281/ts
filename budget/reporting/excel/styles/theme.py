# budget/reporting/excel/styles/theme.py
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


COLORS = {
    "dark_green": "2F6656",
    "light_green": "E7F1ED",
    "total_green": "DCECE6",
    "light_gray": "F7F7F7",
    "summary_fill": "FBFBFB",
    "border_gray": "D9D9D9",
    "text_gray": "666666",
    "white": "FFFFFF",
    "black": "1F1F1F",
    "back_text_green": "1F5E4E",
    "negative_brown": "7A4E4E",
    "delta_red": "FBEAEA",
    "delta_green": "EAF6EE",
    "counterparty_green": "5F7F75",
    "counterparty_gray": "7A7A7A",
    
    # НОВЫЕ ЦВЕТА для YTD и комментариев
    "header_alt":"1B5E20",          # темно-зеленый для YTD заголовков (чуть темнее основного)
    "light_red": "FFEBEE",           # очень светлый красный для подсветки аномалий
    "light_green": "E8F5E9",         # очень светлый зеленый для подсветки аномалий
    "comment_bg": "FFF8E1",          # теплый фон для комментариев
    "comment_border": "FFC107",      # янтарный для рамки комментариев
    

}

FILLS = {
    "header": PatternFill("solid", fgColor=COLORS["dark_green"]),
    "section": PatternFill("solid", fgColor=COLORS["light_green"]),
    "alt": PatternFill("solid", fgColor=COLORS["light_gray"]),
    "total": PatternFill("solid", fgColor=COLORS["total_green"]),
    "summary": PatternFill("solid", fgColor=COLORS["summary_fill"]),
    "back": PatternFill("solid", fgColor=COLORS["light_green"]),
    "delta_red": PatternFill("solid", fgColor=COLORS["delta_red"]),
    "delta_green": PatternFill("solid", fgColor=COLORS["delta_green"]),
    "none": PatternFill(fill_type=None),
    "header_alt": PatternFill("solid", fgColor=COLORS["header_alt"]),
    "delta_red_light": PatternFill("solid", fgColor=COLORS["light_red"]),
    "delta_green_light": PatternFill("solid", fgColor=COLORS["light_green"]),
    "comment": PatternFill("solid", fgColor=COLORS["comment_bg"]),

}


FONTS = {
    "title": Font(name="Roboto", size=16, bold=True, color=COLORS["black"]),
    "subtitle": Font(name="Roboto", size=10, color=COLORS["text_gray"]),
    "section": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "header_white": Font(name="Roboto", size=10, bold=True, color=COLORS["white"]),
    "bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "normal": Font(name="Roboto", size=10, color=COLORS["black"]),
    "total": Font(name="Roboto", size=11, bold=True, color=COLORS["black"]),
    "back": Font(name="Roboto", size=10, bold=True, color=COLORS["back_text_green"], underline="single"),
    "pivot_header": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),

    "negative": Font(name="Roboto", size=10, color=COLORS["negative_brown"]),
    "negative_bold": Font(name="Roboto", size=10, bold=True, color=COLORS["negative_brown"]),
    "negative_total": Font(name="Roboto", size=11, bold=True, color=COLORS["negative_brown"]),
    
    "counterparty": Font(name="Roboto", size=10, italic=True, color=COLORS["counterparty_gray"]),
    "counterparty_negative": Font(name="Roboto", size=10, color=COLORS["counterparty_gray"]),
    "comment": Font(name="Roboto", size=9, italic=True, color=COLORS["text_gray"]),
    "comment_bold": Font(name="Roboto", size=10, bold=True, color=COLORS["black"]),
    "subitem": Font(name="Roboto",size=10,italic=True,color=COLORS["counterparty_gray"],),
}

thin = Side(style="thin", color=COLORS["border_gray"])
medium_dark = Side(style="medium", color=COLORS["black"])
medium_green = Side(style="medium", color=COLORS["dark_green"])

BORDERS = {
    "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
    "bottom_thin": Border(bottom=thin),
    "bottom_medium": Border(bottom=medium_dark),
    "none": Border(),
    "left_sublevel": Border(left=thin),
    "comment_box": Border(left=medium_green, right=medium_green, top=medium_green, bottom=medium_green),
    "comment_top": Border(top=medium_green),

}

ALIGNMENTS = {
    "left": Alignment(horizontal="left", vertical="center"),
    "center": Alignment(horizontal="center", vertical="center"),
    "right": Alignment(horizontal="right", vertical="center"),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "right_wrap": Alignment(horizontal="right", vertical="center", wrap_text=True),
}

FORMATS = {
    "money": '#,##0;(#,##0)',
    "money_int": '#,##0;(#,##0)',
    "date": 'dd.mm.yyyy',
    "percent": '0.0%',
    "percent_int": '0%',          # целые проценты
}