# budget/reporting/excel/styles/helpers.py

from openpyxl.styles import Border, Side
from budget.reporting.excel.styles.theme import (
    FILLS,
    FONTS,
    BORDERS,
    ALIGNMENTS,
    FORMATS,
    COLORS,
)


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_heights(ws, heights: dict):
    for row_idx, height in heights.items():
        ws.row_dimensions[row_idx].height = height


def draw_back_button(ws, cell="A1", text="← SUMMARY", target_sheet="SUMMARY"):
    ws[cell] = text
    ws[cell].hyperlink = f"#'{target_sheet}'!A1"
    ws[cell].font = FONTS["back"]
    ws[cell].alignment = ALIGNMENTS["left"]
    ws[cell].fill = FILLS["back"]
    ws[cell].border = Border(
        left=Side(style="thin", color=COLORS["border_gray"]),
        right=Side(style="thin", color=COLORS["border_gray"]),
        top=Side(style="thin", color=COLORS["border_gray"]),
        bottom=Side(style="thin", color=COLORS["border_gray"]),
    )


def draw_sheet_header(ws, title, subtitle="", note=""):
    ws["A2"] = title
    ws["A3"] = subtitle
    ws["A4"] = note

    ws["A2"].font = FONTS["title"]
    ws["A3"].font = FONTS["subtitle"]
    ws["A4"].font = FONTS["subtitle"]

    ws["A2"].alignment = ALIGNMENTS["left"]
    ws["A3"].alignment = ALIGNMENTS["left"]
    ws["A4"].alignment = ALIGNMENTS["left"]

    for col in range(1, 7):
        ws.cell(row=5, column=col).border = BORDERS["bottom_medium"]


def draw_section_title(ws, row, col_start, col_end, title):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILLS["section"]
        cell.border = BORDERS["bottom_thin"]
        if col == col_start:
            cell.value = title
            cell.font = FONTS["section"]
            cell.alignment = ALIGNMENTS["left"]
        else:
            cell.value = None


def draw_table_header(ws, row, headers, start_col=1, fill_key="header", font_key="header_white"):
    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i)
        cell.value = header
        cell.fill = FILLS[fill_key]
        cell.font = FONTS[font_key]
        cell.alignment = ALIGNMENTS["center_wrap"]
        cell.border = BORDERS["thin"]


def style_data_row(ws, row, values, start_col=1, zebra=False):
    fill = FILLS["alt"] if zebra else FILLS["none"]

    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.fill = fill
        cell.border = BORDERS["thin"]
        cell.font = FONTS["normal"]

        if isinstance(value, (int, float)):
            cell.number_format = FORMATS["money"]
            cell.alignment = ALIGNMENTS["right"]
        else:
            cell.alignment = ALIGNMENTS["left"]


def style_total_row(ws, row, values, start_col=1):
    for i, value in enumerate(values, start=start_col):
        cell = ws.cell(row=row, column=i, value=value)
        cell.fill = FILLS["total"]
        cell.border = BORDERS["bottom_medium"]
        cell.font = FONTS["total"]

        if isinstance(value, (int, float)):
            cell.number_format = FORMATS["money"]
            cell.alignment = ALIGNMENTS["right"]
        else:
            cell.alignment = ALIGNMENTS["left"]


def apply_money(cell):
    cell.number_format = FORMATS["money"]
    cell.alignment = ALIGNMENTS["right"]


def apply_int_money(cell):
    cell.number_format = FORMATS["money_int"]
    cell.alignment = ALIGNMENTS["right"]


def apply_date(cell):
    cell.number_format = FORMATS["date"]
    cell.alignment = ALIGNMENTS["center"]