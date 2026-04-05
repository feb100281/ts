# budget/reporting/excel/sheets/compare_articles_sheet.py

from openpyxl.styles import Alignment, PatternFill

from budget.reporting.excel.styles.helpers import (
    set_column_widths,
    set_row_heights,
    draw_sheet_header,
    draw_back_button,
)
from budget.reporting.excel.styles.theme import (
    FILLS,
    FONTS,
    BORDERS,
    ALIGNMENTS,
    FORMATS,
)


WHITE_SPACER_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF", bgColor="FFFFFF")
SPACER_WIDTH = 2.2
DATA_START_COL = 4  # после A=Статья, B=Прим., C=spacer




def _label_alignment(level=0):
    return Alignment(horizontal="left", vertical="center", indent=level)


def _paint_spacer_cell(cell):
    cell.value = None
    cell.fill = WHITE_SPACER_FILL
    cell.border = BORDERS["none"]


def _paint_vertical_spacer(ws, col_idx, row_from, row_to):
    for row_idx in range(row_from, row_to + 1):
        _paint_spacer_cell(ws.cell(row=row_idx, column=col_idx))


def _draw_separator_row(ws, row_idx, last_col):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = FILLS["none"]
        cell.border = BORDERS["none"]
    ws.row_dimensions[row_idx].height = 8


def _scenario_label(data):
    scenario_raw = (data.get("revenue_param") or {}).get("scenario", "base")
    scenario_map = {
        "base": "Базовый",
        "optimistic": "Оптимистичный",
        "conservative": "Консервативный",
    }
    return scenario_map.get(str(scenario_raw).lower(), str(scenario_raw))


def _version_label(version_data):
    version = version_data["data"]["version"]
    scenario = _scenario_label(version_data["data"])
    number = version.get("number") or "—"
    return f"{number} ({scenario})"


def _get_value_font(value, normal_font, negative_font):
    try:
        if value is not None and float(value) < 0:
            return negative_font
    except (TypeError, ValueError):
        pass
    return normal_font


def _get_delta_fill(value, default_fill):
    try:
        value = float(value or 0)
    except (TypeError, ValueError):
        return default_fill

    if value < 0:
        return FILLS["delta_red"]
    elif value > 0:
        return FILLS["delta_green"]
    return default_fill


def _collect_compare_rows(versions_data, detail_sheet_map):
    row_map = {}

    for version_idx, item in enumerate(versions_data):
        rows = item["data"]["compare_pivot"].get("rows", [])

        for row in rows:
            key = row["path_key"]

            if key not in row_map:
                row_map[key] = {
                    "level": row["level"],
                    "row_type": row["row_type"],
                    "label": row["label"],
                    "activity": row["activity"],
                    "operation": row["operation"],
                    "item": row["item"],
                    "path_key": row["path_key"],
                    "sheet_names": detail_sheet_map.get(row["path_key"], {}) if row["row_type"] == "item" else {},
                    "values": {},
                }

            row_map[key]["values"][version_idx] = {
                "total": float(row.get("total", 0) or 0),
                "months": row.get("months", {}) or {},
                "quarters": row.get("quarters", {}) or {},
            }

    def sort_key(item):
        return (
            item["activity"] or "",
            item["operation"] or "",
            item["level"],
            item["item"] or "",
        )

    return sorted(row_map.values(), key=sort_key)


def _union_month_keys(versions_data):
    all_keys = set()
    all_labels = {}

    for item in versions_data:
        pivot = item["data"].get("compare_pivot") or {}
        for key in pivot.get("months", []):
            all_keys.add(key)
        all_labels.update(pivot.get("month_labels", {}))

    keys = sorted(all_keys)
    return keys, {k: all_labels.get(k, k) for k in keys}


def _union_quarter_keys(versions_data):
    all_keys = set()
    all_labels = {}

    for item in versions_data:
        pivot = item["data"].get("compare_pivot") or {}
        for key in pivot.get("quarters", []):
            all_keys.add(key)
        all_labels.update(pivot.get("quarter_labels", {}))

    keys = sorted(all_keys)
    return keys, {k: all_labels.get(k, k) for k in keys}


def _build_total_compare_columns(versions_data, start_col=DATA_START_COL):
    """
    ARTICLES_COMPARE:
    C = spacer после 'Прим.'
    База | spacer | Версия2 | spacer | Δ | Δ% | Версия3 | spacer | Δ | Δ% ...
    """
    columns = []
    col = start_col

    columns.append({
        "type": "base",
        "version_idx": 0,
        "label": f"База: {_version_label(versions_data[0])}",
        "col": col,
    })
    col += 1

    if len(versions_data) > 1:
        columns.append({
            "type": "spacer",
            "col": col,
        })
        col += 1

    for idx, item in enumerate(versions_data[1:], start=1):
        columns.append({
            "type": "compare",
            "version_idx": idx,
            "label": _version_label(item),
            "col": col,
        })
        col += 1

        columns.append({
            "type": "spacer",
            "col": col,
        })
        col += 1

        columns.append({
            "type": "delta_abs",
            "version_idx": idx,
            "label": "Δ к базе",
            "col": col,
        })
        col += 1

        columns.append({
            "type": "delta_pct",
            "version_idx": idx,
            "label": "Δ %",
            "col": col,
        })
        col += 1

    return columns, col - 1


def _build_period_block_columns(versions_data):
    """
    MONTHS_COMPARE / QUARTERS_COMPARE:
    C = spacer после 'Прим.'
    Внутри блока периода spacer нет:
    База | Версия2 | Δ | Δ% | Версия3 | Δ | Δ% ...
    Между блоками периодов есть spacer.
    """
    columns = []
    offset = 0

    columns.append({
        "type": "base",
        "version_idx": 0,
        "label": f"База: {_version_label(versions_data[0])}",
        "offset": offset,
    })
    offset += 1

    for idx, item in enumerate(versions_data[1:], start=1):
        columns.append({
            "type": "compare",
            "version_idx": idx,
            "label": _version_label(item),
            "offset": offset,
        })
        offset += 1

        columns.append({
            "type": "delta_abs",
            "version_idx": idx,
            "label": "Δ к базе",
            "offset": offset,
        })
        offset += 1

        columns.append({
            "type": "delta_pct",
            "version_idx": idx,
            "label": "Δ %",
            "offset": offset,
        })
        offset += 1

    return columns, offset


def _draw_total_sheet_header(ws, row_idx, compare_cols):
    cell = ws.cell(row=row_idx, column=1, value="Статья")
    cell.fill = FILLS["header"]
    cell.font = FONTS["header_white"]
    cell.border = BORDERS["thin"]
    cell.alignment = ALIGNMENTS["center"]

    note_cell = ws.cell(row=row_idx, column=2, value="Прим.")
    note_cell.fill = FILLS["header"]
    note_cell.font = FONTS["header_white"]
    note_cell.border = BORDERS["thin"]
    note_cell.alignment = ALIGNMENTS["center"]

    _paint_spacer_cell(ws.cell(row=row_idx, column=3))

    for item in compare_cols:
        col = item["col"]
        cell = ws.cell(row=row_idx, column=col)

        if item["type"] == "spacer":
            _paint_spacer_cell(cell)
        else:
            cell.value = item["label"]
            cell.fill = FILLS["header"]
            cell.font = FONTS["header_white"]
            cell.border = BORDERS["thin"]
            cell.alignment = ALIGNMENTS["center_wrap"]


def _draw_period_sheet_header(ws, top_row, block_cols, period_keys, period_labels, block_content_width):
    ws.merge_cells(start_row=top_row, start_column=1, end_row=top_row + 1, end_column=1)
    cell = ws.cell(row=top_row, column=1, value="Статья")
    cell.fill = FILLS["header"]
    cell.font = FONTS["header_white"]
    cell.border = BORDERS["thin"]
    cell.alignment = ALIGNMENTS["center"]

    ws.merge_cells(start_row=top_row, start_column=2, end_row=top_row + 1, end_column=2)
    note_cell = ws.cell(row=top_row, column=2, value="Прим.")
    note_cell.fill = FILLS["header"]
    note_cell.font = FONTS["header_white"]
    note_cell.border = BORDERS["thin"]
    note_cell.alignment = ALIGNMENTS["center"]

    _paint_spacer_cell(ws.cell(row=top_row, column=3))
    _paint_spacer_cell(ws.cell(row=top_row + 1, column=3))

    current_col = DATA_START_COL

    for period_index, period_key in enumerate(period_keys):
        content_start_col = current_col
        content_end_col = content_start_col + block_content_width - 1

        ws.merge_cells(
            start_row=top_row,
            start_column=content_start_col,
            end_row=top_row,
            end_column=content_end_col,
        )

        p_cell = ws.cell(row=top_row, column=content_start_col, value=period_labels.get(period_key, period_key))
        p_cell.fill = FILLS["header"]
        p_cell.font = FONTS["header_white"]
        p_cell.border = BORDERS["thin"]
        p_cell.alignment = ALIGNMENTS["center"]

        for c in range(content_start_col, content_end_col + 1):
            cell = ws.cell(row=top_row, column=c)
            cell.fill = FILLS["header"]
            cell.border = BORDERS["thin"]

        for item in block_cols:
            col_idx = content_start_col + item["offset"]
            cell = ws.cell(row=top_row + 1, column=col_idx)
            cell.value = item["label"]
            cell.fill = FILLS["header"]
            cell.font = FONTS["header_white"]
            cell.border = BORDERS["thin"]
            cell.alignment = ALIGNMENTS["center_wrap"]

        current_col = content_end_col + 1

        if period_index < len(period_keys) - 1:
            _paint_spacer_cell(ws.cell(row=top_row, column=current_col))
            _paint_spacer_cell(ws.cell(row=top_row + 1, column=current_col))
            current_col += 1


def _draw_number_cell(ws, row, col, value, fill, font, negative_font, border, is_delta=False, is_percent=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _get_delta_fill(value, fill) if is_delta else fill
    cell.font = _get_value_font(value, font, negative_font)
    cell.border = border
    cell.alignment = ALIGNMENTS["right"]
    cell.number_format = "0.0%;[Red]-0.0%" if is_percent else FORMATS["money_int"]


def _draw_total_compare_row(ws, row_idx, item, compare_cols):
    row_type = item["row_type"]

    if row_type == "activity":
        fill = FILLS["total"]
        label_font = FONTS["section"]
        value_font = FONTS["bold"]
        negative_value_font = FONTS["negative_bold"]
        border = BORDERS["bottom_medium"]
        height = 21
        label_indent = 1
    elif row_type == "operation":
        fill = FILLS["none"]
        label_font = FONTS["bold"]
        value_font = FONTS["bold"]
        negative_value_font = FONTS["negative_bold"]
        border = BORDERS["thin"]
        height = 19
        label_indent = 2
    else:
        fill = FILLS["none"]
        label_font = FONTS["bold"]
        value_font = FONTS["normal"]
        negative_value_font = FONTS["negative"]
        border = BORDERS["thin"]
        height = 18
        label_indent = 3

    ws.row_dimensions[row_idx].height = height

    label = ws.cell(row=row_idx, column=1, value=item["label"])
    label.fill = fill
    label.font = label_font
    label.border = border
    label.alignment = _label_alignment(label_indent)


    sheet_names = item.get("sheet_names") or {}
    note_value = sheet_names.get("total", "")
    note_cell = ws.cell(row=row_idx, column=2, value=note_value)
    note_cell.fill = fill
    note_cell.border = border
    note_cell.alignment = ALIGNMENTS["center"]

    if note_value:
        note_cell.font = FONTS["back"]
        note_cell.hyperlink = f"#'{note_value}'!A1"
    else:
        note_cell.font = FONTS["normal"]

    _paint_spacer_cell(ws.cell(row=row_idx, column=3))

    base_value = float(item["values"].get(0, {}).get("total", 0) or 0)

    for layout_item in compare_cols:
        col = layout_item["col"]

        if layout_item["type"] == "spacer":
            _paint_spacer_cell(ws.cell(row=row_idx, column=col))
            continue

        if layout_item["type"] == "base":
            value = base_value
            is_delta = False
            is_percent = False
        else:
            compare_value = float(item["values"].get(layout_item["version_idx"], {}).get("total", 0) or 0)

            if layout_item["type"] == "compare":
                value = compare_value
                is_delta = False
                is_percent = False
            elif layout_item["type"] == "delta_abs":
                value = compare_value - base_value
                is_delta = True
                is_percent = False
            else:
                value = (compare_value - base_value) / base_value if abs(base_value) > 0.0001 else None
                is_delta = True
                is_percent = True

        _draw_number_cell(
            ws=ws,
            row=row_idx,
            col=col,
            value=value,
            fill=fill,
            font=value_font,
            negative_font=negative_value_font,
            border=border,
            is_delta=is_delta,
            is_percent=is_percent,
        )


def _draw_period_compare_row(ws, row_idx, item, block_cols, period_keys, period_type, block_content_width):
    row_type = item["row_type"]

    if row_type == "activity":
        fill = FILLS["total"]
        label_font = FONTS["section"]
        value_font = FONTS["bold"]
        negative_value_font = FONTS["negative_bold"]
        border = BORDERS["bottom_medium"]
        height = 21
        label_indent = 1
    elif row_type == "operation":
        fill = FILLS["none"]
        label_font = FONTS["bold"]
        value_font = FONTS["bold"]
        negative_value_font = FONTS["negative_bold"]
        border = BORDERS["thin"]
        height = 19
        label_indent = 2
    else:
        fill = FILLS["none"]
        label_font = FONTS["bold"]
        value_font = FONTS["normal"]
        negative_value_font = FONTS["negative"]
        border = BORDERS["thin"]
        height = 18
        label_indent = 3

    ws.row_dimensions[row_idx].height = height

    label = ws.cell(row=row_idx, column=1, value=item["label"])
    label.fill = fill
    label.font = label_font
    label.border = border
    label.alignment = _label_alignment(label_indent)

 
    sheet_names = item.get("sheet_names") or {}

    if period_type == "months":
        note_value = sheet_names.get("months", "")
    elif period_type == "quarters":
        note_value = sheet_names.get("quarters", "")
    else:
        note_value = ""
    
    
    note_cell = ws.cell(row=row_idx, column=2, value=note_value)
    note_cell.fill = fill
    note_cell.border = border
    note_cell.alignment = ALIGNMENTS["center"]

    if note_value:
        note_cell.font = FONTS["back"]
        note_cell.hyperlink = f"#'{note_value}'!A1"
    else:
        note_cell.font = FONTS["normal"]

    _paint_spacer_cell(ws.cell(row=row_idx, column=3))

    current_col = DATA_START_COL

    for period_index, period_key in enumerate(period_keys):
        base_value = float(item["values"].get(0, {}).get(period_type, {}).get(period_key, 0) or 0)

        for layout_item in block_cols:
            col = current_col + layout_item["offset"]

            if layout_item["type"] == "base":
                value = base_value
                is_delta = False
                is_percent = False
            else:
                compare_value = float(
                    item["values"].get(layout_item["version_idx"], {}).get(period_type, {}).get(period_key, 0) or 0
                )

                if layout_item["type"] == "compare":
                    value = compare_value
                    is_delta = False
                    is_percent = False
                elif layout_item["type"] == "delta_abs":
                    value = compare_value - base_value
                    is_delta = True
                    is_percent = False
                else:
                    value = (compare_value - base_value) / base_value if abs(base_value) > 0.0001 else None
                    is_delta = True
                    is_percent = True

            _draw_number_cell(
                ws=ws,
                row=row_idx,
                col=col,
                value=value,
                fill=fill,
                font=value_font,
                negative_font=negative_value_font,
                border=border,
                is_delta=is_delta,
                is_percent=is_percent,
            )

        current_col += block_content_width

        if period_index < len(period_keys) - 1:
            _paint_spacer_cell(ws.cell(row=row_idx, column=current_col))
            current_col += 1


def _draw_total_row_total_sheet(ws, row_idx, compare_rows, compare_cols):
    fill = FILLS["total"]
    border = BORDERS["bottom_medium"]

    label = ws.cell(row=row_idx, column=1, value="ИТОГО")
    label.fill = fill
    label.border = border
    label.font = FONTS["total"]
    label.alignment = ALIGNMENTS["left"]

    note = ws.cell(row=row_idx, column=2, value=None)
    note.fill = fill
    note.border = border

    _paint_spacer_cell(ws.cell(row=row_idx, column=3))

    activity_rows = [r for r in compare_rows if r["row_type"] == "activity"]
    base_total = sum(float(r["values"].get(0, {}).get("total", 0) or 0) for r in activity_rows)

    for layout_item in compare_cols:
        col = layout_item["col"]

        if layout_item["type"] == "spacer":
            _paint_spacer_cell(ws.cell(row=row_idx, column=col))
            continue

        if layout_item["type"] == "base":
            value = base_total
            is_delta = False
            is_percent = False
        else:
            compare_total = sum(
                float(r["values"].get(layout_item["version_idx"], {}).get("total", 0) or 0)
                for r in activity_rows
            )

            if layout_item["type"] == "compare":
                value = compare_total
                is_delta = False
                is_percent = False
            elif layout_item["type"] == "delta_abs":
                value = compare_total - base_total
                is_delta = True
                is_percent = False
            else:
                value = (compare_total - base_total) / base_total if abs(base_total) > 0.0001 else None
                is_delta = True
                is_percent = True

        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.fill = _get_delta_fill(value, fill) if is_delta else fill
        cell.border = border
        cell.font = _get_value_font(value, FONTS["total"], FONTS["negative_total"])
        cell.alignment = ALIGNMENTS["right"]
        cell.number_format = "0.0%;[Red]-0.0%" if is_percent else FORMATS["money_int"]


def _draw_total_row_period_sheet(ws, row_idx, compare_rows, block_cols, period_keys, period_type, block_content_width):
    fill = FILLS["total"]
    border = BORDERS["bottom_medium"]

    label = ws.cell(row=row_idx, column=1, value="ИТОГО")
    label.fill = fill
    label.border = border
    label.font = FONTS["total"]
    label.alignment = ALIGNMENTS["left"]

    note = ws.cell(row=row_idx, column=2, value=None)
    note.fill = fill
    note.border = border

    _paint_spacer_cell(ws.cell(row=row_idx, column=3))

    activity_rows = [r for r in compare_rows if r["row_type"] == "activity"]
    current_col = DATA_START_COL

    for period_index, period_key in enumerate(period_keys):
        base_total = sum(
            float(r["values"].get(0, {}).get(period_type, {}).get(period_key, 0) or 0)
            for r in activity_rows
        )

        for layout_item in block_cols:
            col = current_col + layout_item["offset"]

            if layout_item["type"] == "base":
                value = base_total
                is_delta = False
                is_percent = False
            else:
                compare_total = sum(
                    float(r["values"].get(layout_item["version_idx"], {}).get(period_type, {}).get(period_key, 0) or 0)
                    for r in activity_rows
                )

                if layout_item["type"] == "compare":
                    value = compare_total
                    is_delta = False
                    is_percent = False
                elif layout_item["type"] == "delta_abs":
                    value = compare_total - base_total
                    is_delta = True
                    is_percent = False
                else:
                    value = (compare_total - base_total) / base_total if abs(base_total) > 0.0001 else None
                    is_delta = True
                    is_percent = True

            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = _get_delta_fill(value, fill) if is_delta else fill
            cell.border = border
            cell.font = _get_value_font(value, FONTS["total"], FONTS["negative_total"])
            cell.alignment = ALIGNMENTS["right"]
            cell.number_format = "0.0%;[Red]-0.0%" if is_percent else FORMATS["money_int"]

        current_col += block_content_width

        if period_index < len(period_keys) - 1:
            _paint_spacer_cell(ws.cell(row=row_idx, column=current_col))
            current_col += 1


def _build_total_sheet(wb, versions_data, compare_rows):
    ws = wb.create_sheet("ARTICLES_COMPARE")
    ws.sheet_view.showGridLines = False

    draw_back_button(ws, cell="A1", text="← SUMMARY", target_sheet="SUMMARY_COMPARE")

    compare_cols, last_col = _build_total_compare_columns(versions_data, start_col=DATA_START_COL)

    widths = {
        "A": 54,
        "B": 10,
        "C": SPACER_WIDTH,
    }
    for item in compare_cols:
        col_letter = ws.cell(row=1, column=item["col"]).column_letter
        widths[col_letter] = 18 if item["type"] != "spacer" else SPACER_WIDTH

    set_column_widths(ws, widths)
    set_row_heights(ws, {
        1: 20,
        2: 26,
        3: 18,
        4: 18,
        5: 10,
        6: 8,
        7: 28,
    })

    draw_sheet_header(
        ws,
        title="СРАВНЕНИЕ БЮДЖЕТОВ",
        subtitle="Постатейное сравнение выбранных версий — итог за период",
        note="",
    )

    _paint_vertical_spacer(ws, 3, 1, max(7, ws.max_row + 50))
    _draw_separator_row(ws, 6, last_col)
    _draw_total_sheet_header(ws, 7, compare_cols)

    row_idx = 8
    prev_activity = None
    prev_operation = None

    for item in compare_rows:
        if item["row_type"] == "activity":
            if prev_activity is not None:
                _draw_separator_row(ws, row_idx, last_col)
                _paint_spacer_cell(ws.cell(row=row_idx, column=3))
                row_idx += 1
            prev_activity = item["label"]
            prev_operation = None

        elif item["row_type"] == "operation":
            if prev_operation is not None:
                _draw_separator_row(ws, row_idx, last_col)
                _paint_spacer_cell(ws.cell(row=row_idx, column=3))
                row_idx += 1
            prev_operation = item["label"]

        _draw_total_compare_row(ws, row_idx, item, compare_cols)
        row_idx += 1

    _draw_separator_row(ws, row_idx, last_col)
    _paint_spacer_cell(ws.cell(row=row_idx, column=3))
    row_idx += 1
    _draw_total_row_total_sheet(ws, row_idx, compare_rows, compare_cols)

    ws.freeze_panes = "D8"


def _build_period_sheet(
    wb,
    sheet_name,
    title,
    subtitle,
    note,
    versions_data,
    compare_rows,
    period_type,
    period_keys,
    period_labels
):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    draw_back_button(ws, cell="A1", text="← SUMMARY", target_sheet="SUMMARY_COMPARE")

    block_cols, block_content_width = _build_period_block_columns(versions_data)
    spacer_count = max(len(period_keys) - 1, 0)
    last_col = 3 + (block_content_width * len(period_keys)) + spacer_count

    widths = {
        "A": 54,
        "B": 10,
        "C": SPACER_WIDTH,
    }

    current_col = DATA_START_COL
    for period_index in range(len(period_keys)):
        for item in block_cols:
            col_idx = current_col + item["offset"]
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            widths[col_letter] = 18

        current_col += block_content_width

        if period_index < len(period_keys) - 1:
            col_letter = ws.cell(row=1, column=current_col).column_letter
            widths[col_letter] = SPACER_WIDTH
            current_col += 1

    set_column_widths(ws, widths)
    set_row_heights(ws, {
        1: 20,
        2: 26,
        3: 18,
        4: 18,
        5: 10,
        6: 8,
        7: 24,
        8: 30,
    })

    draw_sheet_header(
        ws,
        title=title,
        subtitle=subtitle,
        note=note,
    )

    _paint_vertical_spacer(ws, 3, 1, max(9, ws.max_row + 50))
    _draw_separator_row(ws, 6, last_col)
    _draw_period_sheet_header(ws, 7, block_cols, period_keys, period_labels, block_content_width)

    row_idx = 9
    prev_activity = None
    prev_operation = None

    for item in compare_rows:
        if item["row_type"] == "activity":
            if prev_activity is not None:
                _draw_separator_row(ws, row_idx, last_col)
                _paint_spacer_cell(ws.cell(row=row_idx, column=3))
                row_idx += 1
            prev_activity = item["label"]
            prev_operation = None

        elif item["row_type"] == "operation":
            if prev_operation is not None:
                _draw_separator_row(ws, row_idx, last_col)
                _paint_spacer_cell(ws.cell(row=row_idx, column=3))
                row_idx += 1
            prev_operation = item["label"]

        _draw_period_compare_row(
            ws=ws,
            row_idx=row_idx,
            item=item,
            block_cols=block_cols,
            period_keys=period_keys,
            period_type=period_type,
            block_content_width=block_content_width,
        )
        row_idx += 1

    _draw_separator_row(ws, row_idx, last_col)
    _paint_spacer_cell(ws.cell(row=row_idx, column=3))
    row_idx += 1
    _draw_total_row_period_sheet(
        ws=ws,
        row_idx=row_idx,
        compare_rows=compare_rows,
        block_cols=block_cols,
        period_keys=period_keys,
        period_type=period_type,
        block_content_width=block_content_width,
    )

    ws.freeze_panes = "D9"


def build_compare_articles_sheet(wb, versions_data, detail_sheet_map):
    compare_rows = _collect_compare_rows(versions_data, detail_sheet_map)

    _build_total_sheet(wb, versions_data, compare_rows)

    month_keys, month_labels = _union_month_keys(versions_data)
    _build_period_sheet(
        wb=wb,
        sheet_name="MONTHS_COMPARE",
        title="СРАВНЕНИЕ БЮДЖЕТОВ ПО МЕСЯЦАМ",
        subtitle="Постатейное сравнение выбранных версий по месяцам",
        note="",
        versions_data=versions_data,
        compare_rows=compare_rows,
        period_type="months",
        period_keys=month_keys,
        period_labels=month_labels,
    )

    quarter_keys, quarter_labels = _union_quarter_keys(versions_data)
    _build_period_sheet(
        wb=wb,
        sheet_name="QUARTERS_COMPARE",
        title="СРАВНЕНИЕ БЮДЖЕТОВ ПО КВАРТАЛАМ",
        subtitle="Постатейное сравнение выбранных версий по кварталам",
        note="",
        versions_data=versions_data,
        compare_rows=compare_rows,
        period_type="quarters",
        period_keys=quarter_keys,
        period_labels=quarter_labels,
    )