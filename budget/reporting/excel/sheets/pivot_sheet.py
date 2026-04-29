# budget/reporting/excel/sheets/pivot_sheet.py

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from budget.reporting.excel.styles.helpers import (
    set_column_widths,
    set_row_heights,
    draw_sheet_header,
)
from budget.reporting.excel.styles.theme import (
    FILLS,
    FONTS,
    BORDERS,
    ALIGNMENTS,
    FORMATS,
)


def _label_alignment(level=0):
    return Alignment(horizontal="left", vertical="center", indent=level)


def _draw_separator_row(ws, row_idx, last_col):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = FILLS["none"]
        cell.border = BORDERS["none"]
    ws.row_dimensions[row_idx].height = 6


def _hide_columns(ws, months_layout, total_fact_col):
    """Скрывает колонки с фактом, дельтами и YTD план/факт по месяцам"""
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = True

    hidden_types = [
        "fact",
        "delta_rub",
        "delta_pct",
        "ytd_plan",
        "ytd_fact",
    ]

    for item in months_layout:
        if item["type"] in hidden_types:
            col_letter = get_column_letter(item["value_col"])
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].outline_level = 1

    col_letter = get_column_letter(total_fact_col)
    ws.column_dimensions[col_letter].hidden = True
    ws.column_dimensions[col_letter].outline_level = 1


def _build_month_layout(months):
    """Создает layout с раздельными колонками для рублей и процентов"""
    layout = []
    col = 3

    for month in months:
        layout.append({
            "type": "plan",
            "month": month,
            "label": f"{month}\nПлан",
            "value_col": col,
            "hidden": False,
        })
        col += 1

        layout.append({
            "type": "fact",
            "month": month,
            "label": f"{month}\nФакт",
            "value_col": col,
            "hidden": True,
        })
        col += 1

        layout.append({
            "type": "delta_rub",
            "month": month,
            "label": f"{month}\nΔ, руб",
            "value_col": col,
            "hidden": True,
        })
        col += 1

        layout.append({
            "type": "delta_pct",
            "month": month,
            "label": f"{month}\nΔ, %",
            "value_col": col,
            "hidden": True,
        })
        col += 1

        layout.append({
            "type": "ytd_plan",
            "month": month,
            "label": f"{month}\nYTD План",
            "value_col": col,
            "hidden": True,
        })
        col += 1

        layout.append({
            "type": "ytd_fact",
            "month": month,
            "label": f"{month}\nYTD Факт",
            "value_col": col,
            "hidden": True,
        })
        col += 1

        layout.append({
            "type": "ytd_rub",
            "month": month,
            "label": f"{month}\nYTD Δ, руб",
            "value_col": col,
            "hidden": False,
        })
        col += 1

        layout.append({
            "type": "ytd_pct",
            "month": month,
            "label": f"{month}\nYTD Δ, %",
            "value_col": col,
            "hidden": False,
        })
        col += 1

        layout.append({
            "type": "spacer",
            "value_col": col,
            "hidden": False,
        })
        col += 1

    total_plan_col = col
    col += 1
    total_fact_col = col
    col += 1
    total_delta_rub_col = col
    col += 1
    total_delta_pct_col = col

    return (
        layout,
        total_plan_col,
        total_fact_col,
        total_delta_rub_col,
        total_delta_pct_col,
    )


def _draw_header(
    ws,
    row_idx,
    months_layout,
    total_plan_col,
    total_fact_col,
    total_delta_rub_col,
    total_delta_pct_col,
):
    cell = ws.cell(row=row_idx, column=1, value="Статья")
    cell.fill = FILLS["header"]
    cell.font = FONTS["header_white"]
    cell.border = BORDERS["thin"]
    cell.alignment = ALIGNMENTS["center"]

    note_cell = ws.cell(row=row_idx, column=2, value="Прим.")
    note_cell.fill = FILLS["header"]
    note_cell.font = FONTS["header_white"]
    note_cell.border = BORDERS["thin"]
    note_cell.alignment = ALIGNMENTS["center"]

    for item in months_layout:
        col = item["value_col"]

        if item["type"] == "spacer":
            spacer = ws.cell(row=row_idx, column=col, value=None)
            spacer.fill = FILLS["none"]
            spacer.border = BORDERS["none"]
        else:
            cell = ws.cell(row=row_idx, column=col, value=item["label"])
            cell.fill = FILLS["header"]
            cell.font = FONTS["header_white"]
            cell.border = BORDERS["thin"]
            cell.alignment = ALIGNMENTS["center_wrap"]

    total_headers = [
        (total_plan_col, "ИТОГО\nПлан"),
        (total_fact_col, "ИТОГО\nФакт"),
        (total_delta_rub_col, "ИТОГО\nΔ, руб"),
        (total_delta_pct_col, "ИТОГО\nΔ, %"),
    ]

    for col, title in total_headers:
        total_cell = ws.cell(row=row_idx, column=col, value=title)
        total_cell.fill = FILLS["header"]
        total_cell.font = FONTS["header_white"]
        total_cell.border = BORDERS["thin"]
        total_cell.alignment = ALIGNMENTS["center_wrap"]


def _get_delta_fill(delta_rub, default_fill, plan_value=0):
    try:
        delta_rub = float(delta_rub or 0)
        plan_value = float(plan_value or 0)
    except (TypeError, ValueError):
        return default_fill

    if delta_rub == 0:
        return default_fill

    if plan_value == 0:
        return FILLS["delta_green"]

    return FILLS["delta_green"] if delta_rub > 0 else FILLS["delta_red"]


def _format_percent(value_pct):
    if value_pct is None:
        return None

    if value_pct >= 0:
        return f"+{value_pct * 100:.1f}%"

    return f"{value_pct * 100:.1f}%"


def _calculate_ytd_for_row(plan_months, fact_months, months):
    """Расчет YTD: план, факт, дельта в рублях и процентах"""
    ytd_plan = {}
    ytd_fact = {}
    ytd_rub = {}
    ytd_pct = {}

    running_plan = 0
    running_fact = 0
    running_rub = 0

    for month in months:
        plan_val = plan_months.get(month, 0)
        fact_val = fact_months.get(month, 0)

        running_plan += plan_val
        running_fact += fact_val
        running_rub = running_fact - running_plan

        ytd_plan[month] = running_plan
        ytd_fact[month] = running_fact
        ytd_rub[month] = running_rub

        if running_plan != 0:
            ytd_pct[month] = running_rub / abs(running_plan)
        else:
            ytd_pct[month] = None if running_rub == 0 else (1 if running_rub > 0 else -1)

    return ytd_plan, ytd_fact, ytd_rub, ytd_pct


def _draw_value_cell(
    ws,
    row,
    col,
    value,
    fill,
    font,
    border,
    alignment=ALIGNMENTS["right"],
    number_format=None,
):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill
    cell.font = font
    cell.border = border
    cell.alignment = alignment

    if number_format:
        cell.number_format = number_format


def _draw_data_row(
    ws,
    row_idx,
    item,
    months_layout,
    months,
    total_plan_col,
    total_fact_col,
    total_delta_rub_col,
    total_delta_pct_col,
):
    row_type = item["row_type"]

    if row_type == "activity":
        fill = FILLS["section"]
        label_font = FONTS["section"]
        value_font = FONTS["bold"]
        border = BORDERS["bottom_medium"]
        height = 21
        label_indent = 1

    elif row_type == "operation":
        fill = FILLS["none"]
        label_font = FONTS["bold"]
        value_font = FONTS["bold"]
        border = BORDERS["thin"]
        height = 19
        label_indent = 2

    elif row_type == "subitem":
        fill = FILLS["none"]
        label_font = FONTS["subitem"]
        value_font = FONTS["subitem"]
        border = BORDERS["thin"]
        height = 17
        label_indent = 6

    else:
        fill = FILLS["none"]
        label_font = FONTS["normal"]
        value_font = FONTS["normal"]
        border = BORDERS["thin"]
        height = 18
        label_indent = 3

    ws.row_dimensions[row_idx].height = height

    label = ws.cell(row=row_idx, column=1, value=item["label"])
    label.fill = fill
    label.font = label_font
    label.border = border
    label.alignment = _label_alignment(label_indent)

    note = item.get("note", "")
    note_cell = ws.cell(row=row_idx, column=2, value=note)
    note_cell.fill = fill
    note_cell.border = border
    note_cell.alignment = ALIGNMENTS["center"]

    if note and item.get("sheet_name"):
        note_cell.font = FONTS["back"]
        note_cell.hyperlink = f"#'{item['sheet_name']}'!A1"
    else:
        note_cell.font = FONTS["normal"]

    ytd_plan, ytd_fact, ytd_rub, ytd_pct = _calculate_ytd_for_row(
        item["plan_months"],
        item["fact_months"],
        months,
    )

    for layout_item in months_layout:
        col = layout_item["value_col"]

        if layout_item["type"] == "spacer":
            _draw_value_cell(
                ws,
                row_idx,
                col,
                None,
                FILLS["none"],
                value_font,
                BORDERS["none"],
            )
            continue

        month = layout_item["month"]

        if layout_item["type"] == "plan":
            value = item["plan_months"].get(month, 0)
            _draw_value_cell(
                ws,
                row_idx,
                col,
                value,
                fill,
                value_font,
                border,
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "fact":
            value = item["fact_months"].get(month, 0)
            _draw_value_cell(
                ws,
                row_idx,
                col,
                value,
                fill,
                value_font,
                border,
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "delta_rub":
            plan_val = item["plan_months"].get(month, 0)
            fact_val = item["fact_months"].get(month, 0)
            delta_rub = fact_val - plan_val
            fill_color = _get_delta_fill(delta_rub, fill, plan_val)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                delta_rub,
                fill_color,
                value_font,
                border,
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "delta_pct":
            plan_val = item["plan_months"].get(month, 0)
            fact_val = item["fact_months"].get(month, 0)
            delta_rub = fact_val - plan_val

            if plan_val != 0:
                delta_pct = delta_rub / abs(plan_val)
            else:
                delta_pct = None if delta_rub == 0 else (1 if delta_rub > 0 else -1)

            formatted_pct = _format_percent(delta_pct)
            fill_color = _get_delta_fill(delta_rub, fill, plan_val)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                formatted_pct,
                fill_color,
                value_font,
                border,
                ALIGNMENTS["center"],
            )

        elif layout_item["type"] == "ytd_plan":
            value = ytd_plan.get(month, 0)
            _draw_value_cell(
                ws,
                row_idx,
                col,
                value,
                fill,
                value_font,
                border,
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "ytd_fact":
            value = ytd_fact.get(month, 0)
            _draw_value_cell(
                ws,
                row_idx,
                col,
                value,
                fill,
                value_font,
                border,
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "ytd_rub":
            plan_val = ytd_plan.get(month, 0)
            delta_rub = ytd_rub.get(month, 0)
            fill_color = _get_delta_fill(delta_rub, fill, plan_val)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                delta_rub,
                fill_color,
                value_font,
                border,
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "ytd_pct":
            plan_val = ytd_plan.get(month, 0)
            delta_pct = ytd_pct.get(month)
            formatted_pct = _format_percent(delta_pct)
            delta_rub = ytd_rub.get(month, 0)
            fill_color = _get_delta_fill(delta_rub, fill, plan_val)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                formatted_pct,
                fill_color,
                value_font,
                border,
                ALIGNMENTS["center"],
            )

    plan_total = item["plan_total"] or 0
    fact_total = item["fact_total"] or 0
    delta_total_rub = fact_total - plan_total

    if plan_total != 0:
        delta_total_pct = delta_total_rub / abs(plan_total)
    else:
        delta_total_pct = None if delta_total_rub == 0 else (1 if delta_total_rub > 0 else -1)

    total_font = FONTS["bold"] if row_type in ("activity", "operation") else value_font

    _draw_value_cell(
        ws,
        row_idx,
        total_plan_col,
        plan_total,
        fill,
        total_font,
        border,
        ALIGNMENTS["right"],
        FORMATS["money_int"],
    )

    _draw_value_cell(
        ws,
        row_idx,
        total_fact_col,
        fact_total,
        fill,
        total_font,
        border,
        ALIGNMENTS["right"],
        FORMATS["money_int"],
    )

    delta_fill = _get_delta_fill(delta_total_rub, fill, plan_total)

    _draw_value_cell(
        ws,
        row_idx,
        total_delta_rub_col,
        delta_total_rub,
        delta_fill,
        total_font,
        border,
        ALIGNMENTS["right"],
        FORMATS["money_int"],
    )

    formatted_pct = _format_percent(delta_total_pct)

    _draw_value_cell(
        ws,
        row_idx,
        total_delta_pct_col,
        formatted_pct,
        delta_fill,
        total_font,
        border,
        ALIGNMENTS["center"],
    )


def _make_subitem_row(subitem):
    return {
        "row_type": "subitem",
        "level": 3,
        "label": subitem.get("label", "—"),
        "note": "",
        "sheet_name": None,
        "plan_months": subitem.get("plan_months", {}),
        "fact_months": subitem.get("fact_months", {}),
        "plan_total": subitem.get("plan_total", 0),
        "fact_total": subitem.get("fact_total", 0),
    }


def build_pivot_sheet(wb, data):
    ws = wb.create_sheet("БЮДЖЕТ")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showOutlineSymbols = True

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = True

    pivot = data["gl_pivot"]
    months = pivot["months"]
    rows = pivot["rows"]

    detail_lookup = {
        detail["sheet_name"]: detail
        for detail in pivot.get("detail_sheets", [])
        if detail.get("sheet_name")
    }

    (
        months_layout,
        total_plan_col,
        total_fact_col,
        total_delta_rub_col,
        total_delta_pct_col,
    ) = _build_month_layout(months)

    last_col = total_delta_pct_col

    widths = {"A": 52, "B": 8}

    for item in months_layout:
        col_letter = get_column_letter(item["value_col"])

        if item["type"] == "spacer":
            widths[col_letter] = 3
        elif "pct" in item["type"]:
            widths[col_letter] = 12
        else:
            widths[col_letter] = 16

    widths[get_column_letter(total_plan_col)] = 16
    widths[get_column_letter(total_fact_col)] = 16
    widths[get_column_letter(total_delta_rub_col)] = 16
    widths[get_column_letter(total_delta_pct_col)] = 12

    set_column_widths(ws, widths)

    set_row_heights(ws, {
        1: 20,
        2: 26,
        3: 18,
        4: 18,
        5: 10,
        6: 6,
        7: 40,
    })

    draw_sheet_header(
        ws,
        title="БЮДЖЕТ",
        subtitle=f'Версия: {data["version"]["number"]}',
        note="Δ = отклонение. YTD Δ = накопленное отклонение с начала года.",
    )

    _draw_separator_row(ws, 6, last_col)

    _draw_header(
        ws,
        7,
        months_layout,
        total_plan_col,
        total_fact_col,
        total_delta_rub_col,
        total_delta_pct_col,
    )

    row_idx = 8
    prev_activity = None
    prev_operation = None

    for item in rows:
        if item["row_type"] == "activity":
            if prev_activity is not None:
                _draw_separator_row(ws, row_idx, last_col)
                row_idx += 1

            prev_activity = item["label"]
            prev_operation = None

        elif item["row_type"] == "operation":
            if prev_operation is not None:
                _draw_separator_row(ws, row_idx, last_col)
                row_idx += 1

            prev_operation = item["label"]

        _draw_data_row(
            ws,
            row_idx,
            item,
            months_layout,
            months,
            total_plan_col,
            total_fact_col,
            total_delta_rub_col,
            total_delta_pct_col,
        )

        row_idx += 1

        if item["row_type"] == "item" and item.get("sheet_name"):
            detail = detail_lookup.get(item["sheet_name"])

            if detail:
                for subitem in detail.get("rows", []):
                    _draw_data_row(
                        ws,
                        row_idx,
                        _make_subitem_row(subitem),
                        months_layout,
                        months,
                        total_plan_col,
                        total_fact_col,
                        total_delta_rub_col,
                        total_delta_pct_col,
                    )

                    ws.row_dimensions[row_idx].hidden = True
                    ws.row_dimensions[row_idx].outline_level = 1
                    row_idx += 1

    _draw_separator_row(ws, row_idx, last_col)
    row_idx += 1

    gt_label = ws.cell(row=row_idx, column=1, value="ИТОГО")
    gt_label.fill = FILLS["total"]
    gt_label.border = BORDERS["bottom_medium"]
    gt_label.font = FONTS["total"]
    gt_label.alignment = ALIGNMENTS["left"]

    gt_note = ws.cell(row=row_idx, column=2, value=None)
    gt_note.fill = FILLS["total"]
    gt_note.border = BORDERS["bottom_medium"]

    grand_plan_sum = sum(pivot["grand_plan_total"].values())
    grand_fact_sum = sum(pivot["grand_fact_total"].values())
    grand_delta_rub = grand_fact_sum - grand_plan_sum

    if grand_plan_sum != 0:
        grand_delta_pct = grand_delta_rub / abs(grand_plan_sum)
    else:
        grand_delta_pct = None if grand_delta_rub == 0 else (1 if grand_delta_rub > 0 else -1)

    ytd_running_plan = 0
    ytd_running_fact = 0
    ytd_running_rub = 0

    for layout_item in months_layout:
        col = layout_item["value_col"]

        if layout_item["type"] == "spacer":
            _draw_value_cell(
                ws,
                row_idx,
                col,
                None,
                FILLS["none"],
                FONTS["total"],
                BORDERS["none"],
            )
            continue

        month = layout_item["month"]

        if layout_item["type"] == "plan":
            value = pivot["grand_plan_total"].get(month, 0)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                value,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "fact":
            value = pivot["grand_fact_total"].get(month, 0)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                value,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "delta_rub":
            plan_val = pivot["grand_plan_total"].get(month, 0)
            fact_val = pivot["grand_fact_total"].get(month, 0)
            delta_rub = fact_val - plan_val

            _draw_value_cell(
                ws,
                row_idx,
                col,
                delta_rub,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "delta_pct":
            plan_val = pivot["grand_plan_total"].get(month, 0)
            fact_val = pivot["grand_fact_total"].get(month, 0)
            delta_rub = fact_val - plan_val

            if plan_val != 0:
                delta_pct = delta_rub / abs(plan_val)
            else:
                delta_pct = None if delta_rub == 0 else (1 if delta_rub > 0 else -1)

            formatted_pct = _format_percent(delta_pct)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                formatted_pct,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["center"],
            )

        elif layout_item["type"] == "ytd_plan":
            value = pivot["grand_plan_total"].get(month, 0)
            ytd_running_plan += value

            _draw_value_cell(
                ws,
                row_idx,
                col,
                ytd_running_plan,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "ytd_fact":
            value = pivot["grand_fact_total"].get(month, 0)
            ytd_running_fact += value

            _draw_value_cell(
                ws,
                row_idx,
                col,
                ytd_running_fact,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "ytd_rub":
            ytd_running_rub = ytd_running_fact - ytd_running_plan

            _draw_value_cell(
                ws,
                row_idx,
                col,
                ytd_running_rub,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["right"],
                FORMATS["money_int"],
            )

        elif layout_item["type"] == "ytd_pct":
            if ytd_running_plan != 0:
                ytd_pct = ytd_running_rub / abs(ytd_running_plan)
            else:
                ytd_pct = None if ytd_running_rub == 0 else (1 if ytd_running_rub > 0 else -1)

            formatted_pct = _format_percent(ytd_pct)

            _draw_value_cell(
                ws,
                row_idx,
                col,
                formatted_pct,
                FILLS["total"],
                FONTS["total"],
                BORDERS["bottom_medium"],
                ALIGNMENTS["center"],
            )

    _draw_value_cell(
        ws,
        row_idx,
        total_plan_col,
        grand_plan_sum,
        FILLS["total"],
        FONTS["total"],
        BORDERS["bottom_medium"],
        ALIGNMENTS["right"],
        FORMATS["money_int"],
    )

    _draw_value_cell(
        ws,
        row_idx,
        total_fact_col,
        grand_fact_sum,
        FILLS["total"],
        FONTS["total"],
        BORDERS["bottom_medium"],
        ALIGNMENTS["right"],
        FORMATS["money_int"],
    )

    _draw_value_cell(
        ws,
        row_idx,
        total_delta_rub_col,
        grand_delta_rub,
        FILLS["total"],
        FONTS["total"],
        BORDERS["bottom_medium"],
        ALIGNMENTS["right"],
        FORMATS["money_int"],
    )

    formatted_grand_pct = _format_percent(grand_delta_pct)

    _draw_value_cell(
        ws,
        row_idx,
        total_delta_pct_col,
        formatted_grand_pct,
        FILLS["total"],
        FONTS["total"],
        BORDERS["bottom_medium"],
        ALIGNMENTS["center"],
    )

    _hide_columns(ws, months_layout, total_fact_col)

    ws.freeze_panes = "C8"