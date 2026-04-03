from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from budget.reporting.excel.styles.helpers import set_column_widths


THIN = Side(style="thin", color="D9DEE8")
MEDIUM = Side(style="medium", color="AEB7C6")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM_MEDIUM = Border(bottom=MEDIUM)

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="1F2937")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=False, color="6B7280")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="1F2937")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color="1F2937")
FONT_NORMAL = Font(name="Calibri", size=11, color="1F2937")

FILL_HEADER = PatternFill("solid", fgColor="EAF3E6")
FILL_SECTION = PatternFill("solid", fgColor="DDEBDD")
FILL_TOTAL = PatternFill("solid", fgColor="F4F8F2")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

MONEY_FORMAT = '# ##0.00'


def _scenario_label(data):
    scenario_raw = (data.get("revenue_param") or {}).get("scenario", "base")
    scenario_map = {
        "base": "Базовый",
        "optimistic": "Оптимистичный",
        "conservative": "Консервативный",
    }
    return scenario_map.get(str(scenario_raw).lower(), str(scenario_raw))


def _version_label(version_data):
    version = version_data["data"]["version"]
    scenario = _scenario_label(version_data["data"])
    number = version.get("number") or "—"
    return f"{number} ({scenario})"


def _all_months(versions_data):
    month_order = []
    seen = set()

    for item in versions_data:
        months = item["data"]["gl_pivot"].get("months", [])
        for month in months:
            if month not in seen:
                seen.add(month)
                month_order.append(month)

    return month_order


def _write_cell(ws, row, col, value=None, font=None, fill=None, border=None, alignment=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)

    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format

    return cell


def build_compare_summary_sheet(wb, versions_data):
    ws = wb.create_sheet("SUMMARY_COMPARE")
    ws.sheet_view.showGridLines = False

    set_column_widths(ws, {
        "A": 28,
        "B": 20,
        "C": 20,
        "D": 20,
        "E": 20,
        "F": 20,
        "G": 20,
        "H": 20,
    })

    # Заголовок
    ws.merge_cells("A2:H2")
    ws["A2"] = "СРАВНЕНИЕ БЮДЖЕТОВ"
    ws["A2"].font = FONT_TITLE
    ws["A2"].alignment = ALIGN_LEFT

    ws.merge_cells("A3:H3")
    ws["A3"] = "Сравнение выбранных версий бюджета по месяцам и итогам"
    ws["A3"].font = FONT_SUBTITLE
    ws["A3"].alignment = ALIGN_LEFT

    for col in range(1, 9):
        ws.cell(row=4, column=col).border = BORDER_BOTTOM_MEDIUM

    row_idx = 6

    # Блок версий
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    _write_cell(ws, row_idx, 1, "ВЫБРАННЫЕ ВЕРСИИ", FONT_HEADER, FILL_SECTION, BORDER_THIN, ALIGN_LEFT)
    row_idx += 1

    _write_cell(ws, row_idx, 1, "Версия", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)
    _write_cell(ws, row_idx, 2, "Сценарий", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)
    _write_cell(ws, row_idx, 3, "Дата начала", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_CENTER)
    _write_cell(ws, row_idx, 4, "Дата окончания", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_CENTER)
    _write_cell(ws, row_idx, 5, "Статус", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)
    row_idx += 1

    for item in versions_data:
        version = item["data"]["version"]
        scenario = _scenario_label(item["data"])

        _write_cell(ws, row_idx, 1, version.get("number") or "—", FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)
        _write_cell(ws, row_idx, 2, scenario, FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)

        c3 = _write_cell(ws, row_idx, 3, version.get("date_from"), FONT_NORMAL, None, BORDER_THIN, ALIGN_CENTER)
        c4 = _write_cell(ws, row_idx, 4, version.get("date_to"), FONT_NORMAL, None, BORDER_THIN, ALIGN_CENTER)

        if hasattr(c3.value, "year"):
            c3.number_format = "DD.MM.YYYY"
        if hasattr(c4.value, "year"):
            c4.number_format = "DD.MM.YYYY"

        _write_cell(ws, row_idx, 5, "—", FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)
        row_idx += 1

    row_idx += 2

    # Итоги по версиям
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    _write_cell(ws, row_idx, 1, "ИТОГИ ПО ВЕРСИЯМ", FONT_HEADER, FILL_SECTION, BORDER_THIN, ALIGN_LEFT)
    row_idx += 1

    _write_cell(ws, row_idx, 1, "Показатель", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)

    version_col_map = {}
    col_idx = 2
    for item in versions_data:
        label = _version_label(item)
        version_col_map[label] = item
        _write_cell(ws, row_idx, col_idx, label, FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_CENTER)
        col_idx += 1

    row_idx += 1

    metrics = [
        ("Итого план", "grand_plan_sum"),
        ("Итого факт", "grand_fact_sum"),
        ("Итого отклонение", "grand_delta_sum"),
    ]

    for metric_label, metric_key in metrics:
        _write_cell(ws, row_idx, 1, metric_label, FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)

        col_idx = 2
        for item in versions_data:
            pivot = item["data"]["gl_pivot"]
            value = pivot.get(metric_key, 0)
            _write_cell(
                ws, row_idx, col_idx, value,
                FONT_NORMAL, None, BORDER_THIN, ALIGN_RIGHT, MONEY_FORMAT
            )
            col_idx += 1

        row_idx += 1

    row_idx += 2

    # Помесячное сравнение плана
    months = _all_months(versions_data)

    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    _write_cell(ws, row_idx, 1, "ПОМЕСЯЧНОЕ СРАВНЕНИЕ — ПЛАН", FONT_HEADER, FILL_SECTION, BORDER_THIN, ALIGN_LEFT)
    row_idx += 1

    _write_cell(ws, row_idx, 1, "Месяц", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)

    col_idx = 2
    for item in versions_data:
        label = _version_label(item)
        _write_cell(ws, row_idx, col_idx, label, FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_CENTER)
        col_idx += 1
    row_idx += 1

    for month in months:
        _write_cell(ws, row_idx, 1, month, FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)

        col_idx = 2
        for item in versions_data:
            pivot = item["data"]["gl_pivot"]
            value = pivot.get("grand_plan_total", {}).get(month, 0)
            _write_cell(
                ws, row_idx, col_idx, value,
                FONT_NORMAL, None, BORDER_THIN, ALIGN_RIGHT, MONEY_FORMAT
            )
            col_idx += 1
        row_idx += 1

    row_idx += 2

    # Помесячное сравнение факта
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    _write_cell(ws, row_idx, 1, "ПОМЕСЯЧНОЕ СРАВНЕНИЕ — ФАКТ", FONT_HEADER, FILL_SECTION, BORDER_THIN, ALIGN_LEFT)
    row_idx += 1

    _write_cell(ws, row_idx, 1, "Месяц", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)

    col_idx = 2
    for item in versions_data:
        label = _version_label(item)
        _write_cell(ws, row_idx, col_idx, label, FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_CENTER)
        col_idx += 1
    row_idx += 1

    for month in months:
        _write_cell(ws, row_idx, 1, month, FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)

        col_idx = 2
        for item in versions_data:
            pivot = item["data"]["gl_pivot"]
            value = pivot.get("grand_fact_total", {}).get(month, 0)
            _write_cell(
                ws, row_idx, col_idx, value,
                FONT_NORMAL, None, BORDER_THIN, ALIGN_RIGHT, MONEY_FORMAT
            )
            col_idx += 1
        row_idx += 1

    row_idx += 2

    # Помесячное сравнение отклонения
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    _write_cell(ws, row_idx, 1, "ПОМЕСЯЧНОЕ СРАВНЕНИЕ — ОТКЛОНЕНИЕ", FONT_HEADER, FILL_SECTION, BORDER_THIN, ALIGN_LEFT)
    row_idx += 1

    _write_cell(ws, row_idx, 1, "Месяц", FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_LEFT)

    col_idx = 2
    for item in versions_data:
        label = _version_label(item)
        _write_cell(ws, row_idx, col_idx, label, FONT_BOLD, FILL_HEADER, BORDER_THIN, ALIGN_CENTER)
        col_idx += 1
    row_idx += 1

    for month in months:
        _write_cell(ws, row_idx, 1, month, FONT_NORMAL, None, BORDER_THIN, ALIGN_LEFT)

        col_idx = 2
        for item in versions_data:
            pivot = item["data"]["gl_pivot"]
            value = pivot.get("grand_delta_total", {}).get(month, 0)
            _write_cell(
                ws, row_idx, col_idx, value,
                FONT_NORMAL, None, BORDER_THIN, ALIGN_RIGHT, MONEY_FORMAT
            )
            col_idx += 1
        row_idx += 1

    ws.freeze_panes = "A7"