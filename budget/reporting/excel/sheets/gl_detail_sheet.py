# # budget/reporting/excel/sheets/gl_detail_sheet.py

# from copy import copy
# from openpyxl.utils import get_column_letter

# from budget.reporting.excel.styles.helpers import (
#     set_column_widths,
#     set_row_heights,
#     draw_back_button,
#     draw_sheet_header,
# )
# from budget.reporting.excel.styles.theme import (
#     FILLS,
#     FONTS,
#     BORDERS,
#     ALIGNMENTS,
#     FORMATS,
# )


# def _get_delta_fill(delta_rub, default_fill, plan_value=0):
#     try:
#         delta_rub = float(delta_rub or 0)
#         plan_value = float(plan_value or 0)
#     except (TypeError, ValueError):
#         return default_fill

#     if delta_rub == 0:
#         return default_fill

#     if plan_value == 0:
#         return FILLS["delta_green"]

#     return FILLS["delta_green"] if delta_rub > 0 else FILLS["delta_red"]


# def _format_percent(value_pct):
#     if value_pct is None:
#         return None

#     return f"+{value_pct * 100:.1f}%" if value_pct >= 0 else f"{value_pct * 100:.1f}%"


# def _calculate_ytd_for_row(plan_months, fact_months, months):
#     ytd_rub = {}
#     ytd_pct = {}
#     running_rub = 0
#     running_plan = 0

#     for month in months:
#         plan_val = plan_months.get(month, 0)
#         fact_val = fact_months.get(month, 0)

#         running_rub += fact_val - plan_val
#         running_plan += plan_val

#         ytd_rub[month] = running_rub

#         if running_plan != 0:
#             ytd_pct[month] = running_rub / abs(running_plan)
#         else:
#             ytd_pct[month] = None if running_rub == 0 else (1 if running_rub > 0 else -1)

#     return ytd_rub, ytd_pct


# def _hide_fact_and_delta_columns(ws, month_layout, total_fact_col):
#     ws.sheet_properties.outlinePr.summaryBelow = False
#     ws.sheet_properties.outlinePr.summaryRight = True
#     ws.sheet_view.showOutlineSymbols = True

#     for item in month_layout:
#         if item["type"] in ("fact", "delta_rub", "delta_pct"):
#             col_letter = get_column_letter(item["col"])
#             ws.column_dimensions[col_letter].hidden = True
#             ws.column_dimensions[col_letter].outline_level = 1

#     col_letter = get_column_letter(total_fact_col)
#     ws.column_dimensions[col_letter].hidden = True
#     ws.column_dimensions[col_letter].outline_level = 1


# def _get_row_font(value, is_child=False):
#     if is_child:
#         return FONTS["counterparty_negative"] if value < 0 else FONTS["counterparty"]

#     return FONTS["negative"] if value < 0 else FONTS["normal"]


# def _get_label_alignment(is_child=False):
#     alignment = copy(ALIGNMENTS["left"])
#     if is_child:
#         alignment.indent = 2
#     return alignment


# def _draw_detail_value_row(
#     ws,
#     row_idx,
#     row,
#     month_layout,
#     months,
#     total_plan_col,
#     total_fact_col,
#     total_delta_rub_col,
#     total_delta_pct_col,
#     is_child=False,
# ):
#     label_value = row.get("label", "—")

#     label_cell = ws.cell(row=row_idx, column=1, value=label_value)
#     label_cell.fill = FILLS["none"]
#     label_cell.border = BORDERS["thin"]
#     label_cell.font = FONTS["counterparty"] if is_child else FONTS["normal"]
#     label_cell.alignment = _get_label_alignment(is_child=is_child)

#     ytd_rub, ytd_pct = _calculate_ytd_for_row(
#         row["plan_months"],
#         row["fact_months"],
#         months,
#     )

#     for layout_item in month_layout:
#         col_idx = layout_item["col"]

#         if layout_item["type"] == "spacer":
#             cell = ws.cell(row=row_idx, column=col_idx, value=None)
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]
#             continue

#         month = layout_item["month"]

#         if layout_item["type"] == "plan":
#             value = row["plan_months"].get(month, 0)
#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = FILLS["none"]
#             cell.font = _get_row_font(value, is_child=is_child)
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["thin"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "fact":
#             value = row["fact_months"].get(month, 0)
#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = FILLS["none"]
#             cell.font = _get_row_font(value, is_child=is_child)
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["thin"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "delta_rub":
#             plan_val = row["plan_months"].get(month, 0)
#             value = row["delta_months"].get(month, 0)

#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = _get_delta_fill(value, FILLS["none"], plan_val)
#             cell.font = _get_row_font(value, is_child=is_child)
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["thin"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "delta_pct":
#             if is_child:
#                 cell = ws.cell(row=row_idx, column=col_idx, value=None)
#                 cell.fill = FILLS["none"]
#                 cell.border = BORDERS["none"]
#                 continue

#             plan_val = row["plan_months"].get(month, 0)
#             delta_rub = row["delta_months"].get(month, 0)

#             if plan_val != 0:
#                 delta_pct = delta_rub / abs(plan_val)
#             else:
#                 delta_pct = None if delta_rub == 0 else (1 if delta_rub > 0 else -1)

#             cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
#             cell.fill = _get_delta_fill(delta_rub, FILLS["none"], plan_val)
#             cell.font = _get_row_font(delta_rub, is_child=is_child)
#             cell.alignment = ALIGNMENTS["center"]
#             cell.border = BORDERS["thin"]

#         elif layout_item["type"] == "ytd_rub":
#             plan_val = row["plan_months"].get(month, 0)
#             value = ytd_rub.get(month, 0)

#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = _get_delta_fill(value, FILLS["none"], plan_val)
#             cell.font = _get_row_font(value, is_child=is_child)
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["thin"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "ytd_pct":
#             if is_child:
#                 cell = ws.cell(row=row_idx, column=col_idx, value=None)
#                 cell.fill = FILLS["none"]
#                 cell.border = BORDERS["none"]
#                 continue

#             plan_val = row["plan_months"].get(month, 0)
#             delta_pct = ytd_pct.get(month)
#             delta_rub = ytd_rub.get(month, 0)

#             cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
#             cell.fill = _get_delta_fill(delta_rub, FILLS["none"], plan_val)
#             cell.font = _get_row_font(delta_rub, is_child=is_child)
#             cell.alignment = ALIGNMENTS["center"]
#             cell.border = BORDERS["thin"]

#     total_plan_value = row["plan_total"]
#     total_fact_value = row["fact_total"]
#     total_delta_value = row["delta_total"]

#     total_plan_cell = ws.cell(row=row_idx, column=total_plan_col, value=total_plan_value)
#     total_plan_cell.fill = FILLS["none"]
#     total_plan_cell.font = _get_row_font(total_plan_value, is_child=is_child)
#     total_plan_cell.alignment = ALIGNMENTS["right"]
#     total_plan_cell.border = BORDERS["thin"]
#     total_plan_cell.number_format = FORMATS["money_int"]

#     total_fact_cell = ws.cell(row=row_idx, column=total_fact_col, value=total_fact_value)
#     total_fact_cell.fill = FILLS["none"]
#     total_fact_cell.font = _get_row_font(total_fact_value, is_child=is_child)
#     total_fact_cell.alignment = ALIGNMENTS["right"]
#     total_fact_cell.border = BORDERS["thin"]
#     total_fact_cell.number_format = FORMATS["money_int"]

#     total_delta_cell = ws.cell(row=row_idx, column=total_delta_rub_col, value=total_delta_value)
#     total_delta_cell.fill = _get_delta_fill(total_delta_value, FILLS["none"], total_plan_value)
#     total_delta_cell.font = _get_row_font(total_delta_value, is_child=is_child)
#     total_delta_cell.alignment = ALIGNMENTS["right"]
#     total_delta_cell.border = BORDERS["thin"]
#     total_delta_cell.number_format = FORMATS["money_int"]

#     if not is_child:
#         if total_plan_value != 0:
#             total_delta_pct_value = total_delta_value / abs(total_plan_value)
#         else:
#             total_delta_pct_value = None if total_delta_value == 0 else (1 if total_delta_value > 0 else -1)

#         total_delta_pct_cell = ws.cell(
#             row=row_idx,
#             column=total_delta_pct_col,
#             value=_format_percent(total_delta_pct_value),
#         )
#         total_delta_pct_cell.fill = _get_delta_fill(total_delta_value, FILLS["none"], total_plan_value)
#         total_delta_pct_cell.font = _get_row_font(total_delta_value, is_child=is_child)
#         total_delta_pct_cell.alignment = ALIGNMENTS["center"]
#         total_delta_pct_cell.border = BORDERS["thin"]
#     else:
#         empty_cell = ws.cell(row=row_idx, column=total_delta_pct_col, value=None)
#         empty_cell.fill = FILLS["none"]
#         empty_cell.border = BORDERS["none"]

#     ws.row_dimensions[row_idx].height = 18 if not is_child else 17


# def build_gl_detail_sheet(wb, detail):
#     ws = wb.create_sheet(detail["sheet_name"])
#     ws.sheet_view.showGridLines = False
#     ws.sheet_view.showOutlineSymbols = True

#     months = detail["months"]

#     col = 2
#     month_layout = []

#     for month in months:
#         month_layout.append({"type": "plan", "month": month, "label": f"{month}\nПлан", "col": col})
#         col += 1
#         month_layout.append({"type": "fact", "month": month, "label": f"{month}\nФакт", "col": col})
#         col += 1
#         month_layout.append({"type": "delta_rub", "month": month, "label": f"{month}\nΔ, руб", "col": col})
#         col += 1
#         month_layout.append({"type": "delta_pct", "month": month, "label": f"{month}\nΔ, %", "col": col})
#         col += 1
#         month_layout.append({"type": "ytd_rub", "month": month, "label": f"{month}\nYTD Δ, руб", "col": col})
#         col += 1
#         month_layout.append({"type": "ytd_pct", "month": month, "label": f"{month}\nYTD Δ, %", "col": col})
#         col += 1
#         month_layout.append({"type": "spacer", "col": col})
#         col += 1

#     total_plan_col = col
#     col += 1
#     total_fact_col = col
#     col += 1
#     total_delta_rub_col = col
#     col += 1
#     total_delta_pct_col = col

#     widths = {"A": 54}

#     for item in month_layout:
#         col_letter = get_column_letter(item["col"])

#         if item["type"] == "spacer":
#             widths[col_letter] = 3
#         elif "pct" in item["type"]:
#             widths[col_letter] = 12
#         else:
#             widths[col_letter] = 16

#     widths[get_column_letter(total_plan_col)] = 16
#     widths[get_column_letter(total_fact_col)] = 16
#     widths[get_column_letter(total_delta_rub_col)] = 16
#     widths[get_column_letter(total_delta_pct_col)] = 12

#     set_column_widths(ws, widths)

#     set_row_heights(ws, {
#         1: 20,
#         2: 26,
#         3: 18,
#         4: 18,
#         5: 10,
#         6: 40,
#     })

#     draw_back_button(ws, cell="A1", text="← БЮДЖЕТ", target_sheet="БЮДЖЕТ")

#     draw_sheet_header(
#         ws,
#         title=f'РАСШИФРОВКА {detail["note"]}',
#         subtitle=detail["item"],
#         note=f'{detail["activity"]} | {detail["operation"]}',
#     )

#     ws["A6"] = "Статья"
#     ws["A6"].fill = FILLS["header"]
#     ws["A6"].font = FONTS["header_white"]
#     ws["A6"].alignment = ALIGNMENTS["center"]
#     ws["A6"].border = BORDERS["thin"]

#     for item in month_layout:
#         col_idx = item["col"]

#         if item["type"] == "spacer":
#             cell = ws.cell(row=6, column=col_idx, value=None)
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]
#         else:
#             cell = ws.cell(row=6, column=col_idx, value=item["label"])
#             cell.fill = FILLS["header"]
#             cell.font = FONTS["header_white"]
#             cell.alignment = ALIGNMENTS["center_wrap"]
#             cell.border = BORDERS["thin"]

#     for col_idx, title in [
#         (total_plan_col, "ИТОГО\nПлан"),
#         (total_fact_col, "ИТОГО\nФакт"),
#         (total_delta_rub_col, "ИТОГО\nΔ, руб"),
#         (total_delta_pct_col, "ИТОГО\nΔ, %"),
#     ]:
#         cell = ws.cell(row=6, column=col_idx, value=title)
#         cell.fill = FILLS["header"]
#         cell.font = FONTS["header_white"]
#         cell.alignment = ALIGNMENTS["center_wrap"]
#         cell.border = BORDERS["thin"]

#     row_idx = 7

#     for row in detail["rows"]:
#         _draw_detail_value_row(
#             ws=ws,
#             row_idx=row_idx,
#             row=row,
#             month_layout=month_layout,
#             months=months,
#             total_plan_col=total_plan_col,
#             total_fact_col=total_fact_col,
#             total_delta_rub_col=total_delta_rub_col,
#             total_delta_pct_col=total_delta_pct_col,
#             is_child=False,
#         )

#         row_idx += 1

#         for child in row.get("children", []):
#             _draw_detail_value_row(
#                 ws=ws,
#                 row_idx=row_idx,
#                 row=child,
#                 month_layout=month_layout,
#                 months=months,
#                 total_plan_col=total_plan_col,
#                 total_fact_col=total_fact_col,
#                 total_delta_rub_col=total_delta_rub_col,
#                 total_delta_pct_col=total_delta_pct_col,
#                 is_child=True,
#             )

#             ws.row_dimensions[row_idx].hidden = True
#             ws.row_dimensions[row_idx].outline_level = 1
#             row_idx += 1

#     total_label = ws.cell(row=row_idx, column=1, value="ИТОГО")
#     total_label.fill = FILLS["total"]
#     total_label.font = FONTS["total"]
#     total_label.alignment = ALIGNMENTS["left"]
#     total_label.border = BORDERS["bottom_medium"]

#     running_ytd_rub = 0
#     running_ytd_plan = 0
#     ytd_rub_by_month = {}
#     ytd_pct_by_month = {}

#     for month in months:
#         plan_val = detail["total_plan_by_month"].get(month, 0)
#         fact_val = detail["total_fact_by_month"].get(month, 0)

#         running_ytd_rub += fact_val - plan_val
#         running_ytd_plan += plan_val

#         ytd_rub_by_month[month] = running_ytd_rub

#         if running_ytd_plan != 0:
#             ytd_pct_by_month[month] = running_ytd_rub / abs(running_ytd_plan)
#         else:
#             ytd_pct_by_month[month] = None if running_ytd_rub == 0 else (1 if running_ytd_rub > 0 else -1)

#     for layout_item in month_layout:
#         col_idx = layout_item["col"]

#         if layout_item["type"] == "spacer":
#             cell = ws.cell(row=row_idx, column=col_idx, value=None)
#             cell.fill = FILLS["none"]
#             cell.border = BORDERS["none"]
#             continue

#         month = layout_item["month"]

#         if layout_item["type"] == "plan":
#             value = detail["total_plan_by_month"].get(month, 0)
#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = FILLS["total"]
#             cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["bottom_medium"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "fact":
#             value = detail["total_fact_by_month"].get(month, 0)
#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = FILLS["total"]
#             cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["bottom_medium"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "delta_rub":
#             plan_val = detail["total_plan_by_month"].get(month, 0)
#             value = detail["total_delta_by_month"].get(month, 0)
#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = _get_delta_fill(value, FILLS["total"], plan_val)
#             cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["bottom_medium"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "delta_pct":
#             plan_val = detail["total_plan_by_month"].get(month, 0)
#             delta_rub = detail["total_delta_by_month"].get(month, 0)

#             if plan_val != 0:
#                 delta_pct = delta_rub / abs(plan_val)
#             else:
#                 delta_pct = None if delta_rub == 0 else (1 if delta_rub > 0 else -1)

#             cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
#             cell.fill = _get_delta_fill(delta_rub, FILLS["total"], plan_val)
#             cell.font = FONTS["total"]
#             cell.alignment = ALIGNMENTS["center"]
#             cell.border = BORDERS["bottom_medium"]

#         elif layout_item["type"] == "ytd_rub":
#             plan_val = detail["total_plan_by_month"].get(month, 0)
#             value = ytd_rub_by_month.get(month, 0)

#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.fill = _get_delta_fill(value, FILLS["total"], plan_val)
#             cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
#             cell.alignment = ALIGNMENTS["right"]
#             cell.border = BORDERS["bottom_medium"]
#             cell.number_format = FORMATS["money_int"]

#         elif layout_item["type"] == "ytd_pct":
#             plan_val = detail["total_plan_by_month"].get(month, 0)
#             delta_rub = ytd_rub_by_month.get(month, 0)
#             delta_pct = ytd_pct_by_month.get(month)

#             cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
#             cell.fill = _get_delta_fill(delta_rub, FILLS["total"], plan_val)
#             cell.font = FONTS["total"]
#             cell.alignment = ALIGNMENTS["center"]
#             cell.border = BORDERS["bottom_medium"]

#     total_plan_sum = detail["plan_total"]
#     total_fact_sum = detail["fact_total"]
#     total_delta_sum = detail["delta_total"]

#     total_plan_sum_cell = ws.cell(row=row_idx, column=total_plan_col, value=total_plan_sum)
#     total_plan_sum_cell.fill = FILLS["total"]
#     total_plan_sum_cell.font = FONTS["negative_total"] if total_plan_sum < 0 else FONTS["total"]
#     total_plan_sum_cell.alignment = ALIGNMENTS["right"]
#     total_plan_sum_cell.border = BORDERS["bottom_medium"]
#     total_plan_sum_cell.number_format = FORMATS["money_int"]

#     total_fact_sum_cell = ws.cell(row=row_idx, column=total_fact_col, value=total_fact_sum)
#     total_fact_sum_cell.fill = FILLS["total"]
#     total_fact_sum_cell.font = FONTS["negative_total"] if total_fact_sum < 0 else FONTS["total"]
#     total_fact_sum_cell.alignment = ALIGNMENTS["right"]
#     total_fact_sum_cell.border = BORDERS["bottom_medium"]
#     total_fact_sum_cell.number_format = FORMATS["money_int"]

#     total_delta_sum_cell = ws.cell(row=row_idx, column=total_delta_rub_col, value=total_delta_sum)
#     total_delta_sum_cell.fill = _get_delta_fill(total_delta_sum, FILLS["total"], total_plan_sum)
#     total_delta_sum_cell.font = FONTS["negative_total"] if total_delta_sum < 0 else FONTS["total"]
#     total_delta_sum_cell.alignment = ALIGNMENTS["right"]
#     total_delta_sum_cell.border = BORDERS["bottom_medium"]
#     total_delta_sum_cell.number_format = FORMATS["money_int"]

#     if total_plan_sum != 0:
#         total_delta_pct_sum = total_delta_sum / abs(total_plan_sum)
#     else:
#         total_delta_pct_sum = None if total_delta_sum == 0 else (1 if total_delta_sum > 0 else -1)

#     total_delta_pct_cell = ws.cell(
#         row=row_idx,
#         column=total_delta_pct_col,
#         value=_format_percent(total_delta_pct_sum),
#     )
#     total_delta_pct_cell.fill = _get_delta_fill(total_delta_sum, FILLS["total"], total_plan_sum)
#     total_delta_pct_cell.font = FONTS["total"]
#     total_delta_pct_cell.alignment = ALIGNMENTS["center"]
#     total_delta_pct_cell.border = BORDERS["bottom_medium"]

#     _hide_fact_and_delta_columns(
#         ws,
#         month_layout,
#         total_fact_col,
       
#     )

#     ws.freeze_panes = "B7"





# budget/reporting/excel/sheets/gl_detail_sheet.py

from copy import copy
from openpyxl.utils import get_column_letter

from budget.reporting.excel.styles.helpers import (
    set_column_widths,
    set_row_heights,
    draw_back_button,
    draw_sheet_header,
)
from budget.reporting.excel.styles.theme import (
    FILLS,
    FONTS,
    BORDERS,
    ALIGNMENTS,
    FORMATS,
)


def _get_delta_fill(delta_rub, default_fill, plan_value=0):
    try:
        delta_rub = float(delta_rub or 0)
        plan_value = float(plan_value or 0)
    except (TypeError, ValueError):
        return default_fill

    if delta_rub == 0:
        return default_fill

    if plan_value == 0:
        return FILLS["delta_green"]

    return FILLS["delta_green"] if delta_rub > 0 else FILLS["delta_red"]


def _format_percent(value_pct):
    if value_pct is None:
        return None

    return f"+{value_pct * 100:.1f}%" if value_pct >= 0 else f"{value_pct * 100:.1f}%"


def _calculate_ytd_for_row(plan_months, fact_months, months):
    ytd_plan = {}
    ytd_fact = {}
    ytd_rub = {}
    ytd_pct = {}

    running_plan = 0
    running_fact = 0

    for month in months:
        plan_val = plan_months.get(month, 0)
        fact_val = fact_months.get(month, 0)

        running_plan += plan_val
        running_fact += fact_val
        running_rub = running_fact - running_plan

        ytd_plan[month] = running_plan
        ytd_fact[month] = running_fact
        ytd_rub[month] = running_rub

        if running_plan != 0:
            ytd_pct[month] = running_rub / abs(running_plan)
        else:
            ytd_pct[month] = None if running_rub == 0 else (1 if running_rub > 0 else -1)

    return ytd_plan, ytd_fact, ytd_rub, ytd_pct


def _hide_fact_and_delta_columns(ws, month_layout, total_fact_col):
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = True
    ws.sheet_view.showOutlineSymbols = True

    hidden_types = (
        "fact",
        "delta_rub",
        "delta_pct",
        "ytd_plan",
        "ytd_fact",
    )

    for item in month_layout:
        if item["type"] in hidden_types:
            col_letter = get_column_letter(item["col"])
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].outline_level = 1

    col_letter = get_column_letter(total_fact_col)
    ws.column_dimensions[col_letter].hidden = True
    ws.column_dimensions[col_letter].outline_level = 1


def _get_row_font(value, is_child=False):
    if is_child:
        return FONTS["counterparty_negative"] if value < 0 else FONTS["counterparty"]

    return FONTS["negative"] if value < 0 else FONTS["normal"]


def _get_label_alignment(is_child=False):
    alignment = copy(ALIGNMENTS["left"])
    if is_child:
        alignment.indent = 2
    return alignment


def _draw_detail_value_row(
    ws,
    row_idx,
    row,
    month_layout,
    months,
    total_plan_col,
    total_fact_col,
    total_delta_rub_col,
    total_delta_pct_col,
    is_child=False,
):
    label_value = row.get("label", "—")

    label_cell = ws.cell(row=row_idx, column=1, value=label_value)
    label_cell.fill = FILLS["none"]
    label_cell.border = BORDERS["thin"]
    label_cell.font = FONTS["counterparty"] if is_child else FONTS["normal"]
    label_cell.alignment = _get_label_alignment(is_child=is_child)

    ytd_plan, ytd_fact, ytd_rub, ytd_pct = _calculate_ytd_for_row(
        row["plan_months"],
        row["fact_months"],
        months,
    )

    for layout_item in month_layout:
        col_idx = layout_item["col"]

        if layout_item["type"] == "spacer":
            cell = ws.cell(row=row_idx, column=col_idx, value=None)
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]
            continue

        month = layout_item["month"]

        if layout_item["type"] == "plan":
            value = row["plan_months"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["none"]
            cell.font = _get_row_font(value, is_child=is_child)
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["thin"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "fact":
            value = row["fact_months"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["none"]
            cell.font = _get_row_font(value, is_child=is_child)
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["thin"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "delta_rub":
            plan_val = row["plan_months"].get(month, 0)
            value = row["delta_months"].get(month, 0)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = _get_delta_fill(value, FILLS["none"], plan_val)
            cell.font = _get_row_font(value, is_child=is_child)
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["thin"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "delta_pct":
            if is_child:
                cell = ws.cell(row=row_idx, column=col_idx, value=None)
                cell.fill = FILLS["none"]
                cell.border = BORDERS["none"]
                continue

            plan_val = row["plan_months"].get(month, 0)
            delta_rub = row["delta_months"].get(month, 0)

            if plan_val != 0:
                delta_pct = delta_rub / abs(plan_val)
            else:
                delta_pct = None if delta_rub == 0 else (1 if delta_rub > 0 else -1)

            cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
            cell.fill = _get_delta_fill(delta_rub, FILLS["none"], plan_val)
            cell.font = _get_row_font(delta_rub, is_child=is_child)
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]

        elif layout_item["type"] == "ytd_plan":
            value = ytd_plan.get(month, 0)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["none"]
            cell.font = _get_row_font(value, is_child=is_child)
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["thin"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "ytd_fact":
            value = ytd_fact.get(month, 0)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["none"]
            cell.font = _get_row_font(value, is_child=is_child)
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["thin"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "ytd_rub":
            plan_val = ytd_plan.get(month, 0)
            value = ytd_rub.get(month, 0)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = _get_delta_fill(value, FILLS["none"], plan_val)
            cell.font = _get_row_font(value, is_child=is_child)
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["thin"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "ytd_pct":
            if is_child:
                cell = ws.cell(row=row_idx, column=col_idx, value=None)
                cell.fill = FILLS["none"]
                cell.border = BORDERS["none"]
                continue

            plan_val = ytd_plan.get(month, 0)
            delta_pct = ytd_pct.get(month)
            delta_rub = ytd_rub.get(month, 0)

            cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
            cell.fill = _get_delta_fill(delta_rub, FILLS["none"], plan_val)
            cell.font = _get_row_font(delta_rub, is_child=is_child)
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["thin"]

    total_plan_value = row["plan_total"]
    total_fact_value = row["fact_total"]
    total_delta_value = row["delta_total"]

    total_plan_cell = ws.cell(row=row_idx, column=total_plan_col, value=total_plan_value)
    total_plan_cell.fill = FILLS["none"]
    total_plan_cell.font = _get_row_font(total_plan_value, is_child=is_child)
    total_plan_cell.alignment = ALIGNMENTS["right"]
    total_plan_cell.border = BORDERS["thin"]
    total_plan_cell.number_format = FORMATS["money_int"]

    total_fact_cell = ws.cell(row=row_idx, column=total_fact_col, value=total_fact_value)
    total_fact_cell.fill = FILLS["none"]
    total_fact_cell.font = _get_row_font(total_fact_value, is_child=is_child)
    total_fact_cell.alignment = ALIGNMENTS["right"]
    total_fact_cell.border = BORDERS["thin"]
    total_fact_cell.number_format = FORMATS["money_int"]

    total_delta_cell = ws.cell(row=row_idx, column=total_delta_rub_col, value=total_delta_value)
    total_delta_cell.fill = _get_delta_fill(total_delta_value, FILLS["none"], total_plan_value)
    total_delta_cell.font = _get_row_font(total_delta_value, is_child=is_child)
    total_delta_cell.alignment = ALIGNMENTS["right"]
    total_delta_cell.border = BORDERS["thin"]
    total_delta_cell.number_format = FORMATS["money_int"]

    if not is_child:
        if total_plan_value != 0:
            total_delta_pct_value = total_delta_value / abs(total_plan_value)
        else:
            total_delta_pct_value = None if total_delta_value == 0 else (1 if total_delta_value > 0 else -1)

        total_delta_pct_cell = ws.cell(
            row=row_idx,
            column=total_delta_pct_col,
            value=_format_percent(total_delta_pct_value),
        )
        total_delta_pct_cell.fill = _get_delta_fill(total_delta_value, FILLS["none"], total_plan_value)
        total_delta_pct_cell.font = _get_row_font(total_delta_value, is_child=is_child)
        total_delta_pct_cell.alignment = ALIGNMENTS["center"]
        total_delta_pct_cell.border = BORDERS["thin"]
    else:
        empty_cell = ws.cell(row=row_idx, column=total_delta_pct_col, value=None)
        empty_cell.fill = FILLS["none"]
        empty_cell.border = BORDERS["none"]

    ws.row_dimensions[row_idx].height = 18 if not is_child else 17


def build_gl_detail_sheet(wb, detail):
    ws = wb.create_sheet(detail["sheet_name"])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showOutlineSymbols = True

    months = detail["months"]

    col = 2
    month_layout = []

    for month in months:
        month_layout.append({"type": "plan", "month": month, "label": f"{month}\nПлан", "col": col})
        col += 1

        month_layout.append({"type": "fact", "month": month, "label": f"{month}\nФакт", "col": col})
        col += 1

        month_layout.append({"type": "delta_rub", "month": month, "label": f"{month}\nΔ, руб", "col": col})
        col += 1

        month_layout.append({"type": "delta_pct", "month": month, "label": f"{month}\nΔ, %", "col": col})
        col += 1

        month_layout.append({"type": "ytd_plan", "month": month, "label": f"{month}\nYTD План", "col": col})
        col += 1

        month_layout.append({"type": "ytd_fact", "month": month, "label": f"{month}\nYTD Факт", "col": col})
        col += 1

        month_layout.append({"type": "ytd_rub", "month": month, "label": f"{month}\nYTD Δ, руб", "col": col})
        col += 1

        month_layout.append({"type": "ytd_pct", "month": month, "label": f"{month}\nYTD Δ, %", "col": col})
        col += 1

        month_layout.append({"type": "spacer", "col": col})
        col += 1

    total_plan_col = col
    col += 1
    total_fact_col = col
    col += 1
    total_delta_rub_col = col
    col += 1
    total_delta_pct_col = col

    widths = {"A": 54}

    for item in month_layout:
        col_letter = get_column_letter(item["col"])

        if item["type"] == "spacer":
            widths[col_letter] = 3
        elif "pct" in item["type"]:
            widths[col_letter] = 12
        else:
            widths[col_letter] = 16

    widths[get_column_letter(total_plan_col)] = 16
    widths[get_column_letter(total_fact_col)] = 16
    widths[get_column_letter(total_delta_rub_col)] = 16
    widths[get_column_letter(total_delta_pct_col)] = 12

    set_column_widths(ws, widths)

    set_row_heights(ws, {
        1: 20,
        2: 26,
        3: 18,
        4: 18,
        5: 10,
        6: 40,
    })

    draw_back_button(ws, cell="A1", text="← БЮДЖЕТ", target_sheet="БЮДЖЕТ")

    draw_sheet_header(
        ws,
        title=f'РАСШИФРОВКА {detail["note"]}',
        subtitle=detail["item"],
        note=f'{detail["activity"]} | {detail["operation"]}',
    )

    ws["A6"] = "Статья"
    ws["A6"].fill = FILLS["header"]
    ws["A6"].font = FONTS["header_white"]
    ws["A6"].alignment = ALIGNMENTS["center"]
    ws["A6"].border = BORDERS["thin"]

    for item in month_layout:
        col_idx = item["col"]

        if item["type"] == "spacer":
            cell = ws.cell(row=6, column=col_idx, value=None)
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]
        else:
            cell = ws.cell(row=6, column=col_idx, value=item["label"])
            cell.fill = FILLS["header"]
            cell.font = FONTS["header_white"]
            cell.alignment = ALIGNMENTS["center_wrap"]
            cell.border = BORDERS["thin"]

    for col_idx, title in [
        (total_plan_col, "ИТОГО\nПлан"),
        (total_fact_col, "ИТОГО\nФакт"),
        (total_delta_rub_col, "ИТОГО\nΔ, руб"),
        (total_delta_pct_col, "ИТОГО\nΔ, %"),
    ]:
        cell = ws.cell(row=6, column=col_idx, value=title)
        cell.fill = FILLS["header"]
        cell.font = FONTS["header_white"]
        cell.alignment = ALIGNMENTS["center_wrap"]
        cell.border = BORDERS["thin"]

    row_idx = 7

    for row in detail["rows"]:
        _draw_detail_value_row(
            ws=ws,
            row_idx=row_idx,
            row=row,
            month_layout=month_layout,
            months=months,
            total_plan_col=total_plan_col,
            total_fact_col=total_fact_col,
            total_delta_rub_col=total_delta_rub_col,
            total_delta_pct_col=total_delta_pct_col,
            is_child=False,
        )

        row_idx += 1

        for child in row.get("children", []):
            _draw_detail_value_row(
                ws=ws,
                row_idx=row_idx,
                row=child,
                month_layout=month_layout,
                months=months,
                total_plan_col=total_plan_col,
                total_fact_col=total_fact_col,
                total_delta_rub_col=total_delta_rub_col,
                total_delta_pct_col=total_delta_pct_col,
                is_child=True,
            )

            ws.row_dimensions[row_idx].hidden = True
            ws.row_dimensions[row_idx].outline_level = 1
            row_idx += 1

    total_label = ws.cell(row=row_idx, column=1, value="ИТОГО")
    total_label.fill = FILLS["total"]
    total_label.font = FONTS["total"]
    total_label.alignment = ALIGNMENTS["left"]
    total_label.border = BORDERS["bottom_medium"]

    running_ytd_plan = 0
    running_ytd_fact = 0
    ytd_plan_by_month = {}
    ytd_fact_by_month = {}
    ytd_rub_by_month = {}
    ytd_pct_by_month = {}

    for month in months:
        plan_val = detail["total_plan_by_month"].get(month, 0)
        fact_val = detail["total_fact_by_month"].get(month, 0)

        running_ytd_plan += plan_val
        running_ytd_fact += fact_val
        running_ytd_rub = running_ytd_fact - running_ytd_plan

        ytd_plan_by_month[month] = running_ytd_plan
        ytd_fact_by_month[month] = running_ytd_fact
        ytd_rub_by_month[month] = running_ytd_rub

        if running_ytd_plan != 0:
            ytd_pct_by_month[month] = running_ytd_rub / abs(running_ytd_plan)
        else:
            ytd_pct_by_month[month] = None if running_ytd_rub == 0 else (1 if running_ytd_rub > 0 else -1)

    for layout_item in month_layout:
        col_idx = layout_item["col"]

        if layout_item["type"] == "spacer":
            cell = ws.cell(row=row_idx, column=col_idx, value=None)
            cell.fill = FILLS["none"]
            cell.border = BORDERS["none"]
            continue

        month = layout_item["month"]

        if layout_item["type"] == "plan":
            value = detail["total_plan_by_month"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["total"]
            cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["bottom_medium"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "fact":
            value = detail["total_fact_by_month"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["total"]
            cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["bottom_medium"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "delta_rub":
            plan_val = detail["total_plan_by_month"].get(month, 0)
            value = detail["total_delta_by_month"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = _get_delta_fill(value, FILLS["total"], plan_val)
            cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["bottom_medium"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "delta_pct":
            plan_val = detail["total_plan_by_month"].get(month, 0)
            delta_rub = detail["total_delta_by_month"].get(month, 0)

            if plan_val != 0:
                delta_pct = delta_rub / abs(plan_val)
            else:
                delta_pct = None if delta_rub == 0 else (1 if delta_rub > 0 else -1)

            cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
            cell.fill = _get_delta_fill(delta_rub, FILLS["total"], plan_val)
            cell.font = FONTS["total"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["bottom_medium"]

        elif layout_item["type"] == "ytd_plan":
            value = ytd_plan_by_month.get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["total"]
            cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["bottom_medium"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "ytd_fact":
            value = ytd_fact_by_month.get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = FILLS["total"]
            cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["bottom_medium"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "ytd_rub":
            plan_val = ytd_plan_by_month.get(month, 0)
            value = ytd_rub_by_month.get(month, 0)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = _get_delta_fill(value, FILLS["total"], plan_val)
            cell.font = FONTS["negative_total"] if value < 0 else FONTS["total"]
            cell.alignment = ALIGNMENTS["right"]
            cell.border = BORDERS["bottom_medium"]
            cell.number_format = FORMATS["money_int"]

        elif layout_item["type"] == "ytd_pct":
            plan_val = ytd_plan_by_month.get(month, 0)
            delta_rub = ytd_rub_by_month.get(month, 0)
            delta_pct = ytd_pct_by_month.get(month)

            cell = ws.cell(row=row_idx, column=col_idx, value=_format_percent(delta_pct))
            cell.fill = _get_delta_fill(delta_rub, FILLS["total"], plan_val)
            cell.font = FONTS["total"]
            cell.alignment = ALIGNMENTS["center"]
            cell.border = BORDERS["bottom_medium"]

    total_plan_sum = detail["plan_total"]
    total_fact_sum = detail["fact_total"]
    total_delta_sum = detail["delta_total"]

    total_plan_sum_cell = ws.cell(row=row_idx, column=total_plan_col, value=total_plan_sum)
    total_plan_sum_cell.fill = FILLS["total"]
    total_plan_sum_cell.font = FONTS["negative_total"] if total_plan_sum < 0 else FONTS["total"]
    total_plan_sum_cell.alignment = ALIGNMENTS["right"]
    total_plan_sum_cell.border = BORDERS["bottom_medium"]
    total_plan_sum_cell.number_format = FORMATS["money_int"]

    total_fact_sum_cell = ws.cell(row=row_idx, column=total_fact_col, value=total_fact_sum)
    total_fact_sum_cell.fill = FILLS["total"]
    total_fact_sum_cell.font = FONTS["negative_total"] if total_fact_sum < 0 else FONTS["total"]
    total_fact_sum_cell.alignment = ALIGNMENTS["right"]
    total_fact_sum_cell.border = BORDERS["bottom_medium"]
    total_fact_sum_cell.number_format = FORMATS["money_int"]

    total_delta_sum_cell = ws.cell(row=row_idx, column=total_delta_rub_col, value=total_delta_sum)
    total_delta_sum_cell.fill = _get_delta_fill(total_delta_sum, FILLS["total"], total_plan_sum)
    total_delta_sum_cell.font = FONTS["negative_total"] if total_delta_sum < 0 else FONTS["total"]
    total_delta_sum_cell.alignment = ALIGNMENTS["right"]
    total_delta_sum_cell.border = BORDERS["bottom_medium"]
    total_delta_sum_cell.number_format = FORMATS["money_int"]

    if total_plan_sum != 0:
        total_delta_pct_sum = total_delta_sum / abs(total_plan_sum)
    else:
        total_delta_pct_sum = None if total_delta_sum == 0 else (1 if total_delta_sum > 0 else -1)

    total_delta_pct_cell = ws.cell(
        row=row_idx,
        column=total_delta_pct_col,
        value=_format_percent(total_delta_pct_sum),
    )
    total_delta_pct_cell.fill = _get_delta_fill(total_delta_sum, FILLS["total"], total_plan_sum)
    total_delta_pct_cell.font = FONTS["total"]
    total_delta_pct_cell.alignment = ALIGNMENTS["center"]
    total_delta_pct_cell.border = BORDERS["bottom_medium"]

    _hide_fact_and_delta_columns(
        ws,
        month_layout,
        total_fact_col,
    )

    ws.freeze_panes = "B7"